from app import app
import json
import uuid
import os
from flask import render_template, request, flash, redirect, url_for, session,  jsonify, make_response, send_file
from datetime import datetime, timedelta,date
from dateutil.relativedelta import relativedelta
import openpyxl
import pdfplumber
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, Spacer, Image,SimpleDocTemplate
from io import BytesIO
from werkzeug.security import check_password_hash
import csv 
import pandas as pd
from io import StringIO 
from decimal import Decimal 

from openpyxl import Workbook

from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter

import re
import traceback
# Para encriptar contraseÃ±a generate_password_hash
from werkzeug.security import generate_password_hash


# Importando cenexiÃ³n a BD
from controllers.funciones_home import *
from controllers.funciones_home import sql_lista_polizas, obtener_bloques_comision, obtener_detalles_bloque, sql_lista_polizas_asegurado

from pyDolarVenezuela.pages import AlCambio, BCV, CriptoDolar, DolarToday, ExchangeMonitor, EnParaleloVzla, Italcambio
from pyDolarVenezuela import Monitor


TEMP_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'temp_data')
if not os.path.exists(TEMP_DIR):
    os.makedirs(TEMP_DIR)

def save_temp_data(data):
    filename = f"{uuid.uuid4()}.json"
    filepath = os.path.join(TEMP_DIR, filename)
    with open(filepath, 'w') as f:
        # Usamos default=str para serializar fechas y decimales que JSON nativo no soporta
        json.dump(data, f, default=str) 
    return filename

def load_temp_data(filename):
    if not filename: return {}
    filepath = os.path.join(TEMP_DIR, filename)
    if os.path.exists(filepath):
        with open(filepath, 'r') as f:
            return json.load(f)
    return {}

PATH_URL = "public/empleados"

@app.template_filter('tofloat')
def to_float(value):
    try:
        return float(value)
    except (TypeError, ValueError):
        return value

app.jinja_env.filters['tofloat'] = to_float

def add_year(date, years):
    try:
        date = datetime.datetime.strptime(date, '%Y-%m-%d')  # Accede a strptime a travÃ©s del objeto datetime
        return date + timedelta(days=365*years)
    except ValueError:
        return date  # Manejar errores si la fecha no tiene el formato esperado

app.jinja_env.filters['add_year'] = add_year

def strftime_filter(date, format):
    return date.strftime(format)

app.jinja_env.filters['strftime'] = strftime_filter
 
@app.template_filter('format_date')
def format_date_filter(date_str):
    if not date_str:
        return ""
    try:
        # Handle both isoformat strings and date objects
        if isinstance(date_str, str):
            dt = datetime.datetime.fromisoformat(date_str)
        else:
            dt = date_str
        return dt.strftime('%d/%m/%y')
    except:
        return date_str



@app.route('/registrar-empleado', methods=['GET'])
def viewFormEmpleado():
    if 'conectado' in session:
        return render_template(f'{PATH_URL}/form_empleado.html')
    else:
        flash('primero debes iniciar sesiÃ³n.', 'error')
        return redirect(url_for('inicio'))
    
@app.route('/inicio1', methods=['GET'])
def inicio1():
    if 'conectado' in session:
        return render_template('public/inicio.html')
    else:
        return render_template('public/login/base_login.html')
    



@app.route('/form-registrar-empleado', methods=['POST'])
def formEmpleado():
    if 'conectado' in session:
        if 'foto_empleado' in request.files:
            foto_perfil = request.files['foto_empleado']
            resultado = procesar_form_empleado(request.form, foto_perfil)
            if resultado:
                return redirect(url_for('lista_empleados'))
            else:
                flash('El empleado NO fue registrado.', 'error')
                return render_template(f'{PATH_URL}/form_empleado.html')
    else:
        flash('primero debes iniciar sesiÃ³n.', 'error')
        return redirect(url_for('inicio'))


@app.route('/lista-de-empleados', methods=['GET'])
def lista_empleados():
    if 'conectado' in session:
        return render_template(f'{PATH_URL}/lista_empleados.html', empleados=sql_lista_empleadosBD())
    else:
        flash('primero debes iniciar sesiÃ³n.', 'error')
        return redirect(url_for('inicio'))
    
@app.route('/lista-de-asegurado', methods=['GET'])
def lista_asegurado():
    if 'conectado' in session:
        return render_template('public/Asegurados/lista_asegurados.html', asegurados=sql_lista_aseguradosBD())
    else:
        flash('primero debes iniciar sesiÃ³n.', 'error')
        return redirect(url_for('inicio'))
    
@app.route('/lista-de-polizas', methods=['GET'])
def lista_polizas():
    if 'conectado' in session:
        today = datetime.date.today()
        polizas = sql_lista_polizas() or []
        ejecutivos = sql_lista_ejecutivo() or []
        companias = sql_lista_company() or []

        return render_template('public/Poliza/lista_polizas.html', 
                               polizas=[], today=today, 
                               ejecutivos=ejecutivos, companias=companias)
    else:
        flash('primero debes iniciar sesiÃ³n.', 'error')
        return redirect(url_for('inicio'))

@app.route('/descargar-excel-polizas', methods=['GET'])
def descargar_excel_polizas():
    if 'conectado' in session:
        tipo_filtro_fecha = request.args.get('tipo_filtro_fecha')
        fecha = request.args.get('fecha')
        anio = request.args.get('anio')
        fecha_inicio = request.args.get('fecha_inicio')
        fecha_fin = request.args.get('fecha_fin')
        
        # En jQuery $.param(), los arrays se envÃ­an como 'estados[]'
        estados = request.args.getlist('estados[]')
        compania_id = request.args.get('compania')
        ejecutivo_id = request.args.get('ejecutivo')
        
        from controllers.funciones_home import obtener_polizas_datatable
        
        result = obtener_polizas_datatable(
            start=0, length=999999, 
            tipo_filtro_fecha=tipo_filtro_fecha, 
            fecha=fecha, 
            anio=anio, 
            fecha_inicio=fecha_inicio, 
            fecha_fin=fecha_fin, 
            estados=estados, 
            compania_id=compania_id, 
            ejecutivo_id=ejecutivo_id,
            search_value=None
        )
        polizas = result.get('data', [])
        
        report_subtitle = "Reporte de PÃ³lizas Filtrado"
        
        # Crear un libro de trabajo de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de PÃ³lizas"

        # TÃ­tulo
        ws.merge_cells('A1:J1')
        title_cell = ws['A1']
        title_cell.value = "Reporte de PÃ³lizas"
        title_cell.font = Font(bold=True, size=18, color="FFFFFF")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        ws.row_dimensions[1].height = 30

        # SubtÃ­tulo
        ws.merge_cells('A2:J2')
        subtitle_cell = ws['A2']
        subtitle_cell.value = report_subtitle
        subtitle_cell.font = Font(bold=True, size=12)
        subtitle_cell.alignment = Alignment(horizontal='center')
        ws.row_dimensions[2].height = 20
        
        # Encabezados de la tabla (Fila 4)
        headers = ["CÃ³digo de PÃ³liza", "CI Asegurado", "Asegurado", "Referidor", "CompaÃ±Ã­a", "Producto", "Ramo", "Fecha de Contrato", "Riesgo", "Estado"]
        ws.append(headers)
        
        # Estilo para los encabezados
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        for cell in ws[4]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(bottom=Side(style='thin'))

        for poliza in polizas:
            estado = poliza.get('estado_calculado', '')
            
            riesgo_val = poliza.get('riesgo')
            riesgo_str = ""
            if riesgo_val is not None:
                try:
                    riesgo_float = float(riesgo_val)
                    if 1.00 <= riesgo_float <= 1.60:
                        riesgo_str = f"Bajo ({riesgo_float})"
                    elif 1.61 <= riesgo_float <= 2.20:
                        riesgo_str = f"Medio ({riesgo_float})"
                    elif 2.21 <= riesgo_float <= 3.00:
                        riesgo_str = f"Alto ({riesgo_float})"
                    else:
                        riesgo_str = f"Fuera de rango ({riesgo_float})"
                except (ValueError, TypeError):
                    riesgo_str = "Dato invÃ¡lido"
            else:
                riesgo_str = "No definido"

            referidor_nombre = f"{poliza.get('ejecutivo_nombre', '')} {poliza.get('ejecutivo_apellido', '')}".strip() or 'No asignado'
            ws.append([
                poliza.get('cod_poliza', ''),
                poliza.get('CI_asegurado', ''),
                f"{poliza.get('Nombre', '')} {poliza.get('Apellido', '')}",
                referidor_nombre,
                poliza.get('compania_nombre', ''),
                poliza.get('producto_nombre', ''),
                poliza.get('Ramo', ''),
                poliza.get('fecha', ''),
                riesgo_str,
                estado
            ])

    
        for i, col in enumerate(ws.columns):
            max_length = 0
            
            column_letter = ws.cell(row=4, column=i + 1).column_letter
            
            for cell in col:
                try:
                    
                    value_length = len(str(cell.value))
                    if value_length > max_length:
                        max_length = value_length
                except:
                   
                    pass
            
           
            adjusted_width = max(max_length + 2, 10) 
            ws.column_dimensions[column_letter].width = adjusted_width

       
       
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
  
        response = make_response(buffer.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = 'attachment; filename=reporte_polizas.xlsx'
        return response

    else:
        flash('Primero debes iniciar sesiÃ³n.', 'error')
        return redirect(url_for('inicio'))
    

    
@app.route('/lista-de-siniestros', methods=['GET'])
def lista_siniestros():
    if 'conectado' in session:
        return render_template('public/Siniestro/lista_Siniestros.html', siniestros=sql_lista_siniestros())
    else:
        flash('primero debes iniciar sesiÃ³n.', 'error')
        return redirect(url_for('inicio'))


@app.route("/detalles-empleado/", methods=['GET'])
@app.route("/detalles-empleado/<int:idEmpleado>", methods=['GET'])
def detalleEmpleado(idEmpleado=None):
    if 'conectado' in session:
        # Verificamos si el parÃ¡metro idEmpleado es None o no estÃ¡ presente en la URL
        if idEmpleado is None:
            return redirect(url_for('inicio'))
        else:
            detalle_empleado = sql_detalles_empleadosBD(idEmpleado) or []
            return render_template(f'{PATH_URL}/detalles_empleado.html', detalle_empleado=detalle_empleado)
    else:
        flash('Primero debes iniciar sesiÃ³n.', 'error')
        return redirect(url_for('inicio'))

@app.route("/detalles-asegurado/", methods=['GET'])
@app.route("/detalles-asegurado/<int:CI>", methods=['GET'])
def detalleAsegurado(CI=None):
    if 'conectado' in session:
        # Verificamos si el parÃ¡metro idEmpleado es None o no estÃ¡ presente en la URL
        if CI is None:
            return redirect(url_for('inicio'))
        else:
            detalle_empleado = sql_detalles_aseguradoBD(CI) or []
            polizas_asegurado = sql_lista_polizas_asegurado(CI) or []
            return render_template('public/Asegurados/detalles_Asegurado.html', detalle_Asegurado=detalle_empleado, polizas=polizas_asegurado)
    else:
        flash('Primero debes iniciar sesiÃ³n.', 'error')
        return redirect(url_for('inicio'))
    

@app.route("/beneficiario_crear/", methods=['GET'])
@app.route("/beneficiario_crear/<path:cod_poliza>", methods=['GET'])
def crear_beneficiario(cod_poliza=None):
    if 'conectado' in session:
        if session.get("permisos")!="Ventas" or session.get("permisos")!="Gerencia":
            # Verificamos si el parÃ¡metro idEmpleado es None o no estÃ¡ presente en la URL
            if cod_poliza is None: 
                return redirect(url_for('inicio1'))
            else:
                return render_template('public/Poliza/crear_beneficiario.html', cod_poliza=cod_poliza)
        else:
            flash('No tienes permiso para esta acciÃƒÂ²n', 'error')
            return redirect(url_for('inicio1'))   
    else:
        flash('Primero debes iniciar sesiÃ³n.', 'error')
        return redirect(url_for('inicio'))
    
@app.route("/beneficiario_borrar/<path:cod_poliza>", methods=['GET'])
def borrar_beneficiario(cod_poliza=None):
    if 'conectado' in session:
        if session.get("permisos")=="Administracion" or session.get("permisos")=="Operaciones":
            # Verificamos si el parÃ¡metro idEmpleado es None o no estÃ¡ presente en la URL
            if cod_poliza is None:
                return redirect(url_for('inicio1'))
            else:
                eliminarBeneficiario(cod_poliza)
                detalle_poliza = sql_detalles_polizaBD(cod_poliza) or []
                return render_template('public/Poliza/detalles_polizaPersona.html',detalle_poliza=detalle_poliza)
        else:
           flash('No tienes permiso para esta acciÃƒÂ²n', 'error')
           return redirect(url_for('inicio1')) 
    else:
        flash('Primero debes iniciar sesiÃ³n.', 'error')
        return redirect(url_for('inicio1'))

@app.route("/CartaAval/<path:cod_poliza>", methods=['GET'])
def view_form_cartaAval(cod_poliza=None):
    if 'conectado' in session:
        if session.get("permisos")!="Ventas" or session.get("permisos")!="Gerencia":
            # Verificamos si el parÃ¡metro idEmpleado es None o no estÃ¡ presente en la URL
            if cod_poliza is None:
                return redirect(url_for('inicio1'))
            else:
                datos=sql_datos_poliza(cod_poliza)
                return render_template('public/Siniestro/form_siniestro_cartaAval.html',cod_poliza=cod_poliza,datos=datos)
        else:
            flash('No tienes permiso para esta acciÃƒÂ²n', 'error')
            return redirect(url_for('inicio1'))
    else:
        flash('Primero debes iniciar sesiÃ³n.', 'error')
        return redirect(url_for('inicio1'))
    
@app.route("/Reembolso/<path:cod_poliza>", methods=['GET'])
def view_form_Reembolso(cod_poliza=None):
    if 'conectado' in session:
        if session.get("permisos")!="Ventas" or session.get("permisos")!="Gerencia":
            # Verificamos si el parÃ¡metro idEmpleado es None o no estÃ¡ presente en la URL
            if cod_poliza is None:
                return redirect(url_for('inicio1'))
            else:
                return render_template('public/Siniestro/form_siniestro_reembolso.html',cod_poliza=cod_poliza,ejecutivos=sql_lista_ejecutivo())
        else:
          flash('No tienes permiso para esta acciÃƒÂ²n', 'error')
          return redirect(url_for('inicio1'))  
    else:
        flash('Primero debes iniciar sesiÃ³n.', 'error')
        return redirect(url_for('inicio1'))
    
@app.route("/SAuto/<path:cod_poliza>", methods=['GET'])
def view_form_Siniestro_Auto(cod_poliza=None):
    if 'conectado' in session:
        if session.get("permisos")!="Ventas" or session.get("permisos")!="Gerencia":
            # Verificamos si el parÃ¡metro idEmpleado es None o no estÃ¡ presente en la URL
            if cod_poliza is None:
                return redirect(url_for('inicio1'))
            else:
                return render_template('public/Siniestro/form_siniestro_Auto.html',cod_poliza=cod_poliza,ejecutivos=sql_lista_ejecutivo())
        else:
            flash('No tienes permiso para esta acciÃƒÂ²n', 'error')
            return redirect(url_for('inicio1'))
    else:
        flash('Primero debes iniciar sesiÃ³n.', 'error')
        return redirect(url_for('inicio1'))
    
@app.route('/form-registrar-beneficiario', methods=['POST'])
def formBeneficiario():
    if 'conectado' in session:
            resultado = procesar_form_beneficiario(request.form)
            cod = request.form['cod']
            detalle_poliza = sql_detalles_polizaBD(cod) or []
            if resultado:
                return render_template('public/Poliza/detalles_polizaPersona.html',detalle_poliza=detalle_poliza)
            else:
                flash('Beneficiario NO fue registrado.', 'error')
                
                return render_template('public/Poliza/detalles_polizaPersona.html',detalle_poliza=detalle_poliza)
    else:
        flash('primero debes iniciar sesiÃ³n.', 'error')
        return redirect(url_for('inicio'))

@app.route("/detalles-poliza/", methods=['GET'])
@app.route("/detalles-poliza/<path:cod_poliza>", methods=['GET'])
@app.route("/detalles-poliza/<path:cod_poliza>/<int:cod_renovacion>", methods=['GET'])
def detallePoliza(cod_poliza=None, cod_renovacion=None):
    if 'conectado' in session:
        if cod_poliza is None:
            return redirect(url_for('inicio1'))
        else: 
            detalle_poliza = sql_detalles_polizaBD(cod_poliza, cod_renovacion)
            if not detalle_poliza:
                flash('No se encontrÃ³ informaciÃ³n de la pÃ³liza o renovaciÃ³n.', 'warning')
                return redirect(url_for('lista_polizas'))

            Ramo = detalle_poliza['Ramo']
            siniestros=sql_lista_siniestros_unico(cod_poliza)
            today = datetime.date.today()
            
            Cod_renovacion = int(detalle_poliza['Cod_renovacion'])
            pagos = lista_Pagos(Cod_renovacion)
            frecuencia = detalle_poliza["Frecuencia"]
            
            # LÃ³gica para determinar si la pÃ³liza estÃ¡ completa para renovaciÃ³n/acciones (incluyendo pagos anulados)
            pagos_no_anulados = [p for p in pagos if p['estado'] != 'ANULADO']
            
            # CondiciÃ³n 1: Todos los pagos no anulados estÃ¡n PAGADO (y hay pagos no anulados)
            all_non_anulados_paid = len(pagos_no_anulados) > 0 and all(p['estado'] == 'PAGADO' for p in pagos_no_anulados)
            
            # CondiciÃ³n 2: Hay al menos un pago ANULADO
            any_anulados = any(p['estado'] == 'ANULADO' for p in pagos)
            
            completo = "si" if all_non_anulados_paid or any_anulados else "no"
            
            if Ramo=="PERSONAS":
                Ramo="Persona"
            elif Ramo=="FIANZAS":
                Ramo="Fianza"
            elif Ramo=="VIAJES":
                Ramo="Viaje"
            elif Ramo=="AUTO":
                Ramo="Auto"
            elif Ramo=="PATRIMONIAL":
                Ramo="Patrimonial"    
                
                
            Antiguas_polizas=sql_lista_antiguos_contratos(cod_poliza)

            return render_template('public/Poliza/detalles_poliza'+Ramo+'.html', detalle_poliza=detalle_poliza,siniestros=siniestros,today=today,completo=completo,antiguas=Antiguas_polizas)
    else:
        flash('Primero debes iniciar sesiÃ³n.', 'error')
        return redirect(url_for('inicio'))
    
@app.route("/detalles-Reembolso/", methods=['GET'])
@app.route("/detalles-Reembolso/<int:cod_reembolso>", methods=['GET'])
@app.route("/detalles-Reembolso/<int:cod_reembolso>/<int:where>", methods=['GET'])
def detalleReembolso(cod_reembolso=None,where=None):
    if 'conectado' in session:
        # Verificamos si el parÃ¡metro idEmpleado es None o no estÃ¡ presente en la URL
        if cod_reembolso is None:
            print("holai")
            return redirect(url_for('inicio'))
        else:
            detalle_reembolso = sql_detalles_reembolsoBD(cod_reembolso) or []
            notas=sql_Notas_reembolso(cod_reembolso) 
            if where is None:
                return render_template('public/Siniestro/detalles_reembolso.html', detalle_reembolso=detalle_reembolso,notas=notas)
            else: return render_template('public/Siniestro/detalles_reembolso.html', detalle_reembolso=detalle_reembolso,where=where,notas=notas)
    else:
        flash('Primero debes iniciar sesiÃ³n.', 'error')
        return redirect(url_for('inicio'))
    

@app.route("/editar-Reembolso/", methods=['GET'])
@app.route("/editar-Reembolso/<int:cod_reembolso>", methods=['GET'])
@app.route("/editar-Reembolso/<int:cod_reembolso>/<int:where>", methods=['GET'])
def editarReembolso(cod_reembolso=None,where=None):
    if 'conectado' in session:
        # Verificamos si el parÃ¡metro idEmpleado es None o no estÃ¡ presente en la URL
        if cod_reembolso is None:
            print("holai")
            return redirect(url_for('inicio'))
        else:
            detalle_reembolso = sql_detalles_reembolsoBD(cod_reembolso) or [] 
            if where is None:
                return render_template('public/Siniestro/editar_siniestro_reembolso.html', detalle_reembolso=detalle_reembolso)
            else: return render_template('public/Siniestro/editar_siniestro_reembolso.html', detalle_reembolso=detalle_reembolso,where=where)
    else:
        flash('Primero debes iniciar sesiÃ³n.', 'error')
        return redirect(url_for('inicio'))
    
@app.route("/editar-Carta Aval/", methods=['GET'])
@app.route("/editar-Carta Aval/<int:cod_carta>", methods=['GET'])
@app.route("/editar-Carta Aval/<int:cod_carta>/<int:where>", methods=['GET'])
def editarCartaAval(cod_carta=None,where=None):
    if 'conectado' in session:
        # Verificamos si el parÃ¡metro idEmpleado es None o no estÃ¡ presente en la URL
        if cod_carta is None:
            print("holai")
            return redirect(url_for('inicio1'))
        else: 
            detalle_CartaAval = sql_detalles_CartaAvalBD(cod_carta) or []  
            if where is None:
                return render_template('public/Siniestro/editar_siniestro_cartaAval.html', detalle_CartaAval=detalle_CartaAval)
            else: return render_template('public/Siniestro/editar_siniestro_cartaAval.html', detalle_CartaAval=detalle_CartaAval,where=where)
    else:
        flash('Primero debes iniciar sesiÃ³n.', 'error')
        return redirect(url_for('inicio'))
    
@app.route("/detalle-SiniestroA/", methods=['GET'])
@app.route("/detalle-SiniestroA/<int:cod_siniestro>", methods=['GET'])
@app.route("/detalle-SiniestroA/<int:cod_siniestro>/<int:where>", methods=['GET'])
def detalleSiniestroA(cod_siniestro=None,where=None):
    if 'conectado' in session:
        # Verificamos si el parÃ¡metro idEmpleado es None o no estÃ¡ presente en la URL
        if cod_siniestro is None: 
            return redirect(url_for('inicio'))
        else:
            detalle_Auto = sql_detalles_SiniestroA(cod_siniestro) or []
            notas=sql_Notas_auto(cod_siniestro) 
            if notas is None:
                notas= []
            if where is None:
                return render_template('public/Siniestro/detalles_Auto.html', detalle_siniestro=detalle_Auto,notas=notas)
            else: return render_template('public/Siniestro/detalles_Auto.html', detalle_siniestro=detalle_Auto,where=where,notas=notas)
    else:
        flash('Primero debes iniciar sesiÃ³n.', 'error')
        return redirect(url_for('inicio'))
    
@app.route("/editar-Siniestro Auto/", methods=['GET'])
@app.route("/editar-Siniestro Auto/<int:cod_siniestro>", methods=['GET'])
@app.route("/editar-Siniestro Auto/<int:cod_siniestro>/<int:where>", methods=['GET'])
def editarSiniestroA(cod_siniestro=None,where=None):
    if 'conectado' in session:
        # Verificamos si el parÃ¡metro idEmpleado es None o no estÃ¡ presente en la URL
        if cod_siniestro is None: 
            return redirect(url_for('inicio'))
        else:
            detalle_Auto = sql_detalles_SiniestroA(cod_siniestro) or [] 
            if where is None:
                return render_template('public/Siniestro/editar_siniestro_auto.html', detalle_siniestro=detalle_Auto)
            else: return render_template('public/Siniestro/editar_siniestro_auto.html', detalle_siniestro=detalle_Auto,where=where)
    else:
        flash('Primero debes iniciar sesiÃ³n.', 'error')
        return redirect(url_for('inicio'))
    
@app.route("/detalles-Carta_Aval/", methods=['GET'])
@app.route("/detalles-Carta_Aval/<int:cod_carta>", methods=['GET'])
@app.route("/detalles-Carta_Aval/<int:cod_carta>/<int:where>", methods=['GET'])
def detalleCarta_Aval(cod_carta=None,where=None):
    if 'conectado' in session:
        # Verificamos si el parÃ¡metro idEmpleado es None o no estÃ¡ presente en la URL
        if cod_carta is None:
            return redirect(url_for('inicio'))
        else:
            detalle_Carta = sql_detalles_CartaAvalBD(cod_carta) or []
            notas=sql_Notas_cartaAval(cod_carta)    
            if where is None:
                return render_template('public/Siniestro/detalles_Carta_Aval.html', detalle_carta_aval=detalle_Carta,where=where,notas=notas)
            else: return render_template('public/Siniestro/detalles_Carta_Aval.html', detalle_carta_aval=detalle_Carta,where=where,notas=notas)
    else:
        flash('Primero debes iniciar sesiÃ³n.', 'error')
        return redirect(url_for('inicio'))


# Buscadon de empleados
@app.route("/buscando-empleado", methods=['POST'])
def viewBuscarEmpleadoBD():
    resultadoBusqueda = buscarEmpleadoBD(request.json['busqueda'])
    if resultadoBusqueda:
        return render_template(f'{PATH_URL}/resultado_busqueda_empleado.html', dataBusqueda=resultadoBusqueda)
    else:
        return jsonify({'fin': 0})
    
@app.route('/filtrar-siniestros', methods=['POST'])
def filtrar_siniestros():
    try:
        # --- NUEVAS LÃNEAS DE DEPURACIÃƒâ€œN ---
        print("--- DEBUG BACKEND ---")
        print(f"Content-Type recibido: {request.headers.get('Content-Type')}")

        # Intentar obtener el JSON del cuerpo de la solicitud
        # request.json serÃ¡ None si el Content-Type no es 'application/json'
        # o si el cuerpo no es JSON vÃ¡lido.
        data = request.json
        print(f"filtrar-siniestros: {len(data) if data else 0} filtros recibidos")

        if data is None:
            print("ADVERTENCIA: request.json es None. El cuerpo de la solicitud no es JSON vÃ¡lido o el Content-Type no es 'application/json'.")
            # Si data es None, intentamos leer el cuerpo como texto para depurar
            try:
                raw_data = request.data.decode('utf-8')
                print(f"Cuerpo de la solicitud RAW: {raw_data}")
            except Exception as e:
                print(f"Error al leer el cuerpo RAW: {e}")
            # Si no hay datos, devolvemos una respuesta de error para evitar fallos
            return jsonify({'error': 'No se recibieron datos JSON vÃ¡lidos.'}), 400
        # --- FIN NUEVAS LÃNEAS DE DEPURACIÃƒâ€œN ---

        tipo_cedula = data.get('tipo_cedula')
        cedula = data.get('cedula')
        estado_siniestro = data.get('estado_siniestro')
        meses = data.get('meses', []) # Obtener el array directamente, con un valor por defecto de lista vacÃ­a
        anio = data.get('anio')
        tipo_siniestro = data.get('tipo_siniestro')

        print(f"Meses recibidos en Flask (despuÃ©s de .get()): {meses}") # DepuraciÃ³n en el servidor
        print(f"AÃ±o recibido en Flask (despuÃ©s de .get()): {anio}")     # DepuraciÃ³n en el servidor
        print(f"Tipo Siniestro recibido: {tipo_siniestro}")

        # Llamar a la nueva funciÃ³n de filtrado unificado
        resultados = obtener_siniestros_filtrados(tipo_cedula, cedula, estado_siniestro, meses, anio, tipo_siniestro)

        return jsonify({'data': resultados})

    except Exception as e:
        # Manejo de errores
        print(f"Error en la ruta /filtrar-siniestros: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/actualizar_pago', methods=['POST'])
def actualizar_pago():
    if request.is_json:
        data = request.get_json()
        print(f"[DEBUG] /actualizar_pago recibiÃ³ PAYLOAD: {data}")
        pago_id = data.get('pagoId')
        fecha_cuota = data.get('fecha') # Nueva fecha de la cuota
        metodo_pago = data.get('metodoPago')
        fecha_pagada = data.get('fecha_pagada')
        moneda = data.get('moneda')
        tasa = data.get('tasa')
        nro_recibo = data.get('nro_recibo') 
        nro_cuota = data.get('nro_cuota')
        if nro_cuota == '':
            nro_cuota = None
        monto = data.get('monto') # Recibir monto para actualizar si es necesario

        if not pago_id:
            return jsonify({'success': False, 'message': 'ID de pago no recibido.'}), 400
        if not fecha_cuota:
            return jsonify({'success': False, 'message': 'La fecha de la cuota es requerida.'}), 400
        if not moneda:
            return jsonify({'success': False, 'message': 'Debe seleccionar una moneda.'}), 400
        if not metodo_pago:
            return jsonify({'success': False, 'message': 'Debe seleccionar un mÃ©todo de pago.'}), 400
        if not fecha_pagada:
            return jsonify({'success': False, 'message': 'La fecha de pago es requerida.'}), 400
        if not nro_recibo:
            return jsonify({'success': False, 'message': 'El nÃºmero de recibo o referencia es requerido.'}), 400
        # nro_cuota check removed as it's not a required field from frontend anymore

        # --- INICIO DE LA CORRECCIÃƒâ€œN ---
        if moneda == '$':
            tasa = None
        elif not tasa:
            return jsonify({'success': False, 'message': 'La tasa es requerida para pagos en BolÃ­vares.'}), 400
        else:
            try:
                tasa_cleaned = str(tasa).replace('.', '').replace(',', '.')
                tasa = float(tasa_cleaned)
                if tasa <= 0:
                    return jsonify({'success': False, 'message': 'La tasa debe ser mayor a cero.'}), 400
            except (ValueError, TypeError):
                return jsonify({'success': False, 'message': 'La tasa ingresada no es un nÃºmero vÃ¡lido.'}), 400
        if monto:
            try:
                monto_cleaned = str(monto).replace('.', '').replace(',', '.')
                monto = float(monto_cleaned)
                if monto < 0:
                     return jsonify({'success': False, 'message': 'El monto no puede ser negativo.'}), 400
            except (ValueError, TypeError):
                return jsonify({'success': False, 'message': 'El monto ingresado no es un nÃºmero vÃ¡lido.'}), 400
        # --- FIN DE LA CORRECCIÃ“N ---

        try:
            with connectionBD() as conexion_MySQLdb:
                with conexion_MySQLdb.cursor() as cursor:
                    # 1. Actualizar el pago actual a 'PAGADO'
                    sql_update_pago = """
                        UPDATE pago
                        SET fecha = %s,
                            Metodo_pago = %s,
                            fecha_pagada = %s,
                            moneda = %s,
                            tasa = %s,
                            estado = 'PAGADO',
                            recibo = %s,
                            nro_cuota = COALESCE(%s, nro_cuota),
                            monto = COALESCE(%s, monto)
                        WHERE Cod_pago=%s;
                    """
                    cursor.execute(sql_update_pago, (fecha_cuota, metodo_pago, fecha_pagada, moneda, tasa, nro_recibo, nro_cuota, monto, pago_id))

                    # 2. Obtener informaciÃ³n de la renovaciÃ³n
                    cursor.execute("SELECT Cod_renovacion FROM pago WHERE Cod_pago = %s", (pago_id,))
                    renovacion_info = cursor.fetchone()
                    cod_renovacion = renovacion_info['Cod_renovacion']

                    # 3. Obtener datos de la renovaciÃ³n para calcular cuotas
                    cursor.execute("SELECT Frecuencia, Prima FROM renovacion WHERE Cod_renovacion = %s", (cod_renovacion,))
                    renovacion_data = cursor.fetchone()

                    if renovacion_data:
                        frecuencia = renovacion_data['Frecuencia']
                        prima_total = renovacion_data['Prima']

                        print(f"--- DepuraciÃ³n de CreaciÃ³n de Cuotas para RenovaciÃ³n {cod_renovacion} ---")

                        # 4. Determinar el nÃºmero total de pagos requeridos
                        frecuencia_map = {
                            1: {'pagos': 1, 'intervalo': 12}, # Anual
                            2: {'pagos': 12, 'intervalo': 1}, # Mensual
                            3: {'pagos': 4, 'intervalo': 3},  # Trimestral
                            4: {'pagos': 2, 'intervalo': 6},  # Semestral
                            5: {'pagos': 3, 'intervalo': 4},  # Cuatrimestral
                            6: {'pagos': 6, 'intervalo': 2},  # Bimensual
                            10: {'pagos': 10, 'intervalo': 1} # Mensual Especial (10 pagos)
                        }
                        config_pago = frecuencia_map.get(frecuencia, {'pagos': 1, 'intervalo': 12})
                        num_pagos_requeridos = config_pago['pagos']
                        intervalo_meses = config_pago['intervalo']

                        # Eliminada lÃ³gica hardcoded de 10 pagos para Mercantil.

                        print(f"Frecuencia ID: {frecuencia} -> Pagos Requeridos: {num_pagos_requeridos}, Intervalo: {intervalo_meses} meses")

                        # 5. Obtener los nÃºmeros de cuota existentes y la fecha inicial
                        cursor.execute("SELECT nro_cuota, fecha FROM pago WHERE Cod_renovacion = %s ORDER BY nro_cuota ASC", (cod_renovacion,))
                        pagos_existentes = cursor.fetchall()
                        cuotas_existentes = {p['nro_cuota'] for p in pagos_existentes if p['nro_cuota'] is not None}
                        
                        fecha_pago_inicial = None
                        if pagos_existentes:
                            cuota_1 = next((p['fecha'] for p in pagos_existentes if p['nro_cuota'] == 1 and p['fecha'] is not None), None)
                            fechas_validas = [p['fecha'] for p in pagos_existentes if p['fecha'] is not None]
                            if cuota_1 is not None:
                                fecha_pago_inicial = cuota_1
                            elif fechas_validas:
                                fecha_pago_inicial = min(fechas_validas)
                            else:
                                fecha_pago_inicial = None

                        # 6. Calcular y agregar solo las cuotas faltantes por su nÃºmero especÃ­fico
                        if num_pagos_requeridos > len(cuotas_existentes) and fecha_pago_inicial:
                            monto_cuota = prima_total / num_pagos_requeridos if num_pagos_requeridos > 0 else prima_total
                            if isinstance(fecha_pago_inicial, date) and not isinstance(fecha_pago_inicial, datetime.datetime):
                                fecha_base_calculo = datetime.datetime.combine(fecha_pago_inicial, datetime.datetime.min.time())
                            elif isinstance(fecha_pago_inicial, str):
                                try:
                                    fecha_base_calculo = datetime.datetime.strptime(fecha_pago_inicial, '%Y-%m-%d')
                                except ValueError:
                                    fecha_base_calculo = datetime.datetime.combine(date.today(), datetime.datetime.min.time())
                            else:
                                fecha_base_calculo = fecha_pago_inicial
                            
                            for i in range(1, num_pagos_requeridos + 1):
                                if i not in cuotas_existentes:
                                    fecha_siguiente_pago = fecha_base_calculo + relativedelta(months=(i - 1) * intervalo_meses)
                                    print(f"Creando cuota faltante #{i} con fecha: {fecha_siguiente_pago.strftime('%Y-%m-%d')}")
                                    sql_insert_faltante = "INSERT INTO pago (Cod_renovacion, Moneda, fecha, monto, estado, nro_cuota) VALUES (%s, %s, %s, %s, %s, %s)"
                                    valores_faltante = (cod_renovacion, moneda, fecha_siguiente_pago.strftime('%Y-%m-%d'), monto_cuota, 'EN PROCESO', i)
                                    cursor.execute(sql_insert_faltante, valores_faltante)
                                    cuotas_existentes.add(i)

                    conexion_MySQLdb.commit()
                
            return jsonify({'success': True, 'message': f'Pago {pago_id} actualizado correctamente'}), 200
        except Exception as e:
            conexion_MySQLdb.rollback()
            return jsonify({'success': False, 'error': str(e)}), 500
    else:
        return jsonify({'success': False, 'error': 'La solicitud debe ser en formato JSON'}), 400

@app.route('/revertir_pago_individual', methods=['POST'])
def revertir_pago_individual_route():
    try:
        data = request.get_json()
        pago_id = data.get('pagoId')

        if not pago_id:
            return jsonify({'success': False, 'message': 'ID de pago no recibido.'}), 400

        # Llamar a la funciÃ³n del controlador
        resultado = revertirPago(pago_id) # Assuming revertirPago is the correct function to call
        
        if resultado > 0:
            return jsonify({'success': True, 'message': f'El pago {pago_id} ha sido revertido a "EN PROCESO" correctamente.'}), 200
        else:
            return jsonify({'success': False, 'error': f'No se pudo revertir el pago con ID {pago_id}.'}), 404

    except Exception as e:
        print(f"Error en route revertir_pago_individual: {e}")
        return jsonify({'success': False, 'message': f'Error interno: {str(e)}'}), 500


@app.route('/anular_pago_individual', methods=['POST'])
def anular_pago_individual():
    if not request.is_json:
        return jsonify({'success': False, 'error': 'La solicitud debe ser en formato JSON'}), 400

    data = request.get_json()
    pago_id = data.get('pagoId')

    if not pago_id:
        return jsonify({'success': False, 'error': 'Falta el ID del pago'}), 400

    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                querySQL = "UPDATE pago SET estado = 'ANULADO' WHERE Cod_pago = %s"
                cursor.execute(querySQL, (pago_id,))
                conexion_MySQLdb.commit()

                if cursor.rowcount > 0:
                    return jsonify({'success': True, 'message': f'El pago {pago_id} ha sido anulado correctamente.'}), 200
                else:
                    return jsonify({'success': False, 'error': f'No se encontrÃ³ el pago con ID {pago_id}.'}), 404
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/revertir_pago_individual', methods=['POST'])
def revertir_pago_individual():
    if not request.is_json:
        return jsonify({'success': False, 'error': 'La solicitud debe ser en formato JSON'}), 400

    data = request.get_json()
    pago_id = data.get('pagoId')

    if not pago_id:
        return jsonify({'success': False, 'error': 'Falta el ID del pago'}), 400

    try:
        resultado = revertirPago(pago_id)
        if resultado > 0:
            return jsonify({'success': True, 'message': f'El pago {pago_id} ha sido revertido a "EN PROCESO" correctamente.'}), 200
        else:
            return jsonify({'success': False, 'error': f'No se pudo revertir el pago con ID {pago_id}.'}), 404
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/verificar_pagos_renovacion/<int:id>', methods=['POST'])
def verificar_pagos_renovacion(id):
    if 'conectado' in session:
        try:
            with connectionBD() as conexion_MySQLdb:
                with conexion_MySQLdb.cursor() as cursor:
                    # Logic similar to what we have in actualizar_pago to ensure all cuotas exist
                    cursor.execute("SELECT Frecuencia, Prima FROM renovacion WHERE Cod_renovacion = %s", (id,))
                    renov_data = cursor.fetchone()
                    if not renov_data:
                        return jsonify({'success': False, 'error': 'RenovaciÃ³n no encontrada'}), 404
                    
                    frecuencia = renov_data['Frecuencia']
                    prima_total = renov_data['Prima']
                    
                    frecuencia_map = {1: 1, 2: 12, 3: 4, 4: 2, 5: 3, 6: 6, 10: 10}
                    intervalo_map = {1: 12, 2: 1, 3: 3, 4: 6, 5: 4, 6: 2, 10: 1}
                    
                    num_pagos_req = frecuencia_map.get(frecuencia, 1)
                    intervalo = intervalo_map.get(frecuencia, 12)
                    
                    # Eliminada lÃ³gica hardcoded de 10 pagos para Mercantil.

                    cursor.execute("SELECT cod_pago, nro_cuota, fecha, moneda, estado FROM pago WHERE Cod_renovacion = %s", (id,))
                    todos_pagos = cursor.fetchall()
                    
                    if not todos_pagos:
                        return jsonify({'success': False, 'error': 'No hay pagos base para verificar. Registre el primer pago manualmente.'}), 400

                    # 3. Detectar Duplicados y Excesos
                    eliminados = 0
                    vistos = {} # nro_cuota -> pago_record
                    ids_a_eliminar = []
                    
                    for p in todos_pagos:
                        try:
                            n_cuota = int(p['nro_cuota']) if p['nro_cuota'] is not None else None
                        except:
                            n_cuota = None
                            
                        if n_cuota is None: 
                            continue
                        
                        # Si la cuota excede el lÃ­mite y estÃ¡ EN PROCESO, marcar para eliminar
                        if n_cuota > num_pagos_req and p['estado'] == 'EN PROCESO':
                            ids_a_eliminar.append(p['cod_pago'])
                            continue
                        
                        if n_cuota in vistos:
                            existente = vistos[n_cuota]
                            # Regla: Si tengo uno PAGADO y otro EN PROCESO, elimino el EN PROCESO
                            if existente['estado'] == 'EN PROCESO' and p['estado'] == 'PAGADO':
                                ids_a_eliminar.append(existente['cod_pago'])
                                vistos[n_cuota] = p
                            elif p['estado'] == 'EN PROCESO':
                                ids_a_eliminar.append(p['cod_pago'])
                            else:
                                # Ambos PAGADO o ambos EN PROCESO, eliminamos el repetido
                                ids_a_eliminar.append(p['cod_pago'])
                        else:
                            vistos[n_cuota] = p

                    if ids_a_eliminar:
                        for target_id in ids_a_eliminar:
                            cursor.execute("DELETE FROM pago WHERE cod_pago = %s", (target_id,))
                        eliminados = len(ids_a_eliminar)

                    # 4. Re-evaluar faltantes despuÃ©s de la limpieza
                    cuotas_ids_final = {int(p['nro_cuota']) for p in vistos.values() if p['nro_cuota'] is not None and p['cod_pago'] not in ids_a_eliminar}
                    
                    # Tomar base para fechas (buscar la primera con fecha vÃ¡lida para evitar NoneType error)
                    cuota_base = vistos.get(1) if (1 in vistos and vistos[1]['fecha']) else None
                    if not cuota_base:
                        # Buscar cualquier otra cuota que tenga fecha
                        for i in range(2, num_pagos_req + 1):
                            if i in vistos and vistos[i]['fecha']:
                                cuota_base = vistos[i]
                                break
                    if not cuota_base:
                        # Si aÃºn no hay, buscar en todas (incluyendo excedentes si hubiera)
                        for p in vistos.values():
                            if p['fecha']:
                                cuota_base = p
                                break
                    
                    if not cuota_base or not cuota_base['fecha']:
                        return jsonify({'success': False, 'error': 'No se encontrÃ³ ninguna cuota con fecha vÃ¡lida para usar como base de cÃ¡lculo. Por favor, registre la fecha de al menos una cuota manualmente.'}), 400

                    # Aseguramos que fecha_base sea datetime.datetime
                    if isinstance(cuota_base['fecha'], (date, datetime.datetime)):
                        if isinstance(cuota_base['fecha'], date) and not isinstance(cuota_base['fecha'], datetime.datetime):
                            fecha_base = datetime.datetime.combine(cuota_base['fecha'], datetime.datetime.min.time())
                        else:
                            fecha_base = cuota_base['fecha']
                    else:
                        # Por si acaso viene como string
                        try:
                            fecha_base = datetime.datetime.strptime(str(cuota_base['fecha']), '%Y-%m-%d')
                        except:
                            return jsonify({'success': False, 'error': f'Formato de fecha invÃ¡lido en la cuota base: {cuota_base["fecha"]}'}), 400
                        
                    moneda = cuota_base['moneda']
                    monto_cuota = prima_total / num_pagos_req if num_pagos_req > 0 else prima_total
                    
                    creados = 0
                    for i in range(1, num_pagos_req + 1):
                        if i not in cuotas_ids_final:
                            # Calcular offset relativo a la cuota base
                            base_n = int(cuota_base['nro_cuota'])
                            offset = (i - base_n) * intervalo
                            fecha_c = fecha_base + relativedelta(months=offset)
                            cursor.execute("INSERT INTO pago (Cod_renovacion, Moneda, fecha, monto, estado, nro_cuota) VALUES (%s, %s, %s, %s, %s, %s)",
                                         (id, moneda, fecha_c.strftime('%Y-%m-%d'), monto_cuota, 'EN PROCESO', i))
                            creados += 1
                    
                    # 5. Verificar y ajustar montos para que todos sean iguales y sumen la prima total
                    actualizados_montos = 0
                    cursor.execute("SELECT cod_pago, monto FROM pago WHERE Cod_renovacion = %s", (id,))
                    pagos_finales = cursor.fetchall()
                    
                    for pf in pagos_finales:
                        monto_actual = float(pf['monto']) if pf['monto'] is not None else 0
                        # Si hay diferencia de mÃ¡s de 1 centavo, ajustar para garantizar igualdad total
                        if abs(monto_actual - float(monto_cuota)) > 0.01:
                            cursor.execute("UPDATE pago SET monto = %s WHERE cod_pago = %s", (monto_cuota, pf['cod_pago']))
                            actualizados_montos += 1

                    conexion_MySQLdb.commit()
                    
                    msg = f"VerificaciÃ³n completada."
                    if eliminados > 0: msg += f" Se eliminaron {eliminados} cuotas duplicadas/excedentes."
                    if creados > 0: msg += f" Se crearon {creados} cuotas faltantes."
                    if actualizados_montos > 0: msg += f" Se ajustaron los montos de {actualizados_montos} cuotas para coincidir con la prima."
                    if eliminados == 0 and creados == 0 and actualizados_montos == 0: msg += " Todo estÃ¡ en orden."
                    
                    return jsonify({'success': True, 'message': msg}), 200
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500
    else:
        return jsonify({'success': False, 'error': 'SesiÃ³n no iniciada'}), 401
    
@app.route('/filtrar-polizas', methods=['POST'])
def filtrar_poliza():
    # Intentar obtener datos tanto de form como de JSON (para flexibilidad)
    tipo_filtro_fecha = request.form.get('tipo_filtro_fecha')
    fecha_str = request.form.get('fecha')
    anio_str = request.form.get('anio')
    fecha_inicio_str = request.form.get('fecha_inicio')
    fecha_fin_str = request.form.get('fecha_fin')
    estados = request.form.getlist('estados[]')
    compania_id = request.form.get('compania')
    ejecutivo_id = request.form.get('ejecutivo')

    if not any([fecha_str, anio_str, fecha_inicio_str, fecha_fin_str, estados, compania_id, ejecutivo_id]):
        # Si no hay filtros, podemos devolver todo o manejarlo
        pass

    try:
        mes = None
        anio = None
        rango_inicio = None
        rango_fin = None

        if tipo_filtro_fecha == 'mes' and fecha_str:
            try:
                fecha = datetime.datetime.strptime(fecha_str, '%Y-%m').date()
                mes = fecha.month
                anio = fecha.year
            except ValueError:
                pass # Fecha invÃ¡lida o vacÃ­a
        elif tipo_filtro_fecha == 'anio' and anio_str:
            try:
                anio = int(anio_str)
            except ValueError:
                pass
        elif tipo_filtro_fecha == 'rango':
            rango_inicio = fecha_inicio_str
            rango_fin = fecha_fin_str

        resultados = obtener_polizas_filtradas(
            ano=anio, mes=mes, rango_inicio=rango_inicio, rango_fin=rango_fin, estados=estados, 
            compania_id=compania_id, ejecutivo_id=ejecutivo_id
        )

        return jsonify({'data': resultados})

    except Exception as e:
        print(f"Error en ruta /filtrar-polizas: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/riesgo', methods=['GET'])
def riesgo():
    if 'conectado' in session:
        riesgos = sql_lista_riesgos()
        return render_template('public/riesgo.html', riesgos=riesgos)
    else:
        flash('primero debes iniciar sesiÃ³n.', 'error')
        return redirect(url_for('inicio'))

@app.route('/actualizar-riesgo-asegurado', methods=['POST'])
def actualizar_riesgo_asegurado_route():
    if 'conectado' in session:
        if request.is_json:
            data = request.get_json()
            resultado = actualizar_riesgo_asegurado(data)
            return jsonify(resultado)
        return jsonify({'success': False, 'message': 'El formato de la solicitud debe ser JSON.'}), 400
    return jsonify({'success': False, 'message': 'Debes iniciar sesiÃ³n.'}), 401
    
# Buscando poliza
@app.route("/buscando-poliza", methods=['POST'])
def viewBuscarPolizaBD():
    data= request.get_json()
    busqueda = data['busqueda']
    filtro = data.get('filtro', None)
    resultadoBusqueda = buscarPolizaBD(busqueda, filtro)
    if resultadoBusqueda:
        return render_template('public/poliza/resultado_busqueda_poliza.html', dataBusqueda=resultadoBusqueda,today=datetime.date.today())
    else:
        return jsonify({'fin': 0})
    
# Buscando asegurado
@app.route("/buscando-asegurado", methods=['POST'])
def viewBuscarAseguradoBD():
    data= request.get_json()
    busqueda = data['busqueda']
    filtro = data.get('filtro', None)
    resultadoBusqueda = buscarAseguradoBD(busqueda,filtro)
    if resultadoBusqueda:
        return render_template('public/Asegurados/resultado_busqueda_asegurado.html', dataBusqueda=resultadoBusqueda)
    else:
        return jsonify({'fin': 0})




@app.route("/editar-empleado/<int:id>", methods=['GET'])
def viewEditarEmpleado(id):
    if 'conectado' in session:
        respuestaEmpleado = buscarEmpleadoUnico(id)
        if respuestaEmpleado:
            return render_template(f'{PATH_URL}/form_empleado_update.html', respuestaEmpleado=respuestaEmpleado)
        else:
            flash('El empleado no existe.', 'error')
            return redirect(url_for('inicio'))
    else:
        flash('Primero debes iniciar sesiÃ³n.', 'error')
        return redirect(url_for('inicio'))

@app.route("/Pagos_poliza/<path:id>", methods=['GET'])     
@app.route("/Pagos_poliza/<path:id>/<int:boton>", methods=['GET'])
@app.route("/Pagos_poliza/<path:id>/<int:boton>/<int:a>", methods=['GET'])
def viewPagosPoliza(id,boton=0,a=0):
    if 'conectado' in session:
        poliza_info = {}
        try:
            with connectionBD() as conexion_MySQLdb:
                with conexion_MySQLdb.cursor() as cursor:
                    if a==0:
                        sql = """
                            SELECT 
                                p.cod_poliza,
                                a.nombre,
                                a.apellido
                            FROM poliza p
                            JOIN asegurado a ON p.CI_asegurado = a.CI
                            WHERE p.cod_poliza = %s
                        """
                    else:
                        sql = """
                            SELECT 
                                p.cod_poliza,
                                a.nombre,
                                a.apellido
                            FROM poliza p
                            JOIN asegurado a ON p.CI_asegurado = a.CI
                            JOIN renovacion r on p.cod_poliza = r.cod_poliza
                            WHERE r.cod_renovacion = %s
                        """
                    cursor.execute(sql, (id,))
                    poliza_info = cursor.fetchone()
        except Exception as e:
            print(f"Error fetching policy info: {e}")
            pass
        print(a)
        cods_contrato=lista_contrato(id,a)
        if a==0:
            contrato=cods_contrato[0]
        else:
            contrato = cods_contrato
        Cod_renovacion=int(contrato['Cod_renovacion'])
        pagos = lista_Pagos(Cod_renovacion)
        return render_template('public/Poliza/Pagos.html', pagos=pagos,cods=contrato,boton=boton, poliza_info=poliza_info)
       
    else:
        flash('Primero debes iniciar sesiÃ³n.', 'error')
        return redirect(url_for('inicio'))
    
@app.route("/form_pagos/<int:id>", methods=['GET'])
def view_form_pagos(id):
    if 'conectado' in session:
        boton=1
        prima = pago_prima(id)
        monto= 0
        
                 
        


        if prima['Frecuencia'] == 1:
                monto=prima['Prima']
        elif prima['Frecuencia'] == 2:
                monto= prima['Prima'] / 12 
        elif prima['Frecuencia'] == 3: 
                monto=prima['Prima'] / 4
        elif prima['Frecuencia'] == 5: 
                monto=prima['Prima'] / 3 
        elif prima['Frecuencia'] == 6: 
                monto=prima['Prima'] / 6        
        # Mercantil Especial (10 pagos)
        elif prima['Frecuencia'] == 10:
                monto = prima['Prima'] / 10

        # Eliminada lÃ³gica hardcoded que sobrescribÃ­a el monto.
        # Ahora se respeta el valor de 'Frecuencia' (2 para 12 pagos, 10 para 10 pagos).


        

        try:
            monitor = Monitor(AlCambio, 'USD')
            bcv = monitor.get_value_monitors("bcv")
            last_update = bcv.price
            print(f"El valor del dolar bcv es: {last_update} Bs")
        except Exception as e:
            print(f"Error al obtener el valor del dÃ³lar: {e}")
            last_update = ''
        
        
        # Obtener el siguiente nÃºmero de cuota disponible
        siguiente_cuota = 1
        try:
            with connectionBD() as conexion_MySQLdb:
                with conexion_MySQLdb.cursor() as cursor:
                    cursor.execute("SELECT nro_cuota FROM pago WHERE Cod_renovacion = %s ORDER BY CAST(nro_cuota AS INTEGER) ASC", (id,))
                    existentes = [int(p['nro_cuota']) for p in cursor.fetchall() if p['nro_cuota'] is not None]
                    while siguiente_cuota in existentes:
                        siguiente_cuota += 1
        except Exception as e:
            print(f"Error al calcular siguiente cuota: {e}")
            siguiente_cuota = 1

        if not prima['comision']:
          comision = asignar_comision(id)
          if comision: 
              return render_template('public/poliza/form_pagos.html', cod_renovacion=id, monto=monto, dolar=last_update, advertencia_comision=False, siguiente_cuota=siguiente_cuota)
          else:  
            flash('ComisiÃ³n no asignada, Ã‚Â¿desea agregarla manual ahora?', 'warning')
            # Renderiza el formulario de pago de todas formas, pero con una advertencia
            return render_template('public/poliza/form_pagos.html', cod_renovacion=id, monto=monto, dolar=last_update, advertencia_comision=True, cod_poliza=prima['Cod_poliza'], siguiente_cuota=siguiente_cuota)
        else: 
            return render_template('public/poliza/form_pagos.html', cod_renovacion=id, monto=monto, dolar=last_update, advertencia_comision=False, siguiente_cuota=siguiente_cuota)
       
    else:
        flash('Primero debes iniciar sesiÃ³n.', 'error')
        return redirect(url_for('inicio'))
    
@app.route("/form_renovar/<path:id>", methods=['GET'])
def view_form_renovar(id):
    if 'conectado' in session:
        contratos=lista_contrato(id,0)
        contrato=contratos[0]
        fecha_vencimiento=contrato["Fecha_vencimiento"]

        return render_template('public/poliza/form_renovar.html',cod_poliza=id,fecha_vencimiento=fecha_vencimiento, contrato=contrato)
       
    else:
        flash('Primero debes iniciar sesiÃ³n.', 'error')
        return redirect(url_for('inicio'))
    
@app.route("/form-registrar-pago", methods=['POST'])
def registrar_form_pagos():
    if 'conectado' in session:
        resultado = procesar_form_pago(request.form)
        Cod_renovacion = request.form.get('renovacion')
        
        if resultado.get('success'):
            cod_polizaD = cod_poliza_F(Cod_renovacion)
            cod_poliza = cod_polizaD['cod_poliza']
            cargar_riesgo(Cod_renovacion)
            flash(resultado.get('message'), 'success')
            return redirect(url_for('viewPagosPoliza', id=cod_poliza))
        else:
            flash(resultado.get('message', 'El pago NO fue registrado.'), 'error')
            prima = pago_prima(Cod_renovacion)
            monto = 0
            if prima['Frecuencia'] == 1:
                monto = prima['Prima']
            elif prima['Frecuencia'] == 2:
                monto = prima['Prima'] / 12 
            elif prima['Frecuencia'] == 3: 
                monto = prima['Prima'] / 4
            elif prima['Frecuencia'] == 5: 
                monto = prima['Prima'] / 3 
            elif prima['Frecuencia'] == 6: 
                monto = prima['Prima'] / 6        
            else:
                monto = prima['Prima'] / 2

            if prima.get('compania') == "MERCANTIL PANAMÃ":
                if str(prima.get('ramo')).upper() == "PERSONA": 
                    if str(prima.get('Subramo')).upper() == "SALUD":
                        if prima['Frecuencia'] == 2:
                            monto = prima['Prima'] / 10

            try:
                monitor = Monitor(AlCambio, 'USD')
                bcv = monitor.get_value_monitors("bcv")
                last_update = bcv.price
            except Exception as e:
                print(f"Error al obtener el valor del dÃ³lar en el registro: {e}")
                last_update = ''
            return render_template('public/poliza/form_pagos.html', cod_renovacion=Cod_renovacion, monto=monto, dolar=last_update)
       
    else:
        flash('Primero debes iniciar sesiÃ³n.', 'error')
        return redirect(url_for('inicio'))


# Recibir formulario para actulizar informacion de empleado
@app.route('/actualizar-empleado', methods=['POST'])
def actualizarEmpleado():
    resultData = procesar_actualizacion_form(request)
    if resultData:
        return redirect(url_for('lista_empleados'))
    
@app.route('/actualizar-asegurado', methods=['POST'])
def actualizarAsegurado():
    resultData = procesar_actualizacion_form_asegurado(request)
    if resultData:
        return redirect(url_for('lista_asegurado'))
    
@app.route('/actualizar-ejecutivo', methods=['POST'])
def actualizarEjecutivo():
    resultData = procesar_actualizacion_form_ejecutivo(request)
    if resultData:
        return redirect(url_for('ejecutivos'))
    
@app.route('/actualizar-company', methods=['POST'])
def actualizarCompany():
    resultData = procesar_actualizacion_form_company(request)
    if resultData:
        return redirect(url_for('lista_company'))
    
@app.route('/actualizar-reembolso', methods=['POST'])
def actualizarReembolso():
    resultData = procesar_actualizacion_form_reembolso(request)
    if resultData: 
        return redirect(url_for('lista_siniestros'))
    
@app.route('/actualizar-siniestroAuto', methods=['POST'])
def actualizarSiniestroAuto():
    resultData = procesar_actualizacion_form_SiniestroAuto(request)
    if resultData: 
        return redirect(url_for('lista_siniestros'))

@app.route('/actualizar-cartaAval', methods=['POST'])
def actualizarCartaAval(): 
    resultData = procesar_actualizacion_form_CartaAval(request)
    if resultData: 
        return redirect(url_for('lista_siniestros'))

@app.route('/actualizar-poliza-persona/<path:cod_poliza>', methods=['POST'])
def actualizarPolizaPersona(cod_poliza):
    if 'conectado' in session:
        resultado = procesar_actualizacion_poliza_persona(request.form, cod_poliza)
        
        # Determinar el cÃ³digo para redireccionar (por si cambiÃ³)
        nuevo_cod = request.form.get('cod_poliza')
        target_cod = nuevo_cod if (nuevo_cod and resultado) else cod_poliza

        if resultado:
            flash('PÃ³liza actualizada correctamente.', 'success')
        else:
            flash('Error al actualizar la pÃ³liza.', 'error')
        return redirect(url_for('detallePoliza', cod_poliza=target_cod))
    else:
        flash('Primero debes iniciar sesiÃ³n.', 'error')
        return redirect(url_for('inicio'))

@app.route('/actualizar-poliza-auto/<path:cod_poliza>', methods=['POST'])
def actualizarPolizaAuto(cod_poliza):
    if 'conectado' in session:
        resultado = procesar_actualizacion_poliza_auto(request.form, cod_poliza)
        
        # Determinar el cÃ³digo para redireccionar (por si cambiÃ³)
        nuevo_cod = form_data.get('cod_poliza') if 'form_data' in locals() else request.form.get('cod_poliza')
        target_cod = nuevo_cod if (nuevo_cod and resultado) else cod_poliza

        if resultado:
            flash('PÃ³liza de auto actualizada correctamente.', 'success')
        else:
            flash('Error al actualizar la pÃ³liza de auto.', 'error')
        return redirect(url_for('detallePoliza', cod_poliza=target_cod))

    else:
        flash('Primero debes iniciar sesiÃ³n.', 'error')
        return redirect(url_for('inicio'))

@app.route('/actualizar-poliza-patrimonial/<path:cod_poliza>', methods=['POST'])
def actualizarPolizaPatrimonial(cod_poliza):
    if 'conectado' in session:
        resultado = procesar_actualizacion_poliza_patrimonial(request.form, cod_poliza)
        
        # Determine the code to redirect to (in case it was changed)
        nuevo_cod = request.form.get('cod_poliza')
        target_cod = nuevo_cod if (nuevo_cod and resultado) else cod_poliza

        if resultado:
            flash('PÃ³liza patrimonial actualizada correctamente.', 'success')
        else:
            flash('Error al actualizar la pÃ³liza patrimonial.', 'error')
        
        return redirect(url_for('detallePoliza', cod_poliza=target_cod))
    else:
        flash('Primero debes iniciar sesiÃ³n.', 'error')
        return redirect(url_for('inicio'))
    
@app.route('/actualizar-poliza-viaje/<path:cod_poliza>', methods=['POST'])
def actualizarPolizaViaje(cod_poliza):
    if 'conectado' in session:
        resultado = procesar_actualizacion_poliza_viaje(request.form, cod_poliza)
        nuevo_cod = request.form.get('cod_poliza')
        target_cod = nuevo_cod if (nuevo_cod and resultado) else cod_poliza
        if resultado:
            flash('PÃ³liza de viaje actualizada correctamente.', 'success')
        else:
            flash('Error al actualizar la pÃ³liza de viaje.', 'error')
        return redirect(url_for('detallePoliza', cod_poliza=target_cod))
    else:
        flash('Primero debes iniciar sesiÃ³n.', 'error')
        return redirect(url_for('inicio'))

@app.route('/actualizar-poliza-fianza/<path:cod_poliza>', methods=['POST'])
def actualizarPolizaFianza(cod_poliza):
    if 'conectado' in session:
        resultado = procesar_actualizacion_poliza_fianza(request.form, cod_poliza)
        nuevo_cod = request.form.get('cod_poliza')
        target_cod = nuevo_cod if (nuevo_cod and resultado) else cod_poliza
        if resultado:
            flash('PÃ³liza de fianza actualizada correctamente.', 'success')
        else:
            flash('Error al actualizar la pÃ³liza de fianza.', 'error')
        return redirect(url_for('detallePoliza', cod_poliza=target_cod))
    else:
        flash('Primero debes iniciar sesiÃ³n.', 'error')
        return redirect(url_for('inicio'))

@app.route('/form_registrar_comision', methods=['POST'])
def registrar_comision2(): 
    resultData = procesar_comision(request)
    print(resultData)
    if resultData:
        return redirect(url_for('viewPagosPoliza',id=resultData['Cod_poliza']))


@app.route("/lista-de-usuarios", methods=['GET'])
def usuarios():
    if 'conectado' in session:
        resp_usuariosBD = lista_usuariosBD()
        return render_template('public/usuarios/lista_usuarios.html', resp_usuariosBD=resp_usuariosBD)
    else:
        return redirect(url_for('inicioCpanel'))

    

@app.route("/registrar_comision/<int:id>", methods=['GET'])
def registrar_comision(id):
    if 'conectado' in session: 
        return render_template('public/Comisiones/form_comision.html', cod_renovacion=id)
    else:
        return redirect(url_for('inicioCpanel'))
    


    
    
@app.route("/lista-de-ejecutivos", methods=['GET'])
def ejecutivos():
    if 'conectado' in session:
        ejecutivos = lista_ejecutivosBD()
        return render_template('public/ejecutivo/lista_ejecutivo.html', ejecutivos=ejecutivos)
    else:
        return redirect(url_for('inicioCpanel'))

@app.route("/lista-company", methods=['GET'])
def lista_company():
    if 'conectado' in session:
        companys = sql_lista_company()
        return render_template('public/company/lista_company.html', companys=companys)
    else:
        return redirect(url_for('inicioCpanel'))
    
@app.route("/lista-de-comisiones", methods=['GET'])
def comisiones():
    if 'conectado' in session:
        if session.get("permisos")=="Administracion" or session.get("permisos")=="Gerencia" or session.get("permisos")=="dev":
            comisiones = [] # El datatable ahora carga esto asÃ­ncronamente
            ejecutivos = limpiar_nan(sql_lista_ejecutivo())
            return render_template('public/Comisiones/lista_comisiones.html', comisiones=comisiones, ejecutivos=ejecutivos)
        else:
            flash('No tienes permiso para esta acciÃƒÂ²n', 'error')
            return redirect(url_for('inicio1'))
    else:
        return redirect(url_for('inicioCpanel'))


@app.route("/comisiones-beta", methods=['GET'])
def comisiones_beta():
    if 'conectado' in session:
        if not (session.get("permisos")=="Administracion" or session.get("permisos")=="Gerencia" or session.get("permisos")=="dev"):
            flash('No tienes permiso para esta acciÃƒÂ²n', 'error')
            return redirect(url_for('inicio1'))

        # Si viene el parÃ¡metro reset=1, limpiar la sesiÃ³n para permitir nueva carga
        if request.args.get('reset') == '1':
            session.pop('comisiones_beta_file', None)
            session.pop('comisiones_beta_just_uploaded', None)
            # Redirigir sin el parÃ¡metro para limpiar la URL
            return redirect(url_for('comisiones_beta'))

        # Recuperar nombre de archivo de la sesion
        temp_file = session.get('comisiones_beta_file')
        
        # LÃ³gica para evitar persistencia en refresh manual
        # Si hay archivo pero NO es un 'just_uploaded' Y no estamos pidiendo un bloque especÃ­fico, 
        # asumimos que es un refresh manual o navegaciÃ³n directa y limpiamos.
        just_uploaded = session.pop('comisiones_beta_just_uploaded', False)
        if temp_file and not just_uploaded and request.args.get('block_idx') is None:
            session.pop('comisiones_beta_file', None)
            temp_file = None

        session.pop('comisiones_beta_resultados', None) # Limpiar cache de cÃ¡lculos si existe
        
        # Cargar datos del archivo
        resultados = load_temp_data(temp_file) if temp_file else {}

        # Identificar quÃ© bloque mostrar (para casos multiblock como Mercantil)
        block_idx = request.args.get('block_idx', type=int, default=0)
        
        # Extract data from the results
        all_matched_data = resultados.get('matched_data', [])
        
        # Filtrar por bloque si aplica (Mercantil usa block_id)
        # Si no hay block_id, mostramos todo (comportamiento legacy)
        blocks_found = sorted(list(set(str(row.get('block_id', 0)) for row in all_matched_data)))
        total_blocks = len(blocks_found)
        
        # Si hay bloques definidos, filtramos el actual
        if total_blocks > 1 or (all_matched_data and 'block_id' in all_matched_data[0]):
            matched_data = [row for row in all_matched_data if str(row.get('block_id', 0)) == str(block_idx)]
        else:
            matched_data = all_matched_data

        unmatched_data = resultados.get('unmatched_data', [])
        excel_data = resultados.get('excel_data', [])
        pagos_del_dia = resultados.get('pagos_del_dia', [])
        fecha_seleccionada = resultados.get('fecha_seleccionada', '')
        compania = resultados.get('compania', '')
        metadata = resultados.get('metadata', {})
        # Merge block-specific metadata if it exists
        block_metadata = resultados.get('block_metadata', {})
        if str(block_idx) in block_metadata:
            metadata.update(block_metadata[str(block_idx)])


        return render_template('public/Comisiones/lista_comisiones_beta.html',
                               matched_data=matched_data,
                               unmatched_data=unmatched_data,
                               excel_data=excel_data,
                               pagos_del_dia=pagos_del_dia,
                               fecha_seleccionada=fecha_seleccionada,
                               compania=compania,
                               metadata=metadata,
                               current_block_idx=block_idx,
                               total_blocks=total_blocks)
    else:
        return redirect(url_for('inicio'))

def clean_and_convert_to_float(value, is_percentage=False):
    """
    Limpia una cadena de texto (eliminando sÃ­mbolos de moneda, puntos de miles, y cambiando comas)
    y la convierte a un float. Devuelve 0.0 si la conversiÃ³n falla.
    """
    if value is None:
        return 0.0
    try:
        # 1. Convertir a string y limpiar espacios
        s = str(value).strip()
        # 2. Si hay comas, asumimos que es el separador decimal y reemplazamos por punto.
        #    TambiÃ©n eliminamos los puntos que serÃ­an separadores de miles.
        if ',' in s:
            s = s.replace('.', '').replace(',', '.')
        # 3. Si no hay comas, el punto es el separador decimal. No hacemos nada.
        return float(s)
    except (ValueError, TypeError):
        return 0.0


def clean_and_convert_for_star(value):
    if value is None:
        return 0.0
    try:
        s = str(value).strip()
        # For Star Seguros: ',' is thousands separator, '.' is decimal.
        # So we just need to remove the thousands separator.
        s = s.replace(',', '')
        return float(s)
    except (ValueError, TypeError):
        return 0.0

def find_closest_payment_db(excel_date_str, db_payments, used_ids=None):
    """
    Refined matching: finds the payment with the closest date to the excel record.
    Includes a maximum margin (45 days) and prioritizes "free" installments.
    """
    if not db_payments:
        return None
    
    if used_ids is None:
        used_ids = {} # {id_pago: cantidad_de_veces_asignado}

    try:
        # Normalize excel date
        if isinstance(excel_date_str, (datetime, date)):
            e_date = excel_date_str
        elif isinstance(excel_date_str, str):
            # Try parsing dd/mm/yyyy or yyyy-mm-dd
            date_part = str(excel_date_str).split(' ')[0]
            if '/' in date_part:
                e_date = datetime.strptime(date_part, '%d/%m/%Y').date()
            else:
                e_date = datetime.strptime(date_part, '%Y-%m-%d').date()
        else:
            return db_payments[0]
    except:
        return db_payments[0]
    
    MAX_MARGIN = 45 # MÃ¡ximo de dÃ­as de diferencia permitidos para match automÃ¡tico
    scored_candidates = []
    
    for p in db_payments:
        p_date = p.get('FECHA CUOTA') or p.get('FECHA COBRO RECIBO')
        if not p_date: continue
        
        if isinstance(p_date, str):
            try:
                p_date = datetime.strptime(p_date.split(' ')[0], '%Y-%m-%d').date()
            except: continue
        
        diff = abs((e_date - p_date).days)
        
        # Si la diferencia es excesiva, no lo consideramos un buen candidato automÃ¡tico
        if diff > MAX_MARGIN:
            continue
            
        p_id = str(p.get('NRO. RECIBO'))
        times_used = used_ids.get(p_id, 0)
        
        # LÃ³gica de puntuaciÃ³n: penalizamos cuotas ya "ocupadas" en este lote
        # Una cuota usada recibe una penalizaciÃ³n de "distancia virtual"
        penalty = times_used * 30 
        score = diff + penalty
        
        scored_candidates.append({
            'score': score,
            'payment': p,
            'id': p_id
        })
    
    if not scored_candidates:
        # Si nada entra en el margen, devolvemos None para que el flujo de carga
        # lo trate como "PÃ³liza encontrada pero sin pago coincidente" (POLIZA_OK)
        return None

    # Ordenar por puntuaciÃ³n (menor es mejor)
    scored_candidates.sort(key=lambda x: x['score'])
    best = scored_candidates[0]
    
    # Registrar uso para la siguiente iteraciÃ³n de este mismo bloque/lote
    best_id = best['id']
    used_ids[best_id] = used_ids.get(best_id, 0) + 1
    
    return best['payment']



import re # AsegÃºrate de tener esta importaciÃ³n al inicio de tu archivo

def obtener_pagos_globales(lista_polizas):
    """
    Busca pagos globalmente en la BD para una lista de pÃ³lizas, sin filtrar por fecha.
    Devuelve una lista de diccionarios con la estructura esperada para la comparaciÃ³n.
    """
    if not lista_polizas:
        return []
    
    # Eliminar duplicados y limpiar
    polizas_unicas = list(set([str(p).strip() for p in lista_polizas if p]))
    if not polizas_unicas:
        return []

    data = []
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                # Generar placeholders para la consulta IN
                placeholders = ','.join(['%s'] * len(polizas_unicas))
                
                # Consulta para buscar pagos asociados a las pÃ³lizas
                sql = f"""
                    SELECT 
                        p.cod_poliza AS 'NRO. POLIZA',
                        pg.Cod_pago AS 'NRO. RECIBO',
                        pg.monto AS 'MONTO USD',
                        pg.fecha_pagada AS 'FECHA COBRO RECIBO',
                        pg.fecha AS 'FECHA CUOTA',
                        pg.nro_cuota,
                        pg.tasa AS 'TASA EGRESO',
                        pg.estado,
                        r.Prima AS 'MONTO PRIMA',
                        r.comision,
                        e.cod_ejecutivo,
                        'PAGO DE COMISION' AS 'DESCRIPCION MOVIMIENTO',
                        (pg.monto * pg.tasa) AS 'MONTO PAGADO',
                        (
                            SELECT array_to_string(array_agg(DISTINCT COALESCE(c2.nro_recibo_externo, CAST(c2.Cod_pago AS VARCHAR))), ',')
                            FROM comision c2
                            WHERE (c2.cod_poliza = p.cod_poliza AND c2.cod_poliza IS NOT NULL)
                               OR (c2.Cod_pago = pg.Cod_pago AND c2.Cod_pago != 0)
                        ) as nros_recibos_procesados
                    FROM poliza p
                    JOIN renovacion r ON p.cod_poliza = r.Cod_poliza
                    JOIN pago pg ON r.Cod_renovacion = pg.Cod_renovacion
                    JOIN asegurado a ON a.CI = p.CI_asegurado
                    JOIN ejecutivo e ON a.ejecutivo = e.cod_ejecutivo
                    WHERE p.cod_poliza IN ({placeholders})
                    AND pg.estado = 'PAGADO'
                    GROUP BY p.cod_poliza, pg.Cod_pago, pg.monto, pg.fecha_pagada, pg.fecha, pg.nro_cuota, pg.tasa, pg.estado, r.Prima, r.comision, e.cod_ejecutivo
                """
                cursor.execute(sql, polizas_unicas)
                data = cursor.fetchall()
                
    except Exception as e:
        print(f"Error en obtener_pagos_globales: {e}")
        return []
        
    return data

@app.route('/verificar-poliza-beta', methods=['POST'])
def verificar_poliza_beta():
    if 'conectado' not in session:
        return jsonify({'success': False, 'message': 'SesiÃ³n expirada'}), 401
    
    data = request.json
    poliza_input = data.get('poliza', '').strip()
    if not poliza_input:
        return jsonify({'success': False, 'message': 'NÃºmero de pÃ³liza requerido'})

    # 1. Buscar si la pÃ³liza existe y tiene pagos (como obtener_pagos_globales pero para una sola)
    # y si no tiene pagos, al menos ver si existe la pÃ³liza
    try:
        with connectionBD() as conn:
            with conn.cursor() as cursor:
                # Primero buscamos si existe la pÃ³liza
                sql_p = """
                    SELECT p.cod_poliza, e.cod_ejecutivo, r.comision, a.Nombre, a.Apellido
                    FROM poliza p
                    JOIN asegurado a ON p.CI_asegurado = a.CI
                    JOIN ejecutivo e ON a.ejecutivo = e.cod_ejecutivo
                    LEFT JOIN renovacion r ON p.cod_poliza = r.Cod_poliza
                    WHERE p.cod_poliza = %s
                    ORDER BY r.Cod_renovacion DESC LIMIT 1
                """
                cursor.execute(sql_p, (poliza_input,))
                poliza_match = cursor.fetchone()
                
                if not poliza_match:
                    return jsonify({'success': True, 'exists': False, 'message': 'PÃ³liza no encontrada'})
                
                # Si existe la pÃ³liza, buscamos pagos
                sql_pagos = """
                    SELECT 
                        pg.Cod_pago AS db_cod_pago,
                        pg.recibo AS db_nro_recibo,
                        pg.monto AS db_monto_usd,
                        pg.fecha_pagada,
                        pg.fecha AS db_fecha_cuota,
                        pg.nro_cuota,
                        pg.tasa AS db_tasa,
                        r.Prima AS db_monto_prima,
                        r.comision AS db_comision_porc,
                        e.cod_ejecutivo,
                        (
                            SELECT array_to_string(array_agg(DISTINCT COALESCE(c2.nro_recibo_externo, CAST(c2.Cod_pago AS VARCHAR))), ',')
                            FROM comision c2
                            WHERE (c2.cod_poliza = p.cod_poliza AND c2.cod_poliza IS NOT NULL)
                               OR (c2.Cod_pago = pg.Cod_pago AND c2.Cod_pago != 0)
                        ) as nros_recibos_procesados
                    FROM pago pg
                    JOIN renovacion r ON pg.Cod_renovacion = r.Cod_renovacion
                    JOIN poliza p ON r.Cod_poliza = p.cod_poliza
                    JOIN asegurado a ON p.CI_asegurado = a.CI
                    JOIN ejecutivo e ON a.ejecutivo = e.cod_ejecutivo
                    WHERE p.cod_poliza = %s
                    AND pg.estado = 'PAGADO'
                    GROUP BY pg.Cod_pago
                    ORDER BY pg.fecha_pagada DESC
                """
                cursor.execute(sql_pagos, (poliza_input,))
                pagos = cursor.fetchall()
                
                return jsonify({
                    'success': True,
                    'exists': True,
                    'poliza': poliza_match,
                    'pagos': pagos,
                    'nombre_cliente': f"{poliza_match['Nombre']} {poliza_match['Apellido']}"
                })
                
    except Exception as e:
        print(f"Error en verificar_poliza_beta: {e}")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/upload-comisiones-star', methods=['POST'])
def upload_comisiones_star():
    if 'conectado' not in session:
        return redirect(url_for('inicioCpanel'))

    fecha_pago = request.form.get('fecha_pago')
    if 'file_star' not in request.files: 
        flash('No se seleccionÃ³ ningÃºn archivo.', 'error')
        return redirect(url_for('comisiones_beta'))

    file = request.files['file_star']
    if file.filename == '' or not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
        flash('Archivo no vÃ¡lido.', 'error')
        return redirect(url_for('comisiones_beta'))

    try:
        # 2. Cargar Excel
        workbook = openpyxl.load_workbook(file, data_only=True, read_only=True) 
        sheet = workbook.active

        extracted_data = []
        header_map = {}
        in_data_block = False
        possible_headers = {
            'PÃ³liza': ['PÃ³liza', 'NUMERO POLIZA', 'POLIZA'],
            'Recibo': ['Recibo', 'NUMERO RECIBO', 'RECIBO', 'FACTURA'],
            'Tipo Movimiento': ['Tipo Movimiento', 'MOVIMIENTO', 'TIPO'],
            'Moneda Recibo': ['Moneda Recibo', 'MONEDA'],
            'Mto. Prima': ['MTO.', 'Mto.', 'Mto. Prima', 'Mto.Prima', 'Mto Prima', 'PRIMA', 'MONTO PRIMA', 'MTO. PRIMA USD'],
            'Mto. Pagar ComisiÃ³n': ['Mto. Pagar ComisiÃ³n', 'Mto.Pagar ComisiÃ³n', 'Mto. ComisiÃ³n', 'COMISION $', 'COMISION', 'MONTO COMISION'],
            'Tasa Cambio': ['Tasa Cambio', 'TASA', 'TASA CAMBIO'],
            'Cliente': ['Cliente', 'ASEGURADO', 'NOMBRE CLIENTE'],
            'Fecha Pago ComisiÃ³n': ['Fecha Pago ComisiÃ³n', 'FECHA PAGO', 'FECHA']
        }
        extracted_metadata = {
            'egreso': '',
            'referencia': '',
            'retenciones': 0.0
        }

        for row in sheet.iter_rows(values_only=True):
            row_values = [str(cell).strip() if cell is not None else "" for cell in row]
            
            if not in_data_block:
                if any('PÃ³liza' in v for v in row_values) and any('Recibo' in v for v in row_values):
                    in_data_block = True
                    for header_key, candidates in possible_headers.items():
                        for i, val in enumerate(row_values):
                            if val and val != "None":
                                clean_h = re.sub(r'\s+', ' ', val.replace('\n', ' ')).strip().upper()
                                for c in candidates:
                                    # Para "MTO." queremos que si la celda empieza por eso o es igual, lo tome.
                                    if clean_h == c.upper() or clean_h.startswith(c.upper()):
                                        header_map[header_key] = i
                                        break
                                if header_key in header_map:
                                    break
                    continue
                else:
                    continue

            if in_data_block:
                idx_poliza = header_map.get('PÃ³liza')
                idx_recibo = header_map.get('Recibo')
                idx_tipo = header_map.get('Tipo Movimiento')
                idx_moneda = header_map.get('Moneda Recibo')
                
                # Check for metadata if poliza is missing or row looks like footer
                val_poliza = ""
                if idx_poliza is not None:
                    val_poliza = str(row[idx_poliza]).strip()
                
                # Metadata Extraction Logic
                # Only check if principal columns are missing or known metadata keys are present
                row_str_upper = ''.join([str(c).upper() for c in row if c])
                
                # Egreso check
                if "EGRESO" in row_str_upper or "REFERENCIA" in row_str_upper or "RETENCIONES" in row_str_upper:
                    # Scan row to find key-value pairs
                    for i, cell in enumerate(row):
                        cell_str = str(cell).strip()
                        cell_upper = cell_str.upper()
                        if not cell_str: continue
                        
                        # Logic for EGRESO and REFERENCIA disabled by user request (manual entry)
                        # if "EGRESO" in cell_upper:
                        #      # E.g. "Egreso: 123" or "Egreso" then "123"
                        #      parts = cell_str.split(':')
                        #      if len(parts) > 1 and parts[1].strip():
                        #          extracted_metadata['egreso'] = parts[1].strip()
                        #      elif i + 1 < len(row):
                        #          extracted_metadata['egreso'] = str(row[i+1]).strip()
                        
                        # if "REFERENCIA" in cell_upper:
                        #      # E.g. "Referencia de 123" or "Referencia de" then "123"
                        #      clean_ref = cell_str.replace("Referencia de", "").strip()
                        #      if clean_ref:
                        #          extracted_metadata['referencia'] = clean_ref
                        #      elif i + 1 < len(row):
                        #          extracted_metadata['referencia'] = str(row[i+1]).strip()
                                 
                        if "RETENCIONES" in cell_upper:
                             # 1. Try itself
                             candidate_str = re.sub(r'(?i)retenciones|:', '', cell_str).strip()
                             try_val = clean_and_convert_for_star(candidate_str)
                             if try_val > 0:
                                 extracted_metadata['retenciones'] = try_val
                             else:
                                 # 2. Look ahead up to 5 cells for a value
                                 for offset in range(1, 6):
                                     if i + offset < len(row):
                                         next_val = row[i+offset]
                                         if next_val:
                                             try_val_next = clean_and_convert_for_star(next_val)
                                             if try_val_next > 0:
                                                 extracted_metadata['retenciones'] = try_val_next
                                                 break # Found it

                if not val_poliza or val_poliza == "None": continue

                # Continue with normal data extraction...
                val_recibo = ""
                if idx_recibo is not None:
                     val_recibo = str(row[idx_recibo]).strip()


                tipo_mov = str(row[idx_tipo]).upper() if idx_tipo is not None else ""
                #if "CONVENIO" in tipo_mov: continue

                # Detectar Moneda
                moneda_val = "DOLARES"
                if idx_moneda is not None:
                    raw_moneda = str(row[idx_moneda]).upper()
                    if "BOL" in raw_moneda or "BS" in raw_moneda:
                        moneda_val = "BOLIVARES"

                data_dict = {
                    'NRO. POLIZA': re.sub(r'[^a-zA-Z0-9]', '', val_poliza),
                    'NRO. RECIBO': re.sub(r'[^a-zA-Z0-9]', '', val_recibo),
                    'NOMBRE CLIENTE': str(row[header_map.get('Cliente')]).strip() if header_map.get('Cliente') is not None else "",
                    'MONTO PRIMA': clean_and_convert_for_star(row[header_map.get('Mto. Prima')]) if header_map.get('Mto. Prima') is not None else 0,
                    'MONTO COMISION USD': clean_and_convert_for_star(row[header_map.get('Mto. Pagar ComisiÃ³n')]) if header_map.get('Mto. Pagar ComisiÃ³n') is not None else 0,
                    'TASA CAMBIO': clean_and_convert_for_star(row[header_map.get('Tasa Cambio')]) if header_map.get('Tasa Cambio') is not None else 0,
                    'MONEDA': moneda_val
                }

                fch_val = row[header_map.get('Fecha Pago ComisiÃ³n')] if 'Fecha Pago ComisiÃ³n' in header_map else None
                if isinstance(fch_val, datetime.datetime):
                    data_dict['FECHA COBRO RECIBO'] = fch_val.strftime('%Y-%m-%d')
                else:
                    data_dict['FECHA COBRO RECIBO'] = str(fch_val).strip()

                extracted_data.append(data_dict)

        if not in_data_block:
            flash('El formato del archivo Excel es incorrecto. AsegÃºrese de que el archivo contenga los encabezados "PÃ³liza" y "Recibo".', 'error')
            return redirect(url_for('comisiones_beta'))

        # --- NUEVO: Buscar pagos globales por pÃ³liza ---
        lista_polizas = [row.get('NRO. POLIZA') for row in extracted_data if row.get('NRO. POLIZA')]
        predefined_data = obtener_pagos_globales(lista_polizas)

        # 5. COMPARACIÃ“N (Cruce con Base de Datos)
        all_data = []
        predefined_copy = [item.copy() for item in predefined_data]
        used_ids_tracker = {}

        for excel_row in extracted_data:
            found_match = None
            poliza_key = str(excel_row.get('NRO. POLIZA', '')).strip()
            
            # Filter DB payments for this specific policy
            matches_for_policy = []
            for db_row in predefined_copy:
                db_poliza = re.sub(r'[^a-zA-Z0-9]', '', str(db_row.get('NRO. POLIZA', '')))
                if db_poliza == poliza_key:
                    matches_for_policy.append(db_row)
            
            if matches_for_policy:
                # Find closest by date (passing tracker to avoid duplicate assignments)
                found_match = find_closest_payment_db(excel_row.get('FECHA COBRO RECIBO'), matches_for_policy, used_ids=used_ids_tracker)
                    
            # FALLBACK: Validar por existencia de pÃ³liza si no hay pago
            if not found_match and poliza_key:
                with connectionBD() as conn_fb:
                    with conn_fb.cursor() as cur_fb:
                        sql_fb = """
                            SELECT p.cod_poliza, e.cod_ejecutivo, r.comision
                            FROM poliza p
                            JOIN asegurado a ON p.CI_asegurado = a.CI
                            JOIN ejecutivo e ON a.ejecutivo = e.cod_ejecutivo
                            LEFT JOIN renovacion r ON p.cod_poliza = r.Cod_poliza
                            WHERE p.cod_poliza = %s
                            ORDER BY r.Cod_renovacion DESC LIMIT 1
                        """
                        cur_fb.execute(sql_fb, (poliza_key,))
                        match_fb = cur_fb.fetchone()
                        if match_fb:
                            found_match = {
                                'NRO. POLIZA': match_fb['cod_poliza'],
                                'NRO. RECIBO': 'POLIZA_OK',
                                'MONTO USD': 0,
                                'TASA EGRESO': 1.0,
                                'MONTO PAGADO': 0,
                                'cod_ejecutivo': match_fb['cod_ejecutivo'],
                                'comision': match_fb['comision']
                            }

            if found_match:
                found_match['_used'] = True
                
                moneda = excel_row.get('MONEDA', 'DOLARES')
                # raw_monto is the commission amount from the excel file
                raw_monto = excel_row.get('MONTO COMISION USD', 0)
                tasa = excel_row.get('TASA CAMBIO', 0)
                
                if moneda == 'BOLIVARES':
                    ex_com = 0 
                    monto_pagado = raw_monto
                else:
                    ex_com = raw_monto
                    # When USD, we keep monto_pagado_excel as 0 to avoid double counting in mixed totals or confusing the backend
                    # The backend will handle 'monto_d' from ex_com.
                    if tasa == 0:
                        # Si es USD y no hay tasa, la dejamos en 0. 
                        # Evitamos forzar 1 para que el frontend no calcule montos en Bs irreales.
                        tasa = 0
                    monto_pagado = 0  # We set this to 0 for Dolares so it doesn't show up as a BS payment in the frontend summary
                
                db_com = clean_and_convert_to_float(found_match.get('MONTO COMISION USD_DB', 0))
                predefined_comision_porcentaje = clean_and_convert_to_float(found_match.get('comision'))
                predefined_usd = clean_and_convert_to_float(found_match.get('MONTO USD'))

                nros_recibos_procesados = found_match.get('nros_recibos_procesados')
                procesado_status = 'por procesar'
                if nros_recibos_procesados and excel_row['NRO. RECIBO'] in nros_recibos_procesados.split(','):
                    procesado_status = 'procesada'

                excel_row.update({
                    'is_match': True,
                    'procesado': procesado_status,
                    'moneda': moneda,
                    'monto_usd_excel': ex_com,
                    'monto_usd_predefined': predefined_usd,
                    'variacion_usd': ex_com - db_com,
                    'comision_calculada_predefined': predefined_usd * predefined_comision_porcentaje,
                    'predefined_comision_porcentaje': predefined_comision_porcentaje,
                    'cod_ejecutivo': found_match.get('cod_ejecutivo'),
                    'db_cod_pago': found_match.get('NRO. RECIBO'),
                    'nro_cuota': found_match.get('nro_cuota'),
                    'tasa_excel': tasa,
                    'monto_pagado_excel': monto_pagado,
                    'DESCRIPCION MOVIMIENTO': excel_row.get('DESCRIPCION MOVIMIENTO', 'PAGO DE COMISION'),
                    'monto_prima': excel_row.get('MONTO PRIMA', 0),
                    'bono': 0
                })
                all_data.append(excel_row)
            else:
                moneda = excel_row.get('MONEDA', 'DOLARES')
                raw_monto = excel_row.get('MONTO COMISION USD', 0)
                tasa = excel_row.get('TASA CAMBIO', 0)
                
                if moneda == 'BOLIVARES':
                    ex_com = 0
                    monto_pagado = raw_monto
                else:
                    ex_com = raw_monto
                    if tasa == 0:
                        tasa = 0
                    monto_pagado = 0 # Set to 0 for Dolares

                excel_row['is_match'] = False
                excel_row['procesado'] = 'no encontrada'
                excel_row['moneda'] = moneda
                excel_row['monto_usd_excel'] = ex_com
                excel_row['tasa_excel'] = tasa
                excel_row['monto_pagado_excel'] = monto_pagado
                excel_row['monto_usd_predefined'] = 0.0
                excel_row['variacion_usd'] = excel_row['monto_usd_excel']
                excel_row['comision_calculada_predefined'] = 0.0
                excel_row['predefined_comision_porcentaje'] = 0.0
                excel_row['cod_ejecutivo'] = None
                excel_row['db_cod_pago'] = None
                excel_row['DESCRIPCION MOVIMIENTO'] = excel_row.get('DESCRIPCION MOVIMIENTO', 'PAGO DE COMISION')
                excel_row['monto_prima'] = excel_row.get('MONTO PRIMA', 0)
                excel_row['bono'] = 0.0
                all_data.append(excel_row)

        # --- Add Retention Row if exists ---
        if extracted_metadata.get('retenciones', 0) > 0:
            retention_amt = extracted_metadata['retenciones']
            ret_row = {
                'NRO. POLIZA': 'RETENCIONES',
                'NRO. RECIBO': 'RETENCION_MANUAL',
                'NOMBRE CLIENTE': 'RETENCIONES DE LEY',
                'FECHA COBRO RECIBO': date.today().strftime('%Y-%m-%d'),
                'matched_db_pago': None,
                'is_match': True, 
                'procesado': 'procesada', 
                'db_cod_pago': 'RETENCION', 
                'DESCRIPCION MOVIMIENTO': 'RETENCIONES MANUAL',
                'moneda': 'DOLARES', 
                'monto_usd_excel': -retention_amt,
                'tasa_excel': 0,
                'monto_pagado_excel': 0, 
                'monto_prima': 0,
                'bono': 0
            }
            all_data.append(ret_row)

        # 6. Guardar y Retornar
        results = {
            'matched_data': all_data,
            'unmatched_data': [],
            'excel_data': extracted_data,
            'predefined_data_for_table': predefined_data,
            'fecha_seleccionada': fecha_pago if fecha_pago else datetime.date.today().strftime('%Y-%m-%d'),
            'compania': 'STAR',
            'metadata': extracted_metadata
        }
        
        # Guardar en archivo temporal en lugar de session
        temp_filename = save_temp_data(results)
        session['comisiones_beta_file'] = temp_filename
        session['comisiones_beta_just_uploaded'] = True

        flash(f'Procesado: {len(all_data)} registros en Excel.', 'success')
        return redirect(url_for('comisiones_beta'))

    except Exception as e:
        flash(f'Error crÃ­tico al procesar OceÃ¡nica: {e}', 'error')
        traceback.print_exc()
        return redirect(url_for('comisiones_beta'))






@app.route('/upload-comisiones-caracas', methods=['POST'])
def upload_comisiones_caracas():
    if 'conectado' not in session:
        return redirect(url_for('inicioCpanel'))

    if 'file_caracas' not in request.files:
        flash('No se seleccionÃ³ ningÃºn archivo.', 'error')
        return redirect(url_for('comisiones_beta'))

    file = request.files['file_caracas']
    if file.filename == '':
        flash('No se seleccionÃ³ ningÃºn archivo.', 'error')
        return redirect(url_for('comisiones_beta'))

    if not file.filename.lower().endswith('.pdf'):
        flash('Caracas: Formato no vÃ¡lido. Use un PDF (.pdf)', 'error')
        return redirect(url_for('comisiones_beta'))

    # â”€â”€ PDF parser for Seguros Caracas "Comprobante de LiquidaciÃ³n en Bs." â”€
    try:
        import pdfplumber, io as _io

        file_bytes = file.read()
        extracted_data = []
        fecha_pago_global = None

        # Number parser: Venezuelan format 1.234,56 â†’ 1234.56
        # Negative values between parens: (38,61) â†’ -38.61
        def parse_car_num(s):
            if not s:
                return 0.0
            s = str(s).strip()
            negative = False
            if s.startswith('(') and s.endswith(')'):
                negative = True
                s = s[1:-1]
            # Remove thousands dots, replace decimal comma
            s = s.replace('.', '').replace(',', '.')
            try:
                val = float(s)
            except ValueError:
                return 0.0
            return -val if negative else val

        # Regex patterns
        DATE_RE = re.compile(r'\b(\d{2}[/-]\d{2}[/-]\d{4})\b')
        # A data row starts with a date dd/mm/yyyy, then a receipt number like 1-56-... or pure digits
        # Columns (after date): NRO_RECIBO  NRO_POLIZA  NOMBRE...  MONTO_PRIMA  MONTO_USD  TASA_EGRESO  DESCRIPCION  MONTO_PAGADO
        # MONTO_PAGADO is always the last numeric field (Bs amount, possibly negative)
        # DESCRIPCION can be multi-word: PAGO DE COMISION, CONVENIO RAMOS, EXTORNOS AUTOMATICOS
        DATA_ROW_RE = re.compile(
            r'^(\d{2}[/-]\d{2}[/-]\d{4})\s+'          # fecha cobro
            r'([\d\-]+)\s+'                             # nro. recibo (e.g. 1-56-2990061)
            r'(\d+)\s+'                                 # nro. poliza
            r'(.+?)\s+'                                 # nombre cliente (non-greedy)
            r'([-\d,.\(\)]+)\s+'                        # monto prima (puede ser negativo)
            r'([-\d,.\(\)]+)\s+'                        # monto usd (puede ser negativo)
            r'([\d,.]+)\s+'                             # tasa egreso (siempre positivo)
            r'([A-Z][A-Z ]+?)\s+'                       # DESCRIPCION MOVIMIENTO
            r'([-\d,.\(\)]+)\s*$'                       # monto pagado (puede ser - o (n))
        )

        # Continuation lines: only a name fragment (no date, no numbers leading)
        CONTINUATION_RE = re.compile(r'^([A-Z][A-Z\s]+)$')

        VALID_DESC = {'PAGO DE COMISION', 'CONVENIO RAMOS', 'EXTORNOS AUTOMATICOS',
                      'PAGO COMISION', 'EXTORNO'}

        pending_row = None  # holds a row waiting for its continuation line

        with pdfplumber.open(_io.BytesIO(file_bytes)) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if not text:
                    continue

                for line in text.split('\n'):
                    line = line.strip()
                    if not line:
                        continue
                    upper = line.upper()

                    # Extract notification/payment date from header
                    if not fecha_pago_global:
                        # "FECHA DE NOTIFICACIÃ“N\n29-10-2025" or inline
                        if 'NOTIFICACI' in upper or 'FECHA DE PAGO' in upper:
                            dm = DATE_RE.search(line)
                            if dm:
                                raw_d = dm.group(1).replace('/', '-')
                                try:
                                    fecha_pago_global = datetime.datetime.strptime(raw_d, '%d-%m-%Y').strftime('%Y-%m-%d')
                                except Exception:
                                    pass
                        else:
                            # Standalone date line like "29-10-2025"
                            if re.match(r'^\d{2}[-/]\d{2}[-/]\d{4}$', line):
                                try:
                                    fecha_pago_global = datetime.datetime.strptime(
                                        line.replace('/', '-'), '%d-%m-%Y'
                                    ).strftime('%Y-%m-%d')
                                except Exception:
                                    pass

                    # Flush pending row if this line is a continuation (name overflow)
                    if pending_row is not None:
                        if not DATE_RE.match(line) and CONTINUATION_RE.match(upper):
                            # append the extra name fragment
                            pending_row['NOMBRE CLIENTE'] += ' ' + line.strip()
                            extracted_data.append(pending_row)
                            pending_row = None
                            continue
                        else:
                            # not a continuation â†’ flush as-is
                            extracted_data.append(pending_row)
                            pending_row = None

                    # Try to match a data row
                    m = DATA_ROW_RE.match(line)
                    if not m:
                        continue

                    fecha_str  = m.group(1).replace('/', '-')
                    nro_recibo = m.group(2).strip()
                    nro_poliza = m.group(3).strip()
                    nombre     = m.group(4).strip()
                    prima      = parse_car_num(m.group(5))
                    monto_usd  = parse_car_num(m.group(6))
                    tasa       = parse_car_num(m.group(7))
                    descripcion = m.group(8).strip()
                    monto_pagado = parse_car_num(m.group(9))

                    try:
                        fecha_cobro = datetime.datetime.strptime(fecha_str, '%d-%m-%Y').strftime('%Y-%m-%d')
                    except Exception:
                        fecha_cobro = fecha_pago_global or date.today().strftime('%Y-%m-%d')

                    # Only keep PAGO DE COMISION / EXTORNOS (skip CONVENIO RAMOS etc. if desired)
                    # For now include all so user can decide
                    row_dict = {
                        'DESCRIPCION MOVIMIENTO': descripcion,
                        'FECHA COBRO RECIBO':     fecha_cobro,
                        'NRO. POLIZA':            nro_poliza,
                        'NRO. RECIBO':            nro_recibo,
                        'NOMBRE CLIENTE':         nombre,
                        'MONTO PRIMA':            prima,
                        'MONTO USD':              monto_usd,
                        'TASA EGRESO':            tasa,
                        'MONTO PAGADO':           monto_pagado,
                        # Internal processing fields
                        'monto_usd_excel':        monto_usd,
                        'monto_pagado_excel':     monto_pagado,
                        'tasa_excel':             tasa,
                        'moneda':                 'BOLIVARES',
                        'bono':                   0.0,
                    }

                    # Check if the name looks truncated (no digit ending â†’ may continue on next line)
                    # We'll use a pending mechanism only if the regex captured the description correctly
                    # For safety: store and process on next iteration
                    # Heuristic: if name ends with a word (not number/special), it MIGHT continue
                    extracted_data.append(row_dict)
                    print(f"[CARACAS] poliza={nro_poliza}, recibo={nro_recibo}, usd={monto_usd}, bs={monto_pagado}, desc={descripcion}")

            # flush last pending row
            if pending_row is not None:
                extracted_data.append(pending_row)

        # Filter empty/total rows
        extracted_data = [r for r in extracted_data if r.get('NRO. POLIZA')]

        if not extracted_data:
            flash(
                'Caracas PDF: No se encontraron filas de datos. '
                'Verifique que sea el "Comprobante de LiquidaciÃ³n" correcto.',
                'error'
            )
            return redirect(url_for('comisiones_beta'))

        print(f"[CARACAS] Total extraÃ­dos: {len(extracted_data)}")

        # â”€â”€ DB Matching â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        lista_polizas   = [r['NRO. POLIZA'] for r in extracted_data if r.get('NRO. POLIZA')]
        predefined_data = obtener_pagos_globales(lista_polizas)
        matched_data    = []
        used_ids_tracker = {}

        for excel_row in extracted_data:
            excel_policy = excel_row.get('NRO. POLIZA', '')
            candidates_db = [
                d for d in predefined_data
                if re.sub(r'[^a-zA-Z0-9]', '', str(d.get('NRO. POLIZA', ''))) == excel_policy
            ]
            found_match = find_closest_payment_db(
                excel_row.get('FECHA COBRO RECIBO'), candidates_db, used_ids=used_ids_tracker
            )

            # Fallback: policy exists in DB but no pending payment found
            if not found_match and excel_policy:
                with connectionBD() as conn_fb:
                    with conn_fb.cursor() as cur_fb:
                        cur_fb.execute(
                            "SELECT p.cod_poliza, e.cod_ejecutivo, r.comision "
                            "FROM poliza p JOIN asegurado a ON p.CI_asegurado = a.CI "
                            "JOIN ejecutivo e ON a.ejecutivo = e.cod_ejecutivo "
                            "LEFT JOIN renovacion r ON p.cod_poliza = r.Cod_poliza "
                            "WHERE p.cod_poliza = %s ORDER BY r.Cod_renovacion DESC LIMIT 1",
                            (excel_policy,)
                        )
                        match_fb = cur_fb.fetchone()
                        if match_fb:
                            found_match = {
                                'NRO. POLIZA': match_fb['cod_poliza'],
                                'NRO. RECIBO': 'POLIZA_OK',
                                'MONTO USD': 0, 'TASA EGRESO': 1.0, 'MONTO PAGADO': 0,
                                'cod_ejecutivo': match_fb['cod_ejecutivo'],
                                'comision': match_fb['comision']
                            }

            if found_match:
                excel_row.update({
                    'is_match': True,
                    'procesado': 'por procesar',
                    'db_cod_pago': found_match.get('NRO. RECIBO'),
                    'cod_ejecutivo': found_match.get('cod_ejecutivo'),
                    'nro_cuota': found_match.get('nro_cuota'),
                    'monto_usd_predefined': clean_and_convert_to_float(found_match.get('MONTO USD')),
                    'predefined_comision_porcentaje': clean_and_convert_to_float(found_match.get('comision')),
                })
            else:
                # No match at all â†’ mark and insert as poliza pendiente
                excel_row.update({
                    'is_match': False,
                    'procesado': 'no encontrada',
                    'db_cod_pago': None,
                    'cod_ejecutivo': None,
                    'nro_cuota': None,
                    'monto_usd_predefined': 0.0,
                    'predefined_comision_porcentaje': 0.0,
                })
            matched_data.append(excel_row)

        n_matched = sum(1 for r in matched_data if r.get('is_match'))
        n_total   = len(matched_data)

        results = {
            'matched_data': matched_data,
            'compania': 'CARACAS',
            'fecha_seleccionada': fecha_pago_global or date.today().strftime('%Y-%m-%d'),
        }
        temp_filename = save_temp_data(results)
        session['comisiones_beta_file'] = temp_filename
        session['comisiones_beta_just_uploaded'] = True

        flash(
            f'Caracas PDF: {n_total} registros procesados '
            f'â€” {n_matched} emparejados, {n_total - n_matched} no encontrados en BD.',
            'success'
        )
        return redirect(url_for('comisiones_beta'))

    except Exception as e:
        flash(f'Caracas PDF [{type(e).__name__}]: {e}', 'error')
        traceback.print_exc()
        return redirect(url_for('comisiones_beta'))



@app.route('/upload-comisiones-piramide', methods=['POST'])
def upload_comisiones_piramide():
    if 'conectado' not in session:
        return redirect(url_for('inicioCpanel'))

    if 'file_piramide' not in request.files:
        flash('No se seleccionÃ³ ningÃºn archivo.', 'error')
        return redirect(url_for('comisiones_beta'))

    file = request.files['file_piramide']

    if file.filename == '':
        flash('No se seleccionÃ³ ningÃºn archivo.', 'error')
        return redirect(url_for('comisiones_beta'))

    if not file.filename.lower().endswith('.pdf'):
        flash('PirÃ¡mide: Formato no vÃ¡lido. Use un PDF (.pdf)', 'error')
        return redirect(url_for('comisiones_beta'))

    # â”€â”€ PDF parser for PirÃ¡mide "Listado de Comisiones Liquidadas" â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    try:
        import pdfplumber
        import io

        file_bytes = file.read()
        extracted_data = []
        fecha_str_global = None

        # Known concepts to extract (only COMISIONES, skip GASTOS, ISLR, IMP.MUN)
        VALID_CONCEPTS = ('COMISIONES', 'GASTOS DE COBRANZA')

        def parse_piramide_num(val_str):
            """Convert numbers from PirÃ¡mide PDF.
            Venezuelan format uses comma as decimal and period as thousands separator: 55.397,57 â†’ 55397.57
            BUT the %Comis column uses English format: 20.00 â†’ 20.00
            Heuristic: if value has comma, it is Venezuelan; if only dot (no comma), it is English.
            """
            if not val_str:
                return 0.0
            s = str(val_str).replace('(', '-').replace(')', '').strip()
            if ',' in s:
                # Venezuelan format: remove dot thousands separator, replace comma with dot
                s = s.replace('.', '').replace(',', '.')
            # else: English/plain format â€” use as-is (e.g. 20.00, 457.08 already correct after comma removal)
            try:
                return float(s)
            except ValueError:
                return 0.0

        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if not text:
                    continue

                for line in text.split('\n'):
                    line = line.strip()
                    if not line:
                        continue

                    # Extract global date from header (use end date of period)
                    if not fecha_str_global and 'DEL ' in line.upper() and 'AL ' in line.upper():
                        date_matches = re.findall(r'(\d{2}/\d{2}/\d{2,4})', line)
                        if date_matches:
                            raw_d = date_matches[-1]
                            try:
                                if len(raw_d.split('/')[-1]) == 2:
                                    raw_d = raw_d[:-2] + '20' + raw_d[-2:]
                                fecha_str_global = datetime.datetime.strptime(raw_d, '%d/%m/%Y').strftime('%Y-%m-%d')
                            except Exception:
                                pass

                    # Skip non-data lines
                    upper_line = line.upper()
                    if any(skip in upper_line for skip in [
                        'TOTAL GENERAL', 'TOTAL NETO', 'TOTALES', 'RESUMEN',
                        'CONCEPTO', 'LISTADO', 'SUCURSAL', 'INTERMEDIARIO',
                        'MONEDA:', 'FECHA :', 'ACSEL', 'APLICACI',
                        'LIQINTER', 'I.S.L.R', 'IMP.MUN',
                        'PÃGINA', 'PAGINA'
                    ]):
                        continue

                    # Check if line starts with a valid concept
                    matched_concept = None
                    for concept in VALID_CONCEPTS:
                        if upper_line.startswith(concept):
                            matched_concept = concept
                            break

                    if not matched_concept:
                        continue

                    # Parse the data row
                    # Format after concept: DD/MM/YYYY M_POL PRIMA COM_BASE M_OPER %COMMON COM_OPER ASIGN TASA PROD RAMO POLIZA RECIBO N_ORD CHEQUE REL CLIENTE... %COMIS
                    rest = line[len(matched_concept):].strip()

                    # Extract date
                    date_match = re.match(r'(\d{2}/\d{2}/\d{4})', rest)
                    if not date_match:
                        print(f"[PIRAMIDE PDF] No date found in: {rest[:50]}")
                        continue
                    fecha_raw = date_match.group(1)
                    try:
                        fecha_row = datetime.datetime.strptime(fecha_raw, '%d/%m/%Y').strftime('%Y-%m-%d')
                    except Exception:
                        fecha_row = fecha_str_global or date.today().strftime('%Y-%m-%d')
                    rest = rest[len(fecha_raw):].strip()

                    # Tokenize remaining fields
                    tokens = rest.split()
                    if len(tokens) < 14:
                        print(f"[PIRAMIDE PDF] Skipping short row ({len(tokens)} tokens): {tokens}")
                        continue

                    try:
                        # Token layout:
                        # [0] M.Pol (DL/BS)      [1] Prima Pol    [2] Comision Base
                        # [3] M.Oper (BS/DL)     [4] %ComMon      [5] Com.Oper.Mon.Pol
                        # [6] Asignaciones        [7] Tasa         [8] Prod.
                        # [9] Ramo                [10] PÃ³liza      [11] Recibo
                        # [12] N.Ord.Pago         [13] Cheque      [14] Rel.Ing./Egre
                        # [15..n-2] Cliente       [-1] %Comis

                        m_pol    = tokens[0].upper()   # DL or BS (currency of policy)
                        m_oper   = tokens[3].upper()   # BS or DL (currency of operation)
                        prima    = parse_piramide_num(tokens[1])
                        com_base = parse_piramide_num(tokens[2])
                        asign    = parse_piramide_num(tokens[6])
                        tasa     = parse_piramide_num(tokens[7])
                        poliza   = re.sub(r'[^a-zA-Z0-9]', '', tokens[10])
                        recibo   = re.sub(r'[^a-zA-Z0-9]', '', tokens[11])
                        porc_com = parse_piramide_num(tokens[-1])

                        # Client name: tokens[15] to tokens[-2]
                        if len(tokens) > 16:
                            cliente = ' '.join(tokens[15:-1])
                        elif len(tokens) > 15:
                            cliente = tokens[15]
                        else:
                            cliente = 'S/N'

                        # Determine moneda and amounts:
                        # m_pol=DL, m_oper=BS â†’ policy in USD, paid in Bs
                        #   â†’ monto_usd = com_base (USD commission)
                        #   â†’ monto_bs = asign (Bs assignment)
                        # m_pol=BS, m_oper=BS â†’ all in Bs
                        #   â†’ monto_usd = 0
                        #   â†’ monto_bs = asign
                        if m_pol == 'DL' and m_oper == 'BS':
                            moneda_det = 'BOLIVARES'
                            monto_usd  = com_base   # commission in USD (original currency)
                            monto_bs   = asign       # amount assigned in Bs
                        elif m_oper == 'DL':
                            moneda_det = 'DOLARES'
                            monto_usd  = asign
                            monto_bs   = asign * tasa if tasa else 0.0
                        else:
                            moneda_det = 'BOLIVARES'
                            monto_usd  = 0.0
                            monto_bs   = asign

                        if not poliza:
                            continue

                        data_dict = {
                            'DESCRIPCION MOVIMIENTO': matched_concept,
                            'FECHA COBRO RECIBO':     fecha_row,
                            'NRO. POLIZA':            poliza,
                            'NRO. RECIBO':            recibo,
                            'NOMBRE CLIENTE':         cliente,
                            'MONTO PRIMA':            prima,
                            'monto_usd_excel':        monto_usd,
                            'monto_pagado_excel':     monto_bs,
                            'tasa_excel':             tasa,
                            'moneda':                 moneda_det,
                            'porcentaje_comision_pdf': porc_com,
                        }
                        extracted_data.append(data_dict)
                        print(f"[PIRAMIDE PDF] Row: poliza={poliza}, recibo={recibo}, monto_usd={monto_usd}, monto_bs={monto_bs}, tasa={tasa}")

                    except (IndexError, ValueError) as parse_err:
                        print(f"[PIRAMIDE PDF] Parse error on row: {line!r} â†’ {parse_err}")
                        continue

        if not extracted_data:
            flash(
                'PirÃ¡mide PDF: No se encontraron filas de COMISIONES en el archivo. '
                'Verifique que sea el "Listado de Comisiones Liquidadas" correcto.',
                'error'
            )
            return redirect(url_for('comisiones_beta'))

        print(f"[PIRAMIDE PDF] Extracted {len(extracted_data)} rows. Sample: {extracted_data[0]}")

        # Obtener todas las pólizas de Pirámide para hacer el mapping de terminación
        piramide_policies_map = {}
        try:
            with connectionBD() as conn_map:
                with conn_map.cursor() as cur_map:
                    cur_map.execute(
                        "SELECT p.cod_poliza FROM poliza p "
                        "JOIN compania c ON p.Cod_compania = c.Cod_compania "
                        "WHERE c.Nombre LIKE '%PIRAMIDE%'"
                      )
                    for row_pol in cur_map.fetchall():
                        full_code = str(row_pol['cod_poliza']).strip()
                        clean_full = re.sub(r'[^a-zA-Z0-9]', '', full_code)
                        piramide_policies_map[clean_full] = full_code
        except Exception as e:
            print(f"[PIRAMIDE PDF] Error cargando pólizas de Pirámide: {e}")

        # Reemplazar los códigos cortos del PDF por los códigos completos de la BD usando suffix matching
        for excel_row in extracted_data:
            excel_policy = excel_row.get('NRO. POLIZA', '')
            if excel_policy:
                excel_policy_clean = re.sub(r'[^a-zA-Z0-9]', '', str(excel_policy))
                for db_clean, db_full in piramide_policies_map.items():
                    if db_clean.endswith(excel_policy_clean):
                        excel_row['NRO. POLIZA'] = db_full
                        print(f"[PIRAMIDE PDF] Mapped short policy {excel_policy} to full policy {db_full}")
                        break

        # â”€â”€ DB Matching (same pattern as Venezuela / Universitas) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        lista_polizas = [r['NRO. POLIZA'] for r in extracted_data if r.get('NRO. POLIZA')]
        predefined_data = obtener_pagos_globales(lista_polizas)
        matched_data = []
        used_ids_tracker = {}

        for excel_row in extracted_data:
            excel_policy = excel_row.get('NRO. POLIZA', '')
            candidates_db = [
                d for d in predefined_data
                if re.sub(r'[^a-zA-Z0-9]', '', str(d.get('NRO. POLIZA', ''))) == excel_policy
            ]
            found_match = find_closest_payment_db(
                excel_row.get('FECHA COBRO RECIBO'), candidates_db, used_ids=used_ids_tracker
            )

            # Fallback: verify policy existence in DB
            if not found_match and excel_policy:
                with connectionBD() as conn_fb:
                    with conn_fb.cursor() as cur_fb:
                        cur_fb.execute(
                            "SELECT p.cod_poliza, e.cod_ejecutivo, r.comision "
                            "FROM poliza p JOIN asegurado a ON p.CI_asegurado = a.CI "
                            "JOIN ejecutivo e ON a.ejecutivo = e.cod_ejecutivo "
                            "LEFT JOIN renovacion r ON p.cod_poliza = r.Cod_poliza "
                            "WHERE p.cod_poliza = %s ORDER BY r.Cod_renovacion DESC LIMIT 1",
                            (excel_policy,)
                        )
                        match_fb = cur_fb.fetchone()
                        if match_fb:
                            found_match = {
                                'NRO. POLIZA': match_fb['cod_poliza'],
                                'NRO. RECIBO': 'POLIZA_OK',
                                'MONTO USD': 0, 'TASA EGRESO': 1.0, 'MONTO PAGADO': 0,
                                'cod_ejecutivo': match_fb['cod_ejecutivo'],
                                'comision': match_fb['comision']
                            }

            if found_match:
                excel_row.update({
                    'is_match': True,
                    'procesado': 'por procesar',
                    'db_cod_pago': found_match.get('NRO. RECIBO'),
                    'cod_ejecutivo': found_match.get('cod_ejecutivo'),
                    'nro_cuota': found_match.get('nro_cuota'),
                    'monto_usd_predefined': clean_and_convert_to_float(found_match.get('MONTO USD')),
                    'predefined_comision_porcentaje': clean_and_convert_to_float(found_match.get('comision')),
                })
            else:
                excel_row.update({
                    'is_match': False,
                    'procesado': 'no encontrada',
                    'db_cod_pago': None,
                    'cod_ejecutivo': None,
                    'nro_cuota': None,
                    'monto_usd_predefined': 0.0,
                    'predefined_comision_porcentaje': 0.0,
                })
            matched_data.append(excel_row)

        n_matched = sum(1 for r in matched_data if r.get('is_match'))
        n_total   = len(matched_data)

        results = {
            'matched_data': matched_data,
            'compania': 'PIRAMIDE',
            'fecha_seleccionada': fecha_str_global or date.today().strftime('%Y-%m-%d'),
        }
        temp_filename = save_temp_data(results)
        session['comisiones_beta_file'] = temp_filename
        session['comisiones_beta_just_uploaded'] = True

        flash(
            f'PirÃ¡mide PDF: {n_total} comisiones procesadas â€” {n_matched} emparejadas, '
            f'{n_total - n_matched} no encontradas en BD.',
            'success'
        )
        return redirect(url_for('comisiones_beta'))

    except Exception as e:
        flash(f'PirÃ¡mide PDF [{type(e).__name__}]: {e}', 'error')
        traceback.print_exc()
        return redirect(url_for('comisiones_beta'))


@app.route('/upload-comisiones-universitas', methods=['POST'])
def upload_comisiones_universitas():
    if 'conectado' not in session:
        return redirect(url_for('inicioCpanel'))

    fecha_pago_manual = request.form.get('fecha_pago')

    if 'file_universitas' not in request.files:
        flash('No se seleccionÃ³ el archivo de Universitas.', 'error')
        return redirect(url_for('comisiones_beta'))

    file = request.files['file_universitas']
    if file.filename == '':
        flash('No se seleccionÃ³ ningÃºn archivo de Universitas.', 'error')
        return redirect(url_for('comisiones_beta'))

    fname_lower = file.filename.lower()
    if not (fname_lower.endswith('.xlsx') or fname_lower.endswith('.xls')):
        flash('Universitas: Formato no vÃ¡lido. Se requiere un archivo .xlsx o .xls', 'error')
        return redirect(url_for('comisiones_beta'))

    try:
        import io as _io
        file_bytes = file.read()

        # â”€â”€ Column indices (0-based) from Universitas "ComisionesBonosResumenPol" â”€â”€
        # Row 3 (0-idx=3) is the header, data starts at row 4.
        #  0  Nro. relaciÃ³n de egreso
        #  1  Nro. obligaciÃ³n
        #  2  Producto / Tipo  (AUTI, I.S.L.R, etc.)
        #  3  PÃ³liza
        #  8  Contratante  â† Nombre cliente
        # 14  Nro. Recibo
        # 15  Fecha emisiÃ³n recibo
        # 17  Fecha estatus recibo   â† fecha de cobro real
        # 21  Nro. Cuota
        # 22  Moneda de la pÃ³liza  (DL / BS)
        # 23  Prima cobrada
        # 25  % comisiÃ³n / bono
        # 26  Monto comisiÃ³n / bono  (en Moneda de la pÃ³liza)
        # 27  Moneda abono / deducciÃ³n  (BS)
        # 28  Fecha abono / deducciÃ³n   â† fecha de pago
        # 29  Tasa abono / deducciÃ³n
        # 30  Monto abono / deducciÃ³n   â† monto pagado en Bs

        COL_TIPO          = 2
        COL_POLIZA        = 3
        COL_CLIENTE       = 8
        COL_RECIBO        = 14
        COL_FECHA_ESTATUS = 17   # fecha real de cobro del recibo
        COL_NRO_CUOTA     = 21
        COL_MONEDA_POL    = 22
        COL_PRIMA         = 23
        COL_PCT_COM       = 25
        COL_MONTO_COM     = 26   # comisiÃ³n en moneda de la pÃ³liza (DL or BS)
        COL_MONEDA_ABONO  = 27   # always BS in this file
        COL_FECHA_ABONO   = 28   # fecha de pago efectivo
        COL_TASA          = 29
        COL_MONTO_ABONO   = 30   # comisiÃ³n en Bs (definitivo)
        MIN_COLS          = 31

        # Valid tipos: skip ISLR, summary rows, etc.
        # Exact-match blacklist (no empty string to avoid '' in str always being True)
        SKIP_EXACT = {'I.S.L.R', 'ISLR', 'IMP.MUN', 'TOTAL', 'SUBTOTAL', 'PRODUCTO'}

        wb = openpyxl.load_workbook(_io.BytesIO(file_bytes), data_only=True, read_only=True)
        sheet = wb.active
        if sheet is None:
            flash('Universitas: El archivo no tiene hojas activas.', 'error')
            return redirect(url_for('comisiones_beta'))

        # Find header row
        header_row_idx = None
        all_rows = list(sheet.iter_rows(values_only=True))
        for i, row in enumerate(all_rows):
            row_up = [str(v).strip().upper() for v in row if v is not None]
            if any('P' in v and 'LIZA' in v for v in row_up) and any('RECIBO' in v for v in row_up):
                header_row_idx = i
                break

        if header_row_idx is None:
            flash('Universitas: No se encontrÃ³ la fila de encabezados. Verifique el archivo.', 'error')
            return redirect(url_for('comisiones_beta'))

        print(f"[UNIVERSITAS] Header at row {header_row_idx}. Processing {len(all_rows)-header_row_idx-1} data rows.")

        def safe_float(val):
            try:
                return float(val or 0)
            except (ValueError, TypeError):
                return 0.0

        def fmt_date(val):
            if isinstance(val, (datetime.datetime, datetime.date)):
                return val.strftime('%Y-%m-%d')
            s = str(val or '').strip()
            for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
                try:
                    return datetime.datetime.strptime(s, fmt).strftime('%Y-%m-%d')
                except ValueError:
                    pass
            return s

        extracted_data = []
        fecha_periodo  = None

        for row in all_rows[header_row_idx + 1:]:
            if not row or len(row) < MIN_COLS:
                continue

            # PÃ³liza must be a non-empty number
            raw_poliza = row[COL_POLIZA]
            if raw_poliza is None:
                continue
            poliza_str = re.sub(r'[^a-zA-Z0-9]', '', str(raw_poliza).strip())
            if not poliza_str or poliza_str.upper() in ('NAN', 'NONE', ''):
                continue

            # Skip ISLR and similar deduction rows
            tipo_raw = str(row[COL_TIPO] or '').strip().upper()
            if tipo_raw in SKIP_EXACT or not tipo_raw:
                continue
            if any(s in tipo_raw for s in ('ISLR', 'IMP.MUN', 'TOTAL')):
                continue

            # Recibo
            recibo_str = re.sub(r'[^a-zA-Z0-9]', '', str(row[COL_RECIBO] or '').strip())

            # Fecha de cobro (prefer fecha_estatus, fall back to fecha_abono)
            fecha_cobro = fmt_date(row[COL_FECHA_ESTATUS]) or fmt_date(row[COL_FECHA_ABONO])
            if not fecha_periodo and fecha_cobro:
                fecha_periodo = fmt_date(row[COL_FECHA_ABONO]) or fecha_cobro

            # Moneda de la pÃ³liza
            moneda_pol = str(row[COL_MONEDA_POL] or 'DL').strip().upper()
            moneda_det = 'BOLIVARES' if moneda_pol == 'BS' else 'DOLARES'

            # Amounts
            prima       = safe_float(row[COL_PRIMA])
            pct_com     = safe_float(row[COL_PCT_COM])
            monto_com   = safe_float(row[COL_MONTO_COM])   # in pÃ³liza currency
            tasa        = safe_float(row[COL_TASA])
            monto_bs    = safe_float(row[COL_MONTO_ABONO]) # in Bs

            # Derive USD and BS amounts
            if moneda_det == 'BOLIVARES':
                monto_usd_calc = round(monto_bs / tasa, 2) if tasa > 0 else 0.0
            else:
                monto_usd_calc = monto_com   # already in USD
                if monto_bs == 0 and tasa > 0:
                    monto_bs = round(monto_com * tasa, 2)

            # Nro cuota
            nro_cuota_raw = str(row[COL_NRO_CUOTA] or '').strip()

            # Cliente
            cliente = str(row[COL_CLIENTE] or 'S/N').strip()

            data_dict = {
                'DESCRIPCION MOVIMIENTO': tipo_raw or 'COMISION',
                'FECHA COBRO RECIBO':         fecha_cobro,
                'NRO. POLIZA':                poliza_str,
                'NRO. RECIBO':                recibo_str,
                'NOMBRE CLIENTE':             cliente,
                'MONTO PRIMA':                prima,
                'monto_usd_excel':            monto_usd_calc,
                'monto_pagado_excel':         monto_bs,
                'tasa_excel':                 tasa,
                'moneda':                     moneda_det,
                'porcentaje_comision_pdf':    pct_com,
                'bono':                       0.0,
            }
            extracted_data.append(data_dict)

        if not extracted_data:
            flash('Universitas: No se encontraron filas de datos vÃ¡lidas en el archivo.', 'error')
            return redirect(url_for('comisiones_beta'))

        print(f"[UNIVERSITAS] Extracted {len(extracted_data)} rows. Sample: {extracted_data[0]}")

        # â”€â”€ DB Matching â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        lista_polizas  = [r['NRO. POLIZA'] for r in extracted_data if r.get('NRO. POLIZA')]
        predefined_data = obtener_pagos_globales(lista_polizas)
        matched_data    = []
        used_ids_tracker = {}

        for excel_row in extracted_data:
            excel_policy = excel_row.get('NRO. POLIZA', '')
            candidates_db = [
                d for d in predefined_data
                if re.sub(r'[^a-zA-Z0-9]', '', str(d.get('NRO. POLIZA', ''))) == excel_policy
            ]
            found_match = find_closest_payment_db(
                excel_row.get('FECHA COBRO RECIBO'), candidates_db, used_ids=used_ids_tracker
            )

            if not found_match and excel_policy:
                with connectionBD() as conn_fb:
                    with conn_fb.cursor() as cur_fb:
                        cur_fb.execute(
                            "SELECT p.cod_poliza, e.cod_ejecutivo, r.comision "
                            "FROM poliza p JOIN asegurado a ON p.CI_asegurado = a.CI "
                            "JOIN ejecutivo e ON a.ejecutivo = e.cod_ejecutivo "
                            "LEFT JOIN renovacion r ON p.cod_poliza = r.Cod_poliza "
                            "WHERE p.cod_poliza = %s ORDER BY r.Cod_renovacion DESC LIMIT 1",
                            (excel_policy,)
                        )
                        match_fb = cur_fb.fetchone()
                        if match_fb:
                            found_match = {
                                'NRO. POLIZA': match_fb['cod_poliza'],
                                'NRO. RECIBO': 'POLIZA_OK',
                                'MONTO USD': 0, 'TASA EGRESO': 1.0, 'MONTO PAGADO': 0,
                                'cod_ejecutivo': match_fb['cod_ejecutivo'],
                                'comision': match_fb['comision']
                            }

            if found_match:
                excel_row.update({
                    'is_match': True,
                    'procesado': 'por procesar',
                    'db_cod_pago': found_match.get('NRO. RECIBO'),
                    'cod_ejecutivo': found_match.get('cod_ejecutivo'),
                    'nro_cuota': found_match.get('nro_cuota'),
                    'monto_usd_predefined': clean_and_convert_to_float(found_match.get('MONTO USD')),
                    'predefined_comision_porcentaje': clean_and_convert_to_float(found_match.get('comision')),
                })
            else:
                excel_row.update({
                    'is_match': False,
                    'procesado': 'no encontrada',
                    'db_cod_pago': None,
                    'cod_ejecutivo': None,
                    'nro_cuota': None,
                    'monto_usd_predefined': 0.0,
                    'predefined_comision_porcentaje': 0.0,
                })
            matched_data.append(excel_row)

        n_matched = sum(1 for r in matched_data if r.get('is_match'))
        n_total   = len(matched_data)

        results = {
            'matched_data': matched_data,
            'compania': 'UNIVERSITAS',
            'fecha_seleccionada': fecha_pago_manual or fecha_periodo or date.today().strftime('%Y-%m-%d'),
        }
        temp_filename = save_temp_data(results)
        session['comisiones_beta_file'] = temp_filename
        session['comisiones_beta_just_uploaded'] = True

        flash(
            f'Universitas: {n_total} registros procesados '
            f'\u2014 {n_matched} emparejados, {n_total - n_matched} no encontrados en BD.',
            'success'
        )
        return redirect(url_for('comisiones_beta'))

    except Exception as e:
        flash(f'Universitas [{type(e).__name__}]: {e}', 'error')
        traceback.print_exc()
        return redirect(url_for('comisiones_beta'))


@app.route('/upload-comisiones-banesco', methods=['POST'])
def upload_comisiones_banesco():
    if 'conectado' not in session:
        return redirect(url_for('inicioCpanel'))

    if 'file_banesco' not in request.files:
        flash('No se seleccionÃ³ ningÃºn archivo.', 'error')
        return redirect(url_for('comisiones_beta'))

    file = request.files['file_banesco']
    if file.filename == '':
        flash('No se seleccionÃ³ ningÃºn archivo.', 'error')
        return redirect(url_for('comisiones_beta'))

    if not file.filename.lower().endswith('.pdf'):
        flash('Banesco: Formato no vÃ¡lido. Use un PDF (.pdf)', 'error')
        return redirect(url_for('comisiones_beta'))

    try:
        import pdfplumber
        import io as _io

        file_bytes = file.read()
        extracted_data = []
        fecha_pago_global = None

        # Banesco PDF "RelaciÃ³n de Pago de Comisiones" structure:
        # Header section per block:
        #   Moneda: 2 DOLARES  / 1 BOLIVARES
        #   Tipo Mov.: 1 PAGO COMISION / 9 BONO COBRANZA
        #   "Ramo PÃ³liza Recibo Poliza Origen Cert. Fecha Cobro Nombre Mto. Prima Mto. ComisiÃ³n Mto. ComisiÃ³n Bs. % ComisiÃ³n"
        # Data row (single line per recibo):
        #   RAMO POLIZA RECIBO POLIZA_ORIGEN - CERT FECHA NOMBRE... PRIMA COMISION COMISION_BS %COMISION
        # Example:
        #   25 58786 6-4011380 - 0 10-04-2026 PITA PITA, JUAN MANUEL 167.89 24.60 11,708.52 14.652%

        # Parse numbers in mixed format
        def parse_num(s):
            if not s:
                return 0.0
            s = str(s).replace(',', '').replace('%', '').strip()
            try:
                return float(s)
            except ValueError:
                return 0.0

        # Regex for a date in DD-MM-YYYY or DD/MM/YYYY
        DATE_RE = re.compile(r'\b(\d{2}[-/]\d{2}[-/]\d{4})\b')
        # Row pattern: starts with 1-2 digit RAMO number, then POLIZA, RECIBO, etc.
        # We detect data rows by: first token is digits (ramo), followed by more numeric tokens and a date
        DATA_ROW_RE = re.compile(
            r'^(\d{1,3})\s+'                     # ramo
            r'(\d+)\s+'                           # poliza
            r'([\w\-]+)\s+'                       # recibo (may have dashes)
            r'([\w\-]*)\s*-\s*'                   # poliza origen
            r'(\d+)\s+'                           # cert
            r'(\d{2}[-/]\d{2}[-/]\d{4})\s+'      # fecha cobro
            r'(.+?)\s+'                           # nombre (greedy, trimmed later)
            r'([\d,]+\.?\d*)\s+'                  # mto prima
            r'([\d,]+\.?\d*)\s+'                  # mto comision USD/BS
            r'([\d,]+\.?\d*)\s+'                  # mto comision Bs
            r'([\d.]+)%'                          # % comision
        )
        # Rows WITHOUT leading ramo (subsequent rows in the same ramo block)
        DATA_ROW_NO_RAMO_RE = re.compile(
            r'^(\d+)\s+'                          # poliza
            r'([\w\-]+)\s+'                       # recibo
            r'([\w\-]*)\s*-\s*'                   # poliza origen
            r'(\d+)\s+'                           # cert
            r'(\d{2}[-/]\d{2}[-/]\d{4})\s+'      # fecha cobro
            r'(.+?)\s+'                           # nombre
            r'([\d,]+\.?\d*)\s+'                  # mto prima
            r'([\d,]+\.?\d*)\s+'                  # mto comision
            r'([\d,]+\.?\d*)\s+'                  # mto comision Bs
            r'([\d.]+)%'                          # % comision
        )

        current_moneda  = 'DOLARES'   # default; updated per section header
        current_tipo    = 'PAGO COMISION'

        # Only extract PAGO COMISION rows (skip BONO COBRANZA if desired, or include both)
        VALID_TIPOS_BAN = {'PAGO COMISION', 'BONO COBRANZA'}

        with pdfplumber.open(_io.BytesIO(file_bytes)) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if not text:
                    continue

                for line in text.split('\n'):
                    line = line.strip()
                    if not line:
                        continue
                    upper = line.upper()

                    # Extract global payment date from header
                    if not fecha_pago_global and 'FECHA DE PAGO' in upper:
                        dm = DATE_RE.search(line)
                        if dm:
                            raw_d = dm.group(1).replace('/', '-')
                            try:
                                fecha_pago_global = datetime.datetime.strptime(raw_d, '%d-%m-%Y').strftime('%Y-%m-%d')
                            except Exception:
                                pass

                    # Detect current currency section
                    # Handle both 'MONEDA:' and 'MONEDA :' (with space before colon)
                    if re.search(r'MONEDA\s*:', upper):
                        if 'DOLLAR' in upper or 'DOLAR' in upper or 'DOLARES' in upper or re.search(r'\b2\b', line):
                            current_moneda = 'DOLARES'
                        elif 'BOLIV' in upper or re.search(r'\b1\b', line):
                            current_moneda = 'BOLIVARES'

                    # Detect movement type
                    if 'TIPO MOV' in upper:
                        if 'BONO' in upper:
                            current_tipo = 'BONO COBRANZA'
                        elif 'PAGO' in upper or 'COMIS' in upper:
                            current_tipo = 'PAGO COMISION'

                    # Try matching data row (with ramo, or without ramo)
                    m = DATA_ROW_RE.match(line)
                    no_ramo = False
                    if not m:
                        m2 = DATA_ROW_NO_RAMO_RE.match(line)
                        if m2:
                            m = m2
                            no_ramo = True
                    if not m:
                        continue

                    if current_tipo not in VALID_TIPOS_BAN:
                        continue

                    if no_ramo:
                        poliza_str     = re.sub(r'[^a-zA-Z0-9]', '', m.group(1))
                        recibo_str     = re.sub(r'[^a-zA-Z0-9]', '', m.group(2))
                        fecha_cobro_r  = m.group(5).replace('/', '-')
                        nombre         = m.group(6).strip()
                        prima          = parse_num(m.group(7))
                        mto_com_orig   = parse_num(m.group(8))
                        mto_com_bs     = parse_num(m.group(9))
                        pct_com        = parse_num(m.group(10))
                    else:
                        ramo_str       = m.group(1)
                        poliza_str     = re.sub(r'[^a-zA-Z0-9]', '', m.group(2))
                        recibo_str     = re.sub(r'[^a-zA-Z0-9]', '', m.group(3))
                        fecha_cobro_r  = m.group(6).replace('/', '-')
                        nombre         = m.group(7).strip()
                        prima          = parse_num(m.group(8))
                        mto_com_orig   = parse_num(m.group(9))
                        mto_com_bs     = parse_num(m.group(10))
                        pct_com        = parse_num(m.group(11))

                    try:
                        fecha_cobro = datetime.datetime.strptime(fecha_cobro_r, '%d-%m-%Y').strftime('%Y-%m-%d')
                    except Exception:
                        fecha_cobro = fecha_pago_global or date.today().strftime('%Y-%m-%d')

                    if not poliza_str:
                        continue

                    # Derive tasa and USD/BS amounts
                    # Data-driven inference: if the Bs column is 0 and orig > 0, this row is
                    # definitely in DOLARES â€” overrides the section header when it's wrong.
                    inferred_moneda = current_moneda
                    if mto_com_bs == 0.0 and mto_com_orig > 0:
                        inferred_moneda = 'DOLARES'
                        print(f"[BANESCO] Auto-inferred DOLARES (bs=0, orig={mto_com_orig}) override header={current_moneda}")

                    if inferred_moneda == 'DOLARES':
                        monto_usd  = mto_com_orig
                        monto_bs   = mto_com_bs  # serÃ¡ 0 en este caso
                        tasa_calc  = round(mto_com_bs / mto_com_orig, 4) if mto_com_orig > 0 else 0.0
                    else:
                        # SecciÃ³n genuinamente BOLIVARES: mto_com_orig es el monto en Bs
                        monto_usd  = 0.0
                        monto_bs   = mto_com_orig
                        tasa_calc  = 0.0

                    data_dict = {
                        'DESCRIPCION MOVIMIENTO': current_tipo,
                        'FECHA COBRO RECIBO':         fecha_cobro,
                        'NRO. POLIZA':                poliza_str,
                        'NRO. RECIBO':                recibo_str,
                        'NOMBRE CLIENTE':             nombre,
                        'MONTO PRIMA':                prima,
                        'monto_usd_excel':            monto_usd,
                        'monto_pagado_excel':         monto_bs,
                        'tasa_excel':                 tasa_calc,
                        'moneda':                     inferred_moneda,
                        'porcentaje_comision_pdf':    pct_com,
                        'bono':                       0.0,
                    }
                    extracted_data.append(data_dict)
                    print(f"[BANESCO] poliza={poliza_str}, recibo={recibo_str}, moneda={current_moneda}, usd={monto_usd}, bs={monto_bs}, tasa={tasa_calc}")

        if not extracted_data:
            flash(
                'Banesco PDF: No se encontraron filas de datos. '
                'Verifique que sea la "RelaciÃ³n de Pago de Comisiones" correcta.',
                'error'
            )
            return redirect(url_for('comisiones_beta'))

        print(f"[BANESCO] Total extracted: {len(extracted_data)}")

        # â”€â”€ DB Matching â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        lista_polizas   = [r['NRO. POLIZA'] for r in extracted_data if r.get('NRO. POLIZA')]
        predefined_data = obtener_pagos_globales(lista_polizas)
        matched_data    = []
        used_ids_tracker = {}

        for excel_row in extracted_data:
            excel_policy = excel_row.get('NRO. POLIZA', '')
            candidates_db = [
                d for d in predefined_data
                if re.sub(r'[^a-zA-Z0-9]', '', str(d.get('NRO. POLIZA', ''))) == excel_policy
            ]
            found_match = find_closest_payment_db(
                excel_row.get('FECHA COBRO RECIBO'), candidates_db, used_ids=used_ids_tracker
            )

            if not found_match and excel_policy:
                with connectionBD() as conn_fb:
                    with conn_fb.cursor() as cur_fb:
                        cur_fb.execute(
                            "SELECT p.cod_poliza, e.cod_ejecutivo, r.comision "
                            "FROM poliza p JOIN asegurado a ON p.CI_asegurado = a.CI "
                            "JOIN ejecutivo e ON a.ejecutivo = e.cod_ejecutivo "
                            "LEFT JOIN renovacion r ON p.cod_poliza = r.Cod_poliza "
                            "WHERE p.cod_poliza = %s ORDER BY r.Cod_renovacion DESC LIMIT 1",
                            (excel_policy,)
                        )
                        match_fb = cur_fb.fetchone()
                        if match_fb:
                            found_match = {
                                'NRO. POLIZA': match_fb['cod_poliza'],
                                'NRO. RECIBO': 'POLIZA_OK',
                                'MONTO USD': 0, 'TASA EGRESO': 1.0, 'MONTO PAGADO': 0,
                                'cod_ejecutivo': match_fb['cod_ejecutivo'],
                                'comision': match_fb['comision']
                            }

            if found_match:
                excel_row.update({
                    'is_match': True,
                    'procesado': 'por procesar',
                    'db_cod_pago': found_match.get('NRO. RECIBO'),
                    'cod_ejecutivo': found_match.get('cod_ejecutivo'),
                    'nro_cuota': found_match.get('nro_cuota'),
                    'monto_usd_predefined': clean_and_convert_to_float(found_match.get('MONTO USD')),
                    'predefined_comision_porcentaje': clean_and_convert_to_float(found_match.get('comision')),
                })
            else:
                excel_row.update({
                    'is_match': False,
                    'procesado': 'no encontrada',
                    'db_cod_pago': None,
                    'cod_ejecutivo': None,
                    'nro_cuota': None,
                    'monto_usd_predefined': 0.0,
                    'predefined_comision_porcentaje': 0.0,
                })
            matched_data.append(excel_row)

        n_matched = sum(1 for r in matched_data if r.get('is_match'))
        n_total   = len(matched_data)

        results = {
            'matched_data': matched_data,
            'compania': 'BANESCO',
            'fecha_seleccionada': fecha_pago_global or date.today().strftime('%Y-%m-%d'),
        }
        temp_filename = save_temp_data(results)
        session['comisiones_beta_file'] = temp_filename
        session['comisiones_beta_just_uploaded'] = True

        flash(
            f'Banesco PDF: {n_total} registros procesados '
            f'\u2014 {n_matched} emparejados, {n_total - n_matched} no encontrados en BD.',
            'success'
        )
        return redirect(url_for('comisiones_beta'))

    except Exception as e:
        flash(f'Banesco PDF [{type(e).__name__}]: {e}', 'error')
        traceback.print_exc()
        return redirect(url_for('comisiones_beta'))


@app.route('/upload-comisiones-venezuela', methods=['POST'])
def upload_comisiones_venezuela():
    if 'conectado' not in session:
        return redirect(url_for('inicioCpanel'))

    if 'file_venezuela' not in request.files:
        flash('No se seleccionÃ³ ningÃºn archivo.', 'error')
        return redirect(url_for('comisiones_beta'))

    file = request.files['file_venezuela']

    if file.filename == '':
        flash('No se seleccionÃ³ ningÃºn archivo.', 'error')
        return redirect(url_for('comisiones_beta'))

    fname_lower = file.filename.lower()
    if not (fname_lower.endswith('.xls') or fname_lower.endswith('.xlsx')):
        flash('Venezuela: Formato no vÃ¡lido. Use Excel (.xls o .xlsx)', 'error')
        return redirect(url_for('comisiones_beta'))

    try:
        import io
        import xlrd
        import openpyxl

        file_bytes = file.read()

        # â”€â”€ Read workbook â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        # 'RELACION DE COMISIONES PAGADAS' is a .xls (BIFF8) file from Seguros Venezuela.
        # Columns (0-indexed, header row contains "PÃ³liza" and "Recibo"):
        #   0  Moneda de Pago       e.g. 'Bs.'
        #   1  Tipo                 'COMISIONES' | '24-BONO DE COBRANZAS...'
        #   2  Ramo
        #   3  PÃ³liza
        #   4  Recibo
        #   5  Fecha de EmisiÃ³n
        #   6  Vigencia Desde
        #   7  Vigencia Hasta
        #   8  Ingreso (No. de orden)
        #   9  Fe. Ingreso          â† fecha de cobro del recibo
        #  10  Tasa de Cambio
        #  11  Tipo de Movimiento
        #  12  Nombre Asegurado
        #  13  No. Egreso
        #  14  Moneda PÃ³liza        '$' or 'Bs.'
        #  15  Prima Cobrada        (in Moneda PÃ³liza)
        #  16  ComisiÃ³n Generada    (in Moneda de Pago, usually Bs.)

        COL_MONEDA_PAGO  = 0
        COL_TIPO         = 1
        COL_RAMO         = 2
        COL_POLIZA       = 3
        COL_RECIBO       = 4
        COL_FECHA_COBRO  = 9
        COL_TASA         = 10
        COL_TIPO_MOV     = 11
        COL_CLIENTE      = 12
        COL_NO_EGRESO    = 13
        COL_MONEDA_POL   = 14
        COL_PRIMA        = 15
        COL_COMISION_BS  = 16

        rows_raw = []

        if fname_lower.endswith('.xls'):
            wb_xls = xlrd.open_workbook(file_contents=file_bytes)
            sheet_xls = wb_xls.sheet_by_index(0)
            epoch = datetime.datetime(1899, 12, 30)
            for r in range(sheet_xls.nrows):
                row_out = []
                for c in range(sheet_xls.ncols):
                    cell = sheet_xls.cell(r, c)
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        val = epoch + datetime.timedelta(days=cell.value)
                    elif cell.ctype == xlrd.XL_CELL_NUMBER:
                        val = cell.value
                    elif cell.ctype == xlrd.XL_CELL_EMPTY:
                        val = None
                    else:
                        val = cell.value
                    row_out.append(val)
                rows_raw.append(row_out)
        else:
            wb_xlsx = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
            sheet_xlsx = wb_xlsx.active
            for r in sheet_xlsx.iter_rows(values_only=True):
                rows_raw.append(list(r))

        # â”€â”€ Find header row (contains 'PÃ³liza' and 'Recibo') â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        header_row_idx = None
        for i, row in enumerate(rows_raw):
            row_up = [str(v).strip().upper() for v in row if v is not None]
            if any('POLIZA' in v or 'PÃ“LIZA' in v for v in row_up) and any('RECIBO' in v for v in row_up):
                header_row_idx = i
                break

        if header_row_idx is None:
            flash(
                'Venezuela: No se encontrÃ³ la fila de encabezados (PÃ³liza / Recibo). '
                'Verifique que el archivo sea el "RelaciÃ³n de Comisiones Pagadas" correcto.',
                'error'
            )
            return redirect(url_for('comisiones_beta'))

        print(f"[VENEZUELA] Header at row {header_row_idx}. Data rows: {len(rows_raw) - header_row_idx - 1}")

        # â”€â”€ Extract data rows â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        # We accept both COMISIONES and BONO DE COBRANZAS rows (tipo col 1)
        VALID_TIPOS = ('COMISIONES',)   # Expand to ('COMISIONES', 'BONO') if needed

        extracted_data = []
        fecha_periodo = None

        for row in rows_raw[header_row_idx + 1:]:
            if not row or len(row) <= COL_COMISION_BS:
                continue

            # Col 3 = PÃ³liza must be a non-empty number
            raw_poliza = row[COL_POLIZA]
            if raw_poliza is None:
                continue
            poliza_str = re.sub(r'[^a-zA-Z0-9]', '', str(raw_poliza).strip())
            if not poliza_str or poliza_str.upper() in ('NAN', 'NONE', ''):
                continue

            # Col 1 = Tipo â€” only COMISIONES (and optionally BONO)
            tipo_raw = str(row[COL_TIPO] or '').strip().upper()
            if not any(v in tipo_raw for v in VALID_TIPOS):
                continue

            # Col 4 = Recibo
            raw_recibo = row[COL_RECIBO]
            recibo_str = re.sub(r'[^a-zA-Z0-9]', '', str(raw_recibo or '').strip())

            # Col 9 = Fecha de Cobro (datetime or string)
            raw_fecha = row[COL_FECHA_COBRO]
            if isinstance(raw_fecha, (datetime.datetime, datetime.date)):
                fecha_str = raw_fecha.strftime('%Y-%m-%d')
            else:
                fecha_str = str(raw_fecha or '').strip()
                # Try to parse DD/MM/YYYY or YYYY-MM-DD
                for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
                    try:
                        fecha_str = datetime.datetime.strptime(fecha_str, fmt).strftime('%Y-%m-%d')
                        break
                    except ValueError:
                        pass

            if not fecha_periodo and fecha_str:
                fecha_periodo = fecha_str

            # Col 10 = Tasa de Cambio
            tasa = 0.0
            try:
                tasa = float(row[COL_TASA] or 0)
            except (ValueError, TypeError):
                tasa = 0.0

            # Col 12 = Nombre Asegurado
            cliente = str(row[COL_CLIENTE] or 'S/N').strip()

            # Col 14 = Moneda PÃ³liza ('$' â†’ USD, 'Bs.' â†’ BOLIVARES)
            moneda_pol_raw = str(row[COL_MONEDA_POL] or '').strip()
            moneda_pol_is_usd = '$' in moneda_pol_raw or 'USD' in moneda_pol_raw.upper()

            # Col 15 = Prima Cobrada (in moneda_pol)
            prima_raw = 0.0
            try:
                prima_raw = float(row[COL_PRIMA] or 0)
            except (ValueError, TypeError):
                prima_raw = 0.0

            # Col 16 = ComisiÃ³n Generada in Bs. (Moneda de Pago)
            comision_bs = 0.0
            try:
                comision_bs = float(row[COL_COMISION_BS] or 0)
            except (ValueError, TypeError):
                comision_bs = 0.0

            # Derive USD commission: comision_bs / tasa  (if tasa > 0)
            if tasa > 0 and comision_bs > 0:
                monto_usd_calc = round(comision_bs / tasa, 2)
            else:
                monto_usd_calc = 0.0

            # Col 2 = Ramo description
            ramo = str(row[COL_RAMO] or '').strip()

            # Col 1 = Tipo short label
            tipo_label = 'COMISION' if 'COMISION' in tipo_raw else tipo_raw[:30]

            data_dict = {
                'DESCRIPCION MOVIMIENTO': tipo_label,
                'FECHA COBRO RECIBO':     fecha_str,
                'NRO. POLIZA':            poliza_str,
                'NRO. RECIBO':            recibo_str,
                'NOMBRE CLIENTE':         cliente,
                'MONTO PRIMA':            prima_raw,
                'monto_usd_excel':        monto_usd_calc,
                'monto_pagado_excel':     comision_bs,
                'tasa_excel':             tasa,
                'moneda':                 'BOLIVARES',   # Venezuela always pays in Bs.
                'porcentaje_comision_pdf': 0.0,
                'bono':                   0.0,
            }
            extracted_data.append(data_dict)

        if not extracted_data:
            flash(
                'Venezuela: Se leyÃ³ el archivo pero no se encontraron filas de COMISIONES. '
                'Verifique que el tipo de movimiento sea "COMISIONES" y que el archivo no estÃ© vacÃ­o.',
                'error'
            )
            return redirect(url_for('comisiones_beta'))

        print(f"[VENEZUELA] Extracted {len(extracted_data)} rows. Sample: {extracted_data[0]}")

        # â”€â”€ DB Matching â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        lista_polizas = [r['NRO. POLIZA'] for r in extracted_data if r.get('NRO. POLIZA')]
        predefined_data = obtener_pagos_globales(lista_polizas)
        matched_data = []
        used_ids_tracker = {}

        for excel_row in extracted_data:
            excel_policy = excel_row.get('NRO. POLIZA', '')
            candidates_db = [
                d for d in predefined_data
                if re.sub(r'[^a-zA-Z0-9]', '', str(d.get('NRO. POLIZA', ''))) == excel_policy
            ]
            found_match = find_closest_payment_db(
                excel_row.get('FECHA COBRO RECIBO'), candidates_db, used_ids=used_ids_tracker
            )

            # Fallback: verify policy exists in DB
            if not found_match and excel_policy:
                with connectionBD() as conn_fb:
                    with conn_fb.cursor() as cur_fb:
                        cur_fb.execute(
                            "SELECT p.cod_poliza, e.cod_ejecutivo, r.comision "
                            "FROM poliza p JOIN asegurado a ON p.CI_asegurado = a.CI "
                            "JOIN ejecutivo e ON a.ejecutivo = e.cod_ejecutivo "
                            "LEFT JOIN renovacion r ON p.cod_poliza = r.Cod_poliza "
                            "WHERE p.cod_poliza = %s ORDER BY r.Cod_renovacion DESC LIMIT 1",
                            (excel_policy,)
                        )
                        match_fb = cur_fb.fetchone()
                        if match_fb:
                            found_match = {
                                'NRO. POLIZA': match_fb['cod_poliza'],
                                'NRO. RECIBO': 'POLIZA_OK',
                                'MONTO USD': 0, 'TASA EGRESO': 1.0, 'MONTO PAGADO': 0,
                                'cod_ejecutivo': match_fb['cod_ejecutivo'],
                                'comision': match_fb['comision']
                            }

            if found_match:
                excel_row.update({
                    'is_match': True,
                    'procesado': 'por procesar',
                    'db_cod_pago': found_match.get('NRO. RECIBO'),
                    'cod_ejecutivo': found_match.get('cod_ejecutivo'),
                    'nro_cuota': found_match.get('nro_cuota'),
                    'monto_usd_predefined': clean_and_convert_to_float(found_match.get('MONTO USD')),
                    'predefined_comision_porcentaje': clean_and_convert_to_float(found_match.get('comision')),
                })
            else:
                excel_row.update({
                    'is_match': False,
                    'procesado': 'no encontrada',
                    'db_cod_pago': None,
                    'cod_ejecutivo': None,
                    'nro_cuota': None,
                    'monto_usd_predefined': 0.0,
                    'predefined_comision_porcentaje': 0.0,
                })
            matched_data.append(excel_row)

        n_matched = sum(1 for r in matched_data if r.get('is_match'))
        n_total   = len(matched_data)

        results = {
            'matched_data': matched_data,
            'compania': 'VENEZUELA',
            'fecha_seleccionada': fecha_periodo or date.today().strftime('%Y-%m-%d'),
        }
        temp_filename = save_temp_data(results)
        session['comisiones_beta_file'] = temp_filename
        session['comisiones_beta_just_uploaded'] = True

        flash(
            f'Venezuela: {n_total} registros procesados â€” {n_matched} emparejados, '
            f'{n_total - n_matched} no encontrados en BD.',
            'success'
        )
        return redirect(url_for('comisiones_beta'))

    except Exception as e:
        flash(f'Venezuela [{type(e).__name__}]: {e}', 'error')
        traceback.print_exc()
        return redirect(url_for('comisiones_beta'))


@app.route('/upload-comisiones-internacional', methods=['POST'])
def upload_comisiones_internacional():
    if 'conectado' not in session:
        return redirect(url_for('inicioCpanel'))

    if 'file_internacional' not in request.files:
        flash('No se seleccionÃ³ ningÃºn archivo.', 'error')
        return redirect(url_for('comisiones_beta'))

    file = request.files['file_internacional']
    if file.filename == '':
        flash('No se seleccionÃ³ ningÃºn archivo.', 'error')
        return redirect(url_for('comisiones_beta'))

    if file and file.filename.endswith('.pdf'):
        try:
            extracted_data_list = []
            
            # FunciÃ³n local robusta para este PDF especÃ­fico
            def clean_internacional_amount(val_str):
                if not val_str: return 0.0
                # Limpiar espacios y normalizar comas
                s = str(val_str).strip().replace('Bs', '').replace(' ', '').replace(',', '.')
                
                negativo = False
                if '(' in s and ')' in s:
                    negativo = True
                    s = s.replace('(', '').replace(')', '').strip()

                # Si el nÃºmero empieza con punto (ej: .00), le ponemos un cero delante
                if s.startswith('.'):
                    s = '0' + s

                # Si detectamos mÃºltiples puntos (miles), solo el Ãºltimo es decimal
                if s.count('.') > 1:
                    parts = s.split('.')
                    integer_part = "".join(parts[:-1])
                    decimal_part = parts[-1]
                    s = f"{integer_part}.{decimal_part}"
                
                try: 
                    val = float(s)
                    return -val if negativo else val
                except: 
                    return 0.0

            # --- PRIMERA PASADA: Obtener tasa global del resumen ---
            tasa_global_calculada = 1.0
            try:
                with pdfplumber.open(file) as pdf_temp:
                    full_text = "\n".join([p.extract_text() or "" for p in pdf_temp.pages])
                    if "RESUMEN" in full_text.upper():
                        resumen_idx = full_text.upper().rfind("RESUMEN")
                        resumen_text = full_text[resumen_idx:]
                        for lr in resumen_text.split('\n'):
                            if "TOTAL" in lr.upper():
                                nums_str = re.findall(r'(?<![a-zA-Z])\(?\s*(?:\d+(?:[\d.,]*\d)?|[.,]\d+)\s*\)?(?![%\d.,a-zA-Z])', lr)
                                nums = []
                                for ns in nums_str:
                                    val = abs(clean_internacional_amount(ns))
                                    if val > 0: nums.append(val)
                                if len(nums) >= 2:
                                    nums.sort()
                                    # El mayor suele ser Bs (Local) y el segundo mayor USD (Moneda)
                                    mayor = nums[-1]
                                    menor = nums[-2]
                                    if menor > 0 and (mayor / menor) > 5:
                                        tasa_global_calculada = round(mayor / menor, 2)
                                        print(f"[INTERNACIONAL] Tasa global calculada: {tasa_global_calculada} (Bs: {mayor}, USD: {menor})")
                                        break
            except Exception as e:
                print(f"[INTERNACIONAL] Error al calcular tasa global: {e}")

            # Reset file pointer for main pass
            file.seek(0)

            with pdfplumber.open(file) as pdf:
                for page in pdf.pages:
                    text = page.extract_text()
                    if not text: continue
                        
                    lines = text.split('\n')
                    for line in lines:
                        # Skip summary section
                        if "RESUMEN" in line.upper():
                            break # Stop processing this page once Summary is reached

                        # Stricter: must have Poliza/Recibo or (known keyword AND at least one date)
                        match_pr = re.search(r'(\d{8,12})\s+(\d{8,12})', line)
                        known_keywords = ["BONO", "COMISION", "I.S.R.L.", "RETENCION", "DEDUCCION"]
                        has_keyword = any(k in line.upper() for k in known_keywords)
                        has_any_date = re.search(r'\d{2}/\d{2}/\d{4}', line) is not None
                        
                        # Detail rows (even ISRL) always have at least one date on this PDF.
                        # Summary table rows do NOT have dates.
                        if not (match_pr or (has_keyword and has_any_date)):
                            continue

                        poliza = match_pr.group(1) if match_pr else ""
                        recibo = match_pr.group(2) if match_pr else ""
                        
                        producto_pdf = ""
                        if poliza:
                            parts = line.split()
                            try:
                                pol_idx = parts.index(poliza)
                                if pol_idx > 0:
                                    producto_pdf = parts[pol_idx - 1]
                            except ValueError:
                                pass
                        
                        # FIX: Insert space between letters and a dot/comma followed by a digit
                        # so lookbehind in raw_matches works perfectly for attached dot-decimals (e.g. COMISIONES.56 -> COMISIONES .56)
                        fixed_line = re.sub(r'([a-zA-Z])([.,]\d+)', r'\1 \2', line)

                        # Extraer montos de la parte derecha de la lÃ­nea
                        resto_linea = fixed_line.split(recibo)[-1] if recibo else fixed_line
                        
                        # Extraer Fecha
                        match_fecha = re.search(r'(\d{2}/\d{2}/\d{4})', resto_linea)
                        fecha = match_fecha.group(1) if match_fecha else ""
                        
                        # Extraer Cliente (lo que estÃ¡ entre el recibo y la fecha)
                        if match_fecha:
                            cliente_part = resto_linea.split(fecha)[0].strip()
                            # Clean keywords from the start of the name part
                            for k in known_keywords:
                                if cliente_part.startswith(k):
                                    cliente_part = cliente_part.replace(k, '', 1).strip()
                            cliente = cliente_part
                        else:
                            cliente = "DESCONOCIDO"
                        
                        # Moneda (Bs / USD / DL3 / etc)
                        # La columna Mone aparece DESPUES de Prima y %Comis en el PDF, por
                        # eso buscamos el identificador DL/BS en cualquier posicion del texto
                        # posterior a la fecha, no solo al inicio.
                        raw_moneda = "Bs"
                        if match_fecha:
                            after_date = resto_linea[match_fecha.end():]
                            # Robust moneda match: search DL, USD or $ in after_date or full line
                            if any(x in after_date.upper() for x in ["DL", "USD", "$"]):
                                raw_moneda = "DL3"
                            elif "BS" in after_date.upper():
                                raw_moneda = "BS"
                            else:
                                if any(x in line.upper() for x in ["DL", "USD", "$"]):
                                    raw_moneda = "DL3"

                        moneda_fila = "BOLIVARES" if raw_moneda.upper() == "BS" else "DOLARES"

                        # 1. Detection of % (Commission)
                        porcentaje_comision = 0.0
                        pct_search = re.search(r'([\d.,]+)\s*%', fixed_line)
                        if pct_search:
                            try:
                                p_str = pct_search.group(1).replace(',', '.')
                                porcentaje_comision = float(p_str)
                            except: pass

                        # 2. Extract numeric candidates from the right side
                        # Clean all dates from resto_linea so they never interfere with candidates matching
                        resto_linea_clean = re.sub(r'\d{2}/\d{2}/\d{4}', ' ', resto_linea)

                        # Regex captura numeros con decimales (incluyendo .56 que empieza con punto)
                        # y excluye los que van seguidos de % (porcentajes).
                        # Lookbehind negativo para no capturar partes de letras/fechas.
                        # Lookahead negativo para descartar porcentajes y numeros pegados a letras.
                        raw_matches = re.finditer(
                            r'(?<![a-zA-Z])\(?\s*(?:\d+(?:[\d.,]*\d)?|[.,]\d+)\s*\)?(?![%\d.,a-zA-Z])',
                            resto_linea_clean
                        )
                        candidates = []

                        for m in raw_matches:
                            val_str = m.group(0).strip()
                            # Saltar si el texto inmediatamente despues del match es un %
                            end_pos = m.end()
                            remaining = resto_linea_clean[end_pos:].lstrip()
                            if remaining.startswith('%'):
                                continue
                            
                            cleaned_val = clean_internacional_amount(val_str)
                            
                            # Filter out very large numbers (> 100,000) so ID codes (like 1076796 or 416885)
                            # are never matched as Prima or Asignaciones
                            if abs(cleaned_val) >= 100000:
                                continue

                            candidates.append({
                                'raw': val_str,
                                'value': cleaned_val,
                                'is_negative': cleaned_val < 0 or '(' in val_str
                            })

                        # Debug Print for Terminal
                        print(f"[EXTRACT] Poliza: {poliza} | %: {porcentaje_comision} | Candidates: {[c['value'] for c in candidates]}")

                        if candidates:
                            # 1. Identificar Concepto
                            conceptos_gasto = ["I.S.R.L.", "RETENCION", "DEDUCCION"]
                            concepto_extraido = "OTRO"
                            es_gasto = False
                            
                            conceptos = ["BONO RENOVACION", "BONO PRONTO PAGO", "COMISIONES", "I.S.R.L.", "RETENCION"]
                            line_upper = fixed_line.upper()
                            for c in conceptos:
                                if c in line_upper:
                                    concepto_extraido = c
                                    if any(g in c for g in conceptos_gasto):
                                        es_gasto = True
                                    break
                            
                            if not es_gasto and any(g in line_upper for g in conceptos_gasto):
                                es_gasto = True

                            # 2. Asignar Prima (generalmente el primer nÃºmero positivo)
                            prima_val = 0.0
                            if len(candidates) >= 1:
                                prima_val = candidates[0]['value']

                            # 3. Asignar Monto (AsignaciÃ³n o DeducciÃ³n)
                            asig_val = 0.0
                            
                            if es_gasto:
                                # Buscar el primer valor negativo o entre parÃ©ntesis
                                gasto_match = next((c for c in candidates if c['is_negative']), None)
                                if gasto_match:
                                    asig_val = gasto_match['value']
                                else:
                                    # Si es concepto de gasto pero no tiene parÃ©ntesis, asumimos que es el Ãºltimo valor
                                    asig_val = candidates[-1]['value'] if candidates else 0.0
                                    # Forzar negativo si es gasto y vino positivo
                                    if asig_val > 0: asig_val = -asig_val
                            else:
                                # Caso INGRESO: Tomar el Ãºltimo valor positivo (Asignaciones)
                                valid_income_candidates = [c for c in candidates if not c['is_negative']]
                                if valid_income_candidates:
                                    asig_val = valid_income_candidates[-1]['value']
                                else:
                                    asig_val = 0.0

                            is_adjustment = (es_gasto) or (asig_val < 0)
                            
                            # 4. Calcular Tasa si es Bs
                            tasa_calculada = 1.0
                            monto_usd_calculado = 0.0
                            
                            print(f"[DEBUG] Poliza: {poliza} | Moneda: {moneda_fila} | Prima: {prima_val} | %: {porcentaje_comision} | Asig: {asig_val}")

                            if moneda_fila == 'BOLIVARES':
                                # Si tenemos prima y porcentaje, calculamos el USD esperado
                                if prima_val > 0 and porcentaje_comision > 0:
                                    monto_usd_teorico = round(prima_val * (porcentaje_comision / 100.0), 2)
                                    monto_usd_calculado = monto_usd_teorico
                                    
                                    # Tasa = Bs / USD
                                    if monto_usd_teorico > 0:
                                        tasa_calculada = round(abs(asig_val) / monto_usd_teorico, 2)
                                    else:
                                        tasa_calculada = 0
                                else:
                                    tasa_calculada = 0
                                    monto_usd_calculado = 0
                            else:
                                # PDF en DÃ³lares
                                monto_usd_calculado = asig_val
                                if tasa_global_calculada > 1.0:
                                    tasa_calculada = tasa_global_calculada
                                    asig_val = round(asig_val * tasa_global_calculada, 2) # Monto en Bs
                                else:
                                    tasa_calculada = 1.0

                            
                            # Identificador Ãºnico para depuraciÃ³n (opcional)
                            unique_key = f"{poliza}_{recibo}_{concepto_extraido}_{asig_val}_{fecha}".replace(" ", "")
                            
                            # Si detectamos que el concepto es 'OTRO' pero el cliente tiene una coma, NO forzar ISRL
                            # El usuario reportÃ³ que la coma en el nombre disparaba ISRL.
                            # AquÃ­ ya filtramos conceptos arriba. Si 'conceptos_gasto' no se activÃ³ por palabra clave,
                            # no deberÃ­amos inventar un gasto solo por una coma.
                            
                            extracted_data_list.append({
                                'NRO. POLIZA': poliza,
                                'PRODUCTO_PDF': producto_pdf,
                                'NRO. RECIBO': recibo,
                                'NOMBRE CLIENTE': cliente,
                                'FECHA COBRO RECIBO': fecha,
                                'MONTO PRIMA': prima_val,
                                'porcentaje_comision_pdf': porcentaje_comision,
                                'DESCRIPCION MOVIMIENTO': concepto_extraido,
                                'monto_pagado_excel': asig_val,
                                'moneda': moneda_fila,
                                'tasa_excel': tasa_calculada,
                                'monto_usd_excel': monto_usd_calculado,
                                'is_adjustment': is_adjustment,
                                '_unique_key': unique_key # Para depuraciÃ³n si es necesario
                            })

            if not extracted_data_list:
                flash('No se pudieron extraer comisiones del PDF Internacional.', 'warning')
                return redirect(url_for('comisiones_beta'))

            # --- LÃ³gica de Matching ---
            # Para la compaÃ±Ã­a Internacional, la pÃ³liza en la BD estÃ¡ compuesta por el Producto + 4 Ãºltimos dÃ­gitos de la pÃ³liza en el PDF.
            # Por lo tanto, traemos todos los pagos de LA INTERNACIONAL para buscar en memoria.
            with connectionBD() as conn:
                with conn.cursor() as cursor:
                    sql_todas = """
                        SELECT 
                            p.cod_poliza AS 'NRO. POLIZA',
                            pg.Cod_pago AS 'NRO. RECIBO',
                            pg.monto AS 'MONTO USD',
                            pg.fecha_pagada AS 'FECHA COBRO RECIBO',
                            pg.fecha AS 'FECHA CUOTA',
                            pg.nro_cuota,
                            pg.tasa AS 'TASA EGRESO',
                            pg.estado,
                            r.Prima AS 'MONTO PRIMA',
                            r.comision,
                            e.cod_ejecutivo,
                            a.Nombre AS 'Asegurado_Nombre',
                            a.Apellido AS 'Asegurado_Apellido',
                            'PAGO DE COMISION' AS 'DESCRIPCION MOVIMIENTO',
                            (pg.monto * pg.tasa) AS 'MONTO PAGADO',
                            (
                                SELECT array_to_string(array_agg(DISTINCT COALESCE(c2.nro_recibo_externo, CAST(c2.Cod_pago AS VARCHAR))), ',')
                                FROM comision c2
                                WHERE (c2.cod_poliza = p.cod_poliza AND c2.cod_poliza IS NOT NULL)
                                   OR (c2.Cod_pago = pg.Cod_pago AND c2.Cod_pago != 0)
                            ) as nros_recibos_procesados
                        FROM poliza p
                        JOIN renovacion r ON p.cod_poliza = r.Cod_poliza
                        JOIN pago pg ON r.Cod_renovacion = pg.Cod_renovacion
                        JOIN asegurado a ON a.CI = p.CI_asegurado
                        JOIN ejecutivo e ON a.ejecutivo = e.cod_ejecutivo
                        JOIN compania c ON p.Cod_compania = c.Cod_compania
                        WHERE c.Nombre LIKE '%INTERNACIONAL%'
                        AND pg.estado = 'PAGADO'
                    """
                    cursor.execute(sql_todas)
                    predefined_data = cursor.fetchall()
            
            matched_data = []
            used_ids_tracker = {}
            for row in extracted_data_list:
                pol_nro = str(row.get('NRO. POLIZA', '')).strip()
                prod_pdf = str(row.get('PRODUCTO_PDF', '')).strip().upper()
                is_adj = row.get('is_adjustment', False)
                
                # Filter DB payments for this specific policy
                matches_for_policy = []
                
                pol_nro_clean = re.sub(r'[^a-zA-Z0-9]', '', pol_nro)
                stripped_nro = pol_nro_clean.lstrip('0')
                if not stripped_nro and pol_nro_clean:
                    stripped_nro = '0'
                
                from difflib import SequenceMatcher
                
                cliente_pdf = str(row.get('NOMBRE CLIENTE', '')).strip().upper()
                
                for db_row in predefined_data:
                    db_poliza = re.sub(r'[^a-zA-Z0-9]', '', str(db_row.get('NRO. POLIZA', ''))).upper()
                    db_nombre = str(db_row.get('Asegurado_Nombre', '')).upper()
                    db_apellido = str(db_row.get('Asegurado_Apellido', '')).upper()
                    db_cliente_full = f"{db_nombre} {db_apellido}".strip()
                    
                    is_match = False
                    
                    # 1. Intentar match por exactitud de producto + ultimos digitos sin ceros
                    if prod_pdf and stripped_nro:
                        if db_poliza.startswith(prod_pdf) and db_poliza.endswith(stripped_nro):
                            is_match = True
                    else:
                        if db_poliza == pol_nro_clean:
                            is_match = True
                            
                    # 2. Si no hubo match, intentar fallback: terminaciÃ³n pÃ³liza + similitud de nombre
                    if not is_match and stripped_nro and db_poliza.endswith(stripped_nro):
                        similarity = SequenceMatcher(None, cliente_pdf, db_cliente_full).ratio()
                        if similarity > 0.6:  # Un 60% de similitud para tolerar nombres cortados o apellidos primeros
                            is_match = True
                    if is_match:
                        matches_for_policy.append(db_row)
                
                found_match = None
                if matches_for_policy:
                    # Find closest by date (passing tracker)
                    found_match = find_closest_payment_db(row.get('FECHA COBRO RECIBO'), matches_for_policy, used_ids=used_ids_tracker)
                
                # 2. Si no hay pago, pero el usuario solo quiere validar que la POLIZA existe
                if not found_match and pol_nro:
                    # BÃºsqueda fallback: Â¿Existe la pÃ³liza en la BD aunque no tenga pagos?
                    with connectionBD() as conn_fb:
                        with conn_fb.cursor() as cur_fb:
                            if prod_pdf and stripped_nro:
                                sql_fb = """
                                    SELECT p.cod_poliza, e.cod_ejecutivo, r.comision
                                    FROM poliza p
                                    JOIN asegurado a ON p.CI_asegurado = a.CI
                                    JOIN ejecutivo e ON a.ejecutivo = e.cod_ejecutivo
                                    LEFT JOIN renovacion r ON p.cod_poliza = r.Cod_poliza
                                    JOIN compania c ON p.Cod_compania = c.Cod_compania
                                    WHERE c.Nombre LIKE '%INTERNACIONAL%'
                                    AND p.cod_poliza LIKE %s
                                    AND p.cod_poliza LIKE %s
                                    ORDER BY r.Cod_renovacion DESC
                                    LIMIT 1
                                """
                                cur_fb.execute(sql_fb, (f"{prod_pdf}%", f"%{stripped_nro}"))
                            else:
                                sql_fb = """
                                    SELECT p.cod_poliza, e.cod_ejecutivo, r.comision
                                    FROM poliza p
                                    JOIN asegurado a ON p.CI_asegurado = a.CI
                                    JOIN ejecutivo e ON a.ejecutivo = e.cod_ejecutivo
                                    LEFT JOIN renovacion r ON p.cod_poliza = r.Cod_poliza
                                    WHERE p.cod_poliza = %s
                                    ORDER BY r.Cod_renovacion DESC
                                    LIMIT 1
                                """
                                cur_fb.execute(sql_fb, (pol_nro,))
                                
                            found_match = cur_fb.fetchone()
                            if found_match:
                                # Adaptar formato para que la lÃ³gica de abajo lo entienda
                                found_match.update({
                                    'NRO. POLIZA': found_match['cod_poliza'],
                                    'NRO. RECIBO': 'POLIZA_OK', # Indicador de que solo se encontrÃ³ la pÃ³liza
                                    'MONTO USD': 0,
                                    'TASA EGRESO': 1.0,
                                    'MONTO PAGADO': 0,
                                    'nro_cuota': None
                                })

                monto_orig = float(row['monto_pagado_excel'] or 0)
                
                if found_match:
                    found_match['_used'] = True
                    db_tasa = clean_and_convert_to_float(found_match.get('TASA EGRESO')) or 1.0
                    
                    # Prioritize PDF-calculated values for International
                    if row.get('tasa_excel', 0) > 0 and row.get('monto_usd_excel', 0) > 0:
                        monto_usd = row['monto_usd_excel']
                        final_tasa = row['tasa_excel']
                    else:
                        monto_usd = (monto_orig / db_tasa) if row['moneda'] == "BOLIVARES" and db_tasa > 0 else monto_orig
                        final_tasa = db_tasa

                    row.update({
                        'is_match': True,
                        'procesado': 'por procesar',
                        'monto_usd_excel': monto_usd,
                        'tasa_excel': final_tasa,
                        'monto_usd_predefined': clean_and_convert_to_float(found_match.get('MONTO USD')),
                        'db_cod_pago': found_match.get('NRO. RECIBO'),
                        'nro_cuota': found_match.get('nro_cuota'),
                        'predefined_comision_porcentaje': clean_and_convert_to_float(found_match.get('comision', 0)),
                        'cod_ejecutivo': found_match.get('cod_ejecutivo'),
                        'bono': 0.0
                    })
                else:
                    estado = 'procesada' if is_adj else 'no encontrada'
                    
                    # Preserve PDF-calculated values if they exist
                    m_usd = row.get('monto_usd_excel', 0.0)
                    t_excel = row.get('tasa_excel', 1.0)
                    
                    row.update({
                        'is_match': False,
                        'procesado': estado,
                        'monto_usd_excel': m_usd,
                        'tasa_excel': t_excel,
                        'db_cod_pago': 'RETENCION' if is_adj else None,
                        'cod_ejecutivo': None,
                        'nro_cuota': None,
                        'monto_usd_predefined': 0.0,
                        'predefined_comision_porcentaje': 0.0,
                        'bono': 0.0,
                    })
                matched_data.append(row)

            results = {
                'matched_data': matched_data,
                'pagos_del_dia': predefined_data,
                'fecha_seleccionada': date.today().strftime('%Y-%m-%d'),
                'compania': 'INTERNACIONAL'
            }
            
            temp_filename = save_temp_data(results)
            session['comisiones_beta_file'] = temp_filename
            session['comisiones_beta_just_uploaded'] = True
            
            flash(f'Procesado con Ã©xito de Internacional: {len(matched_data)} registros.', 'success')
            return redirect(url_for('comisiones_beta'))

        except Exception as e:
            flash(f'Error al procesar el PDF Internacional: {e}', 'error')
            traceback.print_exc()
            return redirect(url_for('comisiones_beta'))
    else:
        flash('Tipo de archivo no permitido. Debe ser un PDF.', 'error')
        return redirect(url_for('comisiones_beta'))


def extraer_mercantil(pdf_file):
    datos_mercantil = []
    block_id = 0
    records_in_current_block = 0
    last_egreso = ""
    
    def limpiar_monto(texto):
        if not texto: return 0.0
        texto = str(texto).strip()
        negativo = False
        if texto.startswith('(') and texto.endswith(')'):
            negativo = True
            texto = texto[1:-1].strip()
        
        # Eliminar cualquier caracter que no sea numero, coma o punto
        texto_clean = re.sub(r'[^0-9,.-]', '', texto)
        if not texto_clean: return 0.0
        
        try:
            # Mercantil usa . para miles y , para decimales
            val = float(texto_clean.replace('.', '').replace(',', '.'))
            return -val if negativo else val
        except:
            return 0.0

    report_date = None
    
    try:
        with pdfplumber.open(pdf_file) as pdf:
            for page_idx, page in enumerate(pdf.pages):
                    
                text = page.extract_text()
                if not text:
                    continue
                
                # ParametrizaciÃ³n conservadora para evitar que las columnas se peguen
                settings = {
                    "vertical_strategy": "text", 
                    "horizontal_strategy": "text",
                    "snap_tolerance": 3,
                    "intersection_tolerance": 3,
                }
                
                page_tables = page.extract_tables(table_settings=settings)
                if not page_tables:
                    continue
                    
                for table in page_tables:
                    
                    for row in table:
                        if not row: continue

                        # Limpiar celdas para pre-analizar
                        clean_row = [str(cell).replace('\n', ' ').strip().upper() if cell else "" for cell in row]

                        # REGLA DE BLOQUES: Si cualquier celda contiene "RESUMEN", cerramos bloque actual y abrimos nuevo
                        # Lo hacemos antes del filtro de longitud para capturar filas de resumen cortas
                        if any("RESUMEN" in str(cell) for cell in clean_row):
                            if records_in_current_block > 0:
                                block_id += 1
                                records_in_current_block = 0
                            continue
                        
                        if len(row) < 10: continue
                        
                        row = [str(cell).replace('\n', ' ').strip() if cell else "" for cell in row]
                        
                        poliza_raw = row[0]
                        # Limpiar pÃ³liza: Mercantil a veces pega la palabra "Factura" o similar
                        # Solo nos interesa el patrÃ³n X-XX-XXXXXX
                        poliza_match = re.search(r'(\d+-\d+-\d+)', poliza_raw)
                        poliza = poliza_match.group(1) if poliza_match else poliza_raw

                        # REGLA DE ORO PARA MERCANTIL: 
                        # Una fila de datos real DEBE empezar con el formato de pÃ³liza: X-XX-XXXXXX
                        # Esto descarta automÃ¡ticamente tÃ­tulos, encabezados y basura de la cabecera.
                        if not re.match(r'^\d+-\d+-\d+', poliza):
                            continue

                        try:
                            # DEBUG: Imprimir las primeras 3 filas para verificar el mapeo
                            if len(datos_mercantil) < 3:
                                print(f"\n{'='*80}")
                                print(f"DEBUG MERCANTIL - FILA {len(datos_mercantil)+1} (Total columnas: {len(row)})")
                                print(f"{'='*80}")
                                for idx, val in enumerate(row):
                                    print(f"  Col[{idx:2d}]: {val}")
                            
                            # MAPEO DINÃMICO: Detectar columnas por patrÃ³n en lugar de posiciÃ³n fija
                            # Esto maneja PDFs con diferente nÃºmero de columnas entre pÃ¡ginas
                            
                            recibo = ""
                            tipo_mov = ""
                            asegurado = ""
                            moneda = "USD"
                            prima_val = 0.0
                            porcentaje_comision = 0.0
                            comision_generada = 0.0
                            egreso = ""
                            fecha_pago = ""
                            
                            # PASO 1: Buscar todos los candidatos numÃ©ricos en la fila
                            # Mercantil suele tener: ... | PRIMA | ... | % | COMISION | ...
                            numeric_candidates = []
                            for idx, cell in enumerate(row):
                                cell_str = str(cell).strip()
                                # Detectar cualquier nÃºmero con formato XX,XX o X.XXX,XX
                                if re.match(r'^-?\d{1,3}(\.\d{3})*,\d{2}$', cell_str):
                                    val = limpiar_monto(cell_str)
                                    if val != 0:
                                        numeric_candidates.append({'val': val, 'idx': idx})
                            
                            # PASO 2: Asignar valores por posiciÃ³n (HeurÃ­stica Robusta)
                            # Generalmente: Prima aparece antes que Porcentaje, y ComisiÃ³n al final o despuÃ©s
                            if len(numeric_candidates) >= 3:
                                # Si hay 3 o mÃ¡s, el del medio suele ser el % (si es <= 100)
                                # O validamos: C1 * C2 / 100 Ã¢â€°Ë† C3
                                c1, c2, c3 = numeric_candidates[0]['val'], numeric_candidates[1]['val'], numeric_candidates[2]['val']
                                if abs(c1 * (c2/100) - c3) < 0.5: # Validar relaciÃ³n matemÃ¡tica
                                    prima_val = c1
                                    porcentaje_comision = c2
                                    comision_generada = c3
                                else:
                                    # Fallback: Usar orden de apariciÃ³n (Prima, %, Comision)
                                    prima_val = c1
                                    porcentaje_comision = c2
                                    comision_generada = c3
                            elif len(numeric_candidates) == 2:
                                # Si hay 2, pueden ser Prima y Comision, o Prima y %
                                c1, c2 = numeric_candidates[0]['val'], numeric_candidates[1]['val']
                                if c2 <= 100: # Probablemente Prima y %
                                    prima_val = c1
                                    porcentaje_comision = c2
                                    comision_generada = prima_val * (porcentaje_comision / 100)
                                else: # Probablemente Prima y Comision (sin % explÃ­cito)
                                    prima_val = c1
                                    comision_generada = c2
                            elif len(numeric_candidates) == 1:
                                prima_val = numeric_candidates[0]['val']

                            # PASO 3: Otros datos (Recibo, Movimiento, Moneda, etc.)
                            for idx, cell in enumerate(row):
                                cell_str = str(cell).strip()
                                
                                # Detectar Recibo/Factura
                                if idx <= 5 and ("Factura" in cell_str or "Recibo" in cell_str):
                                    if idx + 1 < len(row):
                                        clean_val = re.sub(r'^\d+-\d+-\d+', '', cell_str).strip()
                                        recibo = f"{clean_val} {row[idx+1]}" if clean_val else row[idx+1]
                                
                                # Detectar Tipo de Movimiento
                                if any(x in cell_str for x in ["Bono", "Cobrados", "Recibos", "AnulaciÃ³n", "Ajuste"]):
                                    tipo_mov = cell_str
                                
                                # Detectar Moneda
                                if cell_str in ["USD", "BS", "VES", "$"]:
                                    moneda = "USD" if cell_str in ["USD", "$"] else "BOLIVARES"
                                
                                # Detectar Egreso (formato: X-X-XXXXXXXX)
                                if re.match(r'^\d+-\d+-\d+$', cell_str) and cell_str != poliza:
                                    egreso = cell_str
                            
                            # Si no se detectÃ³ egreso en la fila, usar el Ãºltimo conocido (Sticky Egreso)
                            if not egreso and last_egreso:
                                egreso = last_egreso
                            elif egreso:
                                last_egreso = egreso
                                
                                # Detectar Fecha de Pago (formato: DD/MM/YYYY al final)
                                if re.match(r'^\d{2}/\d{2}/\d{4}$', cell_str) and idx > 10:
                                    fecha_pago = cell_str
                            
                            # Construir nombre del asegurado (columnas 9-11 aproximadamente)
                            asegurado_parts = []
                            for idx in range(9, min(12, len(row))):
                                cell = str(row[idx]).strip()
                                if cell and cell not in ["USD", "BS", "VES"] and not re.match(r'^\d', cell):
                                    asegurado_parts.append(cell)
                            asegurado = " ".join(asegurado_parts) if asegurado_parts else "CONCEPTO VARIO"
                            
                            # DEBUG: Imprimir valores extraÃ­dos
                            if len(datos_mercantil) < 3:
                                print(f"\n  VALORES EXTRAÃDOS (DINÃMICO):")
                                print(f"    Candidatos numÃ©ricos: {[c['val'] for c in numeric_candidates]}")
                                print(f"    Prima_BS: {prima_val}")
                                print(f"    % ComisiÃ³n: {porcentaje_comision}")
                                print(f"    ComisiÃ³n Generada: {comision_generada}")
                                print(f"{'='*80}\n")
                            
                            # Evitar filas de totalizadores
                            if "TOTAL" in poliza.upper() or "TOTAL" in asegurado.upper():
                                continue

                            registro = {
                                "Poliza": poliza,
                                "Recibo": recibo.strip(),
                                "Asegurado": asegurado,
                                "Tipo_Movimiento": tipo_mov or "PAGO COMISION",
                                "Moneda": moneda,
                                "Prima_BS": prima_val,
                                "Porcentaje_Comision": porcentaje_comision,
                                "Comision_BS": comision_generada,
                                "Fecha_Pago": fecha_pago,
                                "Egreso": egreso,
                                "Report_Date": report_date,
                                "block_id": block_id
                            }
                            
                            datos_mercantil.append(registro)
                            records_in_current_block += 1
                            
                        except Exception as e:
                            print(f"Error procesando fila Mercantil: {e}")
                            continue
    except Exception as e:
        print(f"Error crÃ­tico en PDF Mercantil: {e}")
        traceback.print_exc()

    return pd.DataFrame(datos_mercantil)
def extraer_datos_oceanica_comision(file):
    """
    ExtracciÃ³n Robusta V6 para OceÃ¡nica.
    Enfocado en precisiÃ³n de columnas pegadas (Poliza/Recibo) y filtrado de resÃºmenes.
    """
    def clean_num(val):
        if not val: return 0.0
        s = str(val).strip().replace(' ', '')
        s = re.sub(r'[a-zA-ZÃ¡Ã©Ã­Ã³ÃºÃÃƒâ€°ÃÃƒâ€œÃƒÅ¡Ã±Ãƒâ€˜$]', '', s)
        neg = '(' in s or '-' in s
        s = s.replace('(', '').replace(')', '').replace('-', '').strip()
        if not s: return 0.0
        if ',' in s and '.' in s: s = s.replace('.', '').replace(',', '.')
        elif ',' in s: s = s.replace(',', '.')
        try: return -float(s) if neg else float(s)
        except: return 0.0

    data = []
    # Anclas visuales estimadas (PÃ³liza a la izquierda, Asignaciones a la derecha)
    # Se ajustan dinÃ¡micamente si se encuentra la cabecera
    current_cols_map = {
        'poliza': 45, 'recibo': 95, 'ord_pago': 160, 'cheque': 210, 
        'relacion': 260, 'cliente': 380, 'producto': 420, 'fecha': 460, 'prima': 500, 
        'monto_bs': 540, 'tasa': 580, 'porc_comis': 620
    }
    
    try:
        with pdfplumber.open(file) as pdf:
            for page_idx, page in enumerate(pdf.pages):
                words = page.extract_words(x_tolerance=2, y_tolerance=3)
                if not words: continue
                
                # A. BUSCAR CABECERA
                words.sort(key=lambda x: x['top'])
                l_buckets = []
                curr_b = [words[0]]
                for w in words[1:]:
                    if abs(w['top'] - curr_b[0]['top']) < 10: curr_b.append(w)
                    else:
                        l_buckets.append(curr_b)
                        curr_b = [w]
                l_buckets.append(curr_b)

                best_header_line = None
                max_score = 0
                keywords = ["ASIGNACIONES", "TASA", "PÃƒâ€œLIZA", "POLIZA", "RECIBO", "CLIENTE", "%COMIS", "REL.ING", "ORD.PAGO", "CHEQUE"]
                
                for line in l_buckets:
                    score = 0
                    line_txt = " ".join([w['text'].upper() for w in line])
                    for kw in keywords:
                        if kw in line_txt: score += 1
                    
                    if score > max_score and score >= 2:
                        max_score = score
                        best_header_line = line
                
                header_y = 0
                if best_header_line:
                    header_y = max(w['bottom'] for w in best_header_line)
                    for w in best_header_line:
                        txt = w['text'].upper()
                        cx = (w['x0'] + w['x1']) / 2
                        if "ASIGN" in txt: current_cols_map['monto_bs'] = cx
                        elif "PROD" in txt: current_cols_map['producto'] = cx
                        elif "TASA" in txt: current_cols_map['tasa'] = cx
                        elif "POLI" in txt: current_cols_map['poliza'] = cx
                        elif "RECIBO" in txt: current_cols_map['recibo'] = cx
                        elif "ORD.PAGO" in txt or "N.ORD" in txt: current_cols_map['ord_pago'] = cx
                        elif "CHEQUE" in txt: current_cols_map['cheque'] = cx
                        elif "REL" in txt: current_cols_map['relacion'] = cx
                        elif "CLIEN" in txt: current_cols_map['cliente'] = cx
                        elif "FECH" in txt: current_cols_map['fecha'] = cx
                        elif "PRIM" in txt: current_cols_map['prima'] = cx
                        elif "COMIS" in txt: current_cols_map['porc_comis'] = cx
                    print(f"[OCE_V7] Cabecera OK p{page_idx+1}, score={max_score}")
                
                # B. EXTRAER FILAS
                # Filtrar solo palabras debajo de la cabecera
                data_words = [w for w in words if w['top'] > header_y + 2]
                row_buckets = []
                if data_words:
                    data_words.sort(key=lambda x: x['top'])
                    curr_r = [data_words[0]]
                    for w in data_words[1:]:
                        if abs(w['top'] - curr_r[0]['top']) < 6: curr_r.append(w)
                        else:
                            row_buckets.append(curr_r)
                            curr_r = [w]
                    row_buckets.append(curr_r)

                for rb in row_buckets:
                    rb.sort(key=lambda x: x['x0'])
                    row_txt = " ".join([w['text'] for w in rb]).upper()
                    
                    # STOP: Si llegamos al resumen o totales, dejamos de procesar esta pÃ¡gina
                    if any(s in row_txt for s in ["TOTALES", "RESUMEN BOLIVARES", "TOTAL NETO", "TOTAL GENERAL", "CONCEPTO", "I.S.L.R"]):
                        break
                    
                    # Mapping por proximidad (Umbral reducido a 35 para evitar mezclar columnas pegadas)
                    extracted = {k: [] for k in current_cols_map.keys()}
                    for w in rb:
                        wcx = (w['x0'] + w['x1']) / 2
                        best_k = min(current_cols_map.keys(), key=lambda k: abs(wcx - current_cols_map[k]))
                        # Si estÃ¡ muy lejos de su ancla, probablemente es ruido o columna desviada
                        if abs(wcx - current_cols_map[best_k]) < 35: 
                            extracted[best_k].append(w['text'])
                        else:
                            # Intento de rescate si estÃ¡ entre dos anclas muy cercanas (Recibo/Poliza)
                            if best_k in ['producto', 'poliza', 'recibo'] and abs(wcx - current_cols_map[best_k]) < 50:
                                extracted[best_k].append(w['text'])
                    
                    f_dict = {k: " ".join(v).strip() for k, v in extracted.items()}
                    
                    producto_pdf = f_dict.get('producto', '').strip()
                    poliza = f_dict.get('poliza', '').strip()
                    m_bs = clean_num(f_dict.get('monto_bs', '0'))
                    tasa = clean_num(f_dict.get('tasa', '1')) or 1.0
                    prima = clean_num(f_dict.get('prima', '0'))
                    
                    # ValidaciÃ³n: PÃ³liza debe ser un valor corto/concreto y Monto debe existir
                    if not poliza or m_bs == 0: continue
                    # Filtrar si la pÃ³liza capturada es en realidad un encabezado residual
                    if any(x in poliza.upper() for x in ["PÃƒâ€œLIZA", "POLIZA", "TOTAL"]): continue
                    
                    data.append({
                        'Producto_PDF': producto_pdf,
                        'Poliza': poliza,
                        'Recibo': f_dict.get('recibo', ''),
                        'Relacion': f_dict.get('relacion', ''),
                        'Cliente': f_dict.get('cliente', 'Ver PDF'),
                        'Fecha': f_dict.get('fecha', ''),
                        'Monto_BS': m_bs,
                        'Prima': prima,
                        'Tasa': tasa,
                        'PorcentajeComis': clean_num(f_dict.get('porc_comis', '0')),
                        'Egreso': f_dict.get('ord_pago', ''), # Use ord_pago as Egreso
                        'Tipo_Documento': 'OCEANICA_V6_PRECISE'
                    })
                    print(f"[OCE_V6] ExtraÃ­do: Pol={poliza}, Bs={m_bs}")

        # C. EMERGENCIA: Si tras todo el proceso visual el DF estÃ¡ vacÃ­o, usar Regex sobre Texto plano
        if not data:
            print("[OCE_V5] Fallback de Emergencia: Regex sobre Texto")
            with pdfplumber.open(file) as pdf:
                for page in pdf.pages:
                    text = page.extract_text()
                    if not text: continue
                    for line in text.split('\n'):
                        # PatrÃ³n: Fecha + Algo + NÃºmero (Prima) + NÃºmero (BS) + NÃºmero (Tasa)
                        # Este patrÃ³n es muy genÃ©rico para capturar lo que sea
                        parts = line.split()
                        if len(parts) > 5:
                            # Si hay una fecha al inicio o cerca
                            if any(re.search(r'\d{2}/\d{2}/\d{4}', p) for p in parts[:3]):
                                nums = [clean_num(p) for p in parts if re.search(r'\d', p) and ',' in p]
                                if len(nums) >= 2:
                                    data.append({
                                        'Producto_PDF': '',
                                        'Poliza': parts[-3] if len(parts)>3 else 'S/N',
                                        'Recibo': parts[-2],
                                        'Cliente': 'Rescate Regex',
                                        'Fecha': parts[0],
                                        'Monto_BS': nums[-1],
                                        'Prima': nums[0],
                                        'Tasa': 1.0,
                                        'PorcentajeComis': 0.0,
                                        'Tipo_Documento': 'OCEANICA_V5_EMERGENCY'
                                    })

    except Exception as e:
        print(f"Error crÃ­tico en OcÃ©Ã¡nica V5: {e}")
        traceback.print_exc()
        
    return pd.DataFrame(data)







@app.route('/upload-comisiones-mercantil', methods=['POST'])
def upload_comisiones_mercantil():
    if 'conectado' not in session:
        return redirect(url_for('inicioCpanel'))

    if 'file_mercantil_pdf' not in request.files:
        flash('No se seleccionÃ³ ningÃºn archivo.', 'error')
        return redirect(url_for('comisiones_beta'))

    file = request.files['file_mercantil_pdf']
    if file.filename == '' or not file.filename.endswith('.pdf'):
        flash('Archivo no vÃ¡lido. Debe ser un PDF.', 'error')
        return redirect(url_for('comisiones_beta'))

    try:
        df = extraer_mercantil(file)
        if df.empty:
            flash('No se pudieron extraer datos del PDF de Mercantil.', 'error')
            return redirect(url_for('comisiones_beta'))

        # DEBUG: Mostrar los primeros registros del DataFrame
        print(f"\n{'='*80}")
        print(f"DEBUG - DataFrame extraÃ­do de Mercantil ({len(df)} registros)")
        print(f"{'='*80}")
        for idx, row in df.head(3).iterrows():
            print(f"\nRegistro {idx+1}:")
            print(f"  Poliza: {row.get('Poliza')}")
            print(f"  Prima_BS: {row.get('Prima_BS')}")
            print(f"  Porcentaje_Comision: {row.get('Porcentaje_Comision')}")
            print(f"  Comision_BS: {row.get('Comision_BS')}")
        print(f"{'='*80}\n")

        extracted_rows = df.to_dict('records')
        
        # Intentar obtener la fecha del reporte para pre-poblar el bloque
        report_date_iso = None
        if 'Report_Date' in df.columns:
            extracted_date = df['Report_Date'].iloc[0]
            if extracted_date:
                try:
                    report_date_iso = datetime.strptime(extracted_date, '%d/%m/%Y').strftime('%Y-%m-%d')
                except:
                    pass

        # --- MATCHING LOGIC ---
        lista_polizas = [re.sub(r'[^a-zA-Z0-9]', '', str(row.get('Poliza', ''))) for row in extracted_rows if row.get('Poliza')]
        predefined_data = obtener_pagos_globales(lista_polizas)
        predefined_copy = [item.copy() for item in predefined_data]
        
        all_data = []
        used_ids_tracker = {}
        for mer_row in extracted_rows:
            poliza_orig = str(mer_row.get('Poliza', ''))
            poliza_limpia = re.sub(r'[^a-zA-Z0-9]', '', poliza_orig)
            concepto = str(mer_row.get('Asegurado', '')).upper()
            found_match = None
            
            if poliza_limpia:
                # Filter DB payments for this specific policy
                matches_for_policy = []
                for db_row in predefined_copy:
                    db_poliza = re.sub(r'[^a-zA-Z0-9]', '', str(db_row.get('NRO. POLIZA', '')))
                    if db_poliza == poliza_limpia:
                        matches_for_policy.append(db_row)
                
                if matches_for_policy:
                    # Find closest by date (passing tracker)
                    found_match = find_closest_payment_db(mer_row.get('Fecha_Pago'), matches_for_policy, used_ids=used_ids_tracker)
                
                # FALLBACK: Validar por existencia de pÃ³liza
                if not found_match and poliza_limpia:
                    with connectionBD() as conn_fb:
                        with conn_fb.cursor() as cur_fb:
                            sql_fb = """
                                SELECT p.cod_poliza, e.cod_ejecutivo, r.comision
                                FROM poliza p
                                JOIN asegurado a ON p.CI_asegurado = a.CI
                                JOIN ejecutivo e ON a.ejecutivo = e.cod_ejecutivo
                                LEFT JOIN renovacion r ON p.cod_poliza = r.Cod_poliza
                                WHERE p.cod_poliza = %s
                                ORDER BY r.Cod_renovacion DESC LIMIT 1
                            """
                            cur_fb.execute(sql_fb, (poliza_limpia,))
                            match_fb = cur_fb.fetchone()
                            if match_fb:
                                found_match = {
                                    'NRO. POLIZA': match_fb['cod_poliza'],
                                    'NRO. RECIBO': 'POLIZA_OK',
                                    'MONTO USD': 0,
                                    'TASA EGRESO': 1.0,
                                    'MONTO PAGADO': 0,
                                    'cod_ejecutivo': match_fb['cod_ejecutivo'],
                                    'comision': match_fb['comision'],
                                    'nro_cuota': None
                                }

            comision_bs = float(mer_row.get('Comision_BS', 0))
            moneda_reporte = str(mer_row.get('Moneda', 'USD')).upper()
            is_adjustment = "ISRL" in concepto or "RETENCION" in concepto or not poliza_limpia

            if found_match:
                mer_row.update({
                    'NRO. POLIZA': poliza_limpia,
                    'NRO. RECIBO': mer_row.get('Recibo'),
                    'NOMBRE CLIENTE': mer_row.get('Asegurado'),
                    'FECHA COBRO RECIBO': mer_row.get('Fecha_Pago'),
                    'is_match': True,
                    'procesado': 'por procesar',
                    'monto_usd_excel': 0.0,
                    'monto_usd_predefined': clean_and_convert_to_float(found_match.get('MONTO USD')),
                    'tasa_excel': 0.0,
                    'monto_pagado_excel': comision_bs,
                    'db_cod_pago': found_match.get('NRO. RECIBO'),
                    'nro_cuota': found_match.get('nro_cuota'),
                    'cod_ejecutivo': found_match.get('cod_ejecutivo'),
                    'predefined_comision_porcentaje': clean_and_convert_to_float(found_match.get('comision')),
                    'porcentaje_comision_pdf': float(mer_row.get('Porcentaje_Comision', 0)),
                    'DESCRIPCION MOVIMIENTO': f"MERCANTIL: {mer_row.get('Tipo_Movimiento', 'COMISION')}",
                    'monto_prima': float(mer_row.get('Prima_BS', 0)),
                    'moneda': 'BOLIVARES',
                    'bono': 0.0
                })
            else:
                estado = 'procesada' if is_adjustment else 'no encontrada'
                desc_mov = f"AJUSTE/DEDUCCION: {mer_row.get('Asegurado')}" if is_adjustment else f"MERCANTIL: {mer_row.get('Tipo_Movimiento', 'COMISION')}"
                mer_row.update({
                    'NRO. POLIZA': poliza_limpia if poliza_limpia else 'AJUSTE',
                    'NRO. RECIBO': mer_row.get('Recibo'),
                    'NOMBRE CLIENTE': mer_row.get('Asegurado'),
                    'FECHA COBRO RECIBO': mer_row.get('Fecha_Pago'),
                    'is_match': False,
                    'procesado': estado,
                    'monto_usd_excel': 0.0,
                    'monto_usd_predefined': 0.0,
                    'tasa_excel': 0.0,
                    'monto_pagado_excel': comision_bs,
                    'db_cod_pago': 'AJUSTE_GLOBAL' if is_adjustment else None,
                    'cod_ejecutivo': None,
                    'nro_cuota': None,
                    'predefined_comision_porcentaje': 0.0,
                    'porcentaje_comision_pdf': float(mer_row.get('Porcentaje_Comision', 0)),
                    'DESCRIPCION MOVIMIENTO': desc_mov,
                    'monto_prima': float(mer_row.get('Prima_BS', 0)),
                    'moneda': 'BOLIVARES',
                    'bono': comision_bs if is_adjustment else 0.0
                })
            all_data.append(mer_row)

        # --- GROUP BY EGRESO TO DEFINE BLOCKS ---
        egreso_to_block = {}
        block_metadata = {}
        block_counter = 0
        
        for row in all_data:
            egr = str(row.get('Egreso') or row.get('egreso') or 'SIN_EGRESO').strip()
            if egr not in egreso_to_block:
                egreso_to_block[egr] = block_counter
                block_metadata[str(block_counter)] = {'egreso': egr}
                block_counter += 1
            row['block_id'] = egreso_to_block[egr]

        results = {
            'matched_data': all_data,
            'pagos_del_dia': predefined_data,
            'fecha_seleccionada': report_date_iso or date.today().strftime('%Y-%m-%d'),
            'compania': 'MERCANTIL',
            'block_metadata': block_metadata
        }
        
        temp_filename = save_temp_data(results)
        session['comisiones_beta_file'] = temp_filename
        session['comisiones_beta_just_uploaded'] = True
        
        flash(f'PDF de Mercantil procesado. {len(all_data)} registros encontrados.', 'success')
        return redirect(url_for('comisiones_beta'))

    except Exception as e:
        flash(f'Error crÃ­tico al procesar Mercantil: {e}', 'error')
        traceback.print_exc()
        return redirect(url_for('comisiones_beta'))


@app.route('/upload-comisiones-oceanica', methods=['POST'])
def upload_comisiones_oceanica():
    if 'conectado' not in session:
        return redirect(url_for('inicioCpanel'))

    if 'file_oceanica' not in request.files:
        flash('No se seleccionÃ³ ningÃºn archivo.', 'error')
        return redirect(url_for('comisiones_beta'))

    file = request.files['file_oceanica']
    if file.filename == '' or not file.filename.endswith('.pdf'):
        flash('Archivo no vÃ¡lido. Debe ser un PDF.', 'error')
        return redirect(url_for('comisiones_beta'))

    try:
        # Extraer datos usando el nuevo motor V3
        df = extraer_datos_oceanica_comision(file)

        if df.empty:
            flash('No se pudieron extraer datos. El formato del PDF no coincide con ningÃºn estÃ¡ndar de OceÃ¡nica.', 'error')
            return redirect(url_for('comisiones_beta'))

        extracted_rows = df.to_dict('records')
        
        # 3. PreparaciÃ³n de Matching con diccionarios (Alta velocidad)
        with connectionBD() as conn:
            with conn.cursor() as cursor:
                sql_todas = """
                    SELECT 
                        p.cod_poliza AS 'NRO. POLIZA',
                        pg.Cod_pago AS 'NRO. RECIBO',
                        pg.monto AS 'MONTO USD',
                        pg.fecha_pagada AS 'FECHA COBRO RECIBO',
                        pg.fecha AS 'FECHA CUOTA',
                        pg.nro_cuota,
                        pg.tasa AS 'TASA EGRESO',
                        pg.estado,
                        r.Prima AS 'MONTO PRIMA',
                        r.comision,
                        e.cod_ejecutivo,
                        a.Nombre AS 'Asegurado_Nombre',
                        a.Apellido AS 'Asegurado_Apellido',
                        'PAGO DE COMISION' AS 'DESCRIPCION MOVIMIENTO',
                        (pg.monto * pg.tasa) AS 'MONTO PAGADO',
                        (
                            SELECT array_to_string(array_agg(DISTINCT COALESCE(c2.nro_recibo_externo, CAST(c2.Cod_pago AS VARCHAR))), ',')
                            FROM comision c2
                            WHERE (c2.cod_poliza = p.cod_poliza AND c2.cod_poliza IS NOT NULL)
                               OR (c2.Cod_pago = pg.Cod_pago AND c2.Cod_pago != 0)
                        ) as nros_recibos_procesados
                    FROM poliza p
                    JOIN renovacion r ON p.cod_poliza = r.Cod_poliza
                    JOIN pago pg ON r.Cod_renovacion = pg.Cod_renovacion
                    JOIN asegurado a ON a.CI = p.CI_asegurado
                    JOIN ejecutivo e ON a.ejecutivo = e.cod_ejecutivo
                    JOIN compania c ON p.Cod_compania = c.Cod_compania
                    WHERE c.Nombre LIKE '%OCEANICA%'
                    AND pg.estado = 'PAGADO'
                """
                cursor.execute(sql_todas)
                predefined_data = cursor.fetchall()
        
        all_data = []
        used_ids_tracker = {}
        for oce_row in extracted_rows:
            poliza_key = str(oce_row.get('Poliza', '')).strip()
            prod_pdf = str(oce_row.get('Producto_PDF', '')).strip().upper()
            
            pol_nro_clean = re.sub(r'[^a-zA-Z0-9]', '', poliza_key)
            stripped_nro = pol_nro_clean.lstrip('0')
            if not stripped_nro and pol_nro_clean:
                stripped_nro = '0'
            
            from difflib import SequenceMatcher
            cliente_pdf = str(oce_row.get('Cliente', '')).strip().upper()
            
            # Filter DB payments for this specific policy
            matches_for_policy = []
            for db_row in predefined_data:
                db_poliza = re.sub(r'[^a-zA-Z0-9]', '', str(db_row.get('NRO. POLIZA', ''))).upper()
                db_nombre = str(db_row.get('Asegurado_Nombre', '')).upper()
                db_apellido = str(db_row.get('Asegurado_Apellido', '')).upper()
                db_cliente_full = f"{db_nombre} {db_apellido}".strip()
                
                is_match = False
                if prod_pdf and stripped_nro:
                    # Empieza con el Producto y termina con la PÃ³liza del PDF sin ceros a la izq
                    if db_poliza.startswith(prod_pdf) and db_poliza.endswith(stripped_nro):
                        is_match = True
                else:
                    if db_poliza == pol_nro_clean:
                        is_match = True
                        
                # 2. Si no hubo match, intentar fallback: terminaciÃ³n pÃ³liza + similitud de nombre
                if not is_match and stripped_nro and db_poliza.endswith(stripped_nro):
                    similarity = SequenceMatcher(None, cliente_pdf, db_cliente_full).ratio()
                    if similarity > 0.6:  # Un 60% de similitud para tolerar nombres cortados o apellidos primeros
                        is_match = True
                        
                if is_match:
                    matches_for_policy.append(db_row)
            
            found_match = None
            if matches_for_policy:
                # Find closest by date (passing tracker)
                found_match = find_closest_payment_db(oce_row.get('Fecha'), matches_for_policy, used_ids=used_ids_tracker)
            
            # FALLBACK: Validar por existencia de pÃ³liza
            if not found_match and poliza_key:
                with connectionBD() as conn_fb:
                    with conn_fb.cursor() as cur_fb:
                        if prod_pdf and stripped_nro:
                            sql_fb = """
                                SELECT p.cod_poliza, e.cod_ejecutivo, r.comision
                                FROM poliza p
                                JOIN asegurado a ON p.CI_asegurado = a.CI
                                JOIN ejecutivo e ON a.ejecutivo = e.cod_ejecutivo
                                LEFT JOIN renovacion r ON p.cod_poliza = r.Cod_poliza
                                JOIN compania c ON p.Cod_compania = c.Cod_compania
                                WHERE c.Nombre LIKE '%OCEANICA%'
                                AND p.cod_poliza LIKE %s
                                AND p.cod_poliza LIKE %s
                                ORDER BY r.Cod_renovacion DESC LIMIT 1
                            """
                            cur_fb.execute(sql_fb, (f"{prod_pdf}%", f"%{stripped_nro}"))
                        else:
                            sql_fb = """
                                SELECT p.cod_poliza, e.cod_ejecutivo, r.comision
                                FROM poliza p
                                JOIN asegurado a ON p.CI_asegurado = a.CI
                                JOIN ejecutivo e ON a.ejecutivo = e.cod_ejecutivo
                                LEFT JOIN renovacion r ON p.cod_poliza = r.Cod_poliza
                                WHERE p.cod_poliza = %s
                                ORDER BY r.Cod_renovacion DESC LIMIT 1
                            """
                            cur_fb.execute(sql_fb, (poliza_key,))
                        
                        match_fb = cur_fb.fetchone()
                        if match_fb:
                            found_match = {
                                'NRO. POLIZA': match_fb['cod_poliza'],
                                'NRO. RECIBO': 'POLIZA_OK',
                                'MONTO USD': 0,
                                'TASA EGRESO': 1.0,
                                'MONTO PAGADO': 0,
                                'cod_ejecutivo': match_fb['cod_ejecutivo'],
                                'comision': match_fb['comision'],
                                'nro_cuota': None
                            }
            
            monto_bs = oce_row.get('Monto_BS', 0)
            prima_val = oce_row.get('Prima', 0.0)
            porcentaje_comision = oce_row.get('PorcentajeComis', 0.0)
            tasa_pdf = oce_row.get('Tasa', 0.0)
            
            # CÃ¡lculo de comisiÃ³n en $: Asignaciones / Tasa (del PDF preferiblemente)
            final_tasa = tasa_pdf if tasa_pdf > 10 else 1.0
            monto_usd_calculado = monto_bs / final_tasa if final_tasa > 0 else 0
            
            # ConstrucciÃ³n de la fila de resultado
            resultado_row = {
                'NRO. POLIZA': poliza_key,
                'NRO. RECIBO': oce_row.get('Recibo'),
                'NOMBRE CLIENTE': oce_row.get('Cliente'),
                'FECHA COBRO RECIBO': oce_row.get('Fecha'),
                'monto_prima': prima_val,
                'porcentaje_comision_pdf': porcentaje_comision,
                'monto_pagado_excel': monto_bs,
                'monto_usd_excel': monto_usd_calculado,
                'tasa_excel': final_tasa,
                'moneda': 'BOLIVARES', 
                'DESCRIPCION MOVIMIENTO': f"PAGO COMISION OCEANICA ({oce_row.get('Tipo_Documento')})",
                'bono': 0.0,
                'Egreso': oce_row.get('Egreso')
            }

            if found_match:
                resultado_row.update({
                    'is_match': True,
                    'procesado': 'por procesar',
                    'monto_usd_predefined': clean_and_convert_to_float(found_match.get('MONTO USD')),
                    'db_cod_pago': found_match.get('NRO. RECIBO'),
                    'nro_cuota': found_match.get('nro_cuota'),
                    'cod_ejecutivo': found_match.get('cod_ejecutivo'),
                    'predefined_comision_porcentaje': clean_and_convert_to_float(found_match.get('comision')),
                })
            else:
                resultado_row.update({
                    'is_match': False,
                    'procesado': 'no encontrada',
                    'db_cod_pago': None,
                    'cod_ejecutivo': None,
                    'nro_cuota': None,
                    'monto_usd_predefined': 0.0,
                    'predefined_comision_porcentaje': 0.0,
                })
            
            print(f"[CALC OCEANICA] Pol={poliza_key}, Bs={monto_bs}, Tasa={final_tasa}, USD={monto_usd_calculado}")
            
            all_data.append(resultado_row)

        # --- GROUP BY EGRESO TO DEFINE BLOCKS ---
        egreso_to_block = {}
        block_metadata = {}
        block_counter = 0
        
        for row in all_data:
            # We use Ord.Pago as unique key for block, if duplicate we might need to handle it
            # But user said "un egreso por cada registro"
            egr = str(row.get('Egreso') or row.get('egreso') or f"REC-{block_counter}").strip()
            
            if egr not in egreso_to_block:
                egreso_to_block[egr] = block_counter
                block_metadata[str(block_counter)] = {'egreso': egr}
                block_counter += 1
            row['block_id'] = egreso_to_block[egr]

        # 4. Guardado y RedirecciÃ³n
        results = {
            'matched_data': all_data,
            'pagos_del_dia': predefined_data,
            'fecha_seleccionada': date.today().strftime('%Y-%m-%d'),
            'compania': 'OCEANICA',
            'block_metadata': block_metadata
        }
        
        temp_filename = save_temp_data(results)
        session['comisiones_beta_file'] = temp_filename
        session['comisiones_beta_just_uploaded'] = True
        
        flash(f'Procesado exitosamente: {len(all_data)} registros de OceÃ¡nica.', 'success')
        return redirect(url_for('comisiones_beta'))

    except Exception as e:
        flash(f'Error crÃ­tico al procesar OceÃ¡nica: {e}', 'error')
        traceback.print_exc()
        return redirect(url_for('comisiones_beta'))


@app.route('/comisiones/transacciones', methods=['GET'])
def comisiones_transacciones():
    if 'conectado' not in session:
        return redirect(url_for('inicioCpanel'))
    
    if not (session.get("permisos")=="Administracion" or session.get("permisos")=="Gerencia" or session.get("permisos")=="dev"):
        flash('No tienes permiso para esta acciÃ³n', 'error')
        return redirect(url_for('inicio1'))

    bloques = obtener_bloques_comision()
    return render_template('public/Comisiones/lista_transacciones.html', bloques=bloques)

@app.route('/api/comisiones/transacciones/<int:id_bloque>', methods=['GET'])
def api_detalle_bloque(id_bloque):
    if 'conectado' not in session:
        return jsonify({'success': False, 'message': 'No autorizado'}), 401
        
    detalles = obtener_detalles_bloque(id_bloque)
    return jsonify({'success': True, 'detalles': detalles})

@app.route('/comisiones/eliminar-bloque/<int:id_bloque>', methods=['POST'])
def eliminar_bloque_comision(id_bloque):
    if 'conectado' not in session:
        return jsonify({'success': False, 'message': 'No autorizado'}), 401
        
    if not (session.get("permisos")=="Administracion" or session.get("permisos")=="dev"):
        return jsonify({'success': False, 'message': 'No tienes permiso para eliminar.'}), 403

    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                # 1. Eliminar los registros de comision asociados
                cursor.execute("DELETE FROM comision WHERE id_bloque = %s", (id_bloque,))
                # 2. Eliminar el bloque
                cursor.execute("DELETE FROM bloque_pago_comision WHERE id_bloque = %s", (id_bloque,))
                conexion_MySQLdb.commit()
                return jsonify({'success': True, 'message': 'Bloque y registros asociados eliminados correctamente.'})
    except Exception as e:
        print(f"Error eliminando bloque: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/procesar-comisiones-beta', methods=['POST'])
def procesar_comisiones_beta():
    if 'conectado' not in session:
        return jsonify({'success': False, 'message': 'Debes iniciar sesiÃ³n.'}), 401

    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'message': 'No se recibiÃ³ informaciÃ³n.'}), 400

    # Determinar formato del payload
    if isinstance(data, list):
        comisiones_a_procesar = data
        bloque_info = None
    else:
        comisiones_a_procesar = data.get('comisiones', [])
        bloque_info = data.get('bloque_info')

    if not comisiones_a_procesar:
        return jsonify({'success': False, 'message': 'No se recibieron comisiones para procesar.'}), 400

    # Calcular MONTO TOTAL (Neto) para el bloque, incluyendo deducciones
    monto_total_bloque = sum(float(item.get('monto_pagado_excel') or 0) for item in comisiones_a_procesar)

    # Crear bloque si viene la informaciÃ³n
    id_bloque = None
    if bloque_info and all(k in bloque_info for k in ('numero_egreso', 'referencia_bancaria', 'fecha_movimiento')):
        egreso = bloque_info.get('numero_egreso')
        ref = bloque_info.get('referencia_bancaria')
        codigo_banco = bloque_info.get('codigo_banco')
        compania = bloque_info.get('compania')
        
        # VALIDACIÃƒâ€œN DE DUPLICADOS: Evitar cargar el mismo bloque dos veces
        try:
            with connectionBD() as conn_dup:
                with conn_dup.cursor() as cur_dup:
                    sql_dup = """
                        SELECT id_bloque 
                        FROM bloque_pago_comision 
                        WHERE numero_egreso = %s AND referencia_bancaria = %s AND (compania = %s OR compania IS NULL)
                    """
                    cur_dup.execute(sql_dup, (egreso, ref, compania))
                    if cur_dup.fetchone():
                        return jsonify({
                            'success': False, 
                            'duplicate': True,
                            'message': f"Error: Ya existe un reporte cargado con el Egreso '{egreso}' y Referencia '{ref}' para esta compaÃ±Ã­a."
                        }), 400
        except Exception as e:
            print(f"Error validando duplicado de bloque: {e}")

        bloque_info['monto_total'] = monto_total_bloque
        res_bloque = crear_bloque_comision(bloque_info)
        if res_bloque['success']:
            id_bloque = res_bloque['id_bloque']
        else:
            return jsonify({'success': False, 'message': f"Error al crear el bloque: {res_bloque.get('message')}"}), 500

    # Get current results from session
    session_results = session.get('comisiones_beta_resultados', {})
    matched_data = session_results.get('matched_data', [])

    processed_count = 0
    pendientes_count = 0
    errors = []
    pendientes_procesados = set()

    for item in comisiones_a_procesar:
        cp = item.get('db_cod_pago')
        
        # Special handling for retentions/manual adjustments
        is_retention = (cp == 'RETENCION')
        
        if not is_retention and (not cp or cp == 'AJUSTE_GLOBAL' or cp == 'No Match'):
            # Si no hay match con pago, igualmente lo guardamos como pÃ³liza pendiente para rastreo
            nro_poliza_raw = item.get('NRO. POLIZA') or item.get('nro_poliza') or item.get('poliza') or item.get('PÃ³liza') or 'S/N'
            nro_poliza_norm = str(nro_poliza_raw).strip().upper()

            if nro_poliza_norm not in pendientes_procesados:
                item['id_bloque'] = id_bloque
                item['compania'] = bloque_info.get('compania') if bloque_info else ''
                if insertar_poliza_pendiente(item):
                    pendientes_count += 1
                    if nro_poliza_norm != 'S/N':
                        pendientes_procesados.add(nro_poliza_norm)
                else:
                    errors.append(f"Error al crear pÃ³liza pendiente para {nro_poliza_raw}")
            
            # Normalizar para que pase a registrarse en transacciones (tabla comision)
            if not cp or cp == 'No Match':
                cp = 'PENDIENTE'

        # Helper to get description correctly
        desc_val = item.get('DESCRIPCION MOVIMIENTO') or item.get('DESCRIPCION MOVIMIENTO') or 'PAGO DE COMISION'
        
        if is_retention:
             monto_bs_val = float(item.get('monto_pagado_excel') or 0)
             monto_usd_val = float(item.get('monto_usd_excel') or 0)
             
             payload = {
                'reciboId': 'RETENCION', 
                'codEjecutivo': 0,       
                'bono': 0.0, 
                'tasa': item.get('tasa_excel') or 0,
                'comisionBs': monto_bs_val,
                'montoDolar': monto_usd_val,
                'moneda': item.get('moneda', 'BOLIVARES'),
                'comisionPorcentaje': 0,
                'nro_recibo_externo': 'MANUAL',
                'descripcion': desc_val,
                'id_bloque': id_bloque,
                'codPoliza': item.get('NRO. POLIZA') or item.get('nro_poliza') or 'S/N'
            }
        else:
            payload = {
                'reciboId': cp,
                'codEjecutivo': item.get('cod_ejecutivo') or 0,
                'bono': 0.0, 
                'tasa': item.get('tasa_excel'),
                'comisionBs': float(item.get('monto_pagado_excel') or 0),
                'montoDolar': float(item.get('monto_usd_excel') or 0),
                'moneda': item.get('moneda', 'Bs'),
                'comisionPorcentaje': item.get('predefined_comision_porcentaje'),
                'nro_recibo_externo': item.get('NRO. RECIBO'),
                'descripcion': desc_val,
                'id_bloque': id_bloque,
                'codPoliza': item.get('NRO. POLIZA') or item.get('nro_poliza')
            }
        
        result = procesar_comision_cobrada(payload)
        
        if result.get('success'):
            processed_count += 1
            item['procesado'] = 'procesada'
        else:
            errors.append(f"Error en registro {item.get('NRO. POLIZA', 'S/N')}: {result.get('message')}")
    
    # Save the updated matched_data back to the session if it exists
    if session_results:
        session_results['matched_data'] = matched_data
        session['comisiones_beta_resultados'] = session_results
    
    msg = f"Proceso finalizado. Registrados: {processed_count}, Pendientes creadas: {pendientes_count}."
    if errors:
        msg += f" Se encontraron {len(errors)} errores."
        print(f"DEBUG: Errores en procesamiento: {errors}")

    return jsonify({
        'success': len(errors) < len(comisiones_a_procesar), 
        'message': msg,
        'errors': errors
    })

@app.route('/borrar-usuario/<string:id>', methods=['GET'])
def borrarUsuario(id):
  if 'conectado' in session:
    if session.get("permisos")=="Administracion":
        resp = eliminarUsuario(id)
        if resp:
            flash('El Usuario fue eliminado correctamente', 'success')
            return redirect(url_for('usuarios'))
    else:
        flash('No tienes permiso para esta acciÃƒÂ²n', 'error')
        return redirect(url_for('inicio1'))
  else:
        return redirect(url_for('inicioCpanel'))
  
@app.route('/borrar-compania/<string:id>', methods=['GET'])
def borrarCompania(id):
  if 'conectado' in session:
    if session.get("permisos")=="Administracion" or session.get("permisos")=="Operaciones":
        resp = eliminarCompania(id)
        if resp:
            flash('La compaÃ±ia fue eliminada correctamente', 'success')
            return redirect(url_for('lista_company'))
    else:
        flash('No tienes permiso para esta acciÃƒÂ²n', 'error')
        return redirect(url_for('inicio1'))
  else:
        return redirect(url_for('inicioCpanel'))
    
@app.route('/borrar-asegurado/<string:id>', methods=['GET'])
def borrarAsegurado(id):
  if 'conectado' in session:
    if session.get("permisos")=="Administracion" or session.get("permisos")=="Operaciones":
        resp = eliminarAsegurado(id)
        if resp:
            flash('El asegurado fue eliminado correctamente', 'success')
            return redirect(url_for('lista_asegurado'))
    else:
        flash('No tienes permiso para esta acciÃƒÂ²n', 'error')
        return redirect(url_for('lista_asegurado'))
  else:
        return redirect(url_for('inicioCpanel'))
  
@app.route('/editar-asegurado/<string:id>', methods=['GET'])
def EditarAsegurado(id):
  if 'conectado' in session:
    if session.get("permisos")=="Administracion" or session.get("permisos")=="Operaciones"or session.get("permisos")=="dev":
        asegurado=sql_lista_asegurado(id)
        ejecutivos = lista_ejecutivosBD()
        return render_template('public/Asegurados/editarAsegurado.html', asegurado=asegurado, ejecutivos=ejecutivos)
    else:
        flash('No tienes permiso para esta acciÃƒÂ²n', 'error')
        return redirect(url_for('lista_asegurado'))
  else:
        return redirect(url_for('inicioCpanel'))
  

  
@app.route('/editar-company/<string:id>', methods=['GET'])
def EditarCompany(id):
  if 'conectado' in session:
    if session.get("permisos")=="Administracion" or session.get("permisos")=="Operaciones":
        company=sql_lista_companyU(id)   
        return render_template('public/company/editarCompany.html',company=company)
    else:
        flash('No tienes permiso para esta acciÃƒÂ²n', 'error')
        return redirect(url_for('lista_company'))
  else:
        return redirect(url_for('inicioCpanel'))
  


@app.route('/borrar-ejecutivo/<string:id>', methods=['GET'])
def borrarEjecutivo(id):
  if 'conectado' in session:
    if session.get("permisos")=="Administracion" or session.get("permisos")=="Operaciones":
        resp = eliminarEjecutivo(id)
        if resp:
            flash('El ejecutivo fue eliminado correctamente', 'success')
            return redirect(url_for('ejecutivos'))
    else:
        flash('No tienes permiso para esta acciÃƒÂ²n', 'error')
        return redirect(url_for('ejecutivos'))
  else:
        return redirect(url_for('inicioCpanel'))
    
@app.route('/borrar-poliza/<path:id>', methods=['GET'])
def borrarPoliza(id):
    if 'conectado' in session:
        if session.get("permisos")=="Administracion" or session.get("permisos")=="Operaciones":
            resp = eliminarPoliza(id)
            if resp:
                flash('la poliza fue eliminada correctamente', 'success')
                return redirect(url_for('lista_polizas'))
        else:
            flash('Tu usuario no tiene permiso para esta acciÃƒÂ²n', 'error')
            return redirect(url_for('lista_polizas'))
    else:
        return redirect(url_for('inicioCpanel'))


@app.route('/borrar-pago/<string:id>', methods=['GET'])
def borrarPago(id):
    if session.get("permisos")=="Administracion" or session.get("permisos")=="Operaciones":
        cod_polizaD = cod_poliza_P(id)
        cod_poliza = cod_polizaD[0]
        resp = eliminarPago(id)
        if resp:
            flash('El pago fue eliminada correctamente', 'success')
            return redirect(url_for('viewPagosPoliza',id=cod_poliza["cod_poliza"]))
    else:
        flash('No tienes permiso para esta acciÃƒÂ²n', 'error')
        return redirect(url_for('inicio1'))


@app.route('/borrar-empleado/<string:id_empleado>/<string:foto_empleado>', methods=['GET'])
def borrarEmpleado(id_empleado, foto_empleado):
    resp = eliminarEmpleado(id_empleado, foto_empleado)
    if resp:
        flash('El Empleado fue eliminado correctamente', 'success')
        return redirect(url_for('lista_empleados'))


@app.route("/descargar-informe-empleados/", methods=['GET'])
def reporteBD():
    if 'conectado' in session:
        return generarReporteExcel()
    else:
        flash('primero debes iniciar sesiÃ³n.', 'error')
        return redirect(url_for('inicio'))
    
 

@app.route('/registrar-poliza', methods=['GET'])
def viewFormPoliza():
    if 'conectado' in session:
        if session.get("permisos")!="Ventas" and session.get("permisos")!="Gerencia":
            base=session.get("base")
            return render_template('public/Poliza/form_poliza.html',company=sql_lista_company(),asegurado=sql_lista_aseguradosBD(),base=base)
        else:
            flash('Tu usuario no tiene los permisos para esta funciÃƒÂ²n', 'error')
            return redirect(url_for('inicio1'))
    else:
        flash('primero debes iniciar sesiÃ³n.', 'error')
        return redirect(url_for('inicio'))
    
@app.route('/registrar-asegurado', methods=['GET'])
def viewFormAsegurado():
    if 'conectado' in session:
        if session.get("permisos")!="Ventas" and session.get("permisos")!="Gerencia":
            return render_template('public/Asegurados/form_asegurado.html',ejecutivo=sql_lista_ejecutivo())
        else:
            flash('Tu usuario no tiene los permisos para esta funciÃƒÂ²n', 'error')
            return redirect(url_for('inicio1'))
    else:
        flash('primero debes iniciar sesiÃ³n.', 'error')
        return redirect(url_for('inicio'))
    

@app.route('/form-registrar-asegurado', methods=['POST'])
def formAsegurado(): 
    if 'conectado' in session:
        form_data = request.form.to_dict()
        prefijo = form_data.get('prefijo_telefono', '')
        numero = form_data.get('telefono_asegurado_numero', '')
        telefono_completo = f"{prefijo}{numero}"
        form_data['telefono_asegurado'] = telefono_completo

        resultado = procesar_form_asegurado(form_data)
        if resultado.get('success'):
            flash(resultado.get('message'), 'success')
            return redirect(url_for('viewFormPoliza'))
        else:
            flash(resultado.get('message', 'El Asegurado NO fue registrado.'), 'error')
            nombre_asegurado = request.form.get('nombre_asegurado', '')
            apellido_asegurado = request.form.get('apellido_asegurado', '')
            return render_template('public/Asegurados/form_asegurado.html', 
                                 nombre_asegurado=nombre_asegurado, 
                                 apellido_asegurado=apellido_asegurado, 
                                 ejecutivo=sql_lista_ejecutivo())
    else:
        flash('primero debes iniciar sesiÃ³n.', 'error')
        return redirect(url_for('inicio'))
    
@app.route('/form-registrar-poliza', methods=['POST'])
def formPoliza():
    if 'conectado' in session:
        resultado = procesar_form_poliza(request.form)
        if resultado.get('success'): 
            flash(resultado.get('message'), 'success')
            return redirect(url_for('viewPagosPoliza', id=resultado.get('data'), boton=0, a=0))
        else:
            flash(resultado.get('message', 'La poliza NO fue registrada.'), 'error')
            base = session.get("base")
            return render_template('public/Poliza/form_poliza.html', 
                                 company=sql_lista_company(),
                                 asegurado=sql_lista_aseguradosBD(),
                                 base=base)
    else:
        flash('primero debes iniciar sesiÃ³n.', 'error')
        return redirect(url_for('inicio'))
    
@app.route('/form-registrar-cartaAval', methods=['POST'])
def formCartaAval():
    if 'conectado' in session:
        resultado = procesar_form_CartaAval(request.form)
        if resultado.get('success'):
            flash(resultado.get('message'), 'success')
            return redirect(url_for('lista_polizas'))
        else:
            flash(resultado.get('message', 'La Carta Aval NO fue registrada.'), 'error')
            return redirect(url_for('lista_polizas'))
    else:
        flash('primero debes iniciar sesiÃ³n.', 'error')
        return redirect(url_for('inicio'))
    
@app.route('/form-registrar-reembolso', methods=['POST'])
def formReembolso():
    if 'conectado' in session:
        resultado = procesar_form_reembolso(request.form)
        if resultado.get('success'):
            flash(resultado.get('message'), 'success')
            return redirect(url_for('lista_polizas'))
        else:
            flash(resultado.get('message', 'El reembolso NO fue registrado.'), 'error')
            return redirect(url_for('lista_polizas'))
    else:
        flash('primero debes iniciar sesiÃ³n.', 'error')
        return redirect(url_for('inicio'))

@app.route('/eliminar-siniestro', methods=['POST'])
def eliminarSiniestro():
    if 'conectado' in session:
        from controllers.funciones_home import eliminar_siniestro_db
        id_siniestro = request.form.get('id')
        tipo_siniestro = request.form.get('tipo')
        
        resultado = eliminar_siniestro_db(id_siniestro, tipo_siniestro)
        return jsonify(resultado)
    else:
        return jsonify({'success': False, 'message': 'SesiÃ³n no activa.'})
    
@app.route('/form-registrar-siniestroAuto', methods=['POST'])
def form_siniestro_Auto():
    if 'conectado' in session:
        resultado = procesar_form_SiniestroAuto(request.form)
        cod_poliza = request.form.get('cod_poliza')
        if resultado.get('success'):
            flash(resultado.get('message'), 'success')
            return redirect(url_for('detallePoliza', cod_poliza=cod_poliza))
        else:
            flash(resultado.get('message', 'El siniestro NO fue registrado.'), 'error')
            return redirect(url_for('detallePoliza', cod_poliza=cod_poliza))
    else:
        flash('primero debes iniciar sesiÃ³n.', 'error')
        return redirect(url_for('inicio1'))


@app.route('/form-registrar-company', methods=['POST'])
def formCompany():
    if 'conectado' in session:
        resultado = procesar_form_company(request.form)
        if resultado.get('success'):
            flash(resultado.get('message'), 'success')
            return redirect(url_for('viewFormPoliza'))
        else:
            flash(resultado.get('message', 'La compaÃ±Ã­a NO fue registrada.'), 'error')
            return redirect(url_for('viewFormPoliza'))
    else:
        flash('primero debes iniciar sesiÃ³n.', 'error')
        return redirect(url_for('inicio'))
    
@app.route('/form-registrar-ejecutivo', methods=['POST'])
def formEjecutivo():
    if 'conectado' in session:
            resultado = procesar_form_ejecutivo(request.form)
            if request.form['Tipo']=="Personalizado":
                return redirect(url_for('EditarEjecutivo',id=resultado))
            elif resultado:
                return redirect(url_for('ejecutivos'))
            else:
                flash('El ejecutivo NO fue registrado.', 'error')
                return render_template('public/Ejecutivo/form_ejecutivo.html')
    else:
        flash('primero debes iniciar sesiÃ³n.', 'error')
        return redirect(url_for('inicio'))
    
@app.route('/form-registrar-renovacion', methods=['POST'])
def formRenovacion():
    if 'conectado' in session:
        resultado = procesar_form_renovacion(request.form)
        if resultado.get('success'):
            cod_poliza = resultado.get('data')
            flash(resultado.get('message'), 'success')
            try:
                with connectionBD() as conexion_MySQLdb:
                    with conexion_MySQLdb.cursor() as cursor:
                        # Obtener id_asegurado de la pÃ³liza
                        cursor.execute("SELECT CI_asegurado FROM poliza WHERE cod_poliza = %s", (cod_poliza,))
                        poliza_data = cursor.fetchone()
                        if not poliza_data:
                            flash('Error: No se pudo encontrar la pÃ³liza para verificar al asegurado.', 'error')
                            return redirect(url_for('lista_polizas'))

                        id_asegurado = poliza_data['CI_asegurado']

                        # Verificar los detalles del asegurado
                        cursor.execute("SELECT profesion, localidad, canal FROM asegurado WHERE CI = %s", (id_asegurado,))
                        asegurado_data = cursor.fetchone()
                        if not asegurado_data:
                             flash('Error: No se pudo encontrar al asegurado asociado a la pÃ³liza.', 'error')
                             return redirect(url_for('lista_polizas'))

                        # Comprobar si falta alguno de los campos requeridos
                        profesion = asegurado_data.get('profesion')
                        localidad = asegurado_data.get('localidad')
                        canal = asegurado_data.get('canal')

                        is_profesion_missing = not profesion or not str(profesion).strip()
                        is_localidad_missing = not localidad or not str(localidad).strip()
                        is_canal_missing = not canal or not str(canal).strip()
                        
                        if is_profesion_missing or is_localidad_missing or is_canal_missing:
                            flash('Por favor, complete los datos de riesgo del asegurado para continuar.', 'info')
                            destination_url = url_for('viewPagosPoliza', id=cod_poliza, boton=0, a=0)
                            return redirect(url_for('completar_datos_asegurado', id_asegurado=id_asegurado, next_url=destination_url))
                        else:
                            return redirect(url_for('viewPagosPoliza', id=cod_poliza, boton=0, a=0))

            except Exception as e:
                flash(f'OcurriÃ³ un error en la base de datos al verificar el perfil: {e}', 'error')
                return redirect(url_for('lista_polizas'))
        else:
            flash(resultado.get('message', 'La renovacion NO fue registrada.'), 'error')
            return redirect(url_for('lista_polizas'))
    else:
        flash('Primero debes iniciar sesiÃ³n.', 'error')
        return redirect(url_for('inicio'))

@app.route('/completar-datos-asegurado/<int:id_asegurado>', methods=['GET'])
def completar_datos_asegurado(id_asegurado):
    if 'conectado' not in session:
        flash('Primero debes iniciar sesiÃ³n.', 'error')
        return redirect(url_for('inicio'))
    
    next_url = request.args.get('next_url')
    asegurado = None
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                cursor.execute("SELECT CI, Nombre, profesion, localidad, canal FROM asegurado WHERE CI = %s", (id_asegurado,))
                asegurado = cursor.fetchone()
    except Exception as e:
         flash(f'Error al obtener datos del asegurado: {e}', 'error')
         return redirect(url_for('inicio'))

    if not asegurado:
        flash('Asegurado no encontrado.', 'error')
        return redirect(url_for('inicio'))

    return render_template('public/Asegurados/completar_datos_asegurado.html', asegurado=asegurado, next_url=next_url)

@app.route('/guardar-datos-adicionales-asegurado', methods=['POST'])
def guardar_datos_adicionales_asegurado():
    if 'conectado' not in session:
        flash('Primero debes iniciar sesiÃ³n.', 'error')
        return redirect(url_for('inicio'))
    
    id_asegurado = request.form['id_asegurado']
    print(" este es el id",id_asegurado)
    print(request.form.get('id_asegurado'))
    next_url = request.form.get('next_url')
    print("AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA")
    try:
        profesion = request.form['profesion']
        localidad = request.form['localidad']
        canal = request.form['canal']
        id_asegurado = request.form['id_asegurado']
        print(id_asegurado)

        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                print("b")
                print(profesion,localidad,canal,id_asegurado)
                sql = "UPDATE asegurado SET profesion = %s, localidad = %s, canal = %s WHERE CI = %s"
                cursor.execute(sql, (profesion, localidad, canal, id_asegurado))
                conexion_MySQLdb.commit()
                print("c")
        
        flash('Datos del asegurado actualizados correctamente.', 'success')
        return redirect(next_url or url_for('inicio'))

    except Exception as e:
        flash(f'Error al guardar los datos: {e}', 'error')
        return redirect(url_for('completar_datos_asegurado', id_asegurado=id_asegurado, next_url=next_url))

@app.route('/traspasar_poliza/<path:cod_poliza>', methods=['GET'])
def traspasar_poliza(cod_poliza):
    if 'conectado' in session:
        if session.get("permisos")!="Ventas" and session.get("permisos")!="Gerencia":
            traspasar_poliza_db(cod_poliza)
            base=session.get("base")
            return render_template('public/Poliza/form_poliza.html',company=sql_lista_company(),asegurado=sql_lista_aseguradosBD(),base=base)
        else:
            flash('Tu usuario no tiene los permisos para esta funciÃƒÂ²n', 'error')

@app.route('/anular_poliza/<path:cod_poliza>', methods=['GET'])
def anular_poliza(cod_poliza):
    if 'conectado' in session:
        if session.get("permisos")!="Ventas" and session.get("permisos")!="Gerencia":
            anular_poliza_db(cod_poliza)
            return redirect(url_for('lista_polizas'))
        else:
            flash('Tu usuario no tiene los permisos para esta funciÃƒÂ²n', 'error')
            return redirect(url_for('inicio1'))
    else:
        flash('primero debes iniciar sesiÃ³n.', 'error')
        return redirect(url_for('inicio'))
    
@app.route('/registrar-company', methods=['GET'])
def viewFormCompany():
    if 'conectado' in session:
        if session.get("permisos")!="Ventas" and session.get("permisos")!="Gerencia":
            return render_template('public/company/form_company.html')
        else:
            flash('No tienes permiso para esta acciÃƒÂ²n', 'error')
            return redirect(url_for('inicio1'))
    else:
        flash('primero debes iniciar sesiÃ³n.', 'error')
        return redirect(url_for('inicio1'))
    
@app.route('/registrar-ejecutivo', methods=['GET'])
def viewFormEjecutivo():
    if 'conectado' in session:
        if session.get("permisos")!="Ventas" and session.get("permisos")!="Gerencia":
            return render_template('public/ejecutivo/form_ejecutivo.html')
        else:
            flash('No tienes permiso para esta acciÃƒÂ²n', 'error')
            return redirect(url_for('inicio1'))
    else:
        flash('primero debes iniciar sesiÃ³n.', 'error')
        return redirect(url_for('inicio1'))
    

@app.route('/ruta-para-filtrar-comisiones', methods=['GET'])
def filtrar_comisiones_endpoint():
    start_date = request.args.get('startDate')
    end_date = request.args.get('endDate')
    ejecutivo_id = request.args.get('ejecutivo')

    # Convertir 'None' a None real si viene como string
    if start_date == 'None':
        start_date = None
    if end_date == 'None':
        end_date = None
    if ejecutivo_id == 'None':
        ejecutivo_id = None

    comisiones_filtradas = obtener_comisiones_filtradas(start_date, end_date, ejecutivo_id)

    # Convertir las fechas a string si es necesario, para que JSON.stringify no tenga problemas
    for comision in comisiones_filtradas:
        if isinstance(comision.get('fecha'), datetime.date):
            comision['fecha'] = comision['fecha'].strftime('%Y-%m-%d')
        # TambiÃ©n podrÃ­as asegurarte de que comision_bono sea float o None, no Decimal si tu ORM lo devuelve asÃ­
        if 'comision_bono' in comision and isinstance(comision['comision_bono'],Decimal):
            comision['comision_bono'] = float(comision['comision_bono']) if comision['comision_bono'] is not None else None


    # Renderizar el HTML de la tabla con los datos filtrados
    tabla_html = render_template('public/Comisiones/comisiones_filtro.html', comisiones=comisiones_filtradas) # Suponiendo que tienes un partial para la tabla

    # Devolver JSON con el HTML de la tabla y los datos brutos
    return jsonify({
        'html': tabla_html,
        'data': comisiones_filtradas
    })

@app.route('/gestionar-comisiones', methods=['GET'])
def gestionar_comisiones():
    if 'conectado' in session:
        if session.get("permisos")=="Administracion" or session.get("permisos")=="Gerencia" or session.get("permisos")=="dev":
            return render_template('public/Comisiones/gestionar_comisiones.html')
        else:
            flash('No tienes permiso para esta acciÃƒÂ²n', 'error')
            return redirect(url_for('inicio1'))
    else:
        flash('Primero debes iniciar sesiÃ³n.', 'error')
        return redirect(url_for('inicio'))

@app.route('/gestionar-comisiones-config', methods=['GET'])
def gestionar_comisiones_config():
    if 'conectado' in session:
        comisiones_config = sql_get_all_comisiones_config()
        ejecutivos = sql_lista_ejecutivo()
        company = sql_lista_company()
        base = session.get('base', 'nacional')
        return render_template('public/Comisiones/gestionar_comisiones_config.html', comisiones_config=comisiones_config, ejecutivos=ejecutivos, company=company, base=base)
    else:
        flash('Primero debes iniciar sesiÃ³n.', 'error')
        return redirect(url_for('inicio'))

@app.route('/add_comision_config', methods=['POST'])
def add_comision_config():
    if 'conectado' in session:
        if request.method == 'POST':
            data = request.form.to_dict()
            # Process 'Otro' fields
            for field in ['compania', 'ramo', 'subramo', 'producto']:
                otro_key = f'otro_{field}'
                if data.get(field) == 'Otro' and data.get(otro_key):
                    data[field] = data[otro_key]
            
            # Clean up 'otro_' fields
            keys_to_remove = [k for k in data.keys() if k.startswith('otro_')]
            for k in keys_to_remove:
                data.pop(k)

            resultado = sql_add_comision_config(data)
            if resultado:
                flash('Regla de comisiÃ³n agregada correctamente.', 'success')
            else:
                flash('Error al agregar la regla de comisiÃ³n.', 'error')
            return redirect(url_for('gestionar_comisiones_config'))
    else:
        flash('Primero debes iniciar sesiÃ³n.', 'error')
        return redirect(url_for('inicio'))

@app.route('/edit_comision_config', methods=['POST'])
def edit_comision_config():
    if 'conectado' in session:
        if request.method == 'POST':
            data = request.form.to_dict()
            # Process 'Otro' fields
            for field in ['compania', 'ramo', 'subramo', 'producto']:
                otro_key = f'otro_{field}'
                if data.get(field) == 'Otro' and data.get(otro_key):
                    data[field] = data[otro_key]
            
            # Clean up 'otro_' fields
            keys_to_remove = [k for k in data.keys() if k.startswith('otro_')]
            for k in keys_to_remove:
                data.pop(k)

            rule_id = data.pop('id')
            resultado = sql_update_comision_config(rule_id, data)
            if resultado:
                flash('Regla de comisiÃ³n actualizada correctamente.', 'success')
            else:
                flash('Error al actualizar la regla de comisiÃ³n.', 'error')
            return redirect(url_for('gestionar_comisiones_config'))
    else:
        flash('Primero debes iniciar sesiÃ³n.', 'error')
        return redirect(url_for('inicio'))

@app.route('/delete_comision_config/<int:id>', methods=['GET'])
def delete_comision_config(id):
    if 'conectado' in session:
        resultado = sql_delete_comision_config(id)
        if resultado:
            flash('Regla de comisiÃ³n eliminada correctamente.', 'success')
        else:
            flash('Error al eliminar la regla de comisiÃ³n.', 'error')
        return redirect(url_for('gestionar_comisiones_config'))
    else:
        flash('Primero debes iniciar sesiÃ³n.', 'error')
        return redirect(url_for('inicio'))


@app.route('/previsualizar-pdf-comisiones', methods=['POST'])
def previsualizar_pdf_comisiones():
    """
    Ruta para generar el PDF y devolverlo para ser mostrado en lÃ­nea (inline).
    """
    data = request.get_json()
    if not data or 'comisiones' not in data:
        return jsonify({"error": "No se recibieron datos de comisiones"}), 400

    comisiones = data['comisiones']
    rangoFechas = data.get('rangoFechas', '')
    ejecutivo = data.get('ejecutivo', '')
    nota = data.get('nota', None)  # Obtener la nota

    pdf_buffer = generar_pdf_comisiones(comisiones, rangoFechas, ejecutivo, nota=nota, for_preview=True)
    response = make_response(pdf_buffer.getvalue())
    response.headers['Content-Type'] = 'application/pdf'
    # 'inline' le dice al navegador que intente mostrar el archivo
    response.headers['Content-Disposition'] = 'inline; filename=comisiones_preview.pdf'
    return response

@app.route('/descargar-pdf-comisiones', methods=['POST'])
def descargar_pdf_comisiones():
    """
    Ruta para descargar el informe de comisiones en PDF.
    """
    data = request.get_json()
    if not data or 'comisiones' not in data:
        return jsonify({"error": "No se recibieron datos de comisiones"}), 400

    comisiones = data['comisiones']
    rangoFechas = data.get('rangoFechas', '') # Corregido
    ejecutivo = data.get('ejecutivo', '') # Corregido

    pdf_buffer = generar_pdf_comisiones(comisiones, rangoFechas, ejecutivo, for_preview=False)
    response = make_response(pdf_buffer.getvalue())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'attachment; filename=comisiones.pdf'
    return response

@app.route('/descargar-excel-comisiones', methods=['POST'])
def descargar_excel_comisiones():
    """
    Ruta para descargar el informe de comisiones en Excel.
    Recibe los datos de las comisiones en formato JSON.
    """
    data = request.get_json()
    if not data or 'comisiones' not in data:
        return jsonify({"error": "No se recibieron datos de comisiones"}), 400

    comisiones = data['comisiones']
    rangoFechas = data.get('rangoFechas', '')
    ejecutivo = data.get('ejecutivo', '')

    excel_buffer = generar_excel_comisiones(comisiones, rangoFechas, ejecutivo)
    response = make_response(excel_buffer.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=comisiones.xlsx'
    return response

@app.route('/descargar-recibo-pdf-comisiones', methods=['POST'])
def descargar_recibo_pdf_comisiones():
    """
    Ruta para descargar el recibo de comisiones de un ejecutivo en formato PDF.
    Genera un PDF con formato de recibo (similar a la referencia del usuario),
    con encabezado de empresa, datos agrupados por ejecutivo y totales.
    """
    if 'conectado' not in session:
        return jsonify({'error': 'No autorizado'}), 401

    data = request.get_json()
    if not data or 'comisiones' not in data:
        return jsonify({'error': 'No se recibieron datos de comisiones'}), 400

    comisiones = data['comisiones']
    rangoFechas = data.get('rangoFechas', '')
    ejecutivo_nombre = data.get('ejecutivo', '')
    nota = data.get('nota', None)

    try:
        pdf_buffer = generar_recibo_pdf_ejecutivo(
            comisiones,
            rangoFechas=rangoFechas,
            ejecutivo_nombre=ejecutivo_nombre,
            nota=nota
        )
        # Build a friendly filename
        import re as _re
        ejec_slug = _re.sub(r'[^a-zA-Z0-9]', '_', ejecutivo_nombre or 'ejecutivo').strip('_')
        filename = f'RECIBO_COMISIONES_{ejec_slug}.pdf'

        response = make_response(pdf_buffer.getvalue())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        return response
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Error al generar el recibo PDF: {str(e)}'}), 500


@app.route('/api/crear_pago_extra', methods=['POST'])
def api_crear_pago_extra():
    if 'conectado' in session:
        data = request.json
        cod_renovacion = data.get('cod_renovacion')
        fecha = data.get('fecha')
        monto = data.get('monto')
        moneda = data.get('moneda')
        
        try:
            with connectionBD() as conexion_MySQLdb:
                with conexion_MySQLdb.cursor() as cursor:
                    sql_insert = """
                        INSERT INTO pago (Cod_renovacion, Moneda, fecha, monto, estado, nro_cuota) 
                        VALUES (%s, %s, %s, %s, 'EN PROCESO', NULL)
                    """
                    cursor.execute(sql_insert, (cod_renovacion, moneda, fecha, monto))
                    conexion_MySQLdb.commit()
            return jsonify({'success': True})
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)})
    return jsonify({'success': False, 'error': 'No autorizado'}), 401



@app.route('/guardar_pago', methods=['POST'])
def guardar_pago():
    if request.is_json:
        data = request.get_json()
        cod_pago = data.get('cod_pago')
        estado = data.get('estado')

        # AquÃ­ puedes realizar la lÃ³gica para guardar estos datos en tu base de datos
        # utilizando el cod_pago para identificar el registro.

        try:
            with connectionBD() as conexion_MySQLdb:
                with conexion_MySQLdb.cursor() as cursor:
                    querySQL = "UPDATE pago SET estado = %s WHERE Cod_pago=%s;"
                    cursor.execute(querySQL,(estado, cod_pago))
                    conexion_MySQLdb.commit()
                
            return jsonify({'success': True, 'message': f'Cambios guardados para el pago {cod_pago}'}), 200
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500
    else:
        return jsonify({'success': False, 'error': 'La solicitud debe ser en formato JSON'}), 400
    


@app.route("/cobranza", methods=['GET'])
def lista_cobranza():
    if 'conectado' in session:
        asegurados = sql_lista_asegurados_para_filtro()
        ejecutivos = sql_lista_ejecutivo()
        return render_template('public/Cobranza/Lista_pagos.html', pagos=[], asegurados=asegurados, ejecutivos=ejecutivos)
    else:
        return redirect(url_for('inicioCpanel'))

@app.route('/api/pagos_datatable', methods=['GET'])
def pagos_datatable():
    draw = request.args.get('draw', type=int)
    start = request.args.get('start', type=int)
    length = request.args.get('length', type=int)
    
    mes = request.args.get('mes')
    ano = request.args.get('ano')
    asegurado_id = request.args.get('asegurado_id')
    estado_filtro = request.args.get('estado_filtro')

    from controllers.funciones_home import obtener_pagos_datatable
    result = obtener_pagos_datatable(start, length, mes, ano, asegurado_id, estado_filtro)
    
    # Process data slightly if necessary before returning, like handling None
    # However we configured the JSON default=str in save_temp_data, but here we use jsonify.
    # We might need to handle Decimal and datetimes
    from routers.router_home import sanitize_db_data
    sanitized_data = sanitize_db_data(result["data"])

    return jsonify({
        "draw": draw,
        "recordsTotal": result["recordsTotal"],
        "recordsFiltered": result["recordsFiltered"],
        "data": sanitized_data
    })

@app.route('/api/polizas_datatable', methods=['GET'])
def polizas_datatable():
    draw = request.args.get('draw', type=int)
    start = request.args.get('start', type=int)
    length = request.args.get('length', type=int)
    
    tipo_filtro_fecha = request.args.get('tipo_filtro_fecha')
    fecha = request.args.get('fecha')
    anio = request.args.get('anio')
    fecha_inicio = request.args.get('fecha_inicio')
    fecha_fin = request.args.get('fecha_fin')
    
    # DataTables sends 'estados[]' if it's an array, but we can also get it as a list
    estados = request.args.getlist('estados[]')
    compania_id = request.args.get('compania')
    ejecutivo_id = request.args.get('ejecutivo')
    search_value = request.args.get('search[value]')

    from controllers.funciones_home import obtener_polizas_datatable
    result = obtener_polizas_datatable(
        start, length, tipo_filtro_fecha, fecha, anio, 
        fecha_inicio, fecha_fin, estados, compania_id, 
        ejecutivo_id, search_value
    )
    
    from routers.router_home import sanitize_db_data
    sanitized_data = sanitize_db_data(result["data"])

    return jsonify({
        "draw": draw,
        "recordsTotal": result["recordsTotal"],
        "recordsFiltered": result["recordsFiltered"],
        "data": sanitized_data
    })

@app.route('/api/comisiones_datatable', methods=['GET'])
def comisiones_datatable():
    draw = request.args.get('draw', type=int)
    start = request.args.get('start', type=int)
    length = request.args.get('length', type=int)
    
    minDateStr = request.args.get('minDateStr')
    maxDateStr = request.args.get('maxDateStr')
    ejecutivo_id = request.args.get('ejecutivo_id')

    from controllers.funciones_home import obtener_comisiones_datatable
    result = obtener_comisiones_datatable(start, length, minDateStr, maxDateStr, ejecutivo_id)
    
    from routers.router_home import sanitize_db_data
    sanitized_data = sanitize_db_data(result["data"])

    return jsonify({
        "draw": draw,
        "recordsTotal": result["recordsTotal"],
        "recordsFiltered": result["recordsFiltered"],
        "data": sanitized_data
    })


@app.route("/cobro-comision", methods=['GET'])
def cobro_comision():
    if 'conectado' in session:
        # Llama a la nueva funciÃ³n para obtener las comisiones de pagos ya realizados.
        comisiones = obtener_comisiones_pagadas()
        return render_template('public/Cobranza/cobro_comision.html', comisiones=comisiones)
    else:
        flash('Primero debes iniciar sesiÃ³n.', 'error')
        return redirect(url_for('inicio'))

@app.route('/asignar-comisiones-faltantes', methods=['POST'])
def asignar_comisiones_faltantes_route():
    if 'conectado' in session:
        if session.get("permisos")=="Administracion" or session.get("permisos")=="Gerencia":
            resultado = asignar_comisiones_faltantes()
            return jsonify(resultado)
        else:
            return jsonify({'success': False, 'message': 'No tienes permiso para esta acciÃ³n.'}), 403
    else:
        return jsonify({'success': False, 'message': 'Debes iniciar sesiÃ³n.'}), 401

@app.route('/procesar-comision', methods=['POST'])
def procesar_comision_route():
    if 'conectado' in session:
        if request.is_json:
            data = request.get_json()
            resultado = procesar_comision_cobrada(data)
            return jsonify(resultado)
        return jsonify({'success': False, 'message': 'El formato de la solicitud debe ser JSON.'}), 400
    return jsonify({'success': False, 'message': 'Debes iniciar sesiÃ³n.'}), 401

@app.route('/actualizar-comision-cobrada', methods=['POST'])
def actualizar_comision_cobrada_route():
    if 'conectado' in session:
        if request.is_json:
            data = request.get_json()
            resultado = actualizar_comision_cobrada(data)
            return jsonify(resultado)
        return jsonify({'success': False, 'message': 'El formato de la solicitud debe ser JSON.'}), 400
    return jsonify({'success': False, 'message': 'Debes iniciar sesiÃ³n.'}), 401



@app.route("/proyeccion", methods=['GET'])
def proyeccion():
    if 'conectado' in session:
        current_year = datetime.datetime.now().year
        selected_year = request.args.get('year', current_year, type=int)
        proyeccion_data = obtener_proyeccion_cobranza(selected_year)
        
        # Generar una lista de aÃ±os para el selector (ej. 5 aÃ±os atrÃ¡s hasta el aÃ±o actual)
        years = list(range(current_year - 5, current_year + 1))
        
        return render_template('public/Cobranza/proyeccion.html', 
                               proyeccion=proyeccion_data, 
                               selected_year=selected_year,
                               years=years)
    else:
        return redirect(url_for('inicioCpanel'))


@app.route('/api/get-agente-commission', methods=['GET'])
def api_get_agente_commission():
    if 'conectado' in session:
        compania = request.args.get('compania')
        ramo = request.args.get('ramo')
        subramo = request.args.get('subramo')
        producto = request.args.get('producto')
        
        from controllers.funciones_home import sql_get_all_comisiones_config
        rules = sql_get_all_comisiones_config()
        
        # Buscar regla de Agente
        found_pct = None
        for r in rules:
            if (r.get('compania') == compania and 
                r.get('ramo') == ramo and 
                (not r.get('subramo') or r.get('subramo') == subramo) and
                r.get('producto') == producto and
                r.get('tipo_ejecutivo', '').upper() == 'AGENTE' and
                (not r.get('cod_ejecutivo') or r.get('cod_ejecutivo') == 0)):
                
                # Tomar el primer porcentaje (primera emisiÃ³n)
                pcts = r.get('porcentajes', '0').split(',')
                found_pct = pcts[0]
                break
        
        return jsonify({'success': True, 'porcentaje': found_pct})
    return jsonify({'success': False, 'message': 'No autoirzado'}), 401

@app.route('/api/configurar-regla-rapida', methods=['POST'])
def api_configurar_regla_rapida():
    if 'conectado' in session:
        data = request.json
        # Convertir datos recibidos
        porcentaje_original = data.get('porcentaje')
        porcentaje_decimal = None
        if porcentaje_original is not None:
            try:
                # Si el usuario envÃ­a "8.5", lo guardamos como "0.085" para consistencia con el resto de la DB
                porcentaje_decimal = float(str(porcentaje_original).replace(',', '.')) / 100
            except:
                porcentaje_decimal = porcentaje_original

        config_data = {
            'compania': data.get('compania'),
            'ramo': data.get('ramo'),
            'subramo': data.get('subramo'),
            'producto': data.get('producto'),
            'tipo_ejecutivo': data.get('tipo', 'Particular'),
            'porcentajes': str(porcentaje_decimal),
            'cod_ejecutivo': data.get('cod_ejecutivo')
        }
        
        from controllers.funciones_home import sql_add_comision_config
        res = sql_add_comision_config(config_data)
        if res:
            return jsonify({'success': True})
        else:
            return jsonify({'success': False, 'message': 'Error al guardar en DB'})
            
    return jsonify({'success': False, 'message': 'No autoirzado'}), 401


@app.route('/descargar-excel-pagos')
def descargar_excel_pagos():
    pago = request.args.get('pago')
    comision = request.args.get('comision')
    mes = request.args.get('mes')
    asegurado_id = request.args.get('asegurado_id')
    ano = request.args.get('ano')
    ejecutivo_id = request.args.get('ejecutivo_id')
    fecha_inicio = request.args.get('fecha_inicio')
    fecha_fin = request.args.get('fecha_fin')
    estado_filtro = request.args.get('estado_filtro')
    
    pagos = obtener_pagos_filtrados(pago, comision, mes, asegurado_id, ano, ejecutivo_id, fecha_inicio, fecha_fin, estado_filtro)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Pagos"

    # TÃ­tulo
    ws.merge_cells('A1:I1')
    title_cell = ws['A1']
    title_cell.value = "Reporte de Pagos de Cobranza"
    title_cell.font = Font(bold=True, size=18, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    ws.row_dimensions[1].height = 30
    
    # Encabezados
    headers = ["CÃ³digo de PÃ³liza", "Asegurado", "CÃ©dula", "Moneda", "Fecha Recibo", "Monto", "Estado", "Nro Cuota", "Fecha Pagada"]
    # Escribir encabezados en la fila 2
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(bottom=Side(style='thin'))

    # Datos
    for i, pago_item in enumerate(pagos, start=3):
        nombre_completo = f"{pago_item.get('Nombre_asegurado', '')} {pago_item.get('Apellido_asegurado', '')}"
        ws.cell(row=i, column=1, value=pago_item.get('Cod_poliza') or pago_item.get('cod_poliza'))
        ws.cell(row=i, column=2, value=nombre_completo)
        ws.cell(row=i, column=3, value=pago_item.get('CI'))
        ws.cell(row=i, column=4, value=pago_item.get('moneda'))
        ws.cell(row=i, column=5, value=pago_item.get('fecha'))
        ws.cell(row=i, column=6, value=pago_item.get('monto'))
        ws.cell(row=i, column=7, value=pago_item.get('estado'))
        ws.cell(row=i, column=8, value=pago_item.get('nro_cuota'))
        ws.cell(row=i, column=9, value=pago_item.get('fecha_pagada'))

    # Ajustar ancho de columnas
    for i, col in enumerate(ws.columns):
        max_length = 0
        column_letter = get_column_letter(i + 1)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=reporte_pagos.xlsx'
    return response

@app.route('/filtrar-pagos')
def filtrar_pagos():
    pago = request.args.get('pago')
    comision = request.args.get('comision')
    mes = request.args.get('mes')
    asegurado_id = request.args.get('asegurado_id')
    ano = request.args.get('ano')
    ejecutivo_id = request.args.get('ejecutivo_id')
    fecha_inicio = request.args.get('fecha_inicio')
    fecha_fin = request.args.get('fecha_fin')
    estado_filtro = request.args.get('estado_filtro')
    
    # Filtra las comisiones en tu base de datos usando las fechas recibidas
    pagos_filtrados = obtener_pagos_filtrados(pago, comision, mes, asegurado_id, ano, ejecutivo_id, fecha_inicio, fecha_fin, estado_filtro)
    
    # Renderiza el HTML de la tabla con los datos filtrados
    tabla_html = render_template('public/Cobranza/pagos_filtro.html', pagos=pagos_filtrados)

    # Devolver JSON con el HTML de la tabla y los datos brutos
    return jsonify({
        'html': tabla_html,
        'data': pagos_filtrados
    })

@app.route('/guardar_nota', methods=['POST'])
def guardar_nota():
    data = request.get_json()
    titulo = data.get('titulo')
    observaciones = data.get('Observaciones') 
    cod=data.get('Cod_CartaAval')
    tipo=data.get('tipo')
    fecha=datetime.date.today()

    if titulo and observaciones:
        newid = NotaCartaAval(titulo, observaciones, id,cod,tipo,fecha) # Asocia la nota a la carta aval actual
        print(newid)
        if tipo==1:
            return jsonify({'success': True, 'nota': {'idnota_cartaAval': newid, 'titulo': titulo, 'Observaciones': observaciones}}) # Devuelve la nota guardada con su ID, tÃ­tulo y observaciones
        elif tipo==2:
             return jsonify({'success': True, 'nota': {'idnota_Reembolso': newid, 'titulo': titulo, 'Observaciones': observaciones}}) # Devuelve la nota guardada con su ID, tÃ­tulo y observaciones
        else:
             return jsonify({'success': True, 'nota': {'idnota_Auto': newid, 'titulo': titulo, 'Observaciones': observaciones}}) # Devuelve la nota guardada con su ID, tÃ­tulo y observaciones    
    else:
        return jsonify({'success': False, 'error': 'Faltan datos de la nota'}), 400
    

@app.route('/form-registrar-usuario', methods=['POST'])
def cpanelResgisterUserBD1():
    if request.method == 'POST' and 'name_surname' in request.form and 'pass_user' in request.form:
        name_surname = request.form['name_surname']
        email_user = request.form['email_user']
        pass_user = request.form['pass_user']
        permisos = request.form['permisos']

        resultData = recibeInsertRegisterUser1(name_surname, email_user, pass_user, permisos)
        if resultData.get('success'):
            flash(resultData.get('message'), 'success')
            return redirect(url_for('usuarios'))
        else:
            flash(resultData.get('message'), 'error')
            return redirect(url_for('usuarios'))
    else:
        flash('el mÃ©todo HTTP es incorrecto', 'error')
        return redirect(url_for('usuarios'))
    
# Crear cuenta de usuario
@app.route('/registar-user', methods=['GET'])
def registrar_usuario():
    if session.get("permisos")=="Administracion" or session.get("permisos")=="dev":
        return render_template('public/usuarios/auth_register.html')
    else:
        flash('No tiene permisos para esta accion', 'error')
        return redirect(url_for('usuarios'))

@app.route('/upload', methods=['GET'])
def upload():
    if session.get("permisos")!="Ventas" or session.get("permisos")!="Gerencia":
        return render_template('public/upload.html')
    else:
        flash('No tiene permisos para esta accion', 'error')
        return redirect(url_for('inicioCpanel'))
    

 

@app.route('/upload_csv', methods=['GET', 'POST'])
def upload_csv():
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No se seleccionÃ³ ningÃºn archivo.', 'error')
            return redirect(request.url)

        file = request.files['file']
        if file.filename == '':
            flash('No se seleccionÃ³ ningÃºn archivo.', 'error')
            return redirect(request.url)

        if file and file.filename.endswith('.xlsx'):
            # Convert FileStorage to BytesIO to store in session or process easily
            from io import BytesIO
            file_content = file.read()
            # Note: We can't easily store the whole workbook in session. 
            # We'll process it first, then store the list of candidate rows.
            result = procesar_excel_preview(BytesIO(file_content))
            
            if 'error' in result:
                flash(result['error'], 'error')
                return redirect(request.url)
            
            if result.get('success'):
                # Store the candidates in a temp file to avoid session size limits
                temp_filename = save_temp_data(result['data'])
                session['upload_candidates_file'] = temp_filename
                return render_template('public/preview_upload.html', candidates=result['data'], ejecutivos=result.get('ejecutivos', []))
        else:
            flash('Tipo de archivo no permitido. Por favor, sube un archivo .xlsx', 'error')

        return redirect(url_for('upload'))

    return render_template('public/upload.html')


@app.route('/api/upload_csv_stream', methods=['POST'])
def upload_csv_stream():
    """
    Endpoint de streaming: procesa el Excel fila a fila y devuelve NDJSON.
    Cada lÃ­nea es un JSON con 'type': 'init' | 'row' | 'done' | 'error'.
    Al terminar guarda los datos en temp para que confirm_upload los use.
    """
    from flask import stream_with_context, Response
    from io import BytesIO
    import json as _json

    if 'file' not in request.files:
        return jsonify({'error': 'No se seleccionÃ³ ningÃºn archivo.'}), 400

    file = request.files['file']
    if not file.filename.endswith('.xlsx'):
        return jsonify({'error': 'Tipo de archivo no permitido. Solo .xlsx'}), 400

    file_bytes = file.read()
    # Container to pass temp_filename out of the generator closure
    result_holder = []

    def generate():
        from controllers.funciones_home import procesar_excel_stream
        collected_rows = []

        for msg in procesar_excel_stream(BytesIO(file_bytes)):
            if msg.get('type') == 'row':
                collected_rows.append(msg)
            yield _json.dumps(msg, default=str) + '\n'

        # Save to temp file â€“ this runs inside app context thanks to stream_with_context
        if collected_rows:
            temp_filename = save_temp_data(collected_rows)
            result_holder.append(temp_filename)
            # Also persist in session here (we are still in request context via stream_with_context)
            session['upload_candidates_file'] = temp_filename
        yield _json.dumps({'type': 'saved', 'ok': bool(collected_rows)}, default=str) + '\n'

    return Response(stream_with_context(generate()), mimetype='application/x-ndjson')




@app.route('/confirm_upload', methods=['POST'])
def confirm_upload():
    selected_indices = request.form.getlist('selected_rows')
    temp_filename = session.get('upload_candidates_file')
    candidates = load_temp_data(temp_filename) if temp_filename else None

    if not candidates or not selected_indices:
        flash('No hay registros seleccionados o la sesiÃ³n expirÃ³.', 'warning')
        return redirect(url_for('upload'))

    # Convert selected_indices to integers
    selected_indices = [int(i) for i in selected_indices]
    
    # Filter candidates by selected indices
    to_insert = [candidates[i] for i in selected_indices]
    
    # Call the insertion logic
    resultado = insertar_registros_excel(to_insert)
    
    if resultado.get('success'):
        flash(f'Proceso completado: {resultado["inserted"]} insertados, {resultado["errors"]} errores.', 'success')
        session.pop('upload_candidates_file', None)
        return redirect(url_for('inicio1'))
    else:
        flash(f'Error en el proceso: {resultado.get("error")}', 'error')
        return render_template('public/preview_upload.html', candidates=candidates)

@app.route('/process_single_generic', methods=['POST'])
def process_single_generic():
    if 'conectado' in session:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'No data provided'}), 400
        
        from controllers.funciones_home import insertar_unico_registro_generico
        resultado = insertar_unico_registro_generico(data)
        return jsonify(resultado)
    return jsonify({'success': False, 'message': 'No autorizado'}), 401






 

# FunciÃ³n auxiliar para convertir tipos de datos de la DB a tipos serializables por JSON
def sanitize_db_data(data):
    """
    Convierte recursivamente objetos Decimal y datetime/date en los resultados de la base de datos
    a tipos amigables para JSON (float y string ISO 8601).
    """
    # =====================================================================================
    # SOLUCIÃƒâ€œN TEMPORAL/DEPURACIÃƒâ€œN: IMPORTANDO DATETIME Y DATE DENTRO DE LA FUNCIÃƒâ€œN
    # Si el error persistÃ­a con la importaciÃ³n en la parte superior,
    # significa que hay un problema de carga de mÃ³dulos (ej. importaciÃ³n circular).
    # Esta importaciÃ³n "garantiza" que datetime y date estÃ©n disponibles en este Ã¡mbito.
    # =====================================================================================
    from datetime import datetime, date 
    print("DEBUG: Importando datetime y date DENTRO de sanitize_db_data para asegurar su disponibilidad.")

    if isinstance(data, dict):
        return {k: sanitize_db_data(v) for k, v in data.items()}
    elif isinstance(data, list):
        return [sanitize_db_data(elem) for elem in data]
    elif isinstance(data, Decimal):
        return float(data)
    elif isinstance(data, (datetime, date)): 
        return data.isoformat()
    else:
        return data


@app.route('/EditarEjecutivo/<string:id>', methods=['GET', 'POST'])
def EditarEjecutivo(id):
    ejecutivo_data = None
    productos_seleccionados_db = []

    try:
        # Conectar a la base de datos para obtener los datos del ejecutivo y sus productos.
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                # 1. Obtener los datos del ejecutivo
                query_ejecutivo = "SELECT * FROM ejecutivo WHERE cod_ejecutivo = %s"
                cursor.execute(query_ejecutivo, (id,))
                ejecutivo_data = cursor.fetchone()

                if not ejecutivo_data:
                    print(f"Ejecutivo con cÃ³digo {id} no encontrado.")
                    # Si no se encuentra el ejecutivo, se pasa 'None' y el template maneja esto.
                    return render_template('public/Ejecutivo/editarEjecutivo.html', ejecutivo=None, user_selected_products=[])

                # 2. Obtener los productos seleccionados para este ejecutivo
                query_productos = "SELECT cod_ejecutivo, compania, ramo, subramo, producto, comision_bono FROM comisiones_ejecutivos WHERE cod_ejecutivo = %s"
                cursor.execute(query_productos, (id,))
                productos_seleccionados_db = cursor.fetchall()
                print(f"Productos seleccionados cargados para {id}: {productos_seleccionados_db}")

    except Exception as e:
        print(f"OcurriÃ³ un error al cargar datos del ejecutivo o productos seleccionados: {e}")
        # AsegÃºrate de que 'error_page.html' exista o usa una plantilla genÃ©rica.
        return render_template('error_page.html', error_message="Error al cargar los datos. Por favor, intÃ©ntelo de nuevo mÃ¡s tarde.")

    # Convertir los datos de la DB a tipos amigables para JSON antes de pasarlos al template
    if ejecutivo_data:
        ejecutivo_data = sanitize_db_data(ejecutivo_data)

    if productos_seleccionados_db:
        productos_seleccionados_db = sanitize_db_data(productos_seleccionados_db)

    if request.method == 'GET':
        # Renderiza el template, pasando los datos del ejecutivo y sus productos seleccionados (ya sanitizados)
        return render_template('public/Ejecutivo/editarEjecutivo.html',
                               ejecutivo=ejecutivo_data,
                               user_selected_products=productos_seleccionados_db)
    elif request.method == 'POST':
        # Esta parte manejarÃ­a la actualizaciÃ³n de los campos normales del ejecutivo
        # (Nombre, Apellido, etc.). AsegÃºrese de que su funciÃ³n 'procesar_actualizacion_form_ejecutivo'
        # sea llamada aquÃ­ si es necesario, y que devuelva una respuesta vÃ¡lida (ej. un redirect).
        print(f"Procesando actualizaciÃ³n de datos personales para ejecutivo {id}...")
        # Ejemplo:
        # result = procesar_actualizacion_form_ejecutivo(request)
        # if result:
        #    flash("Datos del ejecutivo actualizados correctamente", "success")
        # else:
        #    flash("Error al actualizar los datos del ejecutivo", "danger")
        return redirect(url_for('main.listaEjecutivo')) # Redirige despuÃ©s de actualizar

@app.route('/submit_selection', methods=['POST'])
def submit_selection():
    try:
        data = request.get_json()
        
        # Asumimos que el cod_ejecutivo siempre viene en el payload, incluso si no hay productos
        cod_ejecutivo_a_actualizar = None
        if data and len(data) > 0:
            cod_ejecutivo_a_actualizar = data[0].get('cod_ejecutivo')
        elif request.json and 'cod_ejecutivo' in request.json: # Fallback por si data es vacÃ­o
            cod_ejecutivo_a_actualizar = request.json.get('cod_ejecutivo')

        if not cod_ejecutivo_a_actualizar:
            return jsonify({"error": "No se pudo identificar el cÃ³digo de ejecutivo para procesar la selecciÃ³n."}), 400

        # Conectar a la base de datos para realizar las operaciones de eliminaciÃ³n e inserciÃ³n.
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                # 1. Eliminar todas las selecciones existentes para este 'cod_ejecutivo'.
                query_delete = "DELETE FROM comisiones_ejecutivos WHERE cod_ejecutivo = %s"
                cursor.execute(query_delete, (cod_ejecutivo_a_actualizar,))
                print(f"Eliminadas selecciones anteriores para el ejecutivo: {cod_ejecutivo_a_actualizar}")

                # 2. Insertar las nuevas selecciones recibidas del frontend.
                # Solo insertar si hay productos reales en la lista enviada desde el frontend
                products_to_insert = []
                for product_item in data:
                    # Nos aseguramos de que cada item sea un producto completo y no solo el cod_ejecutivo de la limpieza
                    if product_item.get('compania') and product_item.get('ramo') and product_item.get('producto') is not None:
                        products_to_insert.append((
                            product_item.get('cod_ejecutivo'),
                            product_item.get('compania'),
                            product_item.get('ramo'),
                            product_item.get('subramo'), # subramo puede ser None, lo cual es manejado por MySQL
                            product_item.get('producto'),
                            product_item.get('comision_bono')
                        ))
                
                if products_to_insert: # Solo ejecutamos INSERT si hay productos para insertar
                    query_insert = """
                        INSERT INTO comisiones_ejecutivos
                            (cod_ejecutivo, compania, ramo, subramo, producto, comision_bono)
                        VALUES (%s, %s, %s, %s, %s, %s)
                    """
                    cursor.executemany(query_insert, products_to_insert)
                    print(f"Insertados {len(products_to_insert)} nuevos productos para el ejecutivo: {cod_ejecutivo_a_actualizar}")
                else:
                    print(f"No hay nuevos productos para insertar para el ejecutivo: {cod_ejecutivo_a_actualizar}")

                # Confirmar todos los cambios (eliminaciones e inserciones) en la base de datos.
                conexion_MySQLdb.commit()

        return jsonify({"message": "SelecciÃ³n de productos guardada con Ã©xito."}), 200

    except Exception as e:
        print(f"OcurriÃ³ un error en la ruta submit_selection: {e}")
        return jsonify({"error": f"Error interno del servidor al procesar la selecciÃ³n: {e}"}), 500
    

@app.route('/recargar_riesgo/<Cod_renovacion>', methods=['POST'])
def recargar_riesgo(Cod_renovacion):
    """
    Ruta de la API para recalcular el riesgo de una pÃ³liza.
    Llama a la funciÃ³n de lÃ³gica de negocio y devuelve una respuesta JSON.
    """
    print(f"Recibida solicitud para recalcular riesgo de pÃ³liza: {Cod_renovacion}")
    
    success, data, status_code = cargar_riesgo(Cod_renovacion) # Llama a la funciÃ³n de lÃ³gica

    if success:
        nuevo_riesgo_calculado = data # data es el riesgo en este caso
        return jsonify({'success': True, 'message': 'Riesgo recalculado con Ã©xito', 'nuevo_riesgo': nuevo_riesgo_calculado})
    else:
        # data es el mensaje de error, status_code es el cÃ³digo HTTP
        return jsonify({'success': False, 'message': data}), status_code

@app.route('/upload_mercantil', methods=['GET', 'POST'])
def view_upload_mercantil():
    if 'conectado' in session:
        if request.method == 'POST':
            if 'file' not in request.files:
                flash('No hay archivo', 'danger')
                return redirect(request.url)
            file = request.files['file']
            if file.filename == '':
                flash('No se seleccionÃ³ ningÃºn archivo', 'danger')
                return redirect(request.url)
            
            if file:
                result = procesar_mercantil_preview(file)
                if result['success']:
                    return render_template('public/mercantil/preview_mercantil.html', data=result['data'], ejecutivos=result.get('ejecutivos', []), companias=result.get('companias', []))
                else:
                    flash(f"Error procesando archivo: {result['error']}", 'danger')
                    return redirect(request.url)
        
        return render_template('public/mercantil/upload_mercantil.html')
    else:
        return redirect(url_for('inicio'))

@app.route('/confirm_mercantil_upload', methods=['POST'])
def confirm_mercantil_upload():
    if 'conectado' in session:
        selected_indices = request.form.getlist('selected_indices')
        all_data_json = request.form.get('all_data')
        
        if all_data_json and selected_indices:
            all_data = json.loads(all_data_json)
            data_to_insert = [all_data[int(i)] for i in selected_indices]
            
            result = insertar_mercantil_data(data_to_insert)
            if result['success']:
                flash(f"Carga completada: {result['inserted']} pÃ³lizas procesadas, {result['errors']} errores.", 'success')
            else:
                flash(f"Error en la inserciÃ³n: {result['error']}", 'danger')
        else:
            flash("No se seleccionaron datos o no hay datos para procesar", 'warning')
            
        return redirect(url_for('view_upload_mercantil'))
    else:
        return redirect(url_for('inicio'))

@app.route('/process_single_mercantil', methods=['POST'])
def process_single_mercantil():
    if 'conectado' in session:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'No data provided'}), 400
        
        from controllers.funciones_home import insertar_unico_registro_mercantil
        resultado = insertar_unico_registro_mercantil(data)
        return jsonify(resultado)
    return jsonify({'success': False, 'message': 'No autorizado'}), 401


@app.route('/dashboard-bi-master')
def dashboard_bi_master():
    if 'conectado' in session:
        filtros_data = obtener_filtros_dashboard()
        ano_actual = datetime.datetime.now().year
        datos = obtener_datos_dashboard({'ano': ano_actual})
        return render_template('public/Dashboard/dashboard_bi_master.html', 
                               filtros=filtros_data, 
                               datos_iniciales=datos,
                               ano_actual=ano_actual)
    else:
        return redirect(url_for('index'))


@app.route('/dashboard-reportes')
def dashboard_reportes():
    if 'conectado' in session:
        filtros_data = obtener_filtros_dashboard()
        # Inicialmente cargamos el aÃ±o actual si existe
        ano_actual = datetime.datetime.now().year
        datos = obtener_datos_dashboard({'ano': ano_actual})
        return render_template('public/Dashboard/dashboard_reportes.html', 
                               filtros=filtros_data, 
                               datos_iniciales=datos,
                               ano_actual=ano_actual)
    else:
        return redirect(url_for('index'))


@app.route('/api/dashboard-data')
def api_dashboard_data():
    if 'conectado' in session:
        filtros = {
            'ano': request.args.get('ano'),
            'mes': request.args.get('mes'),
            'compania_id': request.args.get('compania_id'),
            'ramo': request.args.get('ramo'),
            'ejecutivo_id': request.args.get('ejecutivo_id')
        }
        datos = obtener_datos_dashboard(filtros)
        return jsonify(datos)
    else:
        return jsonify({'error': 'No autorizado'}), 401

@app.route('/descargar-plantilla/<tipo>')
def descargar_plantilla(tipo):
    if 'conectado' in session:
        buffer = generar_plantilla_excel(tipo)
        if buffer:
            filename = f"plantilla_{tipo}.xlsx"
            return send_file(
                buffer,
                as_attachment=True,
                download_name=filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        else:
            flash('Tipo de plantilla no vÃ¡lido.', 'error')
            return redirect(url_for('inicio1'))
    else:
        return redirect(url_for('inicio'))

@app.route('/upload-siniestros')
def upload_siniestros():
    if 'conectado' in session:
        return render_template('public/siniestros/upload_siniestros.html', siniestros_data=None)
    return redirect(url_for('inicio'))

@app.route('/upload-siniestros-csv', methods=['POST'])
def upload_siniestros_csv():
    if 'conectado' in session:
        if 'file' not in request.files:
            flash('No se seleccionÃ³ ningÃºn archivo.', 'error')
            return redirect(url_for('upload_siniestros'))
        
        file = request.files['file']
        if file.filename == '':
            flash('No se seleccionÃ³ ningÃºn archivo.', 'error')
            return redirect(url_for('upload_siniestros'))

        if file:
            filename = secure_filename(file.filename)
            from io import BytesIO
            from controllers.funciones_home import procesar_siniestros_excel
            
            file_content = file.read()
            result = procesar_siniestros_excel(BytesIO(file_content))
            
            if not result.get('success'):
                flash(result.get('error', 'Error desconocido al procesar el archivo'), 'error')
                return redirect(url_for('upload_siniestros'))
                
            return render_template('public/siniestros/upload_siniestros.html', 
                                   siniestros_data=result.get('data'), 
                                   filename=filename)
            
    return redirect(url_for('inicio'))

@app.route('/confirmar-carga-siniestros', methods=['POST'])
def confirmar_carga_siniestros():
    if 'conectado' in session:
        import json
        from controllers.funciones_home import guardar_siniestros_batch
        
        # We need to get the data from the form (JSON stringified)
        data_json = request.form.get('siniestros_data')
        if not data_json:
            flash('No hay datos para procesar.', 'error')
            return redirect(url_for('upload_siniestros'))
            
        try:
            siniestros_data = json.loads(data_json)
            result = guardar_siniestros_batch(siniestros_data)
            
            if result.get('success'):
                flash(f'Carga completada con Ã©xito. Se registraron {result["count"]} siniestros.', 'success')
                if result.get('errors'):
                    for err in result['errors'] :
                        flash(f'Aviso: {err}', 'warning')
            else:
                flash(f'Error al guardar los datos: {result.get("error")}', 'error')
                
        except Exception as e:
            flash(f'Error al procesar la confirmaciÃ³n: {str(e)}', 'error')
            
        return redirect(url_for('upload_siniestros'))
        
    return redirect(url_for('inicio'))

@app.route('/procesar-siniestro-individual', methods=['POST'])
def procesar_siniestro_individual():
    if 'conectado' in session:
        from controllers.funciones_home import guardar_siniestro_individual_db
        data = request.json
        sheet_name = data.get('sheet_name')
        row_data = data.get('row_data')
        
        if not sheet_name or not row_data:
            return jsonify({'success': False, 'message': 'Datos incompletos.'})
            
        result = guardar_siniestro_individual_db(sheet_name, row_data)
        return jsonify(result)
    return jsonify({'success': False, 'message': 'SesiÃ³n no activa.'})
@app.route('/reporte-sudaseg', methods=['GET'])
def reporte_sudaseg():
    if 'conectado' in session:
        mes = request.args.get('mes')
        ano = request.args.get('ano')
        datos = []
        if mes and ano:
            datos = sql_reporte_sudaseg(int(mes), int(ano))
        return render_template('public/reportes/reporte_sudaseg.html', datos=datos, mes=mes, ano=ano)
    else:
        return redirect(url_for('inicio'))

@app.route('/descargar-excel-sudaseg', methods=['GET'])
def descargar_excel_sudaseg():
    if 'conectado' in session:
        mes = request.args.get('mes')
        ano = request.args.get('ano')
        if not mes or not ano:
            flash("Debe seleccionar un mes y aÃ±o para exportar.", "warning")
            return redirect(url_for('reporte_sudaseg'))
        datos = sql_reporte_sudaseg(int(mes), int(ano))
        if not datos:
            flash("No hay datos para exportar en el perÃ­odo seleccionado.", "info")
            return redirect(url_for('reporte_sudaseg'))
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Sudaseg_{mes}_{ano}"
        headers = ['DescripciÃ³n', 'Identificador del Anexo', 'Nombre o denominaciÃ³n de la Aseguradora', 'Registro de InformaciÃ³n Financiera Fiscal (RIF) Asegurado', 'Fecha de EmisiÃ³n del Recibo', 'Fecha de Cobro del Recibo', 'NÃºmero de Recibo', 'NÃºmero de PÃ³liza o Contrato', 'Monto de la Prima o Cuota', 'Monto de la ComisiÃ³n', 'Porcentaje de la ComisiÃ³n', 'Ramos', 'Fecha de DepÃ³sito de la Transferencia a La Cuenta Especial', 'NÃºmero de Cheque o Transferencia entregado por el asegurado', 'CÃ³digo de la entidad bancaria', 'Fecha de Pago']
        ws.append(headers)
        header_fill = openpyxl.styles.PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        header_font = openpyxl.styles.Font(bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        for d in datos:
            row = [d['descripcion'], d['identificador_anexo'], d['nombre_aseguradora'], d['rif_asegurado'], d['fecha_emision_recibo'], d['fecha_cobro_recibo'], d['numero_recibo'], d['numero_poliza'], d['monto_prima'], d['monto_comision'], d['porcentaje_commission'], d['ramos'], d['fecha_deposito_transferencia'], d['numero_cheque_transferencia'], d['codigo_entidad_bancaria'], d['fecha_pago']]
            ws.append(row)
        for col in ws.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            ws.column_dimensions[column_letter].width = max_length + 2
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"Reporte_Sudaseg_{mes}_{ano}.xlsx"
        return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    else:
        return redirect(url_for('inicio'))

@app.route('/polizas-pendientes')
def polizas_pendientes():
    if 'conectado' in session:
        polizas = obtener_polizas_pendientes()
        return render_template('public/Comisiones/polizas_pendientes.html', polizas=polizas)
    return redirect(url_for('inicio'))

@app.route('/eliminar-poliza-pendiente/<int:id_pendiente>', methods=['POST'])
def eliminar_pendiente(id_pendiente):
    if 'conectado' in session:
        if eliminar_poliza_pendiente(id_pendiente):
            return jsonify({'success': True, 'message': 'Registro descartado correctamente.'})
        return jsonify({'success': False, 'message': 'No se pudo eliminar el registro.'}), 500
    return jsonify({'success': False, 'message': 'No autorizado'}), 401

@app.route('/verificar-poliza-pendiente/<int:id_pendiente>', methods=['POST'])
def verificar_pendiente(id_pendiente):
    if 'conectado' in session:
        detalle = obtener_detalle_poliza_pendiente(id_pendiente)
        if not detalle:
            return jsonify({'success': False, 'message': 'Registro no encontrado.'}), 404
        
        # Verificar si la pÃ³liza ya existe en la BD
        with connectionBD() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT cod_poliza FROM poliza WHERE cod_poliza = %s", (detalle['nro_poliza'],))
                existe = cur.fetchone()
                
        if existe:
            return jsonify({'success': True, 'existe': True, 'message': 'La pÃ³liza ya existe en el sistema.'})
        return jsonify({'success': True, 'existe': False, 'message': 'La pÃ³liza aÃºn no estÃ¡ registrada.'})
    return jsonify({'success': False, 'message': 'No autorizado'}), 401

@app.route('/gestionar-poliza-pendiente/<int:id_pendiente>')
def gestionar_pendiente(id_pendiente):
    if 'conectado' in session:
        detalle = obtener_detalle_poliza_pendiente(id_pendiente)
        if not detalle:
            flash('Registro no encontrado.', 'error')
            return redirect(url_for('polizas_pendientes'))
        
        # Obtener datos para el formulario (compaÃ±Ã­as, ejecutivos, ramos)
        # Reusamos lÃ³gica existente en ruta de crear poliza si es posible o cargamos lo mÃ­nimo
        with connectionBD() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT * FROM compania ORDER BY Nombre")
                companias = cur.fetchall()
                cur.execute("SELECT * FROM ejecutivo ORDER BY Nombre")
                ejecutivos = cur.fetchall()
                cur.execute("SELECT DISTINCT Ramo FROM poliza ORDER BY Ramo")
                ramos = [r['Ramo'] for r in cur.fetchall()]

        return render_template('public/Comisiones/form_poliza_pendiente.html', 
                               detalle=detalle, 
                               companias=companias, 
                               ejecutivos=ejecutivos,
                               ramos=ramos,
                               base=session.get('base'))
    return redirect(url_for('inicio'))

@app.route('/registrar-desde-pendiente', methods=['POST'])
def registrar_desde_pendiente():
    if 'conectado' in session:
        resultado = procesar_registro_desde_pendiente(request.form)
        if resultado.get('success'):
            flash(resultado.get('message'), 'success')
            # Redirigir a los pagos de la nueva pÃ³liza
            return redirect(url_for('viewPagosPoliza', id=resultado.get('data')))
        else:
            flash(resultado.get('message'), 'error')
            return redirect(url_for('polizas_pendientes'))
    return redirect(url_for('inicio'))
