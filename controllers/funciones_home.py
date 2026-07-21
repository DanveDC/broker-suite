# Para subir archivo tipo foto al servidor
from werkzeug.utils import secure_filename
import uuid
from datetime import datetime, timedelta
from datetime import date;
from dateutil.relativedelta import relativedelta
from io import BytesIO
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib import colors
import unicodedata
import difflib

from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch 
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from flask import session, flash, jsonify

from conexion.conexionBD import connectionBD  # Conexión a BD

import datetime
import re
import os

from os import remove  # Modulo  para remover archivo
from os import path  # Modulo para obtener la ruta o directorio


import openpyxl  # Para generar el excel
# biblioteca o modulo send_file para forzar la descarga
from flask import send_file
from werkzeug.security import check_password_hash

import re
# Para encriptar contraseña generate_password_hash
from werkzeug.security import generate_password_hash

from pyDolarVenezuela.pages import AlCambio, BCV, CriptoDolar, DolarToday, ExchangeMonitor, EnParaleloVzla, Italcambio
import pandas as pd
import numpy as np

def sanitizar_numero(valor, default=0.0):
    """
    Convierte un valor a float de forma robusta, eliminando:
    - Caracteres Unicode invisibles (\xa0, \ufeff, zero-width chars)
    - Símbolos de moneda ($, Bs, USD, etc.)
    - Separadores de miles (puntos o comas según contexto)
    - Espacios de cualquier tipo
    Retorna default si no se puede parsear.
    """
    if valor is None:
        return default
    if isinstance(valor, (int, float)):
        return float(valor)
    s = str(valor)
    # Eliminar BOM, zero-width chars, non-breaking spaces y otros invisibles
    s = re.sub(r'[\u200b\u200c\u200d\ufeff\u00a0\u2000-\u200a\u202f\u205f\u3000]', '', s)
    # Eliminar símbolos de moneda y letras (Bs, USD, $, etc.)
    s = re.sub(r'[^\d.,\-]', '', s)
    s = s.strip()
    if not s:
        return default
    # Detectar formato: si hay punto Y coma, determinar cuál es decimal
    if ',' in s and '.' in s:
        # Si la coma viene después del punto: 1.234,56 → formato europeo
        if s.rfind(',') > s.rfind('.'):
            s = s.replace('.', '').replace(',', '.')
        else:
            # 1,234.56 → formato americano
            s = s.replace(',', '')
    elif ',' in s:
        # Solo comas: podría ser decimal (123,45) o miles (1,234)
        # Si hay exactamente 2 dígitos después de la última coma, tratar como decimal
        parts = s.rsplit(',', 1)
        if len(parts[1]) <= 2:
            s = s.replace(',', '.')
        else:
            s = s.replace(',', '')
    try:
        return float(s)
    except (ValueError, TypeError):
        return default


def limpiar_nan(obj):
    """
    Recursivamente convierte valores NaN de un objeto (dict, list) a None.
    Esto es crucial para que tojson genere 'null' en lugar de 'NaN', que es inválido en JSON.
    """
    if isinstance(obj, dict):
        return {k: limpiar_nan(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [limpiar_nan(v) for v in obj]
    elif isinstance(obj, float) and np.isnan(obj):
        return None
    elif pd.isna(obj): # Cubre NaT y otros tipos de pandas null
        return None
    return obj


def normalizar_texto(texto):
    """
    Normaliza texto: quita acentos, convierte a mayúsculas y limpia espacios.
    """
    if not texto or not isinstance(texto, str):
        return ""
    # Quitar acentos
    texto = ''.join(c for c in unicodedata.normalize('NFD', texto)
                  if unicodedata.category(c) != 'Mn')
    return texto.strip().upper()


def buscar_mejor_match_ejecutivo(nombre_buscado, ejec_map):
    """
    Busca el mejor match para un nombre dado en un mapa de ejecutivos {nombre_normalizado: id}.
    Incluye fuzzy matching si el match exacto falla.
    """
    nombre_norm = normalizar_texto(nombre_buscado)
    if not nombre_norm:
        return None
    
    # 1. Match exacto sobre normalizado
    if nombre_norm in ejec_map:
        return ejec_map[nombre_norm]
    
    # 2. Fuzzy matching (difflib)
    posibilidades = list(ejec_map.keys())
    matches = difflib.get_close_matches(nombre_norm, posibilidades, n=1, cutoff=0.7) # 0.7 para ser más flexible
    if matches:
        return ejec_map[matches[0]]
        
    return None


def buscar_mejor_match_compania(nombre_buscado, cia_map):
    """
    Busca el mejor match para un nombre de compañía dado en un mapa de compañías {nombre_normalizado: id}.
    Incluye búsqueda por subcadena y fuzzy matching si el match exacto falla.
    """
    nombre_norm = normalizar_texto(nombre_buscado)
    if not nombre_norm:
        return None
    
    # 1. Match exacto sobre normalizado
    if nombre_norm in cia_map:
        return cia_map[nombre_norm]
    
    # 2. Búsqueda por subcadena (Contiene)
    for key, cia_id in cia_map.items():
        if key in nombre_norm or nombre_norm in key:
            return cia_id

    # 3. Fuzzy matching (difflib)
    posibilidades = list(cia_map.keys())
    matches = difflib.get_close_matches(nombre_norm, posibilidades, n=1, cutoff=0.6) # Un poco más flexible para compañías
    if matches:
        return cia_map[matches[0]]
        
    return None

def comparar_textos_flex(texto1, texto2):
    """
    Compara dos textos de forma flexible (mayúsculas, acentos y subcadenas).
    """
    t1 = normalizar_texto(texto1)
    t2 = normalizar_texto(texto2)
    if not t1 or not t2:
        return False
    return t1 == t2 or t1 in t2 or t2 in t1


def normalizar_ramo(ramo_texto):
    """
    Normaliza el ramo a los valores estándar de la BD: Persona, Auto, Patrimonial, etc.
    """
    if not ramo_texto:
        return "Auto"
    
    rt = str(ramo_texto).upper()
    
    if any(k in rt for k in ['VIDA', 'SALUD', 'PERSONA']):
        return "Persona"
    if any(k in rt for k in ['PATRIMONIO', 'INCENDIO', 'HOGAR', 'MULTIRIESGO']):
        return "Patrimonial"
    if 'FIANZA' in rt:
        return "Fianza"
    if 'VIAJE' in rt:
        return "Viaje"
    if any(k in rt for k in ['AUTO', 'VEHICULO', 'RCV']):
        return "Auto"
    return "Auto"


def normalize_moneda(val):
    """
    Normaliza la moneda a los valores estándar de la BD: 'Dolares' o 'Bolivares'.
    Maneja acentos y variaciones comunes (USD, BS, etc).
    """
    if not val:
        return "Dolares"
    s = str(val).upper().strip()
    # Quitar acentos para comparación robusta
    import unicodedata
    s_clean = ''.join(c for c in unicodedata.normalize('NFD', s)
                     if unicodedata.category(c) != 'Mn')
    
    if any(k in s_clean for k in ['BOLIVARES', 'BS', 'VES']):
        return "Bolivares"
    return "Dolares"


def clean_monto(val):
    """
    Limpia un valor de monto para convertirlo a float.
    Remueve símbolos de moneda y separadores de miles.
    """
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        import math
        return 0.0 if math.isnan(val) else float(val)
    if isinstance(val, str):
        # Limpiar caracteres no numéricos excepto el punto decimal
        cleaned = val.replace(',', '').replace('$', '').strip()
        try:
            return float(cleaned)
        except:
            return 0.0
    return 0.0
    return "Auto"


def procesar_form_empleado(dataForm, foto_perfil):
    # Formateando Salario
    salario_sin_puntos = re.sub('[^0-9]+', '', dataForm['salario_empleado'])
    # convertir salario a INT
    salario_entero = int(salario_sin_puntos)

    result_foto_perfil = procesar_imagen_perfil(foto_perfil)
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:

                sql = "INSERT INTO tbl_empleados (nombre_empleado, apellido_empleado, sexo_empleado, telefono_empleado, email_empleado, profesion_empleado, foto_empleado, salario_empleado) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)"

                # Creando una tupla con los valores del INSERT
                valores = (dataForm['nombre_empleado'], dataForm['apellido_empleado'], dataForm['sexo_empleado'],
                           dataForm['telefono_empleado'], dataForm['email_empleado'], dataForm['profesion_empleado'], result_foto_perfil, salario_entero)
                cursor.execute(sql, valores)

                conexion_MySQLdb.commit()
                resultado_insert = cursor.rowcount
                return resultado_insert

    except Exception as e:
        return f'Se produjo un error en procesar_form_empleado: {str(e)}'
    
def procesar_form_asegurado(dataForm):
    try:
        # Validación de campos requeridos
        required_fields = ['Cedula_asegurado', 'nombre_asegurado', 'apellido_asegurado', 'email_asegurado']
        for field in required_fields:
            if not dataForm.get(field):
                return {'success': False, 'message': f'El campo {field} es obligatorio.'}

        # convertir cedula a INT
        try:
            cedula = int(dataForm['Cedula_asegurado'])
        except (ValueError, TypeError):
            return {'success': False, 'message': 'La cédula debe ser un número válido.'}

        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                # Comprobar si ya existe
                sql = "SELECT CI FROM asegurado WHERE CI = %s"
                cursor.execute(sql, (cedula,))
                if cursor.fetchone():
                    return {'success': False, 'message': 'Esta cédula ya se encuentra registrada.'}

                # Inserción
                sql = """
                    INSERT INTO asegurado 
                    (CI, Nombre, Nombre2, Apellido, Apellido2, Tipo_CI, Correo, Fecha_nacimiento, Telefono, Ejecutivo, profesion, localidad, canal) 
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """
                valores = (
                    cedula, 
                    dataForm.get('nombre_asegurado'), 
                    dataForm.get('nombre_asegurado2'), 
                    dataForm.get('apellido_asegurado'), 
                    dataForm.get('apellido_asegurado2'),
                    dataForm.get('tipo'), 
                    dataForm.get('email_asegurado'), 
                    dataForm.get('fecha_nacimiento'), 
                    dataForm.get('telefono_asegurado'), 
                    dataForm.get('Ejecutivo'), 
                    dataForm.get('Profesion'), 
                    dataForm.get('localidad'), 
                    dataForm.get('Canal')
                )
                cursor.execute(sql, valores)
                conexion_MySQLdb.commit()
                
                return {'success': True, 'message': 'Asegurado registrado exitosamente.', 'data': cedula}

    except Exception as e:
        return {'success': False, 'message': f'Error de base de datos: {str(e)}'}
    except Exception as e:
        return {'success': False, 'message': f'Error inesperado: {str(e)}'}
    
def procesar_form_company(dataForm):
    try:
        if not dataForm.get('nombre_company'):
            return {'success': False, 'message': 'El nombre de la compañía es obligatorio.'}

        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:           
                sql = "INSERT INTO compania (rif, nombre) VALUES (%s, %s)"
                valores = (dataForm.get('Rif'), dataForm['nombre_company'])
                cursor.execute(sql, valores)
                conexion_MySQLdb.commit()
                return {'success': True, 'message': 'Compañía registrada exitosamente.'}

    except Exception as e:
        return {'success': False, 'message': f'Error inesperado en procesar_form_company: {str(e)}'}
    
def procesar_form_ejecutivo(dataForm):
    try:
        if not dataForm.get('CI') or not dataForm.get('Nombre'):
            return {'success': False, 'message': 'La CI y el nombre son obligatorios.'}

        telefono_completo = f"{dataForm.get('area_code', '')}{dataForm.get('Telefono', '')}"

        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:           
                sql = "INSERT INTO ejecutivo (CI, Nombre, Nombre2, Apellido, Apellido2, Correo, Telefono, Tipo) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)"
                valores = (
                    dataForm['CI'],
                    dataForm['Nombre'],
                    dataForm.get('Nombre2'),
                    dataForm['Apellido'],
                    dataForm.get('Apellido2'),
                    dataForm['Correo'],
                    telefono_completo,
                    dataForm['Tipo']
                )
                cursor.execute(sql, valores)
                conexion_MySQLdb.commit()
                return {'success': True, 'message': 'Ejecutivo registrado exitosamente.', 'data': cursor.lastrowid}

    except Exception as e:
        return {'success': False, 'message': f'Error inesperado en procesar_form_ejecutivo: {str(e)}'}
    


    

def procesar_form_renovacion(dataForm):
    try:
        # Validación de campos requeridos
        required_fields = ['poliza', 'Frecuencia', 'Fvencimiento']
        for field in required_fields:
            if not dataForm.get(field):
                return {'success': False, 'message': f'El campo {field} es obligatorio.'}

        # Limpiar y convertir el valor de 'prima'
        prima_str = dataForm.get('prima', '0')
        try:
            prima_cleaned = str(prima_str).replace('.', '').replace(',', '.')
            prima_float = float(prima_cleaned)
        except (ValueError, TypeError):
            return {'success': False, 'message': 'El valor de la prima no es válido.'}

        # Procesamiento de fechas
        try:
            cadena_fecha = dataForm['Fvencimiento']
            # Formato esperado: "%a, %d %b %Y %H:%M:%S %Z" (ej: "Mon, 01 Jan 2024 00:00:00 GMT")
            # Ajustar si el formato cambia
            formato_entrada = "%a, %d %b %Y %H:%M:%S %Z"
            objeto_datetime = datetime.datetime.strptime(cadena_fecha, formato_entrada)
            fecha_contrato = objeto_datetime.strftime("%Y-%m-%d")
            fecha_vencimiento = (objeto_datetime + relativedelta(years=1)).strftime("%Y-%m-%d")
        except Exception as e:
            return {'success': False, 'message': f'Error al procesar las fechas: {str(e)}'}

        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:           
                sql = "INSERT INTO renovacion (Cod_poliza, Prima, Frecuencia, Fecha_contrato, cobertura, Fecha_vencimiento) VALUES (%s, %s, %s, %s, %s, %s)"
                valores = (dataForm['poliza'], prima_float, dataForm['Frecuencia'], fecha_contrato, dataForm.get('Cobertura'), fecha_vencimiento)
                cursor.execute(sql, valores)
                conexion_MySQLdb.commit()
                
                return {'success': True, 'message': 'Renovación registrada exitosamente.', 'data': dataForm['poliza']}

    except Exception as e:
        return {'success': False, 'message': f'Error inesperado en procesar_form_renovacion: {str(e)}'}    
    
def procesar_form_CartaAval(dataForm):
    try:
        # Validación básica
        if not dataForm.get('cod_poliza'):
            return {'success': False, 'message': 'El código de póliza es obligatorio.'}

        # Pre-procesar los datos
        processed_data = {}
        for key, value in dataForm.items():
            if isinstance(value, str) and value.strip() == '':
                processed_data[key] = None
            else:
                processed_data[key] = value

        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                sql = """
                    INSERT INTO carta_aval (
                        Cod_poliza, Diagnostico, Procedimiento, Estado, Moneda, Monto_solicitado,
                        Monto_aprobado, Fecha_noti, Fecha_apro, Correo,
                        codigo_siniestro, Tipo_Atencion, Monto_aprobadoD,
                        MES, NEGOCIO, TITULAR, RECLAMANTE, REFERIDOR, ANALISTA, OBSERVACION
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """
                
                diag = processed_data.get('Diagnostico')
                proc = processed_data.get('Procedimiento') or diag
                
                valores = (
                    processed_data.get('cod_poliza'),
                    diag,
                    proc,
                    processed_data.get('Status'),
                    normalize_moneda(processed_data.get('Moneda')),
                    clean_monto(processed_data.get('Monto_solicitado')),
                    clean_monto(processed_data.get('Monto_aprobado')),
                    processed_data.get('Fecha_notificada'),
                    processed_data.get('Fecha_aprobada'),
                    processed_data.get('correo'),
                    processed_data.get('codigo_siniestro'),
                    processed_data.get('Tipo_de_Atencion'),
                    clean_monto(processed_data.get('Monto_aprobado_dolares')),
                    processed_data.get('MES'),
                    processed_data.get('NEGOCIO'),
                    processed_data.get('TITULAR'),
                    processed_data.get('RECLAMANTE'),
                    processed_data.get('REFERIDOR'),
                    processed_data.get('ANALISTA'),
                    processed_data.get('OBSERVACION')
                )

                cursor.execute(sql, valores)
                conexion_MySQLdb.commit()
                return {'success': True, 'message': 'Carta Aval registrada exitosamente.'}

    except Exception as e:
        return {'success': False, 'message': f'Error inesperado en procesar_form_CartaAval: {str(e)}'}
    
def eliminar_siniestro_db(id_siniestro, tipo_siniestro):
    try:
        table = None
        id_column = None
        
        tipo_clean = str(tipo_siniestro).lower().strip()
        
        if "reembolso" in tipo_clean or "patrimonial" in tipo_clean or "viaje" in tipo_clean:
            table = "reembolso"
            id_column = "cod_reembolso"
            note_table = "nota_Reembolso"
            note_id_column = "Cod_Reembolso"
        elif "carta aval" in tipo_clean or "carta_aval" in tipo_clean:
            table = "carta_aval"
            id_column = "Cod_CartaAval"
            note_table = "nota_cartaAval"
            note_id_column = "Cod_CartaAval"
        elif "auto" in tipo_clean:
            table = "automovilsiniestro"
            id_column = "Cod_siniestroA"
            note_table = "nota_Auto"
            note_id_column = "Cod_Auto"
            
        if not table or not id_column:
            return {'success': False, 'message': f'Tipo de siniestro "{tipo_siniestro}" no reconocido para eliminación.'}
            
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                # Ensure notes are deleted first to avoid FK constraints
                cursor.execute(f"DELETE FROM {note_table} WHERE {note_id_column} = %s", (id_siniestro,))
                
                sql = f"DELETE FROM {table} WHERE {id_column} = %s"
                cursor.execute(sql, (id_siniestro,))
                conexion_MySQLdb.commit()
                
                if cursor.rowcount > 0:
                    return {'success': True, 'message': 'Siniestro eliminado correctamente.'}
                else:
                    return {'success': False, 'message': 'No se encontró el siniestro para eliminar.'}
                    
    except Exception as e:
        return {'success': False, 'message': f'Error al eliminar siniestro: {str(e)}'}

def procesar_form_reembolso(dataForm):
    try:
        # Validación básica
        if not dataForm.get('cod_poliza'):
            return {'success': False, 'message': 'El código de póliza es obligatorio.'}

        # Pre-procesar los datos
        processed_data = {}
        for key, value in dataForm.items():
            if isinstance(value, str) and value.strip() == '':
                processed_data[key] = None
            else:
                processed_data[key] = value

        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                sql = """
                    INSERT INTO Reembolso (
                        Cod_poliza, Diagnostico, Estado, Fecha_ocurrencia, Fecha_noti,
                        Fecha_max, Moneda, Monto_solicitado, Monto_pagado, Fecha_pago,
                        Correo, codigo_siniestro, monto_dolares, Tipo_Atencion
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """
                # Auto-healing column
                try:
                    cursor.execute("ALTER TABLE reembolso ADD COLUMN Tipo_Atencion VARCHAR(100)")
                except:
                    pass
                
                valores = (
                    processed_data.get('cod_poliza'),
                    processed_data.get('Diagnostico'),
                    processed_data.get('Status'),
                    processed_data.get('Fecha_ocurrencia'),
                    processed_data.get('Fecha_notificada'),
                    processed_data.get('Fecha_maxima'),
                    normalize_moneda(processed_data.get('Moneda')),
                    clean_monto(processed_data.get('Monto_solicitado')),
                    clean_monto(processed_data.get('Monto_aprobado')),
                    processed_data.get('Fecha_pago'),
                    processed_data.get('correo'),
                    processed_data.get('Numero_de_Siniestro'),
                    clean_monto(processed_data.get('monto_dolares')),
                    processed_data.get('Tipo_Atencion')
                )

                cursor.execute(sql, valores)
                conexion_MySQLdb.commit()
                return {'success': True, 'message': 'Reembolso registrado exitosamente.'}

    except Exception as e:
        return {'success': False, 'message': f'Error inesperado en procesar_form_reembolso: {str(e)}'}
    
def procesar_form_SiniestroAuto(dataForm):
    try:
        # Validación básica
        if not dataForm.get('cod_poliza'):
            return {'success': False, 'message': 'El código de póliza es obligatorio.'}

        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:           
                sql = """
                    INSERT INTO AutomovilSiniestro 
                    (Cod_poliza, Fecha_ocurrencia, Fecha_noti, Fecha_inspec, Estado, Monto_orden, Correo, Descripcion) 
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                """
                valores = (
                    dataForm['cod_poliza'],
                    dataForm.get('Fecha_ocurrencia'),
                    dataForm.get('Fecha_notificada'),
                    dataForm.get('Fecha_inspeccion'),
                    dataForm.get('Status'),
                    dataForm.get('Monto_orden'),
                    dataForm.get('correo'),
                    dataForm.get('Descripcion')
                )
                cursor.execute(sql, valores)
                conexion_MySQLdb.commit()
                return {'success': True, 'message': 'Siniestro de auto registrado exitosamente.'}

    except Exception as e:
        return {'success': False, 'message': f'Error inesperado en procesar_form_SiniestroAuto: {str(e)}'}
    
def procesar_form_beneficiario(dataForm):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:           
                    sql = "INSERT INTO beneficiario (Cod_poliza,Nombre,Apellido,Cedula,Parentesco) VALUES (%s, %s, %s, %s, %s)"
                    #Creando una tupla con los valores del INSERT
                    valores = ( dataForm['cod'],dataForm['nombre_beneficiario'],dataForm['Apellido'],dataForm['Cedula'],dataForm['Parentesco'])
                    cursor.execute(sql,valores)
                    conexion_MySQLdb.commit()
                    resultado_insert = cursor.rowcount
                    return resultado_insert
 

    except Exception as e:
        return f'Se produjo un error en procesar_beneficiario: {str(e)}'
    

def procesar_form_pago(dataForm):
    try:
        # Validación de campos requeridos
        required_fields = ['renovacion', 'moneda', 'fecha', 'Metodo', 'Monto']
        for field in required_fields:
            if not dataForm.get(field):
                return {'success': False, 'message': f'El campo {field} es obligatorio.'}

        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                cod_renovacion = dataForm['renovacion']
                moneda = dataForm['moneda']
                fecha_pago_inicial_str = dataForm['fecha']

                # 1. Insertar el primer pago (el que se está registrando)
                sql_insert_pago = "INSERT INTO pago (Cod_renovacion, Moneda, fecha, Metodo_pago, tasa, monto, fecha_pagada, estado, recibo, nro_cuota) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
                tasa_str = dataForm.get('Tasa')
                tasa = None
                if tasa_str and moneda != "$":
                    try:
                        tasa_cleaned = str(tasa_str).replace('.', '').replace(',', '.')
                        tasa = float(tasa_cleaned)
                    except (ValueError, TypeError):
                        tasa = None
                
                nro_recibo = dataForm.get('nro_recibo')
                try:
                    nro_cuota = int(dataForm.get('nro_cuota', 1))
                except (ValueError, TypeError):
                    nro_cuota = 1
                
                valores_primer_pago = (cod_renovacion, moneda, fecha_pago_inicial_str, dataForm['Metodo'], tasa, dataForm['Monto'], fecha_pago_inicial_str, 'PAGADO', nro_recibo, nro_cuota)
                cursor.execute(sql_insert_pago, valores_primer_pago)
                last_insert_id = cursor.lastrowid

                # 2. Obtener datos de la renovación para calcular cuotas
                sql_frecuencia = "SELECT Frecuencia, Prima, p.ramo FROM renovacion JOIN poliza p ON renovacion.cod_poliza = p.cod_poliza WHERE Cod_renovacion = %s"
                cursor.execute(sql_frecuencia, (cod_renovacion,))
                renovacion_data = cursor.fetchone()

                if not renovacion_data:
                    conexion_MySQLdb.rollback()
                    return {'success': False, 'message': 'No se encontraron datos de la renovación.'}

                frecuencia = renovacion_data['Frecuencia']
                prima_total = renovacion_data['Prima']

                # 3. Determinar número de pagos e intervalo
                frecuencia_map = {
                    1: {'pagos': 1, 'intervalo': 12}, # Anual
                    2: {'pagos': 12, 'intervalo': 1}, # Mensual
                    3: {'pagos': 4, 'intervalo': 3},  # Trimestral
                    4: {'pagos': 2, 'intervalo': 6},  # Semestral
                    5: {'pagos': 3, 'intervalo': 4},  # Cuatrimestral
                    6: {'pagos': 6, 'intervalo': 2},  # Bimensual
                    10: {'pagos': 10, 'intervalo': 1} # Mensual Especial (10 pagos)
                }
                config_pago = frecuencia_map.get(frecuencia, {'pagos': 1, 'intervalo': 12})
                num_pagos_total = config_pago['pagos']
                intervalo_meses = config_pago['intervalo']

                # Eliminada lógica hardcoded de 10 pagos para Mercantil.
                # Ahora se maneja a través del valor de frecuencia (10).

                # 4. Calcular monto de cada cuota
                monto_cuota = prima_total / num_pagos_total if num_pagos_total > 0 else prima_total

                # 5. Obtener cuotas existentes para evitar duplicados (asegurando enteros)
                cursor.execute("SELECT nro_cuota FROM pago WHERE Cod_renovacion = %s", (cod_renovacion,))
                cuotas_existentes = set()
                for p in cursor.fetchall():
                    if p['nro_cuota'] is not None:
                        try:
                            cuotas_existentes.add(int(p['nro_cuota']))
                        except (ValueError, TypeError):
                            pass

                # 6. Insertar los pagos restantes como pendientes solo si no existen
                if num_pagos_total > 1:
                    try:
                        fecha_primer_pago = datetime.datetime.strptime(fecha_pago_inicial_str, '%Y-%m-%d')
                        # 'nro_cuota' es el número de la cuota que se está registrando ahora
                        nro_cuota_actual = int(nro_cuota)
                        
                        for i in range(1, num_pagos_total + 1):
                            if i not in cuotas_existentes:
                                # Calcular fecha relativa a la cuota actual
                                # Si registramos la cuota 1, la 2 es +1 mes.
                                # Si registramos la cuota 3, la 4 es +1 mes, la 2 es -1 mes.
                                offset_meses = (i - nro_cuota_actual) * intervalo_meses
                                fecha_siguiente_pago = fecha_primer_pago + relativedelta(months=offset_meses)
                                
                                sql_insert_siguiente_pago = "INSERT INTO pago (Cod_renovacion, Moneda, fecha, monto, estado, nro_cuota) VALUES (%s, %s, %s, %s, %s, %s)"
                                valores_siguiente_pago = (cod_renovacion, moneda, fecha_siguiente_pago.strftime('%Y-%m-%d'), monto_cuota, 'EN PROCESO', i)
                                cursor.execute(sql_insert_siguiente_pago, valores_siguiente_pago)
                                cuotas_existentes.add(i)
                    except Exception as e:
                        conexion_MySQLdb.rollback()
                        return {'success': False, 'message': f'Error al generar cuotas pendientes: {str(e)}'}

                conexion_MySQLdb.commit()
                return {'success': True, 'message': 'Pago y cuotas registrados exitosamente.', 'data': last_insert_id}

    except Exception as e:
        return {'success': False, 'message': f'Error inesperado en procesar_form_pago: {str(e)}'}

    

def asignar_comision(cod_renovacion):
    try:
        with connectionBD() as conexion_MySQLdb: # Asegúrate de que connectionBD() maneja la conexión y el cierre correctamente
            with conexion_MySQLdb.cursor() as cursor:
                    print(f"DEBUG: Intentando asignar comisión para Cod_renovacion: {cod_renovacion}")
                    sql = """SELECT
                                p.Ramo,
                                p.cod_poliza,
                                CASE
                                    WHEN per.Subramo IS NOT NULL THEN per.Subramo
                                    WHEN a.Subramo IS NOT NULL THEN a.Subramo
                                    WHEN pa.Subramo IS NOT NULL THEN pa.Subramo
                                    WHEN f.Subramo IS NOT NULL THEN f.Subramo
                                    WHEN v.Subramo IS NOT NULL THEN v.Subramo
                                    ELSE NULL
                                END AS Subramo,
                                CASE
                                    WHEN per.Producto IS NOT NULL THEN per.Producto
                                    WHEN a.Producto IS NOT NULL THEN a.Producto
                                    WHEN pa.Producto IS NOT NULL THEN pa.Producto
                                    WHEN v.Producto IS NOT NULL THEN v.Producto
                                    WHEN f.Producto IS NOT NULL THEN f.Producto
                                    ELSE NULL
                                END AS Producto,
                                e.Tipo,
                                e.Cod_ejecutivo,
                                com.Nombre as compania  
                            FROM
                                poliza p
                            JOIN
                                renovacion cr ON p.cod_poliza = cr.Cod_poliza
                            JOIN
                                compania com ON p.Cod_compania = com.Cod_compania
                            JOIN
                                asegurado ase ON p.CI_asegurado = ase.CI
                            JOIN
                                ejecutivo e ON ase.Ejecutivo = e.cod_ejecutivo
                            LEFT JOIN
                                Persona per ON p.cod_poliza = per.Cod_poliza
                            LEFT JOIN
                                Auto a ON p.cod_poliza = a.Cod_poliza
                            LEFT JOIN
                                Patrimonio pa ON p.cod_poliza = pa.Cod_poliza
                            LEFT JOIN
                                Viaje v ON p.cod_poliza = v.Cod_poliza
                            LEFT JOIN 
                                Fianza f ON p.cod_poliza = f.Cod_poliza
                            WHERE
                                cr.Cod_renovacion = %s"""
                    #Creando una tupla con los valores del INSERT
                    cursor.execute(sql, (cod_renovacion,))
                    dato = cursor.fetchone()

                    if not dato:
                        print(f"ERROR: No se encontraron datos de póliza/renovación para Cod_renovacion: {cod_renovacion}")
                        return None # No se puede asignar comisión si no hay datos

                    # Extraer campos necesarios, manejando posibles valores None del resultado de la consulta
                    compania = dato.get('compania')
                    subramo = dato.get('Subramo')
                    ramo = dato.get('Ramo')
                    producto = dato.get('Producto')
                    ejecutivo_tipo = dato.get('Tipo')
                    cod_poliza = dato.get('cod_poliza')

                    if not all([compania, ramo, producto]): # Subramo puede ser None para algunos productos
                        print(f"ERROR: Datos incompletos para formar la clave de comisión para Cod_renovacion {cod_renovacion}. Faltan: Compania={compania}, Ramo={ramo}, Producto={producto}")
                        return None
                    clave = (dato['compania'],dato['Subramo'], dato['Ramo'], dato['Producto'])

                    # Obtener todas las configuraciones de comisión de la base de datos
                    comisiones_config_db = sql_get_all_comisiones_config()
                    
                    comision = None # Inicializar comision a None
                    valores_comision = None

                    # Buscar la regla de comisión que coincida
                    for rule in comisiones_config_db:
                        rule_compania = rule.get('compania')
                        rule_ramo = rule.get('ramo')
                        rule_subramo = rule.get('subramo')
                        rule_producto = rule.get('producto')
                        rule_tipo_ejecutivo = rule.get('tipo_ejecutivo')
                        rule_porcentajes_str = rule.get('porcentajes')

                        # Convertir porcentajes de string a lista de floats
                        try:
                            rule_porcentajes = [float(p.strip()) for p in rule_porcentajes_str.split(',')]
                        except (AttributeError, ValueError):
                            rule_porcentajes = [] # En caso de error, lista vacía

                        # Normalizar el tipo de ejecutivo para la comparación
                        exec_type_key = ejecutivo_tipo.upper() if ejecutivo_tipo else None
                        if exec_type_key == 'PERSONALIZADO':
                            exec_type_key = 'AGENTE'

                        # Comparar los campos de la regla con los datos de la póliza
                        # Se asume que si rule_subramo es None, coincide con cualquier subramo o None en dato
                        subramo_match = (rule_subramo is None and subramo is None) or \
                                        (rule_subramo is not None and subramo is not None and rule_subramo.upper() == subramo.upper()) or \
                                        (rule_subramo is None and subramo is not None) # Rule is general, dato has specific subramo
                        
                        # --- NUEVA LÓGICA DE PRIORIDAD POR EJECUTIVO ---
                        # Verificar si la regla tiene un ejecutivo específico
                        rule_cod_ejecutivo = rule.get('cod_ejecutivo')
                        
                        # Obtener el código del ejecutivo de la póliza (dato['Cod_ejecutivo'])
                        poliza_cod_ejecutivo = dato.get('Cod_ejecutivo')
                        
                        ejecutivo_match = False
                        if rule_cod_ejecutivo is not None and rule_cod_ejecutivo != 0:
                             # Si la regla ES específica, debe coincidir el código
                             if str(rule_cod_ejecutivo) == str(poliza_cod_ejecutivo):
                                 ejecutivo_match = True
                             else:
                                 # Regla específica de OTRO ejecutivo -> No nos sirve
                                 ejecutivo_match = False
                        else:
                             # Regla general (no tiene cod_ejecutivo)
                             # Coincide si el tipo de ejecutivo match
                             if rule_tipo_ejecutivo.upper() == exec_type_key:
                                 ejecutivo_match = True
                                 
                             # PERO: Si existe una regla específica para este ejecutivo, deberíamos haberla encontrado primero 
                             # o deberíamos priorizarla. 
                             # Para simplificar: Asumimos que el usuario no crea conflictos ambiguos, pero idealmente
                             # deberíamos buscar la regla MÁS específica.
                             # Esta implementación simple itera ordenadamente.
                        
                        # Nota: Esto asume que iteramos y tomamos el primero que coincida.
                        # Para garantizar prioridad, deberíamos ordenar "comisiones_config_db" primero por especificidad (cod_ejecutivo DESC) o filtrar mejor.
                        # Mejor enfoque: Buscar match exacto de ejecutivo PRIMERO.
                        
                        pass # Placeholder para la lógica reescrita abajo

                    # Re-implementación del bucle de búsqueda para priorizar ejecutivo específico
                    regla_encontrada = None
                    
                    # 1. Buscar regla ESPECÍFICA para el ejecutivo
                    for rule in comisiones_config_db:
                         rule_cod = rule.get('cod_ejecutivo')
                         if rule_cod is not None and rule_cod != 0 and str(rule_cod) == str(dato.get('Cod_ejecutivo')):
                            # Verificar otros campos
                            r_comp = rule.get('compania')
                            r_ramo = rule.get('ramo')
                            r_sub = rule.get('subramo')
                            r_prod = rule.get('producto')
                            
                            sub_match = (r_sub is None and subramo is None) or \
                                        (r_sub is not None and subramo is not None and r_sub.upper() == subramo.upper()) or \
                                        (r_sub is None and subramo is not None)
                                        
                            if (comparar_textos_flex(r_comp, compania) and
                                r_ramo.upper() == ramo.upper() and
                                sub_match and
                                r_prod.upper() == producto.upper()):
                                
                                regla_encontrada = rule
                                break # Prioridad máxima encontrada
                    
                    # 2. Si no se encontró específica, buscar regla GENERAL (por Tipo)
                    if not regla_encontrada:
                        for rule in comisiones_config_db:
                            rule_cod = rule.get('cod_ejecutivo')
                            if rule_cod is None or rule_cod == 0: # Regla general
                                r_comp = rule.get('compania')
                                r_ramo = rule.get('ramo')
                                r_sub = rule.get('subramo')
                                r_prod = rule.get('producto')
                                r_tipo = rule.get('tipo_ejecutivo')
                                
                                sub_match = (r_sub is None and subramo is None) or \
                                            (r_sub is not None and subramo is not None and r_sub.upper() == subramo.upper()) or \
                                            (r_sub is None and subramo is not None)

                                if (comparar_textos_flex(r_comp, compania) and
                                    r_ramo.upper() == ramo.upper() and
                                    sub_match and
                                    r_prod.upper() == producto.upper() and
                                    r_tipo.upper() == exec_type_key):
                                    
                                    regla_encontrada = rule
                                    break

                    if regla_encontrada:
                        rule_porcentajes_str = regla_encontrada.get('porcentajes')
                        try:
                            valores_comision = [float(p.strip()) for p in rule_porcentajes_str.split(',')]
                        except (AttributeError, ValueError):
                            valores_comision = []
                            
                        # Continúa con la lógica de asignación...
                        pass 
                    else:
                        valores_comision = None

                    if valores_comision:
                        # Obtener el número de renovaciones para esta póliza para determinar el índice de comisión
                        sql_count_renewals = "SELECT Cod_renovacion FROM renovacion WHERE cod_poliza=%s ORDER BY Cod_renovacion ASC"
                        cursor.execute(sql_count_renewals, (cod_poliza,))
                        cantidad_renovaciones = cursor.fetchall()
                        
                        # El índice es 0 para la primera renovación, 1 para la segunda, etc.
                        # Si hay más renovaciones que tasas de comisión, se usa la última tasa.
                        indice_comision = len(cantidad_renovaciones) - 1
                        
                        if indice_comision < len(valores_comision):
                            comision = valores_comision[indice_comision]
                            print(f"DEBUG: Comisión seleccionada para {cod_renovacion}: {comision} (Índice: {indice_comision})")
                            
                            sql_update_renovacion = "UPDATE renovacion SET comision = %s WHERE Cod_renovacion = %s"
                            cursor.execute(sql_update_renovacion, (comision, cod_renovacion))
                            conexion_MySQLdb.commit()
                            return comision
                        else:
                            print(f"WARNING: Índice de comisión {indice_comision} fuera de rango para {clave} y tipo {exec_type_key}. Valores disponibles: {valores_comision}. No se asignó comisión.")
                            return None
                    else:
                        print(f"WARNING: No se encontraron valores de comisión para {clave} y tipo {exec_type_key}.")
                        return None

    except Exception as e:
        print(f"ERROR: Se produjo un error inesperado en asignar_comision para Cod_renovacion {cod_renovacion}: {str(e)}")
        return None




def revertir_pago_individual(pago_id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                # Verificar el estado actual del pago
                cursor.execute("SELECT estado, Cod_renovacion, nro_cuota FROM pago WHERE Cod_pago = %s", (pago_id,))
                pago = cursor.fetchone()

                if not pago:
                    return {'success': False, 'message': 'Pago no encontrado.'}

                if pago['estado'] != 'PAGADO':
                    return {'success': False, 'message': 'El pago no está en estado PAGADO, no se puede revertir.'}

                # Revertir el pago a 'EN PROCESO' y limpiar campos de pago
                # Mantenemos el monto y la fecha de cuota originales (o los dejamos como están)
                # Limpiamos fecha_pagada, metodo_pago, recibo, tasa (si aplica, aunque tasa está en la tabla pago)
                
                sql_revert = """
                    UPDATE pago
                    SET estado = 'EN PROCESO',
                        fecha_pagada = NULL,
                        Metodo_pago = NULL,
                        recibo = NULL,
                        tasa = NULL,
                        moneda = NULL -- Opcional, si queremos que reseleccionen moneda
                    WHERE Cod_pago = %s
                """
                cursor.execute(sql_revert, (pago_id,))
                
                # Opcional: Si este pago generó comisiones, ¿deberíamos borrarlas?
                # Por ahora solo revertimos el estado del pago.
                
                conexion_MySQLdb.commit()
                return {'success': True, 'message': 'Pago revertido a EN PROCESO correctamente.'}

    except Exception as e:
        print(f"Error en revertir_pago_individual: {e}")
        return {'success': False, 'message': f'Error interno: {str(e)}'}

def procesar_comision(dataForm):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                               
                    sql= "UPDATE renovacion SET comision = %s WHERE Cod_renovacion =%s"
                    
                    valores = ( dataForm.form.get('comision'),dataForm.form.get('cod_renovacion'))
                    cursor.execute(sql,(valores))
                    sql="select r.Cod_poliza from renovacion r where r.Cod_renovacion=%s limit 1"
                    valores= (dataForm.form.get('cod_renovacion'),)
                    cursor.execute(sql,(valores))
                    print('hey hey3')
                    resultado=cursor.fetchone()
                    conexion_MySQLdb.commit()
                    
                    return resultado
 

    except Exception as e:
        return f'Se produjo un error en procesar_comision2: {str(e)}'
    

def procesar_comision2(cod_pago, referidor, cod_poliza=None):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                # Asegurar columna cod_poliza
                try:
                    cursor.execute("ALTER TABLE comision ADD COLUMN IF NOT EXISTS cod_poliza VARCHAR(100) DEFAULT NULL")
                except:
                    pass

                if cod_poliza:
                    sql = "INSERT INTO comision (Cod_pago, cod_ejecutivo, cod_poliza) VALUES (%s, %s, %s)"
                    valores = (cod_pago, referidor, cod_poliza)
                else:
                    sql = "INSERT INTO comision (Cod_pago, cod_ejecutivo) VALUES (%s, %s)"
                    valores = (cod_pago, referidor)
                
                cursor.execute(sql, valores) 
                conexion_MySQLdb.commit()
                return cursor.rowcount
 

    except Exception as e:
        return f'Se produjo un error en procesar_comision3: {str(e)}'
    






def procesar_form_poliza(dataForm):
    try:
        # 1. Validación de campos requeridos
        required_fields = ['Ramo', 'numero_poliza', 'asegurado', 'fecha_emision', 'company', 'Frecuencia']
        for field in required_fields:
            if not dataForm.get(field):
                return {'success': False, 'message': f'El campo {field} es obligatorio.'}

        Ramo = dataForm['Ramo']
        try:
            fecha_emision = datetime.datetime.strptime(dataForm['fecha_emision'], '%Y-%m-%d')
            fecha_vencimiento = fecha_emision + timedelta(days=365)
        except (ValueError, TypeError):
            return {'success': False, 'message': 'Formato de fecha de emisión inválido.'}
        
        # Limpiar y convertir el valor de 'prima'
        prima_str = dataForm.get('prima', '0')
        try:
            prima_cleaned = str(prima_str).replace('.', '').replace(',', '.')
            prima_float = float(prima_cleaned)
        except (ValueError, TypeError):
            return {'success': False, 'message': 'El valor de la prima no es válido.'}

        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                # Comprobar si ya existe la póliza
                sql = "SELECT cod_poliza FROM poliza WHERE cod_poliza = %s"
                cursor.execute(sql, (dataForm['numero_poliza'],))
                if cursor.fetchone():
                    return {'success': False, 'message': f'La póliza {dataForm["numero_poliza"]} ya se encuentra registrada.'}

                # Obtener Cod_compania
                sql = "SELECT Cod_compania FROM compania WHERE Nombre = %s LIMIT 1"
                cursor.execute(sql, (dataForm['company'],))
                res_company = cursor.fetchone()
                if not res_company:
                    return {'success': False, 'message': f'La compañía "{dataForm["company"]}" no existe.'}
                cod_compania = res_company['Cod_compania']

                # Inserción en tabla poliza
                sql_poliza = """
                    INSERT INTO poliza (cod_poliza, CI_asegurado, Fecha_emision, Cod_compania, Tomador, Ramo, Tipo_venta) 
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                """
                valores_poliza = (
                    dataForm['numero_poliza'], 
                    dataForm['asegurado'], 
                    dataForm['fecha_emision'],
                    cod_compania, 
                    dataForm.get('Tomador'), 
                    Ramo, 
                    dataForm.get('Tipo_venta')
                )
                cursor.execute(sql_poliza, valores_poliza)

                # Inserción en tabla renovacion
                sql_renov = """
                    INSERT INTO renovacion (Cod_poliza, Prima, Frecuencia, Fecha_contrato, cobertura, Fecha_vencimiento, comision, riesgo) 
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                """
                valores_renov = (
                    dataForm['numero_poliza'], 
                    prima_float, 
                    dataForm['Frecuencia'], 
                    dataForm['fecha_emision'], 
                    dataForm.get('Cobertura'), 
                    fecha_vencimiento, 
                    None, 
                    None
                )
                cursor.execute(sql_renov, valores_renov)
                # cod_renovacion = cursor.lastrowid # No se usa inmediatamente pero se obtenía antes

                # Inserciones específicas por Ramo
                if Ramo == 'Persona':
                    sql = 'INSERT INTO Persona (Cod_poliza, Producto, Subramo) VALUES (%s, %s, %s)'
                    valores = (dataForm['numero_poliza'], dataForm.get('producto'), dataForm.get('SubRamo'))
                    cursor.execute(sql, valores)
                elif Ramo == 'Auto':
                    sql = 'INSERT INTO Auto (Cod_poliza, modelo, Producto, placa, año, marca, Subramo) VALUES (%s, %s, %s, %s, %s, %s, %s)'
                    valores = (dataForm['numero_poliza'], dataForm.get('Modelo'), dataForm.get('producto'), dataForm.get('placa'), dataForm.get('Ano'), dataForm.get('Marca'), dataForm.get('SubRamo'))
                    cursor.execute(sql, valores)
                elif Ramo == 'Patrimonial':
                    sql = 'INSERT INTO Patrimonio (Cod_poliza, direccion, Producto, Subramo) VALUES (%s, %s, %s, %s)'
                    valores = (dataForm['numero_poliza'], dataForm.get('direccion'), dataForm.get('producto'), dataForm.get('SubRamo'))
                    cursor.execute(sql, valores)
                elif Ramo == 'Fianza':
                    sql = 'INSERT INTO Fianza (Cod_poliza, Producto, Subramo) VALUES (%s, %s, %s)'
                    valores = (dataForm['numero_poliza'], dataForm.get('producto'), dataForm.get('SubRamo'))
                    cursor.execute(sql, valores)
                elif Ramo == 'Viaje':
                    sql = 'INSERT INTO Viaje (Cod_poliza, cod_pasaporte, Producto, Subramo) VALUES (%s, %s, %s, %s)'
                    valores = (dataForm['numero_poliza'], dataForm.get('cod_pasaporte'), dataForm.get('producto'), dataForm.get('SubRamo'))
                    cursor.execute(sql, valores)
                
                conexion_MySQLdb.commit()
                return {'success': True, 'message': 'Póliza registrada exitosamente.', 'data': dataForm['numero_poliza']}

    except Exception as e:
        return {'success': False, 'message': f'Error de base de datos: {str(e)}'}
    except Exception as e:
        return {'success': False, 'message': f'Error inesperado: {str(e)}'}

def procesar_excel_preview(file):
    import datetime
    from dateutil.relativedelta import relativedelta
    import openpyxl
    
    try:
        workbook = openpyxl.load_workbook(file, data_only=True, read_only=True)
        
        # Encontrar la hoja que contenga la columna de cédula (Búsqueda en todas las páginas)
        sheet = None
        header = []
        
        # Lista de nombres posibles para la columna de cédula
        posibles_cedulas = ['CEDULA DEL ASEGURADO', 'CEDULA', 'CI', 'DOCUMENTO', 'ID', 'CÉDULA']
        
        for s_name in workbook.sheetnames:
            s = workbook[s_name]
            if s.max_row < 1: continue
            
            # Cabeceras de la primera fila
            row_header = [str(cell.value).strip().upper() if cell.value else "" for cell in s[1]]
            
            # Verificamos si esta hoja parece tener los datos correctos
            if any(h in row_header for h in posibles_cedulas):
                sheet = s
                header = row_header
                break
        
        if not sheet:
            return {'error': 'No se encontró una página con el formato adecuado. Asegúrese de que el archivo tenga una columna llamada "CEDULA" o "CEDULA DEL ASEGURADO".'}

        # Encontrar índices de columnas críticas (búsqueda flexible)
        def find_col(possible_names):
            for name in possible_names:
                if name.upper() in header:
                    return header.index(name.upper())
            return None

        idx_cedula = find_col(posibles_cedulas)
        idx_asegurado = find_col(['ASEGURADO', 'NOMBRE', 'NOMBRE COMPLETO', 'CLIENTE', 'PERSONA'])
        idx_poliza = find_col(['POLIZA', 'PÓLIZA', 'NRO POLIZA', 'NUMERO POLIZA', 'CÓDIGO PÓLIZA'])
        idx_ramo = find_col(['RAMO', 'PRODUCTO', 'TIPO SEG'])
        idx_prima = find_col(['PRIMA TOTAL (SIN IGTF)', 'PRIMA', 'PRIMA TOTAL', 'MONTO PRIMA'])
        idx_fecha = find_col(['FECHA EMISION', 'FECHA', 'FECHA VIGENCIA', 'DESDE'])
        idx_referidor = find_col(['REFERIDOR', 'CORREDOR', 'EJECUTIVO', 'COBRADOR', 'PRODUCCION', 'ASESOR', 'AGENTE', 'VENDEDOR'])
        idx_cuota = find_col(['NRO CUOTA', 'CUOTA', 'CUOTAS', 'NRO. CUOTA'])
        idx_recibo = find_col(['NRO. RECIBO/ FACTURA', 'RECIBO', 'NRO RECIBO', 'REFERENCIA', 'FACTURA', 'NRO. RECIBO / FACTURA'])
        idx_tipo_venta = find_col(['TIPO VENTA', 'TIPO_VENTA', 'N TIPO DE PÓLIZA'])
        idx_correo = find_col(['CORREO', 'EMAIL', 'CORREO ELECTRÓNICO', 'CORREO ELECTRONICO'])
        idx_telefono = find_col(['TELEFONO', 'TELÉFONO', 'CELULAR', 'PHONE'])
        idx_moneda = find_col(['$/BS', 'MONEDA', 'DIVISA'])

        preview_data = []
        
        with connectionBD() as connection:
            with connection.cursor() as cursor:
                # OPTIMIZATION: Pre-fetch DB keys for O(1) rapid lookup to eliminate N-queries issue
                cursor.execute("SELECT CI FROM asegurado")
                asegurados_db = {str(r['CI']) for r in cursor.fetchall()}

                cursor.execute("SELECT cod_poliza FROM poliza")
                polizas_db = {}
                for r in cursor.fetchall():
                    c_pol = str(r['cod_poliza'])
                    c_clean = c_pol.replace("-", "").replace(" ", "")
                    polizas_db[c_pol] = c_pol
                    polizas_db[c_clean] = c_pol
                    if c_clean.startswith("0"):
                        polizas_db[c_clean[1:]] = c_pol

                cursor.execute("SELECT Cod_poliza, Fecha_contrato FROM renovacion")
                renovaciones_db = set()
                for r in cursor.fetchall():
                    dt = r['Fecha_contrato']
                    if hasattr(dt, 'date'): dt = dt.date()
                    elif isinstance(dt, str):
                        try: dt = datetime.datetime.strptime(dt[:10], '%Y-%m-%d').date()
                        except: pass
                    renovaciones_db.add((str(r['Cod_poliza']), dt))

                for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    # Ignorar filas completamente vacías
                    if all(cell is None for cell in row):
                        continue
                        
                    row_dict = dict(zip(header, row))
                    
                    # Validación inicial de campos críticos
                    cedula_val = row[idx_cedula] if idx_cedula is not None else None
                    nombre_val = row[idx_asegurado] if idx_asegurado is not None else None
                    poliza_val = row[idx_poliza] if idx_poliza is not None else None
                    
                    poliza_raw = str(row[idx_poliza]).strip() if idx_poliza is not None and row[idx_poliza] is not None else ""
                    if poliza_raw:
                        poliza_clean = poliza_raw.replace("-", "").replace(" ", "")
                        poliza_sin_cero = poliza_clean[1:] if poliza_clean.startswith("0") else poliza_clean
                        poliza_match = polizas_db.get(poliza_raw) or polizas_db.get(poliza_clean) or polizas_db.get(poliza_sin_cero)
                        if poliza_match:
                            poliza_raw = poliza_match

                    cedula_raw = str(cedula_val).strip()
                    
                    # Parsing Cedula
                    tipo_ci = 'V'
                    cedula = cedula_raw
                    if ' ' in cedula_raw:
                        parts = cedula_raw.split(' ')
                        tipo_ci = parts[0].upper()
                        cedula = parts[1]
                    elif '-' in cedula_raw:
                        parts = cedula_raw.split('-')
                        tipo_ci = parts[0].upper()
                        cedula = parts[1]
                    elif len(cedula_raw) > 2 and any(c.isalpha() for c in cedula_raw[:2]):
                         if cedula_raw[0].isalpha():
                             tipo_ci = cedula_raw[0].upper()
                             cedula = "".join(filter(str.isdigit, cedula_raw))

                    cedula = "".join(filter(str.isdigit, str(cedula)))

                    # Exec / Referidor Detection
                    ejecutivos = lista_ejecutivosBD() or []
                    ejec_map = {normalizar_texto(f"{ej['Nombre']} {ej['Apellido']}"): ej['cod_ejecutivo'] for ej in ejecutivos}
                    ejec_id_to_name = {ej['cod_ejecutivo']: f"{ej['Nombre']} {ej['Apellido']}".strip() for ej in ejecutivos}
                    
                    referidor_excel = str(row[idx_referidor]).strip() if idx_referidor is not None and row[idx_referidor] is not None else ""
                    cod_ejecutivo = buscar_mejor_match_ejecutivo(referidor_excel, ejec_map)
                    
                    missing_executive = False
                    if not cod_ejecutivo:
                        missing_executive = True
                        cod_ejecutivo = None
                        ejecutivo_nombre = "Seleccionar ejecutivo"
                    else:
                        ejecutivo_nombre = ejec_id_to_name.get(cod_ejecutivo, "Oficina/Otros")

                    # Date handling
                    fecha_emision_raw = row[idx_fecha] if idx_fecha is not None else None
                    objeto_datetime_emision = None
                    try:
                        if isinstance(fecha_emision_raw, datetime.datetime):
                            objeto_datetime_emision = fecha_emision_raw.date()
                        elif isinstance(fecha_emision_raw, (int, float)):
                            objeto_datetime_emision = datetime.datetime.fromtimestamp((fecha_emision_raw - 25569) * 86400).date()
                        elif isinstance(fecha_emision_raw, str) and fecha_emision_raw.strip():
                            for fmt in ["%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y", "%d-%m-%Y"]:
                                try:
                                    objeto_datetime_emision = datetime.datetime.strptime(fecha_emision_raw.strip(), fmt).date()
                                    break
                                except: continue
                    except:
                        pass
                    
                    # Prima handling
                    prima_raw = row[idx_prima] if idx_prima is not None else 0
                    prima_val = 0.0
                    try:
                        if prima_raw is not None:
                            cleaned_prima = str(prima_raw).replace('$', '').replace(' ', '').replace(',', '.')
                            prima_val = float(cleaned_prima)
                    except:
                        prima_val = 0.0

                    # Status Detection Logic
                    status = 'ready'
                    reason = 'Nuevo registro'
                    
                    if not cedula or len(cedula) < 5:
                        status = 'invalid'
                        reason = 'Cédula no válida, faltante o muy corta'
                        
                    elif not poliza_raw or poliza_raw.lower() == 'none':
                        status = 'invalid'
                        reason = 'Número de póliza faltante'
                        
                    elif not objeto_datetime_emision:
                        status = 'invalid'
                        reason = 'Fecha de emisión ausente o formato inválido'
                        
                    elif prima_val <= 0:
                        status = 'invalid'
                        reason = 'Monto de prima faltante o inválido'

                    else:
                        # 1. Check Insured
                        if cedula in asegurados_db:
                            status = 'insured_exists'
                            reason = 'Asegurado ya registrado'

                        # 2. Check Policy
                        if poliza_raw in polizas_db:
                            # 3. Check Renewal
                            if (poliza_raw, objeto_datetime_emision) in renovaciones_db:
                                status = 'duplicate'
                                reason = 'Póliza y renovación ya existen'
                            else:
                                if status != 'duplicate':
                                    status = 'new_renewal'
                                    reason = 'Nueva renovación para póliza existente'
                    
                    # Extracción segura de datos para evitar KeyError o Nullos
                    nro_cuota_val = row[idx_cuota] if idx_cuota is not None and row[idx_cuota] is not None else 1
                    tipo_venta_val = row[idx_tipo_venta] if idx_tipo_venta is not None and row[idx_tipo_venta] is not None else 3
                    
                    row_info = {
                        'row': row_idx,
                        'status': status,
                        'reason': reason,
                        'cedula': cedula,
                        'tipo_ci': tipo_ci,
                        'nombre_completo': str(row[idx_asegurado]).strip() if idx_asegurado is not None and row[idx_asegurado] is not None else "Desconocido",
                        'poliza': poliza_raw,
                        'ramo': normalizar_ramo(str(row[idx_ramo])) if idx_ramo is not None and row[idx_ramo] is not None else "Auto",
                        'prima': prima_val,
                        'fecha_emision': objeto_datetime_emision.isoformat() if objeto_datetime_emision else None,
                        'ejecutivo': cod_ejecutivo,
                        'ejecutivo_nombre': ejecutivo_nombre,
                        'referidor_original': referidor_excel,
                        'missing_executive': missing_executive,
                        'raw_data': row_dict,
                        'nro_cuota': nro_cuota_val,
                        'recibo': str(row[idx_recibo]).strip() if idx_recibo is not None and row[idx_recibo] is not None else None,
                        'tipo_venta': tipo_venta_val,
                        'correo': str(row[idx_correo]).strip() if idx_correo is not None and row[idx_correo] is not None else None,
                        'telefono': str(row[idx_telefono]).strip() if idx_telefono is not None and row[idx_telefono] is not None else None,
                        'moneda': '$'
                    }
                    preview_data.append(row_info)
        
        return {'success': True, 'data': limpiar_nan(preview_data), 'ejecutivos': ejecutivos}

    except Exception as e:
        return {'error': str(e)}


def procesar_excel_stream(file):
    """
    Generador que procesa un archivo Excel fila por fila y hace yield de cada registro
    como un dict serializable. Permite al frontend recibir los datos en streaming (NDJSON).
    Primero hace yield de un mensaje 'init' con los ejecutivos y el total de filas,
    luego de cada fila procesada.
    """
    import datetime
    import openpyxl

    try:
        workbook = openpyxl.load_workbook(file, data_only=True, read_only=True)

        posibles_cedulas = ['CEDULA DEL ASEGURADO', 'CEDULA', 'CI', 'DOCUMENTO', 'ID', 'CÉDULA']

        sheet = None
        header = []
        for s_name in workbook.sheetnames:
            s = workbook[s_name]
            if s.max_row < 1:
                continue
            row_header = [str(cell.value).strip().upper() if cell.value else "" for cell in s[1]]
            if any(h in row_header for h in posibles_cedulas):
                sheet = s
                header = row_header
                break

        if not sheet:
            yield {'type': 'error', 'message': 'No se encontró una página con el formato adecuado.'}
            return

        def find_col(possible_names):
            for name in possible_names:
                if name.upper() in header:
                    return header.index(name.upper())
            return None

        idx_cedula      = find_col(posibles_cedulas)
        idx_asegurado   = find_col(['ASEGURADO', 'NOMBRE', 'NOMBRE COMPLETO', 'CLIENTE', 'PERSONA'])
        idx_poliza      = find_col(['POLIZA', 'PÓLIZA', 'NRO POLIZA', 'NUMERO POLIZA', 'CÓDIGO PÓLIZA'])
        idx_ramo        = find_col(['RAMO', 'PRODUCTO', 'TIPO SEG'])
        idx_prima_usd   = find_col(['PRIMA TOTAL (SIN IGTF)', 'PRIMA TOTAL', 'PRIMA NETO'])
        idx_prima       = find_col(['MONTO PRIMA']) # Evitar 'PRIMA' a secas porque suele ser Bs.
        idx_fecha       = find_col(['FECHA EMISION', 'FECHA', 'FECHA VIGENCIA', 'DESDE'])
        idx_referidor   = find_col(['REFERIDOR', 'CORREDOR', 'EJECUTIVO', 'COBRADOR', 'PRODUCCION', 'ASESOR', 'AGENTE', 'VENDEDOR'])
        idx_cuota       = find_col(['NRO CUOTA', 'CUOTA', 'CUOTAS', 'NRO. CUOTA'])
        idx_recibo      = find_col(['NRO. RECIBO/ FACTURA', 'RECIBO', 'NRO RECIBO', 'REFERENCIA', 'FACTURA', 'NRO. RECIBO / FACTURA'])
        idx_tipo_venta  = find_col(['TIPO VENTA', 'TIPO_VENTA', 'N TIPO DE PÓLIZA'])
        idx_correo      = find_col(['CORREO', 'EMAIL', 'CORREO ELECTRÓNICO', 'CORREO ELECTRONICO'])
        idx_telefono    = find_col(['TELEFONO', 'TELÉFONO', 'CELULAR', 'PHONE'])
        idx_pago        = find_col(['INICIAL / FRACCION (SIN IGTF)', 'INICIAL USD', 'PAGO USD'])
        idx_tasa        = find_col(['TASA DE CAMBIO', 'TASA', 'CAMBIO'])

        # Pre-fetch ejecutivos y datos de BD UNA sola vez
        ejecutivos = lista_ejecutivosBD() or []
        ejec_map = {normalizar_texto(f"{ej['Nombre']} {ej['Apellido']}"): ej['cod_ejecutivo'] for ej in ejecutivos}
        ejec_id_to_name = {ej['cod_ejecutivo']: f"{ej['Nombre']} {ej['Apellido']}".strip() for ej in ejecutivos}

        with connectionBD() as connection:
            with connection.cursor() as cursor:
                cursor.execute("SELECT CI FROM asegurado")
                asegurados_db = {str(r['CI']) for r in cursor.fetchall()}
                cursor.execute("SELECT cod_poliza FROM poliza")
                polizas_db = {str(r['cod_poliza']) for r in cursor.fetchall()}
                cursor.execute("SELECT Cod_poliza, Fecha_contrato FROM renovacion")
                renovaciones_db = set()
                for r in cursor.fetchall():
                    dt = r['Fecha_contrato']
                    if hasattr(dt, 'date'): dt = dt.date()
                    elif isinstance(dt, str):
                        try: dt = datetime.datetime.strptime(dt[:10], '%Y-%m-%d').date()
                        except: pass
                    renovaciones_db.add((str(r['Cod_poliza']), dt))

        # Contar filas no vacías para el progreso
        total_rows = sum(1 for row in sheet.iter_rows(min_row=2, values_only=True)
                         if not all(cell is None for cell in row))

        # Primer mensaje: metadatos (ejecutivos + total)
        ejecutivos_clean = [{'cod_ejecutivo': ej['cod_ejecutivo'], 'Nombre': ej.get('Nombre',''), 'Apellido': ej.get('Apellido','')} for ej in ejecutivos]
        yield {'type': 'init', 'ejecutivos': ejecutivos_clean, 'total': total_rows}

        global_idx = 0
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if all(cell is None for cell in row):
                continue

            row_dict = dict(zip(header, row))

            cedula_val = row[idx_cedula] if idx_cedula is not None else None
            poliza_raw = str(row[idx_poliza]).strip() if idx_poliza is not None and row[idx_poliza] is not None else ""
            cedula_raw = str(cedula_val).strip() if cedula_val is not None else ""

            tipo_ci = 'V'
            cedula = cedula_raw
            if ' ' in cedula_raw:
                parts = cedula_raw.split(' ')
                tipo_ci = parts[0].upper()
                cedula = parts[1]
            elif '-' in cedula_raw:
                parts = cedula_raw.split('-')
                tipo_ci = parts[0].upper()
                cedula = parts[1]
            elif len(cedula_raw) > 2 and any(c.isalpha() for c in cedula_raw[:2]):
                if cedula_raw[0].isalpha():
                    tipo_ci = cedula_raw[0].upper()
                    cedula = "".join(filter(str.isdigit, cedula_raw))
            cedula = "".join(filter(str.isdigit, str(cedula)))

            referidor_excel = str(row[idx_referidor]).strip() if idx_referidor is not None and row[idx_referidor] is not None else ""
            cod_ejecutivo = buscar_mejor_match_ejecutivo(referidor_excel, ejec_map)
            missing_executive = not cod_ejecutivo
            ejecutivo_nombre = ejec_id_to_name.get(cod_ejecutivo, "Seleccionar ejecutivo") if cod_ejecutivo else "Seleccionar ejecutivo"

            fecha_emision_raw = row[idx_fecha] if idx_fecha is not None else None
            objeto_datetime_emision = None
            try:
                if isinstance(fecha_emision_raw, datetime.datetime):
                    objeto_datetime_emision = fecha_emision_raw.date()
                elif isinstance(fecha_emision_raw, (int, float)):
                    objeto_datetime_emision = datetime.datetime.fromtimestamp((fecha_emision_raw - 25569) * 86400).date()
                elif isinstance(fecha_emision_raw, str) and fecha_emision_raw.strip():
                    for fmt in ["%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y", "%d-%m-%Y"]:
                        try:
                            objeto_datetime_emision = datetime.datetime.strptime(fecha_emision_raw.strip(), fmt).date()
                            break
                        except: continue
            except: pass

            # Columnas de dinero: forzar priorización de USD
            prima_raw = row[idx_prima_usd] if idx_prima_usd is not None else row[idx_prima] if idx_prima is not None else 0
            prima_val = sanitizar_numero(prima_raw, 0.0)

            pago_raw  = row[idx_pago] if idx_pago is not None else prima_raw
            pago_val  = sanitizar_numero(pago_raw, prima_val)

            status = 'ready'
            reason = 'Nuevo registro'

            if not cedula or len(cedula) < 5:
                status = 'invalid'; reason = 'Cédula no válida, faltante o muy corta'
            elif not poliza_raw or poliza_raw.lower() == 'none':
                status = 'invalid'; reason = 'Número de póliza faltante'
            elif not objeto_datetime_emision:
                status = 'invalid'; reason = 'Fecha de emisión ausente o formato inválido'
            elif prima_val <= 0:
                status = 'invalid'; reason = 'Monto de prima faltante o inválido'
            else:
                if cedula in asegurados_db:
                    status = 'insured_exists'; reason = 'Asegurado ya registrado'
                if poliza_raw in polizas_db:
                    if (poliza_raw, objeto_datetime_emision) in renovaciones_db:
                        status = 'duplicate'; reason = 'Póliza y renovación ya existen'
                    elif status != 'duplicate':
                        status = 'new_renewal'; reason = 'Nueva renovación para póliza existente'

            nro_cuota_val = row[idx_cuota] if idx_cuota is not None and row[idx_cuota] is not None else 1
            tipo_venta_val = row[idx_tipo_venta] if idx_tipo_venta is not None and row[idx_tipo_venta] is not None else 3

            # Serializar raw_data (convertir fechas/Decimal a str)
            raw_dict = {}
            for k, v in row_dict.items():
                if v is None:
                    raw_dict[k] = None
                elif k == 'TASA DE CAMBIO' or (idx_tasa is not None and k == header[idx_tasa]):
                    raw_dict[k] = sanitizar_numero(v, 0.0)
                else:
                    raw_dict[k] = str(v)

            row_info = {
                'type': 'row',
                'index': global_idx,
                'row': row_idx,
                'status': status,
                'reason': reason,
                'cedula': cedula,
                'tipo_ci': tipo_ci,
                'nombre_completo': str(row[idx_asegurado]).strip() if idx_asegurado is not None and row[idx_asegurado] is not None else "Desconocido",
                'poliza': poliza_raw,
                'ramo': normalizar_ramo(str(row[idx_ramo])) if idx_ramo is not None and row[idx_ramo] is not None else "Auto",
                'prima': prima_val,
                'monto_pago': pago_val,
                'fecha_emision': objeto_datetime_emision.isoformat() if objeto_datetime_emision else None,
                'ejecutivo': cod_ejecutivo,
                'ejecutivo_nombre': ejecutivo_nombre,
                'referidor_original': referidor_excel,
                'missing_executive': missing_executive,
                'raw_data': raw_dict,
                'nro_cuota': str(nro_cuota_val) if nro_cuota_val is not None else '1',
                'recibo': str(row[idx_recibo]).strip() if idx_recibo is not None and row[idx_recibo] is not None else None,
                'tipo_venta': str(tipo_venta_val) if tipo_venta_val is not None else '3',
                'correo': str(row[idx_correo]).strip() if idx_correo is not None and row[idx_correo] is not None else None,
                'telefono': str(row[idx_telefono]).strip() if idx_telefono is not None and row[idx_telefono] is not None else None,
                'moneda': '$'
            }
            yield row_info
            global_idx += 1

        yield {'type': 'done', 'total_processed': global_idx}

    except Exception as e:
        yield {'type': 'error', 'message': str(e)}


def insertar_registros_excel(data_to_insert):
    import datetime
    from dateutil.relativedelta import relativedelta
    inserted_count = 0
    error_count = 0
    
    try:
        with connectionBD() as connection:
            with connection.cursor() as cursor:
                for item in data_to_insert:
                    try:
                        raw = item['raw_data']
                        cedula = item['cedula']
                        tipo_ci = item['tipo_ci']
                        cod_poliza = item['poliza']
                        fecha_emision_raw = item['fecha_emision']
                        fecha_emision_dt = datetime.datetime.strptime(fecha_emision_raw, '%Y-%m-%d')
                        fecha_vencimiento = fecha_emision_dt + relativedelta(years=1)
                        
                        # 1. Asegurado (Crear o Actualizar si es necesario - por ahora solo creamos si no existe)
                        if item['status'] == 'ready':
                            # Solo insertar si realmente no existe (doble check)
                            cursor.execute("SELECT CI FROM asegurado WHERE CI = %s", (cedula,))
                            if not cursor.fetchone():
                                primer_nombre, segundo_nombre, primer_apellido, segundo_apellido = separar_nombre_completo(raw.get('ASEGURADO'))
                                sql_ins_asegurado = """
                                    INSERT INTO asegurado (CI, Nombre, Nombre2, Apellido, Apellido2, Tipo_CI, Correo, Telefono, Ejecutivo)
                                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                                """
                                cursor.execute(sql_ins_asegurado, (cedula, primer_nombre, segundo_nombre, primer_apellido, segundo_apellido, tipo_ci, item.get('correo'), item.get('telefono'), 3))

                        # 2. Póliza
                        cursor.execute("SELECT cod_poliza FROM poliza WHERE cod_poliza = %s", (cod_poliza,))
                        poliza_existe = cursor.fetchone()
                        
                        if not poliza_existe:
                            # Capturar Tipo_venta si existe en raw
                            tipo_venta = item.get('tipo_venta') or raw.get('TIPO VENTA') or 3
                            sql_ins_poliza = "INSERT INTO poliza (cod_poliza, CI_asegurado, Fecha_emision, Cod_compania, Tomador, Ramo, Tipo_venta) VALUES (%s, %s, %s, %s, %s, %s, %s)"
                            cursor.execute(sql_ins_poliza, (cod_poliza, cedula, fecha_emision_raw, 4, raw.get('TOMADOR'), "Auto", tipo_venta))
                            # Auto record 
                            cursor.execute("INSERT INTO Auto (Cod_poliza, Producto, Subramo) VALUES (%s, %s, %s) ON CONFLICT DO NOTHING", (cod_poliza, "RCV", "RCV"))

                        # 3. Renovación (Lógica Multi-Contrato)
                        renovacion_id = None
                        
                        # Buscar renovaciones existentes para esta póliza
                        cursor.execute("SELECT Cod_renovacion, Fecha_contrato, Fecha_vencimiento FROM renovacion WHERE Cod_poliza = %s", (cod_poliza,))
                        renovaciones = cursor.fetchall()
                        
                        fecha_pago_dt = fecha_emision_dt # Asumimos fecha de emisión como fecha de pago inicial
                        
                        # Intentar encontrar una renovación que cubra la fecha (con holgura de 1 mes antes y después)
                        for ren in renovaciones:
                            start_date = ren['Fecha_contrato'] 
                            # Asegurarse que start_date sea datetime.date para comparar
                            if isinstance(start_date, datetime.datetime): start_date = start_date.date()
                            
                            end_date = ren['Fecha_vencimiento']
                            if isinstance(end_date, datetime.datetime): end_date = end_date.date()
                            
                            # Rango válido: Inicio real - 1 mes <= Fecha Pago <= Fin real + 1 mes
                            valid_start = start_date - relativedelta(months=1)
                            valid_end = end_date + relativedelta(months=1)
                            
                            if valid_start <= fecha_pago_dt.date() <= valid_end:
                                renovacion_id = ren['Cod_renovacion']
                                break
                        
                        if not renovacion_id:
                            # Si no existe renovación que cubra la fecha, creamos una nueva
                            # Esto cubre el caso de "pagos de este año en el contrato que toca" (nuevo)
                            # y también "pagos de un contrato anterior" si no existe en BD.
                            
                            # Pero primero, verificar si es un pago muy antiguo para evitar duplicar por error
                            # Por ahora, creamos la renovación basada en la fecha del pago (asumiendo que es fecha inicio contrato)
                            
                            sql_ins_renovacion = "INSERT INTO renovacion (Cod_poliza, Prima, Frecuencia, Fecha_contrato, Fecha_vencimiento, comision) VALUES (%s, %s, %s, %s, %s, %s)"
                            frecuencia = 1 if raw.get('FRECUENCIA') == 'ANUAL' else None
                            prima_sanitizada = sanitizar_numero(raw.get('PRIMA TOTAL (SIN IGTF)') or item.get('prima'), 0.0)
                            cursor.execute(sql_ins_renovacion, (cod_poliza, prima_sanitizada, frecuencia, fecha_emision_raw, fecha_vencimiento.date(), 0.15))
                            renovacion_id = cursor.lastrowid


                        # 4. Pago (Robustez y Gaps)
                        monto_cobro = raw.get('BS. INICIAL / BS. FRACCION') or 0
                        
                        # Determinar Nro de Cuota (si está disponible en raw o calcular)
                        # Generic upload usually implies "Inicial" so quota 1, usually.
                        # If raw has quota info, use it. Generic Excel might not have it explicitly named "NRO CUOTA".
                        # Assuming generic is mostly Single payments or Inicial.
                        
                        # Check existance of this specific payment (Dup check by date and amount)
                        cursor.execute("SELECT Cod_pago FROM pago WHERE Cod_renovacion = %s AND fecha = %s AND monto = %s", (renovacion_id, fecha_emision_raw, monto_cobro))
                        pago_existente = cursor.fetchone()
                        
                        if not pago_existente:
                            nro_cuota = item.get('nro_cuota') or 1
                            recibo = item.get('recibo') or item.get('nro_recibo')
                            
                            tasa_val = sanitizar_numero(raw.get('TASA DE CAMBIO'), 0.0)
                            sql_ins_pago = "INSERT INTO pago (Cod_renovacion, moneda, fecha, Metodo_pago, tasa, monto, fecha_pagada, estado, nro_cuota, recibo) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
                            cursor.execute(sql_ins_pago, (renovacion_id, item.get('moneda'), fecha_emision_raw, "Pago Directo", tasa_val, monto_cobro, fecha_emision_raw, "PAGADO", nro_cuota, recibo))
                            last_pago_id = cursor.lastrowid
                            
                            # Commission processing
                            procesar_comision2(last_pago_id, 3)
                            inserted_count += 1
                        else:
                            # Already exists, skip or count as error/dup
                            print(f"Pago duplicado detectado para póliza {cod_poliza}")

                    except Exception as e:
                        print(f"Error insertando fila Genérica: {e}")
                        error_count += 1
                        continue
                
                connection.commit()
                return {'success': True, 'inserted': inserted_count, 'errors': error_count}

    except Exception as e:
        return {'success': False, 'error': str(e)}

def insertar_unico_registro_generico(item):
    """
    Inserta un registro genérico (asegurado, póliza, renovación, pago) en la BD.
    Devuelve un dict con:
        success  : bool
        message  : str resumen
        report   : list de dicts {step, status, detail} para mostrar en UI
        ids      : dict con los IDs reales insertados/encontrados
    """
    import datetime
    from dateutil.relativedelta import relativedelta

    report = []   # pasos ejecutados
    ids    = {}   # IDs resultantes

    def step(name, status, detail):
        report.append({'step': name, 'status': status, 'detail': str(detail)})

    try:
        with connectionBD() as connection:
            with connection.cursor() as cursor:

                # ── Extraer campos del item ─────────────────────────────────
                raw              = item.get('raw_data') or {}
                cedula           = str(item.get('cedula', '')).strip()
                tipo_ci          = str(item.get('tipo_ci', 'V')).strip().upper()
                cod_poliza       = str(item.get('poliza', '')).strip()
                fecha_emision_raw= str(item.get('fecha_emision', '')).strip()[:10]
                moneda           = item.get('moneda', '$')
                ramo_norm        = item.get('ramo', 'Auto')
                correo           = item.get('correo') or None
                telefono         = item.get('telefono') or None
                nro_cuota        = item.get('nro_cuota') or 1
                recibo           = item.get('recibo') or item.get('nro_recibo') or None

                # Ejecutivo: convertir a int o None (evita FK errors con string vacío)
                rawejec = item.get('ejecutivo')
                try:
                    cod_ejecutivo = int(rawejec) if rawejec not in (None, '', 'None') else None
                except (ValueError, TypeError):
                    cod_ejecutivo = None

                # Tipo venta: int
                try:
                    tipo_venta = int(item.get('tipo_venta') or raw.get('TIPO VENTA') or 3)
                except (ValueError, TypeError):
                    tipo_venta = 3

                # Fecha
                try:
                    fecha_emision_dt  = datetime.datetime.strptime(fecha_emision_raw, '%Y-%m-%d')
                    fecha_vencimiento = fecha_emision_dt + relativedelta(years=1)
                except ValueError as e:
                    return {'success': False, 'message': f'Fecha de emisión inválida: {fecha_emision_raw}', 'report': report, 'ids': ids}

                # Prima del raw (para renovación/pago) — sanitización robusta
                raw_prima = sanitizar_numero(
                    raw.get('PRIMA TOTAL (SIN IGTF)') or item.get('prima'),
                    default=0.0
                )

                # Monto cobro: Priorizar el valor pre-parseado del stream o la columna específica USD
                monto_cobro = sanitizar_numero(
                    item.get('monto_pago') or raw.get('INICIAL / FRACCION (SIN IGTF)') or raw_prima,
                    default=raw_prima
                )

                # ── 1. Asegurado ────────────────────────────────────────────
                # Resolver nombre desde raw_data['ASEGURADO'] (editado) o nombre_completo
                nombre_raw = raw.get('ASEGURADO') or item.get('nombre_completo') or 'Desconocido'
                try:
                    primer_nombre, segundo_nombre, primer_apellido, segundo_apellido = separar_nombre_completo(nombre_raw)
                except Exception:
                    primer_nombre = nombre_raw; segundo_nombre = primer_apellido = segundo_apellido = None

                sql_ins_aseg = """
                    INSERT INTO asegurado (CI, Nombre, Nombre2, Apellido, Apellido2, Tipo_CI, Correo, Telefono, Ejecutivo)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (CI) DO UPDATE SET
                        Correo=COALESCE(EXCLUDED.Correo, asegurado.Correo),
                        Telefono=COALESCE(EXCLUDED.Telefono, asegurado.Telefono),
                        Ejecutivo=COALESCE(EXCLUDED.Ejecutivo, asegurado.Ejecutivo)
                """
                cursor.execute(sql_ins_aseg, (
                    cedula, primer_nombre, segundo_nombre,
                    primer_apellido, segundo_apellido,
                    tipo_ci, correo, telefono, cod_ejecutivo
                ))
                ids['cedula'] = cedula
                step('Asegurado', 'ok', f'CI {tipo_ci}-{cedula} ({nombre_raw})')

                # ── 2. Póliza ───────────────────────────────────────────────
                cursor.execute("SELECT cod_poliza FROM poliza WHERE cod_poliza = %s", (cod_poliza,))
                poliza_existe = cursor.fetchone()

                if not poliza_existe:
                    tomador = raw.get('TOMADOR') or nombre_raw
                    sql_ins_pol = (
                        "INSERT INTO poliza (cod_poliza, CI_asegurado, Fecha_emision, Cod_compania, Tomador, Ramo, Tipo_venta) "
                        "VALUES (%s, %s, %s, %s, %s, %s, %s)"
                    )
                    cursor.execute(sql_ins_pol, (cod_poliza, cedula, fecha_emision_raw, 4, tomador, ramo_norm, tipo_venta))

                    # Subramo
                    try:
                        if ramo_norm == 'Auto':
                            cursor.execute("INSERT INTO Auto (Cod_poliza, Producto, Subramo) VALUES (%s,%s,%s) ON CONFLICT DO NOTHING", (cod_poliza, 'RCV', 'RCV'))
                        elif ramo_norm == 'Persona':
                            cursor.execute("INSERT INTO Persona (Cod_poliza, Producto, Subramo) VALUES (%s,%s,%s) ON CONFLICT DO NOTHING", (cod_poliza, 'SALUD', 'SALUD'))
                        elif ramo_norm == 'Patrimonial':
                            cursor.execute("INSERT INTO patrimonio (Cod_poliza, Direccion) VALUES (%s,%s) ON CONFLICT DO NOTHING", (cod_poliza, tomador))
                        elif ramo_norm == 'Fianza':
                            cursor.execute("INSERT INTO fianza (Cod_poliza) VALUES (%s) ON CONFLICT DO NOTHING", (cod_poliza,))
                        elif ramo_norm == 'Viaje':
                            cursor.execute("INSERT INTO viaje (Cod_poliza) VALUES (%s) ON CONFLICT DO NOTHING", (cod_poliza,))
                    except Exception as e_sub:
                        step('Subramo', 'warn', f'No insertado: {e_sub}')

                    ids['poliza'] = cod_poliza
                    step('Póliza', 'ok', f'{cod_poliza} ({ramo_norm}) — NUEVA')
                else:
                    ids['poliza'] = cod_poliza
                    step('Póliza', 'info', f'{cod_poliza} — ya existía en BD')

                # ── 3. Renovación ───────────────────────────────────────────
                renovacion_id = None
                cursor.execute(
                    "SELECT Cod_renovacion, Fecha_contrato, Fecha_vencimiento FROM renovacion WHERE Cod_poliza = %s",
                    (cod_poliza,)
                )
                renovaciones = cursor.fetchall()
                fecha_pago_date = fecha_emision_dt.date()

                for ren in renovaciones:
                    sd = ren['Fecha_contrato']
                    ed = ren['Fecha_vencimiento']
                    if isinstance(sd, datetime.datetime): sd = sd.date()
                    if isinstance(ed, datetime.datetime): ed = ed.date()
                    if isinstance(sd, str): 
                        try: sd = datetime.datetime.strptime(sd[:10], '%Y-%m-%d').date()
                        except: continue
                    if isinstance(ed, str):
                        try: ed = datetime.datetime.strptime(ed[:10], '%Y-%m-%d').date()
                        except: continue
                    if (sd - relativedelta(months=1)) <= fecha_pago_date <= (ed + relativedelta(months=1)):
                        renovacion_id = ren['Cod_renovacion']
                        break

                if not renovacion_id:
                    frecuencia = 1 if str(raw.get('FRECUENCIA', '')).upper() == 'ANUAL' else None
                    cursor.execute(
                        "INSERT INTO renovacion (Cod_poliza, Prima, Frecuencia, Fecha_contrato, Fecha_vencimiento, comision) "
                        "VALUES (%s, %s, %s, %s, %s, %s)",
                        (cod_poliza, raw_prima, frecuencia, fecha_emision_raw, fecha_vencimiento.date(), 0.15)
                    )
                    renovacion_id = cursor.lastrowid
                    ids['renovacion'] = renovacion_id
                    step('Renovación', 'ok', f'ID {renovacion_id} ({fecha_emision_raw} → {fecha_vencimiento.date()})')
                else:
                    ids['renovacion'] = renovacion_id
                    step('Renovación', 'info', f'ID {renovacion_id} — ya existía')

                # ── 4. Pago ─────────────────────────────────────────────────
                tasa_val = sanitizar_numero(raw.get('TASA DE CAMBIO'), 0.0)
                cursor.execute(
                    "SELECT Cod_pago FROM pago WHERE Cod_renovacion = %s AND fecha = %s",
                    (renovacion_id, fecha_emision_raw)
                )
                pago_existe = cursor.fetchone()

                if not pago_existe:
                    cursor.execute(
                        "INSERT INTO pago (Cod_renovacion, moneda, fecha, Metodo_pago, tasa, monto, fecha_pagada, estado, nro_cuota, recibo) "
                        "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
                        (renovacion_id, moneda, fecha_emision_raw, 'Pago Directo',
                         tasa_val, monto_cobro, fecha_emision_raw, 'PAGADO', nro_cuota, recibo)
                    )
                    pago_id = cursor.lastrowid
                    ids['pago'] = pago_id
                    step('Pago', 'ok', f'ID {pago_id}, monto {moneda} {monto_cobro}')

                    # Comisión (no bloquea el commit si falla)
                    try:
                        procesar_comision2(pago_id, 3)
                        step('Comisión', 'ok', 'Procesada')
                    except Exception as e_com:
                        step('Comisión', 'warn', f'No procesada: {e_com}')
                else:
                    ids['pago'] = pago_existe['Cod_pago']
                    step('Pago', 'info', f'ID {pago_existe["Cod_pago"]} — ya existía para esta fecha')

                # ── COMMIT ──────────────────────────────────────────────────
                connection.commit()

                # ── Verificación post-commit ────────────────────────────────
                # Confirmar que la póliza realmente quedó en BD
                cursor.execute(
                    "SELECT p.cod_poliza, p.CI_asegurado, p.Ramo, p.Fecha_emision, "
                    "       a.Nombre, a.Apellido, "
                    "       r.Cod_renovacion, r.Prima "
                    "FROM poliza p "
                    "LEFT JOIN asegurado a ON a.CI = p.CI_asegurado "
                    "LEFT JOIN renovacion r ON r.Cod_poliza = p.cod_poliza AND r.Cod_renovacion = %s "
                    "WHERE p.cod_poliza = %s",
                    (ids.get('renovacion'), cod_poliza)
                )
                verificacion = cursor.fetchone()

                if verificacion:
                    step('Verificación BD', 'ok',
                         f"Póliza {cod_poliza} confirmada — Aseg: {verificacion.get('Nombre','')} {verificacion.get('Apellido','')}")
                    return {
                        'success': True,
                        'message': f'Registro insertado y verificado correctamente',
                        'report': report,
                        'ids': ids,
                        'verificado': True
                    }
                else:
                    step('Verificación BD', 'error',
                         f'La póliza {cod_poliza} NO se encontró en BD después del commit')
                    return {
                        'success': False,
                        'message': f'El commit reportó éxito pero la póliza {cod_poliza} no se encuentra en BD. Revisar logs.',
                        'report': report,
                        'ids': ids,
                        'verificado': False
                    }

    except Exception as e:
        import traceback
        step('Error crítico', 'error', traceback.format_exc())
        return {
            'success': False,
            'message': str(e),
            'report': report,
            'ids': ids
        }

def procesar_mercantil_preview(file):

    """
    Procesa el archivo Excel de Mercantil Panama para generar una vista previa robusta.
    Basado en el script 'generador de update ultimo generador.py'.
    """
   

    try:
        df = pd.read_excel(file, engine='openpyxl', dtype=str)
        df.columns = [str(col).strip().upper() for col in df.columns]
        
        # Mapeo de columnas flexible (basado en COLUMNAS_MATRIZ_CORTE_FLEXIBLE del script de referencia)
        MAPEO_COLS = {
            'POLIZA': ['POLIZA'],
            'SEGUROS': ['SEGUROS', 'COMPAÑIA', 'COMPAÑÍA'],
            'RAMO': ['RAMO'],
            'PRODUCTO': ['PRODUCTO'],
            'SUB_RAMO': ['SUB RAMO', 'PLAN'],
            'TOMADOR': ['TOMADOR', 'CLIENTE'],
            'CORREDOR': ['CORREDOR / REFERIDOR', 'CORREDOR', 'REFERIDOR'],
            'ID': ['ID', 'CEDULA', 'RUT', 'DOCUMENTO'],
            'NACIONALIDAD': ['NACIONALIDAD', 'TIPO ID'],
            'PRIMER_NOMBRE': ['PRIMER NOMBRE'],
            'SEGUNDO_NOMBRE': ['SEGUNDO NOMBRE'],
            'PRIMER_APELLIDO': ['PRIMER APELLIDO'],
            'SEGUNDO_APELLIDO': ['SEGUNDO APELLIDO'],
            'SUMA_ASEGURADA': ['SUMA ASEGURADA', 'COBERTURA'],
            'FECHA_EMISION': ['FECHA EMISION', 'FECHA VIGENCIA', 'VIGENCIA'],
            'PRIMA': ['PRIMA TOTAL (SIN IGTF)', 'PRIMA'],
            'INICIAL': ['INICIAL / FRACCION (SIN IGTF)', 'INICIAL', 'FRACCION'],
            'FRECUENCIA': ['FRECUENCIA'],
            'MONEDA': ['$/BS', 'MONEDA'],
            'METODO': ['METODO DE PAGO', 'METODO'],
            'TASA': ['TASA DE CAMBIO', 'TASA'],
            'ESTATUS': ['ESTATUS', 'ESTADO DE PAGO'],
            'NRO_CUOTA': ['NRO CUOTAS', 'CUOTA', 'NRO. CUOTA'],
            'PLACA': ['PLACA'],
            'MARCA': ['MARCA'],
            'MODELO': ['MODELO'],
            'AÑO': ['AÑO', 'ANIO'],
            'CORREO': ['CORREO ASEGURADO', 'EMAIL', 'CORREO'],
            'TELEFONO': ['TELEFONO ASEGURADO', 'TELEFONO', 'TELÉFONO'],
            'FECHA_COBRO': ['FECHA DE COBRO', 'FECHA COBRO', 'FECHA PAGADA'],
            'REFERIDOR': ['REFERIDOR'],
            'TIPO_VENTA': ['N TIPO DE PÓLIZA', 'TIPO DE POLIZA', 'TIPO VENTA'],
            'RECIBO': ['NRO. RECIBO/ FACTURA', 'NRO. RECIBO', 'RECIBO', 'REFERENCIA']
        }

        # Función para encontrar la columna real en el DF
        def find_col(possible_names):
            for name in possible_names:
                if name.upper() in df.columns:
                    return name.upper()
            return None

        col_map = {k: find_col(v) for k, v in MAPEO_COLS.items()}
        
        # Columnas críticas
        if not col_map['POLIZA']:
            return {'success': False, 'error': "No se encontró la columna 'POLIZA'"}

        # Eliminar filas sin póliza y limpiar espacios
        df[col_map['POLIZA']] = df[col_map['POLIZA']].astype(str).str.strip()
        df = df[df[col_map['POLIZA']] != 'nan']
        
        # Agrupar por póliza
        grupos = df.groupby(col_map['POLIZA'])
        
        preview_data = []
        
        # Obtener ejecutivos y pólizas existentes para validación
        ejecutivos = lista_ejecutivosBD() or []
        # Crear mapa de nombres para búsqueda flexible (simplificado) y mapa de ID a Nombre
        ejec_map = {}
        ejec_id_to_name = {}
        for ej in ejecutivos:
            nombre_full = f"{ej['Nombre']} {ej['Apellido']}".strip().upper()
            ejec_map[nombre_full] = ej['cod_ejecutivo']
            ejec_id_to_name[ej['cod_ejecutivo']] = f"{ej['Nombre']} {ej['Apellido']}".strip()

        # Obtener compañías para validación
        with connectionBD() as connection:
            with connection.cursor() as cursor:
                cursor.execute("SELECT Cod_compania, Nombre FROM compania")
                cias_db_list = cursor.fetchall()
                # Mapa normalizado para búsqueda y mapa de ID a Nombre
                cia_map = {normalizar_texto(c['Nombre']): c['Cod_compania'] for c in cias_db_list}
                cia_id_to_name = {c['Cod_compania']: c['Nombre'] for c in cias_db_list}

                # Cache de asegurados existentes
                cursor.execute("SELECT CI FROM asegurado")
                asegurados_db = {str(r['CI']) for r in cursor.fetchall()}
                
                # Cache de pólizas existentes
                cursor.execute("SELECT cod_poliza FROM poliza")
                polizas_db = {}
                for r in cursor.fetchall():
                    c_pol = str(r['cod_poliza'])
                    c_clean = c_pol.replace("-", "").replace(" ", "")
                    polizas_db[c_pol] = c_pol
                    polizas_db[c_clean] = c_pol
                    if c_clean.startswith("0"):
                        polizas_db[c_clean[1:]] = c_pol

        for cod_poliza, group in grupos:
            main_row = group.iloc[0]
            
            poliza_raw = str(cod_poliza).strip()
            poliza_clean = poliza_raw.replace("-", "").replace(" ", "")
            poliza_sin_cero = poliza_clean[1:] if poliza_clean.startswith("0") else poliza_clean
            poliza_match = polizas_db.get(poliza_raw) or polizas_db.get(poliza_clean) or polizas_db.get(poliza_sin_cero)
            if poliza_match:
                cod_poliza = poliza_match
            
            # Datos del Asegurado
            ci_raw = str(main_row.get(col_map['ID'], '')).strip()
            # Limpiar CI de puntos/guiones
            ci = "".join(filter(str.isdigit, ci_raw)) if ci_raw else ""
            if not ci: ci = ci_raw # Fallback si no es numérico
            
            nacionalidad = str(main_row.get(col_map['NACIONALIDAD'], 'V')).strip().upper()[:1]
            if nacionalidad not in ['V', 'E', 'J', 'G', 'P', 'M']: nacionalidad = 'V'

            nombre1 = main_row.get(col_map['PRIMER_NOMBRE'])
            apellido1 = main_row.get(col_map['PRIMER_APELLIDO'])
            
            # Si no hay nombres explícitos, intentar del Tomador
            if pd.isna(nombre1) and pd.isna(apellido1):
                tomador = str(main_row.get(col_map['TOMADOR'], ''))
                nombre1, _, apellido1, _ = separar_nombre_completo(tomador)

            # Determinar estatus
            status = 'ready'
            if ci in asegurados_db:
                status = 'insured_exists'
            if str(cod_poliza) in polizas_db:
                status = 'policy_exists' # O 'new_renewal' dependiendo de la lógica deseada

            # Mapeo de Ramo
            ramo_raw = str(main_row.get(col_map['RAMO'], '')).strip()
            ramo_normalizado = normalizar_ramo(ramo_raw)

            # Preparar fecha de emisión para comparación de pagos
            fecha_emision_str = str(main_row.get(col_map['FECHA_EMISION'], ''))
            try:
                if ' ' in fecha_emision_str: fecha_emision_str = fecha_emision_str.split(' ')[0]
                if '-' not in fecha_emision_str and '/' in fecha_emision_str:
                    parts = fecha_emision_str.split('/')
                    if len(parts[0]) == 4: fecha_emision_str = f"{parts[0]}-{parts[1]}-{parts[2]}"
                    else: fecha_emision_str = f"{parts[2]}-{parts[1]}-{parts[0]}"
                fecha_e_dt = datetime.datetime.strptime(fecha_emision_str, '%Y-%m-%d')
            except:
                fecha_e_dt = datetime.datetime.now()

            # Buscar Ejecutivo (Referidor)
            referidor_excel = str(main_row.get(col_map['REFERIDOR'] or col_map['CORREDOR'], '')).strip()
            cod_ejecutivo = buscar_mejor_match_ejecutivo(referidor_excel, ejec_map)
            
            missing_executive = False
            if not cod_ejecutivo:
                missing_executive = True
                cod_ejecutivo = 3 # Default 3 (Otros/Oficina) if not found
            
            ejecutivo_nombre = ejec_id_to_name.get(cod_ejecutivo, "Oficina/Otros")

            # Buscar Compañía
            compania_excel = str(main_row.get(col_map['SEGUROS'], '')).strip()
            cod_compania = buscar_mejor_match_compania(compania_excel, cia_map)
            
            missing_company = False
            if not cod_compania:
                missing_company = True
                # No ponemos default todavía, el usuario debe elegir o usaremos 1 (Mercantil) como fallback en inserción
            
            compania_nombre = cia_id_to_name.get(cod_compania, "No Encontrada")
            
            # Tipo de Venta
            tipo_venta = str(main_row.get(col_map['TIPO_VENTA'], 'NUEVA')).strip().upper()

            # Consultar pagos existentes para esta póliza en la BD
            existentes_db = []
            with connectionBD() as conn:
                with conn.cursor() as cur:
                    cur.execute("""
                        SELECT p.Cod_pago, p.monto, p.nro_cuota, p.estado, r.Fecha_contrato 
                        FROM pago p 
                        JOIN renovacion r ON p.Cod_renovacion = r.Cod_renovacion 
                        WHERE r.Cod_poliza = %s
                    """, (str(cod_poliza),))
                    existentes_db = cur.fetchall()

            pagos = []
            has_new = False
            has_update = False

            for _, row in group.iterrows():
                f_cobro_val = row.get(col_map['FECHA_COBRO'])
                f_cobro_raw = str(f_cobro_val).split(' ')[0] if pd.notna(f_cobro_val) and str(f_cobro_val).lower() != 'nan' else None
                
                monto_excel = row.get(col_map['INICIAL'])
                cuota_excel = row.get(col_map['NRO_CUOTA'])
                recibo_excel = row.get(col_map['RECIBO'])
                estado_excel = str(row.get(col_map['ESTATUS'], 'PAGADO')).strip().upper()

                es_anterior = False
                try:
                    f_cobro_dt = datetime.datetime.strptime(f_cobro_raw, '%Y-%m-%d')
                    # Detección inteligente: solo es anterior si la fecha es MUY lejana (> 90 días)
                    # o si hay evidencia de conflicto (mismo nro de cuota con distinto monto)
                    if f_cobro_dt < (fecha_e_dt - datetime.timedelta(days=90)):
                        es_anterior = True
                    elif f_cobro_dt < fecha_e_dt:
                        # Si está en el margen de 90 días, solo marcar si hay conflicto de montos para la misma cuota
                        evidencia_conflicto = False
                        # 1. Conflicto con BD
                        for pdb in existentes_db:
                            if str(pdb['nro_cuota']) == str(cuota_excel) and abs(float(pdb['monto']) - float(monto_excel or 0)) > 0.01:
                                evidencia_conflicto = True
                                break
                        # 2. Conflicto en el mismo Excel
                        if not evidencia_conflicto:
                            cuotas_iguales = group[group[col_map['NRO_CUOTA']].astype(str) == str(cuota_excel)]
                            if len(cuotas_iguales) > 1:
                                for _, r_comp in cuotas_iguales.iterrows():
                                    if abs(float(r_comp.get(col_map['INICIAL']) or 0) - float(monto_excel or 0)) > 0.01:
                                        evidencia_conflicto = True
                                        break
                        if evidencia_conflicto:
                            es_anterior = True
                except:
                    pass

                pago_status = 'new'
                # Buscar coincidencia exacto para evitar duplicados
                found = False
                for pdb in existentes_db:
                    if str(pdb['nro_cuota']) == str(cuota_excel) and abs(float(pdb['monto']) - float(monto_excel or 0)) < 0.01:
                        found = True
                        if pdb['estado'] != estado_excel and estado_excel == 'PAGADO':
                            pago_status = 'update'
                            has_update = True
                        else:
                            pago_status = 'exists'
                        break
                
                if not found:
                    pago_status = 'new'
                    has_new = True

                pagos.append({
                    'monto': monto_excel,
                    'moneda': row.get(col_map['MONEDA']),
                    'tasa': row.get(col_map['TASA']),
                    'metodo': row.get(col_map['METODO']),
                    'estado': estado_excel,
                    'fecha_cobro': f_cobro_raw,
                    'nro_cuota': cuota_excel,
                    'recibo': recibo_excel,
                    'es_anterior': es_anterior,
                    'pago_status': pago_status
                })

            # Determinar estatus final de la fila
            if status == 'ready':
                pass # Mantiene ready
            elif has_new:
                status = 'new_payments'
            elif has_update:
                status = 'updates_needed'
            else:
                status = 'up_to_date'

            item = {
                'poliza': str(cod_poliza),
                'status': status,
                'asegurado': {
                    'ci': ci,
                    'tipo_ci': nacionalidad,
                    'nombre': nombre1 or '',
                    'nombre2': main_row.get(col_map['SEGUNDO_NOMBRE']) or '',
                    'apellido': apellido1 or '',
                    'apellido2': main_row.get(col_map['SEGUNDO_APELLIDO']) or '',
                    'correo': main_row.get(col_map['CORREO']) or '',
                    'telefono': main_row.get(col_map['TELEFONO']) or ''
                },
                'poliza_data': {
                    'compania': main_row.get(col_map['SEGUROS']),
                    'cod_compania': cod_compania,
                    'compania_nombre': compania_nombre,
                    'missing_company': missing_company,
                    'ramo': ramo_normalizado,
                    'ramo_original': ramo_raw,
                    'tomador': main_row.get(col_map['TOMADOR']),
                    'fecha_emision': str(main_row.get(col_map['FECHA_EMISION'], '')),
                    'prima': main_row.get(col_map['PRIMA']),
                    'frecuencia': main_row.get(col_map['FRECUENCIA']),
                    'cobertura': main_row.get(col_map['SUMA_ASEGURADA']),
                    'sub_ramo': main_row.get(col_map['SUB_RAMO']),
                    'producto': main_row.get(col_map['PRODUCTO'])
                },
                'ramo_especifico': {
                    'placa': main_row.get(col_map['PLACA']),
                    'marca': main_row.get(col_map['MARCA']),
                    'modelo': main_row.get(col_map['MODELO']),
                    'año': main_row.get(col_map['AÑO'])
                },
                'pagos': pagos,
                'ejecutivo': cod_ejecutivo,
                'ejecutivo_nombre': ejecutivo_nombre,
                'referidor_original': referidor_excel,
                'missing_executive': missing_executive,
                'tipo_venta': tipo_venta
            }
            preview_data.append(item)

        return {'success': True, 'data': limpiar_nan(preview_data), 'ejecutivos': ejecutivos, 'companias': cias_db_list}

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return {'success': False, 'error': str(e)}

def insertar_mercantil_data(data_to_insert):
    """
    Inserta los registros procesados de Mercantil Panama.
    Realiza inserciones jerárquicas en múltiples tablas.
    """
    import datetime
    from dateutil.relativedelta import relativedelta
    inserted_count = 0
    error_count = 0
    
    try:
        with connectionBD() as connection:
            with connection.cursor() as cursor:
                # Mapeo de compañías (simulado, debería ser dinámico)
                cursor.execute("SELECT Cod_compania, Nombre FROM compania")
                cias_db = {r['Nombre'].upper(): r['Cod_compania'] for r in cursor.fetchall()}

                # Mapeo de frecuencias
                # Mapeo: '1': Anual, '2': Mensual, '3': Trimestral, '4': Semestral, '5': Unica, '6': Cuatrimestral.
                frec_map = {'ANUAL': 1, 'MENSUAL': 2, 'TRIMESTRAL': 3, 'SEMESTRAL': 4, 'UNICA': 5, 'CUATRIMESTRAL': 6}

                for item in data_to_insert:
                    try:
                        asig = item['asegurado']
                        pol = item['poliza_data']
                        ramo_esp = item['ramo_especifico']
                        
                        # 1. Asegurado
                        # Usamos ON CONFLICT para actualizar datos si el asegurado ya existe
                        sql_asegurado = """
                            INSERT INTO asegurado (CI, Tipo_CI, Nombre, Nombre2, Apellido, Apellido2, Correo, Telefono, Ejecutivo)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                            ON CONFLICT (CI) DO UPDATE SET Correo=EXCLUDED.Correo, Telefono=EXCLUDED.Telefono, Ejecutivo=EXCLUDED.Ejecutivo
                        """
                        cursor.execute(sql_asegurado, (asig['ci'], asig['tipo_ci'], asig['nombre'], asig['nombre2'], asig['apellido'], asig['apellido2'], asig['correo'], asig['telefono'], item.get('ejecutivo', 3)))

                        # 2. Póliza
                        # Usar el ID enviado desde el preview si existe, sino buscar de nuevo
                        cod_compania = pol.get('cod_compania')
                        if not cod_compania:
                             cod_compania = cias_db.get(str(pol['compania']).upper(), 1) # Default 1 if not found

                        # Limpiar fecha de emisión
                        fecha_e = pol['fecha_emision']
                        if ' ' in fecha_e: fecha_e = fecha_e.split(' ')[0]
                        if '-' not in fecha_e and '/' in fecha_e:
                            parts = fecha_e.split('/')
                            if len(parts[0]) == 4: fecha_e = f"{parts[0]}-{parts[1]}-{parts[2]}"
                            else: fecha_e = f"{parts[2]}-{parts[1]}-{parts[0]}"

                        # 2. Póliza
                        # Intentamos insertar o actualizar siempre para asegurar la integridad
                        sql_poliza = """
                            INSERT INTO poliza (cod_poliza, CI_asegurado, Fecha_emision, Cod_compania, Tomador, Ramo, Tipo_venta)
                            VALUES (%s, %s, %s, %s, %s, %s, %s)
                            ON CONFLICT (cod_poliza) DO UPDATE SET Cod_compania=EXCLUDED.Cod_compania, Tomador=EXCLUDED.Tomador, Ramo=EXCLUDED.Ramo, Tipo_venta=EXCLUDED.Tipo_venta
                        """
                        cursor.execute(sql_poliza, (item['poliza'], asig['ci'], fecha_e, cod_compania, pol['tomador'], pol['ramo'], 3))

                        # 2.1 Ramo Específico (Insertar si no existe)
                        if pol['ramo'] == 'Auto':
                            cursor.execute("""
                                INSERT INTO Auto (Cod_poliza, Placa, Marca, Modelo, año, Producto, Subramo)
                                VALUES (%s, %s, %s, %s, %s, %s, %s)
                                ON CONFLICT DO NOTHING
                            """, (item['poliza'], ramo_esp['placa'], ramo_esp['marca'], ramo_esp['modelo'], ramo_esp['año'], pol.get('producto') or "RCV", pol['sub_ramo']))
                        elif pol['ramo'] == 'Persona':
                            cursor.execute("""
                                INSERT INTO Persona (Cod_poliza, Producto, Subramo)
                                VALUES (%s, %s, %s)
                                ON CONFLICT DO NOTHING
                            """, (item['poliza'], pol.get('producto') or "SALUD", pol['sub_ramo']))
                        elif pol['ramo'] == 'Patrimonial':
                            cursor.execute("""
                                INSERT INTO patrimonio (Cod_poliza, Direccion)
                                VALUES (%s, %s)
                                ON CONFLICT DO NOTHING
                            """, (item['poliza'], pol['tomador']))
                        elif pol['ramo'] == 'Fianza':
                            cursor.execute("INSERT INTO fianza (Cod_poliza) VALUES (%s) ON CONFLICT DO NOTHING", (item['poliza'],))
                        elif pol['ramo'] == 'Viaje':
                            cursor.execute("INSERT INTO viaje (Cod_poliza) VALUES (%s) ON CONFLICT DO NOTHING", (item['poliza'],))

                        # 3. Renovación
                        renovacion_id = None
                        try:
                            fecha_e_dt = datetime.datetime.strptime(fecha_e, '%Y-%m-%d')
                            fecha_v = fecha_e_dt + relativedelta(years=1)
                        except:
                            fecha_e_dt = datetime.date.today()
                            fecha_v = fecha_e_dt + relativedelta(years=1)

                        frec_val = frec_map.get(str(pol['frecuencia']).upper(), 1)

                        # Buscar renovacion existente que coincida con la fecha de contrato (fecha_emision)
                        # Margen de error de dias? No, Mercantil debe ser exacta con la fecha de emision de poliza/renovacion.
                        # Pero por seguridad buscamos coincidencia exacta de fecha_contrato para esa poliza.
                        cursor.execute("SELECT Cod_renovacion FROM renovacion WHERE Cod_poliza = %s AND Fecha_contrato = %s", (item['poliza'], fecha_e))
                        ren_existente = cursor.fetchone()

                        if ren_existente:
                            renovacion_id = ren_existente['Cod_renovacion']
                        else:
                            # Crear nueva renovación
                            sql_renovacion = """
                                INSERT INTO renovacion (Cod_poliza, Prima, Frecuencia, Fecha_contrato, Fecha_vencimiento, cobertura, comision)
                                VALUES (%s, %s, %s, %s, %s, %s, %s)
                            """
                            cursor.execute(sql_renovacion, (item['poliza'], pol['prima'], frec_val, fecha_e, fecha_v, pol['cobertura'], 0.15))
                            renovacion_id = cursor.lastrowid

                        # 4. Pagos
                        frec_val = frec_map.get(str(pol['frecuencia']).upper(), 1)
                        for pago in item['pagos']:
                            f_cobro_raw = pago.get('fecha_cobro')
                            
                            try:
                                if f_cobro_raw and str(f_cobro_raw).lower() != 'nan':
                                    f_cobro_dt = datetime.datetime.strptime(str(f_cobro_raw).split(' ')[0], '%Y-%m-%d')
                                else:
                                    f_cobro_raw = None # Asegurar NULL para la BD
                                    f_cobro_dt = fecha_e_dt # Fallback solo para lógica de renovación
                            except:
                                f_cobro_dt = fecha_e_dt
                            
                            # Determinar si el pago pertenece a la renovación actual o a la anterior
                            # Si la fecha de cobro es anterior a la fecha de emisión (con 15 días de gracia), asumimos renovación anterior
                            if f_cobro_dt < (fecha_e_dt - datetime.timedelta(days=15)):
                                target_start_dt = fecha_e_dt - relativedelta(years=1)
                            else:
                                target_start_dt = fecha_e_dt
                                
                            target_start_str = target_start_dt.strftime('%Y-%m-%d')
                            
                            # Buscar o crear la renovación correspondiente
                            cursor.execute("SELECT Cod_renovacion FROM renovacion WHERE Cod_poliza = %s AND Fecha_contrato = %s", (item['poliza'], target_start_str))
                            res_ren = cursor.fetchone()
                            
                            if res_ren:
                                target_renov_id = res_ren['Cod_renovacion']
                            else:
                                target_venc_dt = target_start_dt + relativedelta(years=1)
                                sql_new_ren = """
                                    INSERT INTO renovacion (Cod_poliza, Prima, Frecuencia, Fecha_contrato, Fecha_vencimiento, cobertura, comision)
                                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                                """
                                cursor.execute(sql_new_ren, (item['poliza'], pol['prima'], frec_val, target_start_str, target_venc_dt.strftime('%Y-%m-%d'), pol['cobertura'], 0.15))
                                target_renov_id = cursor.lastrowid

                            nro_cuota = pago.get('nro_cuota', 1)

                            # Buscar pago existente para decidir entre INSERT o UPDATE
                            cursor.execute("SELECT Cod_pago, estado FROM pago WHERE Cod_renovacion = %s AND nro_cuota = %s AND monto = %s", (target_renov_id, nro_cuota, pago['monto']))
                            pago_db = cursor.fetchone()

                            status_pago = pago.get('estado') or 'PAGADO'

                            if not pago_db:
                                sql_pago = """
                                    INSERT INTO pago (Cod_renovacion, moneda, fecha, Metodo_pago, tasa, monto, fecha_pagada, estado, nro_cuota, recibo)
                                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                                """
                                cursor.execute(sql_pago, (target_renov_id, pago['moneda'], f_cobro_raw, pago['metodo'] or 'PAGO DIRECTO', pago['tasa'], pago['monto'], f_cobro_raw, status_pago, nro_cuota, pago.get('recibo')))
                                procesar_comision2(cursor.lastrowid, 3, item['poliza'])
                            elif str(pago_db['estado']).strip().upper() != str(status_pago).strip().upper():
                                # Actualizar estado si cambió
                                sql_update_pago = "UPDATE pago SET estado = %s, fecha_pagada = %s WHERE Cod_pago = %s"
                                cursor.execute(sql_update_pago, (status_pago, f_cobro_raw or datetime.date.today(), pago_db['Cod_pago']))
                            else:
                                print(f"Pago omitido por duplicidad exacta: {item['poliza']} cuota {nro_cuota}")

                        inserted_count += 1
                    except Exception as e:
                        print(f"Error insertando registro Mercantil {item['poliza']}: {e}")
                        error_count += 1
                
                connection.commit()
                return {'success': True, 'inserted': inserted_count, 'errors': error_count}

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return {'success': False, 'error': str(e)}

def insertar_unico_registro_mercantil(item):
 
    import datetime
    from dateutil.relativedelta import relativedelta
    try:
        with connectionBD() as connection:
            with connection.cursor() as cursor:
                cursor.execute("SELECT Cod_compania, Nombre FROM compania")
                cias_db = {r['Nombre'].upper(): r['Cod_compania'] for r in cursor.fetchall()}
                frec_map = {'ANUAL': 1, 'MENSUAL': 2, 'TRIMESTRAL': 3, 'SEMESTRAL': 4, 'UNICA': 5, 'CUATRIMESTRAL': 6}

                asig = item['asegurado']
                pol = item['poliza_data']
                ramo_esp = item['ramo_especifico']
                
                # 1. Asegurado
                sql_asegurado = """
                    INSERT INTO asegurado (CI, Tipo_CI, Nombre, Nombre2, Apellido, Apellido2, Correo, Telefono, Ejecutivo)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (CI) DO UPDATE SET Correo=EXCLUDED.Correo, Telefono=EXCLUDED.Telefono, Ejecutivo=EXCLUDED.Ejecutivo
                """
                cursor.execute(sql_asegurado, (asig['ci'], asig['tipo_ci'], asig['nombre'], asig['nombre2'], asig['apellido'], asig['apellido2'], asig['correo'], asig['telefono'], item.get('ejecutivo', 3)))

                # 2. Póliza
                cod_compania = pol.get('cod_compania')
                if not cod_compania:
                     cod_compania = cias_db.get(str(pol['compania']).upper(), 1)

                fecha_e = pol['fecha_emision']
                if ' ' in fecha_e: fecha_e = fecha_e.split(' ')[0]
                if '-' not in fecha_e and '/' in fecha_e:
                    parts = fecha_e.split('/')
                    if len(parts[0]) == 4: fecha_e = f"{parts[0]}-{parts[1]}-{parts[2]}"
                    else: fecha_e = f"{parts[2]}-{parts[1]}-{parts[0]}"

                # Siempre intentar asegurar que la póliza exista (Upsert)
                sql_poliza = """
                    INSERT INTO poliza (cod_poliza, CI_asegurado, Fecha_emision, Cod_compania, Tomador, Ramo, Tipo_venta)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (cod_poliza) DO UPDATE SET Cod_compania=EXCLUDED.Cod_compania, Tomador=EXCLUDED.Tomador, Ramo=EXCLUDED.Ramo, Tipo_venta=EXCLUDED.Tipo_venta
                """
                cursor.execute(sql_poliza, (item['poliza'], asig['ci'], fecha_e, cod_compania, pol['tomador'], pol['ramo'], 3))

                # Ramo Específico (Siempre intentar asegurar que exista)
                if pol['ramo'] == 'Auto':
                    cursor.execute("INSERT INTO Auto (Cod_poliza, Placa, Marca, Modelo, año, Producto, Subramo) VALUES (%s, %s, %s, %s, %s, %s, %s) ON CONFLICT DO NOTHING",
                                   (item['poliza'], ramo_esp['placa'], ramo_esp['marca'], ramo_esp['modelo'], ramo_esp['año'], pol.get('producto') or "RCV", pol['sub_ramo']))
                elif pol['ramo'] == 'Persona':
                    cursor.execute("INSERT INTO Persona (Cod_poliza, Producto, Subramo) VALUES (%s, %s, %s) ON CONFLICT DO NOTHING", (item['poliza'], pol.get('producto') or "SALUD", pol['sub_ramo']))
                elif pol['ramo'] == 'Patrimonial':
                    cursor.execute("INSERT INTO patrimonio (Cod_poliza, Direccion) VALUES (%s, %s) ON CONFLICT DO NOTHING", (item['poliza'], pol['tomador']))
                elif pol['ramo'] == 'Fianza':
                    cursor.execute("INSERT INTO fianza (Cod_poliza) VALUES (%s) ON CONFLICT DO NOTHING", (item['poliza'],))
                elif pol['ramo'] == 'Viaje':
                    cursor.execute("INSERT INTO viaje (Cod_poliza) VALUES (%s) ON CONFLICT DO NOTHING", (item['poliza'],))

                # 3. Renovación
                try:
                    fecha_e_dt = datetime.datetime.strptime(fecha_e, '%Y-%m-%d')
                    fecha_v = fecha_e_dt + relativedelta(years=1)
                except:
                    fecha_e_dt = datetime.date.today()
                    fecha_v = fecha_e_dt + relativedelta(years=1)

                frec_val = frec_map.get(str(pol['frecuencia']).upper(), 1)
                
                # 4. Pagos con lógica de múltiple renovación
                for pago in item['pagos']:
                    f_cobro_raw = pago.get('fecha_cobro')
                    
                    try:
                        if f_cobro_raw and str(f_cobro_raw).lower() != 'nan':
                            f_cobro_dt = datetime.datetime.strptime(str(f_cobro_raw).split(' ')[0], '%Y-%m-%d')
                        else:
                            f_cobro_raw = None
                            f_cobro_dt = fecha_e_dt
                    except:
                        f_cobro_dt = fecha_e_dt
                    
                    # Determinar si el pago pertenece a la renovación actual o a la anterior
                    if f_cobro_dt < (fecha_e_dt - datetime.timedelta(days=15)):
                        target_start_dt = fecha_e_dt - relativedelta(years=1)
                    else:
                        target_start_dt = fecha_e_dt
                        
                    target_start_str = target_start_dt.strftime('%Y-%m-%d')
                    
                    # Buscar o crear la renovación correspondiente
                    cursor.execute("SELECT Cod_renovacion FROM renovacion WHERE Cod_poliza = %s AND Fecha_contrato = %s", (item['poliza'], target_start_str))
                    res_ren = cursor.fetchone()
                    
                    if res_ren:
                        target_renov_id = res_ren['Cod_renovacion']
                    else:
                        target_venc_dt = target_start_dt + relativedelta(years=1)
                        sql_new_ren = """
                            INSERT INTO renovacion (Cod_poliza, Prima, Frecuencia, Fecha_contrato, Fecha_vencimiento, cobertura, comision)
                            VALUES (%s, %s, %s, %s, %s, %s, %s)
                        """
                        cursor.execute(sql_new_ren, (item['poliza'], pol['prima'], frec_val, target_start_str, target_venc_dt.strftime('%Y-%m-%d'), pol['cobertura'], 0.15))
                        target_renov_id = cursor.lastrowid

                    nro_cuota = pago.get('nro_cuota', 1)

                    # Buscar pago existente
                    cursor.execute("SELECT Cod_pago, estado FROM pago WHERE Cod_renovacion = %s AND nro_cuota = %s AND monto = %s", (target_renov_id, nro_cuota, pago['monto']))
                    pago_db = cursor.fetchone()
                    status_pago = pago.get('estado') or 'PAGADO'

                    if not pago_db:
                        sql_pago = """
                            INSERT INTO pago (Cod_renovacion, moneda, fecha, Metodo_pago, tasa, monto, fecha_pagada, estado, nro_cuota, recibo)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """
                        cursor.execute(sql_pago, (target_renov_id, pago['moneda'], f_cobro_raw, pago['metodo'] or 'PAGO DIRECTO', pago['tasa'], pago['monto'], f_cobro_raw, status_pago, nro_cuota, pago.get('recibo')))
                        procesar_comision2(cursor.lastrowid, 3)
                    elif str(pago_db['estado']).strip().upper() != str(status_pago).strip().upper() and str(status_pago).strip().upper() == 'PAGADO':
                        # Actualizar solo si el nuevo estado es PAGADO y el anterior era distinto (ej: PROCESO)
                        sql_update_pago = "UPDATE pago SET estado = %s, fecha_pagada = %s WHERE Cod_pago = %s"
                        cursor.execute(sql_update_pago, (status_pago, f_cobro_raw or datetime.date.today(), pago_db['Cod_pago']))

                connection.commit()
                return {'success': True, 'message': 'Póliza Mercantil insertada correctamente'}
    except Exception as e:
        return {'success': False, 'message': str(e)}


def procesar_imagen_perfil(foto):
    try:
        # Nombre original del archivo
        filename = secure_filename(foto.filename)
        extension = os.path.splitext(filename)[1]

        # Creando un string de 50 caracteres
        nuevoNameFile = (uuid.uuid4().hex + uuid.uuid4().hex)[:100]
        nombreFile = nuevoNameFile + extension

        # Construir la ruta completa de subida del archivo
        basepath = os.path.abspath(os.path.dirname(__file__))
        upload_dir = os.path.join(basepath, f'../static/fotos_empleados/')

        # Validar si existe la ruta y crearla si no existe
        if not os.path.exists(upload_dir):
            os.makedirs(upload_dir)
            # Dando permiso a la carpeta
            os.chmod(upload_dir, 0o755)

        # Construir la ruta completa de subida del archivo
        upload_path = os.path.join(upload_dir, nombreFile)
        foto.save(upload_path)

        return nombreFile

    except Exception as e:
        print("Error al procesar archivo:", e)
        return []


# Lista de Empleados
def sql_lista_empleadosBD():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                querySQL = (f"""
                    SELECT 
                        e.id_empleado,
                        e.nombre_empleado, 
                        e.apellido_empleado,
                        e.salario_empleado,
                        e.foto_empleado,
                        CASE
                            WHEN e.sexo_empleado = 1 THEN 'Masculino'
                            ELSE 'Femenino'
                        END AS sexo_empleado
                    FROM tbl_empleados AS e
                    ORDER BY e.id_empleado DESC
                    """)
                cursor.execute(querySQL,)
                empleadosBD = cursor.fetchall()
        return empleadosBD
    except Exception as e:
        print(
            f"Errro en la función sql_lista_empleadosBD: {e}")
        return None
    

def sql_datos_poliza(cod_poliza):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                querySQL = ("""
                    SELECT 
                        a.Nombre AS nombre_asegurado,
                        a.Apellido AS apellido_asegurado,
                        a.CI AS CI_asegurado,
                        a.Nombre2 AS segundo_nombre,
                        a.Apellido2 AS segundo_apellido,
                        c.Nombre AS nombre_compania,
                        c.Cod_compania AS cod_compania
                    FROM poliza AS p
                    JOIN asegurado AS a ON p.CI_asegurado = a.CI
                    JOIN compania AS c ON p.Cod_compania = c.Cod_compania
                    WHERE p.cod_poliza = %s 
                    """)
                cursor.execute(querySQL,(cod_poliza,))
                empleadosBD = cursor.fetchone()
        return empleadosBD
    except Exception as e:
        print(
            f"Error en la función sql_datos_poliza: {e}")
        return None
    



# Lista de Empleados
def sql_lista_aseguradosBD():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                querySQL = (f"""
                    SELECT 
                        a.CI,
                        a.Nombre,
                        a.Nombre2, 
                        a.Apellido,
                        a.Apellido2,
                        a.Correo,
                        a.Telefono,
                        a.Fecha_nacimiento,
                        a.Tipo_CI,
                        e.Nombre as Nombre_ejecutivo,
                        e.Apellido as Apellido_ejecutivo
                    FROM asegurado AS a
                    LEFT JOIN ejecutivo e ON a.Ejecutivo = e.cod_ejecutivo
                    """)
                cursor.execute(querySQL,)
                empleadosBD = cursor.fetchall()
        return empleadosBD
    except Exception as e:
        print(
            f"Errro en la función sql_lista_aseguradosBD: {e}")
        return None
    

    
def sql_lista_polizas():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                querySQL = (f"""
                    SELECT
                        p.cod_poliza,
                        p.CI_asegurado,
                        DATE_FORMAT(r.Fecha_contrato, '%Y-%m-%d') AS fecha,
                        p.Ramo,
                        a.Nombre,
                        a.Apellido,
                        a.Tipo_CI,
                        r.riesgo,
                        r.Fecha_vencimiento,
                        r.estado,
                        c.nombre as compania_nombre,
                        e.Nombre as ejecutivo_nombre,
                        e.Apellido as ejecutivo_apellido,
                        CASE
                            WHEN p.Ramo IN ('Persona', 'PERSONA', 'PERSONAS') THEN pe.Producto
                            WHEN p.Ramo IN ('Auto', 'AUTO') THEN au.Producto
                            WHEN p.Ramo IN ('Patrimonial', 'PATRIMONIAL') THEN pa.Producto
                            WHEN p.Ramo IN ('Fianza', 'FIANZA', 'FIANZAS') THEN f.Producto
                            WHEN p.Ramo IN ('Viaje', 'VIAJE', 'VIAJES') THEN v.Producto
                            ELSE ''
                        END as producto_nombre,
                        CASE
                            WHEN COUNT(pay.Cod_pago) > 0 THEN 'Sí'
                            ELSE 'No'
                        END AS Tiene_pago
                    FROM
                        poliza p
                    JOIN (
                        SELECT
                            r2.cod_poliza,
                            r2.Cod_renovacion,
                            r2.Fecha_contrato,
                            r2.Fecha_vencimiento,
                            r2.riesgo,
                            r2.estado
                        FROM renovacion r2
                        INNER JOIN (
                            SELECT cod_poliza, MAX(Fecha_contrato) AS max_fecha
                            FROM renovacion
                            GROUP BY cod_poliza
                        ) AS latest ON r2.cod_poliza = latest.cod_poliza
                            AND r2.Fecha_contrato = latest.max_fecha
                    ) AS r ON p.cod_poliza = r.cod_poliza
                    JOIN
                        asegurado a ON p.CI_asegurado = a.CI
                    LEFT JOIN
                        ejecutivo e ON a.Ejecutivo = e.cod_ejecutivo
                    LEFT JOIN
                        compania c ON p.Cod_compania = c.Cod_compania
                    LEFT JOIN
                        Persona pe ON p.cod_poliza = pe.Cod_poliza AND p.Ramo IN ('Persona', 'PERSONA', 'PERSONAS')
                    LEFT JOIN
                        Auto au ON p.cod_poliza = au.Cod_poliza AND p.Ramo IN ('Auto', 'AUTO')
                    LEFT JOIN
                        Patrimonio pa ON p.cod_poliza = pa.Cod_poliza AND p.Ramo IN ('Patrimonial', 'PATRIMONIAL')
                    LEFT JOIN
                        Fianza f ON p.cod_poliza = f.Cod_poliza AND p.Ramo IN ('Fianza', 'FIANZA', 'FIANZAS')
                    LEFT JOIN
                        Viaje v ON p.cod_poliza = v.Cod_poliza AND p.Ramo IN ('Viaje', 'VIAJE', 'VIAJES')
                    LEFT JOIN
                        pago pay ON r.Cod_renovacion = pay.Cod_renovacion
                    GROUP BY
                        p.cod_poliza,
                        p.CI_asegurado,
                        r.Fecha_contrato,
                        p.Ramo,
                        r.Fecha_vencimiento,
                        r.riesgo,
                        r.estado,
                        a.Nombre,
                        a.Apellido,
                        a.Tipo_CI,
                        c.nombre,
                        producto_nombre,
                        e.Nombre,
                        e.Apellido
                    ORDER BY
                        p.cod_poliza;""")
                cursor.execute(querySQL,)
                empleadosBD = cursor.fetchall()
        return empleadosBD if empleadosBD is not None else []
    except Exception as e:
        print(f"Error en la funciÃ³n sql_lista_polizas: {e}")
        return []

# Lista de siniestros
def sql_lista_siniestros():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                querySQL = (f"""
                        SELECT
                            p.Cod_poliza,
                            'Carta Aval' AS tipo_siniestro,
                            DATE_FORMAT(ca.Fecha_noti, '%d-%m-%Y') AS fecha_inicio,
                            ca.Estado AS estado_siniestro,
                            ca.Cod_CartaAval AS codigo, -- Esta columna se mapea a 'codigo'
                            a.Nombre AS nombre_asegurado,
                            a.Apellido AS apellido_asegurado,
                            a.CI AS CI_asegurado,
                            a.Nombre2 AS segundo_nombre,
                            a.Apellido2 AS segundo_apellido,
                            c.Nombre AS nombre_compania,
                            ca.codigo_siniestro,
                            ca.Monto_solicitado AS monto  
                        FROM
                            Carta_aval ca
                        INNER JOIN
                            poliza p ON ca.Cod_poliza = p.Cod_poliza
                        INNER JOIN
                            asegurado AS a ON p.CI_asegurado = a.CI
                        INNER JOIN
                            compania AS c ON p.Cod_compania = c.Cod_compania

                        UNION ALL

                        SELECT
                            p.Cod_poliza,
                            'Reembolso' AS tipo_siniestro,
                            DATE_FORMAT(r.Fecha_ocurrencia, '%d-%m-%Y') AS fecha_inicio,
                            r.estado AS estado_siniestro,
                            r.cod_reembolso AS codigo, -- Asegúrate de que este 'codigo' esté presente en todas las uniones
                            a.Nombre AS nombre_asegurado,
                            a.Apellido AS apellido_asegurado,
                            a.CI AS CI_asegurado,
                            a.Nombre2 AS segundo_nombre,
                            a.Apellido2 AS segundo_apellido,
                            c.Nombre AS nombre_compania,
                            r.codigo_siniestro,
                            r.Monto_solicitado AS monto   
                        FROM
                            Reembolso r
                        INNER JOIN
                            poliza p ON r.Cod_poliza = p.Cod_poliza
                        INNER JOIN
                            asegurado AS a ON p.CI_asegurado = a.CI
                        INNER JOIN
                            compania AS c ON p.Cod_compania = c.Cod_compania

                        UNION ALL

                        SELECT
                            p.Cod_poliza,
                            'Siniestro Auto' AS tipo_siniestro,
                            DATE_FORMAT(sa.Fecha_ocurrencia, '%d-%m-%Y') AS fecha_inicio,
                            sa.estado AS estado_siniestro,
                            sa.Cod_siniestroA AS codigo, -- Asegúrate de que este 'codigo' esté presente en todas las uniones
                            a.Nombre AS nombre_asegurado,
                            a.Apellido AS apellido_asegurado,
                            a.CI AS CI_asegurado,
                            a.Nombre2 AS segundo_nombre,
                            a.Apellido2 AS segundo_apellido,
                            c.Nombre AS nombre_compania,
                            sa.codigo_siniestro,
                            sa.Monto_orden AS monto 
                        FROM
                            AutomovilSiniestro sa
                        INNER JOIN
                            poliza p ON sa.Cod_poliza = p.Cod_poliza
                        INNER JOIN
                            asegurado AS a ON p.CI_asegurado = a.CI
                        INNER JOIN
                            compania AS c ON p.Cod_compania = c.Cod_compania;""")
                cursor.execute(querySQL,)
                empleadosBD = cursor.fetchall()
        return empleadosBD
    except Exception as e:
        print(
            f"Errro en la función sql_lista_siniestro: {e}")
        return None
    
def sql_lista_siniestros_unico(cod_poliza):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                querySQL = ("""
                        SELECT
                            p.Cod_poliza,
                            'Carta Aval' AS tipo_siniestro,
                            DATE_FORMAT(ca.Fecha_noti, '%d-%m-%Y') AS fecha_inicio,
                            ca.Estado AS estado_siniestro,
                            ca.Cod_CartaAval as codigo
                        FROM
                            Carta_aval ca
                        INNER JOIN
                            poliza p ON ca.Cod_poliza = p.Cod_poliza
                        WHERE ca.Cod_poliza=%s
                        UNION ALL
                        SELECT
                            p.Cod_poliza,
                            'Reembolso' AS tipo_siniestro,
                            DATE_FORMAT(r.Fecha_ocurrencia, '%d-%m-%Y') AS fecha_inicio,
                            r.estado AS estado_siniestro,
                            r.cod_reembolso as codigo
                        FROM
                            Reembolso r
                        INNER JOIN
                            poliza p ON r.Cod_poliza = p.Cod_poliza
                        WHERE r.Cod_poliza=%s
                        UNION ALL
                        SELECT
                            p.Cod_poliza,
                            'Siniestro Auto' AS tipo_siniestro,
                            DATE_FORMAT(a.Fecha_ocurrencia, '%d-%m-%Y') AS fecha_inicio,  
                            a.estado AS estado_siniestro,
                            a.Cod_siniestroA as codigo
                        FROM
                            AutomovilSiniestro a
                        INNER JOIN
                            poliza p ON a.Cod_poliza = p.Cod_poliza
                       WHERE a.Cod_poliza=%s
                    """)
                cursor.execute(querySQL,(cod_poliza,cod_poliza,cod_poliza))
                siniestrosBD = cursor.fetchall()
        return siniestrosBD
    except Exception as e:
        print(
            f"Errro en la función sql_lista_siniestro_unico: {e}")
        return None

# Detalles del Empleado
def sql_detalles_empleadosBD(idEmpleado):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                querySQL = ("""
                    SELECT 
                        e.id_empleado,
                        e.nombre_empleado, 
                        e.apellido_empleado,
                        e.salario_empleado,
                        CASE
                            WHEN e.sexo_empleado = 1 THEN 'Masculino'
                            ELSE 'Femenino'
                        END AS sexo_empleado,
                        e.telefono_empleado, 
                        e.email_empleado,
                        e.profesion_empleado,
                        e.foto_empleado,
                        DATE_FORMAT(e.fecha_registro, '%Y-%m-%d %h:%i %p') AS fecha_registro
                    FROM tbl_empleados AS e
                    WHERE id_empleado =%s
                    ORDER BY e.id_empleado DESC
                    """)
                cursor.execute(querySQL, (idEmpleado,))
                empleadosBD = cursor.fetchone()
        return empleadosBD
    except Exception as e:
        print(
            f"Error en la función sql_detalles_empleadosBD: {e}")
        return None
    
def sql_detalles_aseguradoBD(CI):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                querySQL = ("""
                    SELECT 
                        a.CI,
                        a.Nombre,
                        a.Nombre2, 
                        a.Apellido,
                        a.Apellido2,
                        a.Correo,
                        a.Telefono,
                        a.Fecha_nacimiento,
                        a.profesion,
                        a.localidad,
                        a.canal,
                        e.Nombre as Nombre_ejecutivo,
                        e.Apellido as Apellido_ejecutivo
                    FROM asegurado AS a
                    LEFT JOIN ejecutivo e ON a.Ejecutivo = e.cod_ejecutivo
                    where a.CI=%s
                    """)
                cursor.execute(querySQL, (CI,))
                empleadosBD = cursor.fetchone()
        return empleadosBD
    except Exception as e:
        print(
            f"Error en la función sql_detalles_aseguradoBD: {e}")
        return None

def sql_lista_polizas_asegurado(CI):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                querySQL = ("""
                    SELECT p.cod_poliza as Cod_poliza, p.Ramo, DATE_FORMAT(p.Fecha_emision, '%d-%m-%Y') as Fecha_emision, c.Nombre as nombre_compania
                    FROM poliza p
                    LEFT JOIN compania c ON p.Cod_compania = c.Cod_compania
                    WHERE p.CI_asegurado = %s
                """)
                cursor.execute(querySQL, (CI,))
                return cursor.fetchall()
    except Exception as e:
        print(f"Error en sql_lista_polizas_asegurado: {e}")
        return []
    
def sql_detalles_reembolsoBD(cod_reembolso):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                querySQL = ("""
                    SELECT *
                    FROM Reembolso
                    WHERE cod_reembolso = %s
                    """)
                cursor.execute(querySQL, (cod_reembolso,))
                empleadosBD = cursor.fetchone()
        return empleadosBD
    except Exception as e:
        print(
            f"Error en la función sql_detalles_empleadosBD: {e}")
        return None
    
def sql_detalles_CartaAvalBD(cod_carta):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                querySQL = ("""
                    SELECT *
                    FROM Carta_aval
                    WHERE Cod_CartaAval = %s
                    """)
                cursor.execute(querySQL, (cod_carta,))
                empleadosBD = cursor.fetchone()
        return empleadosBD
    except Exception as e:
        print(
            f"Error en la función sql_detalles_CartaAvalBD: {e}")
        return None
    
def sql_Notas_cartaAval(cod_carta):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                querySQL = ("""
                    SELECT *
                    FROM nota_cartaAval
                    WHERE Cod_CartaAval = %s
                    """)
                cursor.execute(querySQL, (cod_carta,))
                notas = cursor.fetchall()
        return notas
    except Exception as e:
        print(
            f"Error en la función sql_detalles_CartaAvalnotasBD: {e}")
        return None
    
def sql_Notas_reembolso(cod_reembolso):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                querySQL = ("""
                    SELECT *
                    FROM nota_Reembolso
                    WHERE Cod_Reembolso = %s
                    """)
                cursor.execute(querySQL, (cod_reembolso,))
                notas = cursor.fetchall()
        return notas
    except Exception as e:
        print(
            f"Error en la función sql_detalles_ReembolsonotasBD: {e}")
        return None
    
def sql_Notas_auto(cod_auto):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                querySQL = ("""
                    SELECT *
                    FROM nota_Auto
                    WHERE Cod_Auto = %s
                    """)
                cursor.execute(querySQL, (cod_auto,))
                notas = cursor.fetchall()
        return notas
    except Exception as e:
        print(
            f"Error en la función sql_detalles_autoNotasBD: {e}")
        return None

def sql_detalles_SiniestroA(cod_auto):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                querySQL = ("""
                    SELECT *
                    FROM AutomovilSiniestro
                    WHERE Cod_siniestroA = %s
                    """)
                cursor.execute(querySQL, (cod_auto,))
                empleadosBD = cursor.fetchone()
        return empleadosBD
    except Exception as e:
        print(
            f"Error en la función sql_detalles_siniestroABD: {e}")
        return None
        
def sql_detalles_polizaBD(cod_poliza, cod_renovacion=None):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                sql= 'SELECT Ramo from poliza WHERE cod_poliza=%s'
                cursor.execute(sql,(cod_poliza,)) 
                R = cursor.fetchone()
                print(R)
                if not R:
                    return None
                Ramo = R['Ramo']
                print(Ramo)
                if Ramo == "Auto" or Ramo=="AUTO":
                    querySQL = ("""
                        SELECT 
                                p.cod_poliza,
                                p.CI_asegurado,
                                DATE_FORMAT(r.Fecha_contrato, '%Y-%m-%d') AS fecha_registro,
                                p.Cod_compania,
                                p.Tomador, 
                                p.Ramo,
                                e.Nombre,
                                e.Apellido,
                                a.modelo,
                                a.Producto,
                                a.placa,
                                a.año,
                                a.marca,
                                r.Fecha_vencimiento,
                                a.Subramo,
                                r.riesgo,
                                r.Cod_renovacion,
                                r.Prima,
                                r.cobertura,
                                r.Frecuencia,
                                ase.Nombre as NombreA,
                                ase.Nombre2,
                                ase.Apellido as ApellidoA,
                                ase.Apellido2,
                                c.Nombre as NombreC
                                
                            FROM
                                poliza p
                            JOIN
                                asegurado as ase ON p.CI_asegurado = ase.CI
                            JOIN
                                ejecutivo e ON ase.Ejecutivo = e.cod_ejecutivo
                            INNER JOIN Auto a ON p.cod_poliza = a.cod_poliza
                            INNER JOIN renovacion r on p.cod_poliza = r.cod_poliza
                            INNER JOIN compania c on p.Cod_compania = c.Cod_compania
                            WHERE p.cod_poliza = %s
                        """)
                elif Ramo == "Persona" or Ramo == "PERSONAS":
                    querySQL = ("""
                        SELECT 
                                p.cod_poliza,
                                p.CI_asegurado,
                                DATE_FORMAT(r.Fecha_contrato, '%Y-%m-%d') AS fecha_registro,
                                p.Cod_compania,
                                p.Tomador,
                                p.Ramo,
                                r.cobertura,
                                r.frecuencia,
                                e.Nombre as Nombre1,
                                e.Apellido as Apellido1,
                                pe.Subramo,
                                pe.Producto,
                                b.Nombre,
                                b.Apellido,
                                b.Cedula,
                                b.Parentesco,
                                r.Fecha_vencimiento,
                                r.Cod_renovacion,
                                r.Prima,
                                r.riesgo,
                                r.Frecuencia,
                                a.Nombre as NombreA,
                                a.Nombre2,
                                a.Apellido as ApellidoA,
                                a.Apellido2,
                                c.Nombre as NombreC
                                
                            FROM
                                poliza p
                            JOIN
                                asegurado a ON p.CI_asegurado = a.CI
                            JOIN
                                ejecutivo e ON a.Ejecutivo = e.cod_ejecutivo
                            INNER JOIN Persona pe ON p.cod_poliza = pe.cod_poliza
                            INNER JOIN renovacion r on p.cod_poliza = r.cod_poliza
                            INNER JOIN compania c on p.Cod_compania = c.Cod_compania
                            LEFT JOIN beneficiario b on p.cod_poliza = b.cod_poliza
                            WHERE p.cod_poliza = %s
                        """)
                elif Ramo == "Patrimonial" or Ramo == "PATRIMONIAL":
                    querySQL = ("""
                        SELECT 
                                p.cod_poliza,
                                p.CI_asegurado,
                                DATE_FORMAT(r.Fecha_contrato, '%Y-%m-%d') AS fecha_registro,
                                p.Cod_compania,
                                p.Tomador,
                                p.Ramo,
                                e.Nombre,
                                e.Apellido,
                                pa.direccion,
                                pa.Producto,
                                r.Fecha_vencimiento,
                                pa.Subramo,
                                r.riesgo,
                                r.Cod_renovacion,
                                r.Frecuencia,
                                a.Nombre as NombreA,
                                a.Nombre2,
                                a.Apellido as ApellidoA,
                                a.Apellido2,
                                c.Nombre as NombreC
                            FROM
                                poliza p
                            JOIN
                                asegurado a ON p.CI_asegurado = a.CI
                            JOIN
                                ejecutivo e ON a.Ejecutivo = e.cod_ejecutivo
                            INNER JOIN patrimonio pa ON p.cod_poliza = pa.cod_poliza
                            INNER JOIN renovacion r on p.cod_poliza = r.cod_poliza
                            INNER JOIN compania c on p.Cod_compania = c.Cod_compania
                            WHERE p.cod_poliza = %s 
                        """)
                elif Ramo == "Fianza" or Ramo =="FIANZAS":
                    querySQL = ("""
                        SELECT 
                                p.cod_poliza,
                                p.CI_asegurado,
                                DATE_FORMAT(r.Fecha_contrato, '%Y-%m-%d') AS fecha_registro,
                                p.Cod_compania,
                                p.Tomador,
                                p.Ramo,
                                e.Nombre,
                                e.Apellido,
                                f.Producto,
                                r.Fecha_vencimiento,
                                f.Subramo,
                                r.riesgo,
                                r.Cod_renovacion,
                                r.Prima,
                                r.cobertura,
                                r.Frecuencia,
                                a.Nombre as NombreA,
                                a.Nombre2,
                                a.Apellido as ApellidoA,
                                a.Apellido2,
                                c.Nombre as NombreC
                            FROM
                                poliza p
                            JOIN
                                asegurado a ON p.CI_asegurado = a.CI
                            JOIN
                                ejecutivo e ON a.Ejecutivo = e.cod_ejecutivo
                            INNER JOIN Fianza f ON p.cod_poliza = f.cod_poliza
                            INNER JOIN renovacion r on p.cod_poliza = r.cod_poliza
                            INNER JOIN compania c on p.Cod_compania = c.Cod_compania
                            WHERE p.cod_poliza = %s 
                        """)
                elif Ramo == "Viaje" or Ramo=="VIAJES":
                    querySQL = ("""
                        SELECT 
                                p.cod_poliza,
                                p.CI_asegurado,
                                DATE_FORMAT(r.Fecha_contrato, '%Y-%m-%d') AS fecha_registro,
                                p.Cod_compania,
                                p.Tomador,
                                p.Ramo,
                                e.Nombre,
                                e.Apellido,
                                v.cod_pasaporte,
                                v.Producto,
                                r.Fecha_vencimiento,
                                v.Subramo,
                                r.riesgo,
                                r.Cod_renovacion,
                                r.Frecuencia,
                                a.Nombre as NombreA,
                                a.Nombre2,
                                a.Apellido as ApellidoA,
                                a.Apellido2,
                                c.Nombre as NombreC
                            FROM
                                poliza p
                            JOIN
                                asegurado a ON p.CI_asegurado = a.CI
                            JOIN
                                ejecutivo e ON a.Ejecutivo = e.cod_ejecutivo
                            INNER JOIN Viaje v ON p.cod_poliza = v.cod_poliza
                            INNER JOIN renovacion r on p.cod_poliza = r.cod_poliza
                            INNER JOIN compania c on p.Cod_compania = c.Cod_compania
                            WHERE p.cod_poliza = %s 
                        """)
                
                params = [cod_poliza]
                if cod_renovacion:
                    querySQL += " AND r.Cod_renovacion = %s"
                    params.append(cod_renovacion)
                else:
                    # Si no se proporciona cod_renovacion, obtenemos el más reciente
                    querySQL += " ORDER BY r.Fecha_contrato DESC LIMIT 1"

                cursor.execute(querySQL, tuple(params))
                PolizaBD = cursor.fetchone()

        return PolizaBD
    except Exception as e:
        print(
            f"Error en la función sql_detalles_polizaBD: {e}")
        return None


def obtener_siniestros_filtrados(tipo_cedula, cedula, estado_siniestro, meses, anio, tipo_siniestro=None):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as mycursor:
                
                base_query_parts = {
                    'carta_aval': """
                        SELECT
                            p.Cod_poliza, 'Carta Aval' AS tipo_siniestro,
                            DATE_FORMAT(ca.Fecha_noti, '%d-%m-%Y') AS fecha_inicio,
                            ca.Estado AS estado_siniestro, ca.Cod_CartaAval AS codigo,
                            a.Nombre AS nombre_asegurado, a.Apellido AS apellido_asegurado,
                            a.CI AS CI_asegurado, a.Nombre2 AS segundo_nombre,
                            a.Apellido2 AS segundo_apellido, c.Nombre AS nombre_compania,
                            ca.codigo_siniestro, ca.Monto_solicitado AS monto
                        FROM Carta_aval ca
                        INNER JOIN poliza p ON ca.Cod_poliza = p.Cod_poliza
                        INNER JOIN asegurado AS a ON p.CI_asegurado = a.CI
                        INNER JOIN compania AS c ON p.Cod_compania = c.Cod_compania
                    """,
                    'reembolso': """
                        SELECT
                            p.Cod_poliza, 'Reembolso' AS tipo_siniestro,
                            DATE_FORMAT(r.Fecha_ocurrencia, '%d-%m-%Y') AS fecha_inicio,
                            r.estado AS estado_siniestro, r.cod_reembolso AS codigo,
                            a.Nombre AS nombre_asegurado, a.Apellido AS apellido_asegurado,
                            a.CI AS CI_asegurado, a.Nombre2 AS segundo_nombre,
                            a.Apellido2 AS segundo_apellido, c.Nombre AS nombre_compania,
                            r.codigo_siniestro, r.Monto_solicitado AS monto
                        FROM Reembolso r
                        INNER JOIN poliza p ON r.Cod_poliza = p.Cod_poliza
                        INNER JOIN asegurado AS a ON p.CI_asegurado = a.CI
                        INNER JOIN compania AS c ON p.Cod_compania = c.Cod_compania
                    """,
                    'auto': """
                        SELECT
                            p.Cod_poliza, 'Siniestro Auto' AS tipo_siniestro,
                            DATE_FORMAT(sa.Fecha_ocurrencia, '%d-%m-%Y') AS fecha_inicio,
                            sa.estado AS estado_siniestro, sa.Cod_siniestroA AS codigo,
                            a.Nombre AS nombre_asegurado, a.Apellido AS apellido_asegurado,
                            a.CI AS CI_asegurado, a.Nombre2 AS segundo_nombre,
                            a.Apellido2 AS segundo_apellido, c.Nombre AS nombre_compania,
                            sa.codigo_siniestro, sa.Monto_orden AS monto
                        FROM AutomovilSiniestro sa
                        INNER JOIN poliza p ON sa.Cod_poliza = p.Cod_poliza
                        INNER JOIN asegurado AS a ON p.CI_asegurado = a.CI
                        INNER JOIN compania AS c ON p.Cod_compania = c.Cod_compania
                    """
                }

                params = []
                
                # Build WHERE clauses and params for each part of the UNION
                final_query_parts = []
                items_to_process = [
                    ('carta_aval', 'ca.Fecha_noti', 'ca.Estado'), 
                    ('reembolso', 'r.Fecha_ocurrencia', 'r.estado'), 
                    ('auto', 'sa.Fecha_ocurrencia', 'sa.estado')
                ]

                # Filter items_to_process based on tipo_siniestro if provided
                if tipo_siniestro:
                    if tipo_siniestro.lower() == 'carta aval':
                        items_to_process = [('carta_aval', 'ca.Fecha_noti', 'ca.Estado')]
                    elif tipo_siniestro.lower() == 'reembolso':
                        items_to_process = [('reembolso', 'r.Fecha_ocurrencia', 'r.estado')]
                    elif tipo_siniestro.lower() == 'siniestro auto':
                        items_to_process = [('auto', 'sa.Fecha_ocurrencia', 'sa.estado')]

                for key, date_col, status_col in items_to_process:
                    
                    query_part = base_query_parts[key]
                    where_conditions = []
                    part_params = []

                    if cedula and tipo_cedula:
                        where_conditions.append("a.Tipo_CI = %s AND a.CI = %s")
                        part_params.extend([tipo_cedula, cedula])
                    
                    if estado_siniestro:
                        where_conditions.append(f"{status_col} = %s")
                        part_params.append(estado_siniestro)

                    if anio and meses:
                        placeholders = ', '.join(['%s'] * len(meses))
                        where_conditions.append(f"EXTRACT(YEAR FROM {date_col}) = %s AND EXTRACT(MONTH FROM {date_col}) IN ({placeholders})")
                        print("prime")
                        part_params.append(anio)
                        part_params.extend(meses)
                    elif anio:
                        print("rajois")
                        where_conditions.append(f"EXTRACT(YEAR FROM {date_col}) = %s")
                        part_params.append(anio)
                    elif meses:
                        print("alo")
                        placeholders = ', '.join(['%s'] * len(meses))
                        where_conditions.append(f"EXTRACT(MONTH FROM {date_col}) IN ({placeholders})")
                        part_params.extend(meses)

                    if where_conditions:
                        query_part += " WHERE " + " AND ".join(where_conditions)
                    
                    final_query_parts.append(f"({query_part})")
                    params.extend(part_params)

                final_query = " UNION ALL ".join(final_query_parts)
                
                mycursor.execute(final_query, tuple(params))
                resultado_busqueda = mycursor.fetchall()
                return resultado_busqueda
    except Exception as e:
        print(f"Ocurrió un error en def obtener_siniestros_filtrados: {e}")
        return []


def obtener_registros_por_mes_poliza(ano, mes):
    """Legacy function, now uses obtener_polizas_filtradas"""
    return obtener_polizas_filtradas(ano=ano, mes=mes)

def obtener_polizas_filtradas(ano=None, mes=None, rango_inicio=None, rango_fin=None, estados=None, compania_id=None, ejecutivo_id=None):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as mycursor:
                querySQL = """
                    SELECT
                        p.cod_poliza,
                        p.CI_asegurado,
                        a.Tipo_CI,
                        a.Nombre,
                        a.Apellido,
                        a.Nombre2,
                        a.Apellido2,
                        r.riesgo,
                        r.estado,
                        DATE_FORMAT(r.Fecha_contrato, '%Y-%m-%d') as fecha, 
                        p.Ramo,
                        r.Fecha_vencimiento,
                        c.nombre as compania_nombre,
                        e.Nombre as ejecutivo_nombre,
                        e.Apellido as ejecutivo_apellido,
                        CASE
                            WHEN p.Ramo IN ('Persona', 'PERSONA', 'PERSONAS') THEN pe.Producto
                            WHEN p.Ramo IN ('Auto', 'AUTO') THEN au.Producto
                            WHEN p.Ramo IN ('Patrimonial', 'PATRIMONIAL') THEN pa.Producto
                            WHEN p.Ramo IN ('Fianza', 'FIANZA', 'FIANZAS') THEN f.Producto
                            WHEN p.Ramo IN ('Viaje', 'VIAJE', 'VIAJES') THEN v.Producto
                            ELSE ''
                        END as producto_nombre,
                        CASE
                            WHEN COUNT(pay.Cod_pago) > 0 THEN 'Sí'
                            ELSE 'No'
                        END AS Tiene_pago
                    FROM
                        poliza p
                    JOIN (
                        SELECT
                            r2.cod_poliza,
                            r2.Cod_renovacion,
                            r2.Fecha_contrato,
                            r2.Fecha_vencimiento,
                            r2.riesgo,
                            r2.estado
                        FROM renovacion r2
                        INNER JOIN (
                            SELECT cod_poliza, MAX(Cod_renovacion) AS max_cod
                            FROM renovacion
                            GROUP BY cod_poliza
                        ) AS latest ON r2.cod_poliza = latest.cod_poliza
                            AND r2.Cod_renovacion = latest.max_cod
                    ) AS r ON p.cod_poliza = r.cod_poliza
                    JOIN
                        asegurado a ON p.CI_asegurado = a.CI
                    LEFT JOIN
                        ejecutivo e ON a.Ejecutivo = e.cod_ejecutivo
                    LEFT JOIN
                        compania c ON p.Cod_compania = c.Cod_compania
                    LEFT JOIN
                        Persona pe ON p.cod_poliza = pe.Cod_poliza AND p.Ramo IN ('Persona', 'PERSONA', 'PERSONAS')
                    LEFT JOIN
                        Auto au ON p.cod_poliza = au.Cod_poliza AND p.Ramo IN ('Auto', 'AUTO')
                    LEFT JOIN
                        Patrimonio pa ON p.cod_poliza = pa.Cod_poliza AND p.Ramo IN ('Patrimonial', 'PATRIMONIAL')
                    LEFT JOIN
                        Fianza f ON p.cod_poliza = f.Cod_poliza AND p.Ramo IN ('Fianza', 'FIANZA', 'FIANZAS')
                    LEFT JOIN
                        Viaje v ON p.cod_poliza = v.Cod_poliza AND p.Ramo IN ('Viaje', 'VIAJE', 'VIAJES')
                    LEFT JOIN
                        pago pay ON r.Cod_renovacion = pay.Cod_renovacion
                    WHERE 1=1
                """
                
                params = []
                if ano and mes:
                    querySQL += " AND EXTRACT(YEAR FROM r.Fecha_contrato) = %s AND EXTRACT(MONTH FROM r.Fecha_contrato) = %s"
                    params.extend([ano, mes])
                elif ano:
                    querySQL += " AND EXTRACT(YEAR FROM r.Fecha_contrato) = %s"
                    params.append(ano)
                elif rango_inicio and rango_fin:
                    querySQL += " AND r.Fecha_contrato BETWEEN %s AND %s"
                    params.extend([rango_inicio, rango_fin])
                elif rango_inicio:
                    querySQL += " AND r.Fecha_contrato >= %s"
                    params.append(rango_inicio)
                elif rango_fin:
                    querySQL += " AND r.Fecha_contrato <= %s"
                    params.append(rango_fin)
                
                if compania_id and compania_id != "":
                    querySQL += " AND p.Cod_compania = %s"
                    params.append(compania_id)
                
                if ejecutivo_id and ejecutivo_id != "":
                    querySQL += " AND a.Ejecutivo = %s"
                    params.append(ejecutivo_id)

                querySQL += """ GROUP BY
                        p.cod_poliza, p.CI_asegurado, a.Tipo_CI, r.Fecha_contrato, p.Ramo, r.Fecha_vencimiento, r.riesgo, r.estado, c.nombre, producto_nombre, e.Nombre, e.Apellido
                    ORDER BY
                        p.cod_poliza;
                """
                mycursor.execute(querySQL, tuple(params))
                resultados = mycursor.fetchall()
                
                # Filtrado por estado en Python
                today = datetime.date.today()
                
                # Calcular el estado para todos los resultados primero
                for res in resultados:
                    actual_estado = ""
                    ramo_upper = (res['Ramo'] or "").upper()
                    is_special_ramo = (ramo_upper in ['FIANZAS', 'FIANZA', 'VIAJES', 'VIAJE'])
                    
                    vencimiento = res['Fecha_vencimiento']
                    tiene_pago = res['Tiene_pago']
                    res_estado = (res['estado'] or "").lower()
                    
                    if res_estado == 'anulada':
                        actual_estado = "Anulada"
                    elif res_estado == 'traspasada':
                        actual_estado = "Traspasada"
                    elif vencimiento <= today and is_special_ramo:
                        actual_estado = "Finalizada"
                    elif vencimiento > today and tiene_pago == "Sí":
                        actual_estado = "Vigente"
                    elif vencimiento > today and tiene_pago == "No":
                        actual_estado = "Pendiente"
                    else:
                        actual_estado = "Vencida"
                    
                    res['estado_calculado'] = actual_estado
                    # Add compatibility fields for frontend
                    res['vencida'] = "no" if actual_estado == "Vigente" else "pendiente" if actual_estado == "Pendiente" else "si"

                if estados and len(estados) > 0 and "" not in estados:
                    resultados = [res for res in resultados if res['estado_calculado'] in estados]
                    
                return resultados

    except Exception as e:
        print(f"Error en obtener_polizas_filtradas: {e}")
        return []


def obtener_comisiones_filtradas(start, end, ejecutivo):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as mycursor:
                if not start and not ejecutivo:
                    querySQL = ("""
                        SELECT
                            p.Ramo,
                            p.cod_poliza,
                            pag.Cod_pago,
                            e.Nombre,
                            e.Apellido,
                            pag.monto,
                            a.Nombre AS N_asegu,
                            a.Apellido AS A_Asegu,
                            cr.comision,
                            cr.Prima,
                            pag.fecha,
                            pag.moneda,
                            pag.tasa,
                            coe.comision_bono as bono
                        FROM
                            pago pag
                        JOIN
                            renovacion cr ON pag.Cod_renovacion = cr.cod_renovacion
                        JOIN
                            poliza p ON cr.cod_poliza = p.cod_poliza
                        JOIN
                            asegurado a ON p.CI_asegurado = a.CI
                        JOIN
                            ejecutivo e ON a.Ejecutivo = e.cod_ejecutivo
                        JOIN
                            comision co ON e.cod_ejecutivo = co.cod_ejecutivo AND pag.Cod_pago = co.Cod_pago
                        LEFT JOIN 
                            comisiones_ejecutivos coe ON e.cod_ejecutivo = coe.cod_ejecutivo AND p.Ramo = 'Auto'
                        """)
                    mycursor.execute(querySQL,)
                elif not ejecutivo and start:
                    querySQL = ("""
                        SELECT
                            p.Ramo,
                            p.cod_poliza,
                            pag.Cod_pago,
                            e.Nombre,
                            e.Apellido,
                            pag.monto,
                            a.Nombre as N_asegu,
                            a.Apellido as A_Asegu,
                            cr.comision,
                            cr.Prima,
                            pag.fecha,
                            pag.moneda,
                            pag.tasa,
                            coe.comision_bono as bono
                        FROM
                            pago pag
                        JOIN
                            renovacion cr ON pag.Cod_renovacion = cr.cod_renovacion
                        JOIN
                            poliza p ON cr.cod_poliza = p.cod_poliza
                        JOIN
                            asegurado a ON p.CI_asegurado = a.CI
                        JOIN
                            ejecutivo e ON a.Ejecutivo = e.cod_ejecutivo
                        JOIN 
                            comision co ON e.cod_ejecutivo = co.cod_ejecutivo AND pag.Cod_pago = co.Cod_pago
                        LEFT JOIN -- Changed to LEFT JOIN
                            comisiones_ejecutivos coe ON e.cod_ejecutivo = coe.cod_ejecutivo AND p.Ramo = 'Auto'
                        WHERE
                            pag.fecha BETWEEN %s AND %s""")
                    mycursor.execute(querySQL, ( start, end))                 
                elif not start and ejecutivo:
                    querySQL = ("""
                        SELECT
                            p.Ramo,
                            p.cod_poliza,
                            pag.Cod_pago,
                            e.Nombre,
                            e.Apellido,
                            pag.monto,
                            a.Nombre as N_asegu,
                            a.Apellido as A_Asegu,
                            cr.comision,
                            cr.Prima,
                            pag.fecha,
                            pag.moneda,
                            pag.tasa,
                            coe.comision_bono as bono
                        FROM
                            pago pag
                        JOIN
                            renovacion cr ON pag.Cod_renovacion = cr.cod_renovacion
                        JOIN
                            poliza p ON cr.cod_poliza = p.cod_poliza
                        JOIN
                            asegurado a ON p.CI_asegurado = a.CI
                        JOIN
                            ejecutivo e ON a.Ejecutivo = e.cod_ejecutivo
                        JOIN
                            comision co ON e.cod_ejecutivo = co.cod_ejecutivo AND pag.Cod_pago = co.Cod_pago
                        LEFT JOIN -- Changed to LEFT JOIN
                            comisiones_ejecutivos coe ON e.cod_ejecutivo = coe.cod_ejecutivo AND p.Ramo = 'Auto'
                        WHERE
                            e.cod_ejecutivo = %s""")
                    mycursor.execute(querySQL, ( ejecutivo,))
                else:
                    querySQL = ("""
                        SELECT
                            p.Ramo,
                            p.cod_poliza,
                            pag.Cod_pago,
                            e.Nombre,
                            e.Apellido,
                            pag.monto,
                            a.Nombre as N_asegu,
                            a.Apellido as A_Asegu,
                            cr.comision,
                            cr.Prima,
                            pag.fecha,
                            pag.moneda,
                            pag.tasa,
                            coe.comision_bono as bono
                        FROM
                            pago pag
                        JOIN
                            renovacion cr ON pag.Cod_renovacion = cr.cod_renovacion
                        JOIN
                            poliza p ON cr.cod_poliza = p.cod_poliza
                        JOIN
                            asegurado a ON p.CI_asegurado = a.CI
                        JOIN
                            ejecutivo e ON a.Ejecutivo = e.cod_ejecutivo
                        JOIN
                            comision co ON e.cod_ejecutivo = co.cod_ejecutivo AND pag.Cod_pago = co.Cod_pago
                        LEFT JOIN 
                            comisiones_ejecutivos coe ON e.cod_ejecutivo = coe.cod_ejecutivo AND p.Ramo = 'Auto'
                        WHERE
                            pag.fecha BETWEEN %s AND %s
                        AND e.cod_ejecutivo = %s
                                """)
                    mycursor.execute(querySQL, (start, end, ejecutivo))

                resultado_busqueda = mycursor.fetchall()
                return resultado_busqueda

    except Exception as e:
        print(f"Ocurrió un error en def filtrar comision: {e}")
        return []



# Funcion Empleados Informe (Reporte)
def empleadosReporte():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                querySQL = ("""
                    SELECT 
                        e.id_empleado,
                        e.nombre_empleado, 
                        e.apellido_empleado,
                        e.salario_empleado,
                        e.email_empleado,
                        e.telefono_empleado,
                        e.profesion_empleado,
                        DATE_FORMAT(e.fecha_registro, '%d de %b %Y %h:%i %p') AS fecha_registro,
                        CASE
                            WHEN e.sexo_empleado = 1 THEN 'Masculino'
                            ELSE 'Femenino'
                        END AS sexo_empleado
                    FROM tbl_empleados AS e
                    ORDER BY e.id_empleado DESC
                    """)
                cursor.execute(querySQL,)
                empleadosBD = cursor.fetchall()
        return empleadosBD
    except Exception as e:
        print(
            f"Error en la función empleadosReporte: {e}")
        return None
    

def sql_lista_antiguos_contratos(cod_poliza):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                querySQL = ("""
                    SELECT
                        p.cod_poliza,
                        p.CI_asegurado,
                        DATE_FORMAT(r.Fecha_contrato, '%Y-%m-%d') as fecha, 
                        p.Ramo,
                        r.Fecha_vencimiento,
                        r.Cod_renovacion,
                        CASE
                            WHEN COUNT(pa.Cod_pago) > 0 THEN 'Sí'
                            ELSE 'No'
                        END AS Tiene_pago
                    FROM
                        poliza p
                    JOIN
                        renovacion r ON p.cod_poliza = r.cod_poliza
                    LEFT JOIN
                        pago pa ON r.Cod_renovacion = pa.Cod_renovacion
                    where p.cod_poliza=%s
                    GROUP BY
                        p.cod_poliza, p.CI_asegurado, r.Fecha_contrato, p.Ramo, r.Fecha_vencimiento,r.Cod_renovacion
                    order by r.Fecha_contrato desc; 
                    """)
                cursor.execute(querySQL,(cod_poliza,))
                polizasBD = cursor.fetchall()
        return polizasBD
    except Exception as e:
        print(
            f"sql_lista_antiguos_contrato: {e}")
        return None


def generarReporteExcel():
    dataEmpleados = empleadosReporte()
    wb = openpyxl.Workbook()
    hoja = wb.active

    # Agregar la fila de encabezado con los títulos
    cabeceraExcel = ("Nombre", "Apellido", "Sexo",
                     "Telefono", "Email", "Profesión", "Salario", "Fecha de Ingreso")

    hoja.append(cabeceraExcel)

    # Formato para números en moneda colombiana y sin decimales
    formato_moneda_colombiana = '#,##0'

    # Agregar los registros a la hoja
    for registro in dataEmpleados:
        nombre_empleado = registro['nombre_empleado']
        apellido_empleado = registro['apellido_empleado']
        sexo_empleado = registro['sexo_empleado']
        telefono_empleado = registro['telefono_empleado']
        email_empleado = registro['email_empleado']
        profesion_empleado = registro['profesion_empleado']
        salario_empleado = registro['salario_empleado']
        fecha_registro = registro['fecha_registro']

        # Agregar los valores a la hoja
        hoja.append((nombre_empleado, apellido_empleado, sexo_empleado, telefono_empleado, email_empleado, profesion_empleado,
                     salario_empleado, fecha_registro))

        # Itera a través de las filas y aplica el formato a la columna G
        for fila_num in range(2, hoja.max_row + 1):
            columna = 7  # Columna G
            celda = hoja.cell(row=fila_num, column=columna)
            celda.number_format = formato_moneda_colombiana

    fecha_actual = datetime.datetime.now()
    archivoExcel = f"Reporte_empleados_{fecha_actual.strftime('%Y_%m_%d')}.xlsx"
    carpeta_descarga = "../static/downloads-excel"
    ruta_descarga = os.path.join(os.path.dirname(
        os.path.abspath(__file__)), carpeta_descarga)

    if not os.path.exists(ruta_descarga):
        os.makedirs(ruta_descarga)
        # Dando permisos a la carpeta
        os.chmod(ruta_descarga, 0o755)

    ruta_archivo = os.path.join(ruta_descarga, archivoExcel)
    wb.save(ruta_archivo)

    # Enviar el archivo como respuesta HTTP
    return send_file(ruta_archivo, as_attachment=True)


def buscarEmpleadoBD(search):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as mycursor:
                querySQL = ("""
                        SELECT 
                            e.id_empleado,
                            e.nombre_empleado, 
                            e.apellido_empleado,
                            e.salario_empleado,
                            CASE
                                WHEN e.sexo_empleado = 1 THEN 'Masculino'
                                ELSE 'Femenino'
                            END AS sexo_empleado
                        FROM tbl_empleados AS e
                        WHERE e.nombre_empleado LIKE %s 
                        ORDER BY e.id_empleado DESC
                    """)
                search_pattern = f"%{search}%"  # Agregar "%" alrededor del término de búsqueda
                mycursor.execute(querySQL, (search_pattern,))
                resultado_busqueda = mycursor.fetchall()
                return resultado_busqueda

    except Exception as e:
        print(f"Ocurrió un error en def buscarEmpleadoBD: {e}")
        return []

def buscarPolizaBD(search,filtro):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as mycursor:
                if filtro == 'Cod':
                    querySQL = ("""
                            SELECT
                                p.cod_poliza,
                                p.CI_asegurado,
                                a.Nombre,
                                a.Nombre2,
                                a.Apellido,
                                a.Apellido2,
                                DATE_FORMAT(r.Fecha_contrato, '%Y-%m-%d') as fecha,
                                p.Ramo,
                                r.Fecha_vencimiento,
                                r.riesgo,
                                CASE
                                    WHEN COUNT(pa.Cod_pago) > 0 THEN 'Sí'
                                    ELSE 'No'
                                END AS Tiene_pago
                            FROM
                                poliza p
                            JOIN (
                                SELECT r2.cod_poliza, r2.Cod_renovacion, r2.Fecha_contrato, r2.Fecha_vencimiento, r2.riesgo
                                FROM renovacion r2
                                JOIN (
                                    SELECT cod_poliza, MAX(Cod_renovacion) as max_cod
                                    FROM renovacion
                                    GROUP BY cod_poliza
                                ) r3 ON r2.Cod_renovacion = r3.max_cod
                            ) r ON p.cod_poliza = r.cod_poliza
                            JOIN
                                asegurado a ON p.CI_asegurado = a.CI
                            LEFT JOIN
                                pago pa ON r.Cod_renovacion = pa.Cod_renovacion
                            WHERE p.cod_poliza=%s
                            GROUP BY
                                p.cod_poliza, p.CI_asegurado, r.Fecha_contrato,r.riesgo, p.Ramo, r.Fecha_vencimiento
                            ORDER BY
                                p.cod_poliza;""")
                else:
                    querySQL = ("""
                            SELECT
                                p.cod_poliza,
                                p.CI_asegurado,
                                a.Nombre,
                                a.Nombre2,
                                a.Apellido,
                                a.Apellido2,
                                DATE_FORMAT(r.Fecha_contrato, '%Y-%m-%d') as fecha,
                                p.Ramo,
                                r.Fecha_vencimiento,
                                r.riesgo,
                                CASE
                                    WHEN COUNT(pa.Cod_pago) > 0 THEN 'Sí'
                                    ELSE 'No'
                                END AS Tiene_pago
                            FROM
                                poliza p
                            JOIN (
                                SELECT r2.cod_poliza, r2.Cod_renovacion, r2.Fecha_contrato, r2.Fecha_vencimiento, r2.riesgo
                                FROM renovacion r2
                                JOIN (
                                    SELECT cod_poliza, MAX(Cod_renovacion) as max_cod
                                    FROM renovacion
                                    GROUP BY cod_poliza
                                ) r3 ON r2.Cod_renovacion = r3.max_cod
                            ) r ON p.cod_poliza = r.cod_poliza
                            JOIN
                                asegurado a ON p.CI_asegurado = a.CI
                            LEFT JOIN
                                pago pa ON r.Cod_renovacion = pa.Cod_renovacion
                            WHERE p.CI_asegurado = %s
                            GROUP BY
                                p.cod_poliza, p.CI_asegurado, r.Fecha_contrato,r.riesgo, p.Ramo, r.Fecha_vencimiento
                            ORDER BY
                                p.cod_poliza;""")

                search_pattern = f"{search}"  # Agregar "%" alrededor del término de búsqueda
                mycursor.execute(querySQL, (search_pattern,))
                resultado_busqueda = mycursor.fetchall()
                return resultado_busqueda
    except Exception as e:
        print(f"Ocurrió un error en def buscarPolizasBD: {e}")
        return []
    
def buscarAseguradoBD(search,filtro):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as mycursor:
                querySQL = ("""
                    SELECT 
                        a.CI,
                        a.Nombre,
                        a.Nombre2, 
                        a.Apellido,
                        a.Apellido2,
                        a.Correo,
                        a.Telefono,
                        a.Fecha_nacimiento,
                        a.profesion,
                        a.localidad,
                        a.canal,
                        a.Tipo_CI
                    FROM asegurado AS a
                    where a.CI=%s and a.Tipo_CI=%s""")
                print(filtro)
                search_pattern = f"{search}"  # Agregar "%" alrededor del término de búsqueda
                print(search_pattern)
                mycursor.execute(querySQL, (search_pattern,filtro))
                resultado_busqueda = mycursor.fetchall()
                return resultado_busqueda
    except Exception as e:
        print(f"Ocurrió un error en def buscarAseguradoBD: {e}")
        return []

def buscarEmpleadoUnico(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as mycursor:
                querySQL = ("""
                        SELECT 
                            e.id_empleado,
                            e.nombre_empleado, 
                            e.apellido_empleado,
                            e.sexo_empleado,
                            e.telefono_empleado,
                            e.email_empleado,
                            e.profesion_empleado,
                            e.salario_empleado,
                            e.foto_empleado
                        FROM tbl_empleados AS e
                        WHERE e.id_empleado =%s LIMIT 1
                    """)
                mycursor.execute(querySQL, (id,))
                empleado = mycursor.fetchone()
                return empleado

    except Exception as e:
        print(f"Ocurrió un error en def buscarEmpleadoUnico: {e}")
        return []


def procesar_actualizacion_form(data):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                nombre_empleado = data.form['nombre_empleado']
                apellido_empleado = data.form['apellido_empleado']
                sexo_empleado = data.form['sexo_empleado']
                telefono_empleado = data.form['telefono_empleado']
                email_empleado = data.form['email_empleado']
                profesion_empleado = data.form['profesion_empleado']

                salario_sin_puntos = re.sub(
                    '[^0-9]+', '', data.form['salario_empleado'])
                salario_empleado = int(salario_sin_puntos)
                id_empleado = data.form['id_empleado']

                if data.files['foto_empleado']:
                    file = data.files['foto_empleado']
                    fotoForm = procesar_imagen_perfil(file)

                    querySQL = """
                        UPDATE tbl_empleados
                        SET 
                            nombre_empleado = %s,
                            apellido_empleado = %s,
                            sexo_empleado = %s,
                            telefono_empleado = %s,
                            email_empleado = %s,
                            profesion_empleado = %s,
                            salario_empleado = %s,
                            foto_empleado = %s
                        WHERE id_empleado = %s
                    """
                    values = (nombre_empleado, apellido_empleado, sexo_empleado,
                              telefono_empleado, email_empleado, profesion_empleado,
                              salario_empleado, fotoForm, id_empleado)
                else:
                    querySQL = """
                        UPDATE tbl_empleados
                        SET 
                            nombre_empleado = %s,
                            apellido_empleado = %s,
                            sexo_empleado = %s,
                            telefono_empleado = %s,
                            email_empleado = %s,
                            profesion_empleado = %s,
                            salario_empleado = %s
                        WHERE id_empleado = %s
                    """
                    values = (nombre_empleado, apellido_empleado, sexo_empleado,
                              telefono_empleado, email_empleado, profesion_empleado,
                              salario_empleado, id_empleado)

                cursor.execute(querySQL, values)
                conexion_MySQLdb.commit()

        return cursor.rowcount or []
    except Exception as e:
        print(f"Ocurrió un error en procesar_actualizacion_form: {e}")
        return None
    

def procesar_actualizacion_form_asegurado(data):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                nombre_empleado = data.form['nombre_asegurado']
                apellido_empleado = data.form['apellido_asegurado']
                nombre2 = data.form['nombre2_asegurado']
                apellido2 = data.form['apellido2_asegurado']
                
                prefijo = data.form.get('prefijo_telefono', '')
                numero = data.form.get('telefono_asegurado_numero', '')
                telefono_empleado = f"{prefijo}{numero}"

                email_empleado = data.form['email_asegurado'] 
                fecha=data.form['Fecha'] 
                profesion=data.form['Profesion'] 
                localidad=data.form['localidad'] 
                canal=data.form['Canal'] 
                ejecutivo = data.form.get('Ejecutivo')
 
                id_empleado = data.form['CI']

                id_empleado = int(id_empleado)

                 
                querySQL = """
                        UPDATE asegurado
                        SET 
                            Nombre = %s,
                            Nombre2 = %s,
                            Apellido = %s,
                            Apellido2 = %s,
                            Telefono = %s,
                            Correo = %s,
                            Fecha_nacimiento=%s,
                            profesion=%s,
                            localidad=%s,
                            canal=%s,
                            Ejecutivo=%s
                        WHERE CI = %s
                    """
                values = (nombre_empleado,nombre2, apellido_empleado,apellido2,
                              telefono_empleado, email_empleado,fecha,profesion, localidad, canal, ejecutivo, id_empleado)

                cursor.execute(querySQL, values)
                conexion_MySQLdb.commit()

        return cursor.rowcount 
    except Exception as e:
        print(f"Ocurrió un error en procesar_actualizacion_form_asegurado: {e}")
        return None

def actualizar_riesgo_asegurado(data):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                ci = data.get('ci')
                profesion = data.get('profesion')
                localidad = data.get('localidad')
                canal = data.get('canal')

                if not ci:
                    return {'success': False, 'message': 'CI no proporcionado.'}

                sql = """
                    UPDATE asegurado
                    SET
                        profesion = %s,
                        localidad = %s,
                        canal = %s
                    WHERE CI = %s;
                """
                valores = (profesion, localidad, canal, ci)
                cursor.execute(sql, valores)
                conexion_MySQLdb.commit()

                if cursor.rowcount > 0:
                    return {'success': True, 'message': 'Datos actualizados correctamente.'}
                else:
                    return {'success': False, 'message': 'No se encontró el asegurado o no hubo cambios.'}

    except Exception as e:
        # En un entorno de producción, sería bueno registrar el error.
        # import logging
        # logging.error(f"Error en actualizar_riesgo_asegurado: {e}")
        return {'success': False, 'message': f'Error en el servidor: {str(e)}'}
    
def procesar_actualizacion_form_company(data):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                nombre = data.form['nombre_Company']
                rif = data.form['Rif']
                cod = data.form['cod']
                cod = int(cod)

                 
                querySQL = """
                        UPDATE compania
                        SET 
                            Nombre = %s,
                            rif = %s
                        WHERE Cod_compania = %s
                    """
                values = (nombre,rif, cod)

                cursor.execute(querySQL, values)
                conexion_MySQLdb.commit()

        return cursor.rowcount 
    except Exception as e:
        print(f"Ocurrió un error en procesar_actualizacion_form_company: {e}")
        return None
    



def procesar_actualizacion_form_reembolso(data):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:  
                cod_reembolso = data.form['cod_reembolso']
                cod_reembolso = int(cod_reembolso)  

                diagnostico = data.form['Diagnostico']
                estado = data.form['Estado']
                fecha_ocurrencia = data.form['Fecha_ocurrencia']
                fecha_noti = data.form['Fecha_noti']
                fecha_max = data.form.get('Fecha_max') # Use .get() for optional fields
                moneda = data.form['Moneda']
                monto_solicitado = data.form['Monto_solicitado']
                monto_pagado = data.form['Monto_pagado']
                monto_dolares = data.form.get('monto_dolares')
                tipo_atencion = data.form.get('Tipo_Atencion')
                fecha_pago = data.form.get('Fecha_pago') # Use .get() for optional fields
                correo = data.form['Correo']
                # Numero_siniestro is explicitly excluded from update as per request

                # Convert date strings to datetime objects for database insertion if your DB expects it
                # Or ensure your database driver handles string dates correctly (MySQLdb usually does)
                fecha_ocurrencia_dt = datetime.strptime(fecha_ocurrencia, '%Y-%m-%d').date()
                fecha_noti_dt = datetime.strptime(fecha_noti, '%Y-%m-%d').date()
                fecha_max_dt = datetime.strptime(fecha_max, '%Y-%m-%d').date() if fecha_max else None
                fecha_pago_dt = datetime.strptime(fecha_pago, '%Y-%m-%d').date() if fecha_pago else None


                querySQL = """
                    UPDATE reembolso
                    SET
                        Diagnostico = %s,
                        Estado = %s,
                        Fecha_ocurrencia = %s,
                        Fecha_noti = %s,
                        Fecha_max = %s,
                        Moneda = %s,
                        Monto_solicitado = %s,
                        Monto_pagado = %s,
                        Monto_dolares = %s,
                        Fecha_pago = %s,
                        Correo = %s,
                        Tipo_Atencion = %s
                    WHERE cod_reembolso = %s
                """
                values = (
                    diagnostico,
                    estado,
                    fecha_ocurrencia_dt,
                    fecha_noti_dt,
                    fecha_max_dt,
                    moneda,
                    monto_solicitado,
                    monto_pagado,
                    monto_dolares,
                    fecha_pago_dt,
                    correo,
                    tipo_atencion,
                    cod_reembolso
                )
                print(values)

                cursor.execute(querySQL, values)
                conexion_MySQLdb.commit()

        return cursor.rowcount 
    except Exception as e:
        print(f"Ocurrió un error en procesar_actualizacion_form_reembolso: {e}")
        return None




def procesar_actualizacion_form_CartaAval(data):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor: 
                print("inicio")
                cod_carta_aval = data.form['Cod_CartaAval']
                cod_carta_aval = int(cod_carta_aval) 
          

                diagnostico = data.form['Diagnostico']
                estado = data.form['Estado']
                moneda = data.form['Moneda']
                monto_solicitado = data.form['Monto_solicitado']
                monto_aprobado = data.form['Monto_aprobado']
                fecha_noti = data.form['Fecha_noti']
                fecha_apro = data.form.get('Fecha_apro',None)  
                correo = data.form['Correo']
                tipo_atencion = data.form['Tipo_Atencion'] 
                monto_aprobadoD=data.form['Monto_aprobadoD']
                mes = data.form.get('MES')
                negocio = data.form.get('NEGOCIO')
                titular = data.form.get('TITULAR')
                reclamante = data.form.get('RECLAMANTE')
                referidor = data.form.get('REFERIDOR')
                analista = data.form.get('ANALISTA')
                observacion = data.form.get('OBSERVACION')
                codigo_siniestro = data.form.get('codigo_siniestro')

                print("aver")
                querySQL = """
                    UPDATE carta_aval
                    SET
                        Diagnostico = %s,
                        Procedimiento = %s,
                        Estado = %s,
                        Moneda = %s,
                        Monto_solicitado = %s,
                        Monto_aprobado = %s,
                        Fecha_noti = %s,
                        Fecha_apro = %s,
                        Correo = %s,
                        Tipo_Atencion = %s,
                        Monto_aprobadoD = %s,
                        MES = %s,
                        NEGOCIO = %s,
                        TITULAR = %s,
                        RECLAMANTE = %s,
                        REFERIDOR = %s,
                        ANALISTA = %s,
                        OBSERVACION = %s,
                        codigo_siniestro = %s
                    WHERE Cod_CartaAval = %s
                """
                values = (
                    diagnostico,
                    data.form.get('Procedimiento') or diagnostico,
                    estado,
                    normalize_moneda(moneda),
                    clean_monto(monto_solicitado),
                    clean_monto(monto_aprobado),
                    fecha_noti,
                    fecha_apro,
                    correo,
                    tipo_atencion,
                    clean_monto(monto_aprobadoD), 
                    mes,
                    negocio,
                    titular,
                    reclamante,
                    referidor,
                    analista,
                    observacion,
                    codigo_siniestro,
                    cod_carta_aval # This is for the WHERE clause
                )
                print("antes de ejecutar")
                print(f"Executing query: {querySQL}")
                print(f"With values: {values}")

                cursor.execute(querySQL, values)
                conexion_MySQLdb.commit()

        return cursor.rowcount # Returns the number of rows updated
    except Exception as e:
        print(f"Ocurrió un error en procesar_actualizacion_form_CartaAval: {e}")
        return None
    
def procesar_actualizacion_form_SiniestroAuto(data):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor: 
                print(data.form)
                cod_siniestro = data.form['Cod_siniestroA']
                print("Cod_siniestro:", cod_siniestro)
                cod_siniestro = int(cod_siniestro)  
                print("aver9")
                fecha_ocurrencia = data.form['Fecha_ocurrencia']
                print("aver8")
                fecha_noti = data.form['Fecha_noti']
                print("aver7")
                fecha_inspec = data.form.get('Fecha_inspec', None)  # Use .get() for optional fields
                print("aver6")
                Estado = data.form['Estado'] 
                print("aver4")
                Monto_orden = data.form['Monto_orden']  
                print("aver3")
                correo = data.form['Correo']
                print("aver2")
                descripcion = data.form['Descripcion']
                print("aver")
                numero_siniestro = data.form['Numero_siniestro']
                print("prueba")

                

                querySQL = """
                    UPDATE automovilsiniestro
                    SET
                        Fecha_ocurrencia = %s,
                        Fecha_noti = %s,
                        Fecha_inspec = %s, 
                        Estado = %s,
                        Monto_orden = %s, 
                        Correo = %s,
                        Descripcion = %s,
                        codigo_siniestro = %s
                    WHERE Cod_siniestroA = %s
                """
                values = (
                    fecha_ocurrencia,
                    fecha_noti,
                    fecha_inspec, 
                    Estado,
                    Monto_orden,
                    correo,
                    descripcion,
                    numero_siniestro,
                    cod_siniestro
                )

                cursor.execute(querySQL, values)
                conexion_MySQLdb.commit()

        return cursor.rowcount 
    except Exception as e:
        print(f"Ocurrió un error en procesar_actualizacion_form_SiniestroAuto: {e}")
        return None

def procesar_actualizacion_form_ejecutivo(data):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                nombre = data.form['nombre_ejecutivo']
                nombre2 = data.form['nombre2_ejecutivo']
                apellido = data.form['apellido_ejecutivo']
                apellido2 = data.form['apellido2_ejecutivo']
                telefono = data.form['telefono_ejecutivo']
                correo = data.form['email_ejecutivo']
                cod = data.form['cod_ejecutivo']
                cod = int(cod)

                 
                querySQL = """
                        UPDATE ejecutivo
                        SET 
                            Nombre = %s,
                            nombre2 = %s,
                            Apellido = %s,
                            Apellido2 = %s,
                            Telefono=%s,
                            Correo=%s
                        WHERE cod_ejecutivo = %s
                    """
                values = (nombre,nombre2,apellido,apellido2,telefono,correo, cod)

                cursor.execute(querySQL, values)
                conexion_MySQLdb.commit()

        return cursor.rowcount 
    except Exception as e:
        print(f"Ocurrió un error en procesar_actualizacion_form_ejecutivo: {e}")
        return None


def recalcular_pagos_por_cambio_prima(cursor, cod_renovacion, nueva_prima, frecuencia):
    """
    Recalcula los montos de los pagos existentes para que sumen la nueva prima,
    dividiéndolos de manera uniforme según la frecuencia.
    """
    frecuencia_map = {
        1: {'pagos': 1, 'intervalo': 12}, 
        2: {'pagos': 12, 'intervalo': 1}, 
        3: {'pagos': 4, 'intervalo': 3},  
        4: {'pagos': 2, 'intervalo': 6},  
        5: {'pagos': 3, 'intervalo': 4},  
        6: {'pagos': 6, 'intervalo': 2},  
        10: {'pagos': 10, 'intervalo': 1}
    }
    
    try:
        f_key = int(frecuencia)
        config = frecuencia_map.get(f_key, {'pagos': 1, 'intervalo': 12})
    except:
        config = {'pagos': 1, 'intervalo': 12}
        
    num_pagos_total = config['pagos']
    if num_pagos_total <= 0:
        num_pagos_total = 1
        
    monto_cuota = float(nueva_prima) / num_pagos_total
    
    # Actualizar los pagos existentes que pertenecen a esta renovación
    sql_update_pagos = "UPDATE pago SET monto = %s WHERE Cod_renovacion = %s"
    cursor.execute(sql_update_pagos, (monto_cuota, cod_renovacion))

def procesar_actualizacion_poliza_persona(dataForm, cod_poliza):

    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                # Capturar datos del formulario
                nuevo_cod_poliza = dataForm.get('cod_poliza')
                frecuencia_form = dataForm.get('frecuencia')
                tomador_form = dataForm.get('tomador_poliza')
                cobertura_form = dataForm.get('cobertura')
                fecha_emision_form = dataForm.get('fecha_emision')
                subramo_form = dataForm.get('subramo')
                producto_form = dataForm.get('producto')

                prima_str = dataForm.get('prima', '0')
                try:
                    prima_form = float(prima_str)
                except ValueError:
                    prima_cleaned = prima_str.replace('.', '').replace(',', '.')
                    prima_form = float(prima_cleaned)

                # 1. Validar unicidad del código si cambia
                if nuevo_cod_poliza and nuevo_cod_poliza != cod_poliza:
                    sql_check_poliza = "SELECT cod_poliza FROM poliza WHERE cod_poliza = %s"
                    cursor.execute(sql_check_poliza, (nuevo_cod_poliza,))
                    if cursor.fetchone():
                        flash('El código de póliza ya existe.', 'error')
                        return None

                # Obtener la renovación más reciente
                sql_get_renovacion = "SELECT Cod_renovacion, Prima, Frecuencia FROM renovacion WHERE cod_poliza = %s ORDER BY Cod_renovacion DESC LIMIT 1"
                cursor.execute(sql_get_renovacion, (cod_poliza,))
                renovacion_actual = cursor.fetchone()

                if not renovacion_actual:
                    flash('No se encontró renovación.', 'error')
                    return None

                cod_renovacion = renovacion_actual['Cod_renovacion']
                
                # 2. Recalcular pagos si cambia prima o frecuencia
                if (str(prima_form) != str(renovacion_actual['Prima'])) or (str(frecuencia_form) != str(renovacion_actual['Frecuencia'])):
                    recalcular_pagos_por_cambio_prima(cursor, cod_renovacion, prima_form, frecuencia_form)

                # 3. Calcular nueva fecha de vencimiento si cambia la emisión
                fecha_vencimiento_form = None
                if fecha_emision_form:
                    fecha_dt = datetime.datetime.strptime(fecha_emision_form, '%Y-%m-%d')
                    fecha_vencimiento_form = (fecha_dt + datetime.timedelta(days=365)).strftime('%Y-%m-%d')

                # 4. Actualizar renovacion
                sql_renovacion_update = """
                    UPDATE renovacion 
                    SET Prima = %s, Frecuencia = %s, cobertura = %s, Fecha_contrato = %s, Fecha_vencimiento = %s 
                    WHERE Cod_renovacion = %s
                """
                cursor.execute(sql_renovacion_update, (prima_form, frecuencia_form, cobertura_form, fecha_emision_form, fecha_vencimiento_form, cod_renovacion))
                 
                # 5. Actualizar poliza y tablas relacionadas
                if nuevo_cod_poliza and nuevo_cod_poliza != cod_poliza:
                    # Actualizar primero las tablas hijas para asegurar consistencia
                    # Usamos el nombre de columna cod_poliza o Cod_poliza según se use en cada tabla
                    cursor.execute("UPDATE renovacion SET cod_poliza = %s WHERE cod_poliza = %s", (nuevo_cod_poliza, cod_poliza))
                    cursor.execute("UPDATE Persona SET cod_poliza = %s WHERE cod_poliza = %s", (nuevo_cod_poliza, cod_poliza))
                    cursor.execute("UPDATE beneficiario SET Cod_poliza = %s WHERE Cod_poliza = %s", (nuevo_cod_poliza, cod_poliza))
                    cursor.execute("UPDATE carta_aval SET Cod_poliza = %s WHERE Cod_poliza = %s", (nuevo_cod_poliza, cod_poliza))
                    cursor.execute("UPDATE Reembolso SET Cod_poliza = %s WHERE Cod_poliza = %s", (nuevo_cod_poliza, cod_poliza))
                    
                    sql_poliza_update = "UPDATE poliza SET cod_poliza = %s, Tomador = %s, Fecha_emision = %s WHERE cod_poliza = %s"
                    valores_poliza = (nuevo_cod_poliza, tomador_form, fecha_emision_form, cod_poliza)
                else:
                    sql_poliza_update = "UPDATE poliza SET Tomador = %s, Fecha_emision = %s WHERE cod_poliza = %s"
                    valores_poliza = (tomador_form, fecha_emision_form, cod_poliza)
                
                cursor.execute(sql_poliza_update, valores_poliza)

                # 6. Actualizar Persona (Subramo y Producto)
                target_cod = nuevo_cod_poliza if (nuevo_cod_poliza and nuevo_cod_poliza != cod_poliza) else cod_poliza
                
                if subramo_form or producto_form:
                     # Ahora target_cod es el nuevo código si hubo cambio
                     sql_persona_update = "UPDATE Persona SET Subramo = %s, Producto = %s WHERE cod_poliza = %s"
                     cursor.execute(sql_persona_update, (subramo_form, producto_form, target_cod))

                conexion_MySQLdb.commit()
                return True
    except Exception as e:
        print(f'Error en actualizar Persona: {str(e)}')
        flash(f'Error: {str(e)}', 'error')
        return None
    except Exception as e:
        print(f'Se produjo un error en procesar_actualizacion_poliza_persona: {str(e)}')
        flash(f'Se produjo un error en el servidor: {str(e)}', 'error')
        return None

def procesar_actualizacion_poliza_auto(form_data, cod_poliza):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                # Capturar datos del formulario
                nuevo_cod_poliza = form_data.get('cod_poliza')
                frecuencia_form = form_data.get('frecuencia')
                tomador_form = form_data.get('tomador_poliza')
                cobertura_form = form_data.get('cobertura')
                fecha_emision_form = form_data.get('fecha_emision')

                prima_str = form_data.get('prima', '0')
                try:
                    prima_form = float(prima_str)
                except ValueError:
                    prima_cleaned = prima_str.replace('.', '').replace(',', '.')
                    prima_form = float(prima_cleaned)

                # 1. Validar unicidad del código si cambia
                if nuevo_cod_poliza and nuevo_cod_poliza != cod_poliza:
                    sql_check_poliza = "SELECT cod_poliza FROM poliza WHERE cod_poliza = %s"
                    cursor.execute(sql_check_poliza, (nuevo_cod_poliza,))
                    if cursor.fetchone():
                        flash('El código de póliza ya existe.', 'error')
                        return None

                # Obtener la renovación más reciente
                sql_get_renovacion = "SELECT Cod_renovacion, Prima, Frecuencia FROM renovacion WHERE cod_poliza = %s ORDER BY Cod_renovacion DESC LIMIT 1"
                cursor.execute(sql_get_renovacion, (cod_poliza,))
                renovacion_actual = cursor.fetchone()

                if not renovacion_actual:
                    flash('No se encontró renovación.', 'error')
                    return None

                cod_renovacion = renovacion_actual['Cod_renovacion']

                # 2. Recalcular pagos si cambia prima o frecuencia
                if (str(prima_form) != str(renovacion_actual['Prima'])) or (str(frecuencia_form) != str(renovacion_actual['Frecuencia'])):
                    recalcular_pagos_por_cambio_prima(cursor, cod_renovacion, prima_form, frecuencia_form)

                # 3. Calcular nueva fecha de vencimiento
                fecha_vencimiento_form = None
                if fecha_emision_form:
                    fecha_dt = datetime.datetime.strptime(fecha_emision_form, '%Y-%m-%d')
                    fecha_vencimiento_form = (fecha_dt + datetime.timedelta(days=365)).strftime('%Y-%m-%d')

                # 4. Actualizar renovacion
                sql_renovacion_update = """
                    UPDATE renovacion 
                    SET Prima = %s, Frecuencia = %s, cobertura = %s, Fecha_contrato = %s, Fecha_vencimiento = %s 
                    WHERE Cod_renovacion = %s
                """
                cursor.execute(sql_renovacion_update, (prima_form, frecuencia_form, cobertura_form, fecha_emision_form, fecha_vencimiento_form, cod_renovacion))
                 
                # 5. Actualizar poliza y tablas relacionadas
                if nuevo_cod_poliza and nuevo_cod_poliza != cod_poliza:
                    cursor.execute("UPDATE renovacion SET cod_poliza = %s WHERE cod_poliza = %s", (nuevo_cod_poliza, cod_poliza))
                    cursor.execute("UPDATE Auto SET cod_poliza = %s WHERE cod_poliza = %s", (nuevo_cod_poliza, cod_poliza))
                    cursor.execute("UPDATE AutomovilSiniestro SET Cod_poliza = %s WHERE Cod_poliza = %s", (nuevo_cod_poliza, cod_poliza))
                    
                    sql_poliza_update = "UPDATE poliza SET cod_poliza = %s, Tomador = %s, Fecha_emision = %s WHERE cod_poliza = %s"
                    valores_poliza = (nuevo_cod_poliza, tomador_form, fecha_emision_form, cod_poliza)
                else:
                    sql_poliza_update = "UPDATE poliza SET Tomador = %s, Fecha_emision = %s WHERE cod_poliza = %s"
                    valores_poliza = (tomador_form, fecha_emision_form, cod_poliza)
                
                cursor.execute(sql_poliza_update, valores_poliza)

                # 6. Actualizar Auto (Subramo y Producto)
                target_cod = nuevo_cod_poliza if (nuevo_cod_poliza and nuevo_cod_poliza != cod_poliza) else cod_poliza
                subramo_form = form_data.get('subramo')
                producto_form = form_data.get('producto')

                if subramo_form and producto_form:
                    sql_auto_update = "UPDATE Auto SET Subramo = %s, Producto = %s WHERE cod_poliza = %s"
                    cursor.execute(sql_auto_update, (subramo_form, producto_form, target_cod))

                conexion_MySQLdb.commit()
                return True
    except Exception as e:
        print(f'Error en actualizar Auto: {str(e)}')
        flash(f'Error: {str(e)}', 'error')
        return None
    except Exception as e:
        print(f'Se produjo un error en procesar_actualizacion_poliza_auto: {str(e)}')
        flash(f'Se produjo un error en el servidor: {str(e)}', 'error')
        return None

def procesar_actualizacion_poliza_patrimonial(dataForm, cod_poliza):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                # Capturar datos del formulario
                nuevo_cod_poliza = dataForm.get('cod_poliza')
                tomador_form = dataForm.get('tomador_poliza')
                fecha_emision_form = dataForm.get('fecha_emision')
                subramo_form = dataForm.get('subramo')
                producto_form = dataForm.get('producto')

                # 1. Validar unicidad del código si cambia
                if nuevo_cod_poliza and nuevo_cod_poliza != cod_poliza:
                    sql_check_poliza = "SELECT cod_poliza FROM poliza WHERE cod_poliza = %s"
                    cursor.execute(sql_check_poliza, (nuevo_cod_poliza,))
                    if cursor.fetchone():
                        flash('El código de póliza ya existe.', 'error')
                        return None

                # Obtener la renovación más reciente
                sql_get_renovacion = "SELECT Cod_renovacion, Prima, Frecuencia FROM renovacion WHERE cod_poliza = %s ORDER BY Cod_renovacion DESC LIMIT 1"
                cursor.execute(sql_get_renovacion, (cod_poliza,))
                renovacion_actual = cursor.fetchone()

                if not renovacion_actual:
                    flash('No se encontró renovación.', 'error')
                    return None

                cod_renovacion = renovacion_actual['Cod_renovacion']
                
                # 2. Recalcular pagos si cambia prima o frecuencia
                if (str(prima_form) != str(renovacion_actual['Prima'])) or (str(frecuencia_form) != str(renovacion_actual['Frecuencia'])):
                    recalcular_pagos_por_cambio_prima(cursor, cod_renovacion, prima_form, frecuencia_form)

                # 3. Calcular nueva fecha de vencimiento si cambia la emisión
                fecha_vencimiento_form = None
                if fecha_emision_form:
                    fecha_dt = datetime.datetime.strptime(fecha_emision_form, '%Y-%m-%d')
                    fecha_vencimiento_form = (fecha_dt + datetime.timedelta(days=365)).strftime('%Y-%m-%d')

                # 4. Actualizar renovacion
                sql_renovacion_update = """
                    UPDATE renovacion 
                    SET Prima = %s, Frecuencia = %s, Fecha_contrato = %s, Fecha_vencimiento = %s
                    WHERE Cod_renovacion = %s
                """
                cursor.execute(sql_renovacion_update, (prima_form, frecuencia_form, fecha_emision_form, fecha_vencimiento_form, cod_renovacion))
                 
                # 5. Actualizar poliza y tablas relacionadas
                if nuevo_cod_poliza and nuevo_cod_poliza != cod_poliza:
                    cursor.execute("UPDATE renovacion SET cod_poliza = %s WHERE cod_poliza = %s", (nuevo_cod_poliza, cod_poliza))
                    cursor.execute("UPDATE patrimonio SET cod_poliza = %s WHERE cod_poliza = %s", (nuevo_cod_poliza, cod_poliza))
                    
                    sql_poliza_update = "UPDATE poliza SET cod_poliza = %s, Tomador = %s, Fecha_emision = %s WHERE cod_poliza = %s"
                    valores_poliza = (nuevo_cod_poliza, tomador_form, fecha_emision_form, cod_poliza)
                else:
                    sql_poliza_update = "UPDATE poliza SET Tomador = %s, Fecha_emision = %s WHERE cod_poliza = %s"
                    valores_poliza = (tomador_form, fecha_emision_form, cod_poliza)
                
                cursor.execute(sql_poliza_update, valores_poliza)

                 # 6. Actualizar Patrimonio (Subramo y Producto)
                target_cod = nuevo_cod_poliza if (nuevo_cod_poliza and nuevo_cod_poliza != cod_poliza) else cod_poliza
                
                if subramo_form or producto_form:
                     sql_patrimonio_update = "UPDATE patrimonio SET Subramo = %s, Producto = %s WHERE cod_poliza = %s"
                     cursor.execute(sql_patrimonio_update, (subramo_form, producto_form, target_cod))

                conexion_MySQLdb.commit()
                return True
    except Exception as e:
        print(f'Error en actualizar Patrimonial: {str(e)}')
        flash(f'Error: {str(e)}', 'error')
        return None

def procesar_actualizacion_poliza_viaje(dataForm, cod_poliza):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                # Capturar datos del formulario
                nuevo_cod_poliza = dataForm.get('cod_poliza')
                frecuencia_form = dataForm.get('frecuencia')
                tomador_form = dataForm.get('tomador_poliza')
                cobertura_form = dataForm.get('cobertura')
                fecha_emision_form = dataForm.get('fecha_emision')
                subramo_form = dataForm.get('subramo')
                producto_form = dataForm.get('producto')

                prima_str = dataForm.get('prima', '0')
                try:
                    prima_form = float(prima_str)
                except ValueError:
                    prima_cleaned = prima_str.replace('.', '').replace(',', '.')
                    prima_form = float(prima_cleaned)

                # 1. Validar unicidad del código si cambia
                if nuevo_cod_poliza and nuevo_cod_poliza != cod_poliza:
                    sql_check_poliza = "SELECT cod_poliza FROM poliza WHERE cod_poliza = %s"
                    cursor.execute(sql_check_poliza, (nuevo_cod_poliza,))
                    if cursor.fetchone():
                        flash('El código de póliza ya existe.', 'error')
                        return None

                # Obtener la renovación más reciente
                sql_get_renovacion = "SELECT Cod_renovacion, Prima, Frecuencia FROM renovacion WHERE cod_poliza = %s ORDER BY Cod_renovacion DESC LIMIT 1"
                cursor.execute(sql_get_renovacion, (cod_poliza,))
                renovacion_actual = cursor.fetchone()

                if not renovacion_actual:
                    flash('No se encontró renovación.', 'error')
                    return None

                cod_renovacion = renovacion_actual['Cod_renovacion']
                
                # 2. Recalcular pagos si cambia prima o frecuencia
                if (str(prima_form) != str(renovacion_actual['Prima'])) or (str(frecuencia_form) != str(renovacion_actual['Frecuencia'])):
                    recalcular_pagos_por_cambio_prima(cursor, cod_renovacion, prima_form, frecuencia_form)

                # 3. Calcular nueva fecha de vencimiento
                fecha_vencimiento_form = None
                if fecha_emision_form:
                    fecha_dt = datetime.datetime.strptime(fecha_emision_form, '%Y-%m-%d')
                    fecha_vencimiento_form = (fecha_dt + datetime.timedelta(days=365)).strftime('%Y-%m-%d')

                # 4. Actualizar renovacion
                sql_renovacion_update = """
                    UPDATE renovacion 
                    SET Prima = %s, Frecuencia = %s, cobertura = %s, Fecha_contrato = %s, Fecha_vencimiento = %s 
                    WHERE Cod_renovacion = %s
                """
                cursor.execute(sql_renovacion_update, (prima_form, frecuencia_form, cobertura_form, fecha_emision_form, fecha_vencimiento_form, cod_renovacion))
                 
                # 5. Actualizar poliza y tablas relacionadas
                if nuevo_cod_poliza and nuevo_cod_poliza != cod_poliza:
                    cursor.execute("UPDATE renovacion SET cod_poliza = %s WHERE cod_poliza = %s", (nuevo_cod_poliza, cod_poliza))
                    cursor.execute("UPDATE Viaje SET cod_poliza = %s WHERE cod_poliza = %s", (nuevo_cod_poliza, cod_poliza))
                    
                    sql_poliza_update = "UPDATE poliza SET cod_poliza = %s, Tomador = %s, Fecha_emision = %s WHERE cod_poliza = %s"
                    valores_poliza = (nuevo_cod_poliza, tomador_form, fecha_emision_form, cod_poliza)
                else:
                    sql_poliza_update = "UPDATE poliza SET Tomador = %s, Fecha_emision = %s WHERE cod_poliza = %s"
                    valores_poliza = (tomador_form, fecha_emision_form, cod_poliza)
                
                cursor.execute(sql_poliza_update, valores_poliza)

                # 6. Actualizar Viaje
                target_cod = nuevo_cod_poliza if (nuevo_cod_poliza and nuevo_cod_poliza != cod_poliza) else cod_poliza
                if subramo_form and producto_form:
                    sql_viaje_update = "UPDATE Viaje SET Subramo = %s, Producto = %s WHERE cod_poliza = %s"
                    cursor.execute(sql_viaje_update, (subramo_form, producto_form, target_cod))

                conexion_MySQLdb.commit()
                return True
    except Exception as e:
        print(f'Error en actualizar Viaje: {str(e)}')
        flash(f'Error: {str(e)}', 'error')
        return None

def procesar_actualizacion_poliza_fianza(dataForm, cod_poliza):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                # Capturar datos del formulario
                nuevo_cod_poliza = dataForm.get('cod_poliza')
                frecuencia_form = dataForm.get('frecuencia')
                tomador_form = dataForm.get('tomador_poliza')
                cobertura_form = dataForm.get('cobertura')
                fecha_emision_form = dataForm.get('fecha_emision')
                subramo_form = dataForm.get('subramo')
                producto_form = dataForm.get('producto')

                prima_str = dataForm.get('prima', '0')
                try:
                    prima_form = float(prima_str)
                except ValueError:
                    prima_cleaned = prima_str.replace('.', '').replace(',', '.')
                    prima_form = float(prima_cleaned)

                # 1. Validar unicidad del código si cambia
                if nuevo_cod_poliza and nuevo_cod_poliza != cod_poliza:
                    sql_check_poliza = "SELECT cod_poliza FROM poliza WHERE cod_poliza = %s"
                    cursor.execute(sql_check_poliza, (nuevo_cod_poliza,))
                    if cursor.fetchone():
                        flash('El código de póliza ya existe.', 'error')
                        return None

                # Obtener la renovación más reciente
                sql_get_renovacion = "SELECT Cod_renovacion, Prima, Frecuencia FROM renovacion WHERE cod_poliza = %s ORDER BY Cod_renovacion DESC LIMIT 1"
                cursor.execute(sql_get_renovacion, (cod_poliza,))
                renovacion_actual = cursor.fetchone()

                if not renovacion_actual:
                    flash('No se encontró renovación.', 'error')
                    return None

                cod_renovacion = renovacion_actual['Cod_renovacion']
                
                # 2. Recalcular pagos si cambia prima o frecuencia
                if (str(prima_form) != str(renovacion_actual['Prima'])) or (str(frecuencia_form) != str(renovacion_actual['Frecuencia'])):
                    recalcular_pagos_por_cambio_prima(cursor, cod_renovacion, prima_form, frecuencia_form)

                # 3. Calcular nueva fecha de vencimiento
                fecha_vencimiento_form = None
                if fecha_emision_form:
                    fecha_dt = datetime.datetime.strptime(fecha_emision_form, '%Y-%m-%d')
                    fecha_vencimiento_form = (fecha_dt + datetime.timedelta(days=365)).strftime('%Y-%m-%d')

                # 4. Actualizar renovacion
                sql_renovacion_update = """
                    UPDATE renovacion 
                    SET Prima = %s, Frecuencia = %s, cobertura = %s, Fecha_contrato = %s, Fecha_vencimiento = %s 
                    WHERE Cod_renovacion = %s
                """
                cursor.execute(sql_renovacion_update, (prima_form, frecuencia_form, cobertura_form, fecha_emision_form, fecha_vencimiento_form, cod_renovacion))
                 
                # 5. Actualizar poliza y tablas relacionadas
                if nuevo_cod_poliza and nuevo_cod_poliza != cod_poliza:
                    cursor.execute("UPDATE renovacion SET cod_poliza = %s WHERE cod_poliza = %s", (nuevo_cod_poliza, cod_poliza))
                    cursor.execute("UPDATE Fianza SET cod_poliza = %s WHERE cod_poliza = %s", (nuevo_cod_poliza, cod_poliza))
                    
                    sql_poliza_update = "UPDATE poliza SET cod_poliza = %s, Tomador = %s, Fecha_emision = %s WHERE cod_poliza = %s"
                    valores_poliza = (nuevo_cod_poliza, tomador_form, fecha_emision_form, cod_poliza)
                else:
                    sql_poliza_update = "UPDATE poliza SET Tomador = %s, Fecha_emision = %s WHERE cod_poliza = %s"
                    valores_poliza = (tomador_form, fecha_emision_form, cod_poliza)
                
                cursor.execute(sql_poliza_update, valores_poliza)

                # 6. Actualizar Fianza
                target_cod = nuevo_cod_poliza if (nuevo_cod_poliza and nuevo_cod_poliza != cod_poliza) else cod_poliza
                if subramo_form and producto_form:
                    sql_fianza_update = "UPDATE Fianza SET Subramo = %s, Producto = %s WHERE cod_poliza = %s"
                    cursor.execute(sql_fianza_update, (subramo_form, producto_form, target_cod))

                conexion_MySQLdb.commit()
                return True
    except Exception as e:
        print(f'Error en actualizar Fianza: {str(e)}')
        flash(f'Error: {str(e)}', 'error')
        return None


# Lista de Usuarios creados
def lista_usuariosBD():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                querySQL = "SELECT id, name_surname, email_user, created_user,permisos FROM users"
                cursor.execute(querySQL,)
                usuariosBD = cursor.fetchall()
        return usuariosBD
    except Exception as e:
        print(f"Error en lista_usuariosBD : {e}")
        return []
    
def get_grace_days(company_name):
    """
    Retorna los días de gracia según la compañía de seguros.
    Basado en tabla de periodos de gracia.
    """
    if not company_name:
        return 15
    name = company_name.upper()
    # 10 días
    if any(k in name for k in ['CARACAS', 'INTERNACIONAL', 'OCEANICA']):
        return 10
    # 30 días
    if any(k in name for k in ['ESTAR', 'STAR', 'MERCANTIL', 'LOYAL', 'VUMI', 'REAL', 'REGIONAL', 'VENEZUELA', 'EVER']):
        return 30
    # Por defecto 15 días (Banesco, Mapfre, Uniseguros, Qualitas, Universitas, Pirámides, etc)
    return 15

def lista_Pagos(cod):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                # Asegurar columnas nuevas en comision antes de referenciarlas
                for col_name, col_type in [("monto_pago", "FLOAT"), ("bono", "FLOAT")]:
                    try:
                        cursor.execute(f"ALTER TABLE comision ADD COLUMN {col_name} {col_type} DEFAULT NULL")
                    except:
                        pass  # Ya existe, ignorar

                querySQL = """
                    SELECT 
                        p.Cod_renovacion, p.cod_pago, p.moneda, p.fecha, r.cod_poliza, r.frecuencia, r.prima, 
                        p.monto, p.estado, p.fecha_pagada, p.nro_cuota, p.recibo, p.Metodo_pago, p.tasa,
                        co.Nombre as Nombre_compania,
                        (SELECT GROUP_CONCAT(CONCAT('PAGO: ', COALESCE(c.monto_pago, 0), '$ | RECIBIDA: ', COALESCE(c.monto_d, 0), '$ | EJEC: ', COALESCE(c.bono, 0), '$ - ', COALESCE(c.Estado, 'COBRADA')) SEPARATOR '||')
                         FROM comision c WHERE c.Cod_pago = p.cod_pago) as comisiones_info
                    FROM pago p
                    INNER JOIN renovacion r ON p.Cod_renovacion = r.Cod_renovacion
                    INNER JOIN poliza po ON r.cod_poliza = po.cod_poliza
                    INNER JOIN compania co ON po.Cod_compania = co.Cod_compania
                    WHERE r.Cod_renovacion=%s 
                    ORDER BY CAST(p.nro_cuota AS INTEGER) ASC
                """
                cursor.execute(querySQL, (cod,))
                PagosBD = cursor.fetchall()
                
                # Inyectar días de gracia
                for p in PagosBD:
                    p['dias_gracia'] = get_grace_days(p.get('Nombre_compania'))
        return PagosBD
    except Exception as e:
        print(f"Error en lista_Pagos : {e}")
        return []
    
def cobranza():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                # Asegurar columnas nuevas en comision antes de referenciarlas
                for col_name, col_type in [("monto_pago", "FLOAT"), ("bono", "FLOAT")]:
                    try:
                        cursor.execute(f"ALTER TABLE comision ADD COLUMN {col_name} {col_type} DEFAULT NULL")
                    except:
                        pass

                querySQL = """
                    SELECT 
                        co.Nombre as Nombre_compania,
                        po.cod_poliza,
                        p.Cod_renovacion, 
                        p.cod_pago, 
                        p.moneda, 
                        p.fecha,
                        r.Cod_poliza,
                        r.frecuencia,
                        r.prima,
                        p.monto,
                        p.recibo,
                        p.estado,
                        p.nro_cuota,
                        p.Metodo_pago,
                        p.fecha_pagada,
                        p.tasa,
                        a.CI,
                        a.Nombre as Nombre_asegurado,
                        a.Apellido as Apellido_asegurado,
                        (SELECT GROUP_CONCAT(CONCAT('PAGO: ', COALESCE(c.monto_pago, 0), '$ | RECIBIDA: ', COALESCE(c.monto_d, 0), '$ | EJEC: ', COALESCE(c.bono, 0), '$ - ', COALESCE(c.Estado, 'COBRADA')) SEPARATOR '||')
                         FROM comision c WHERE c.Cod_pago = p.cod_pago) as comisiones_info
                    FROM pago p
                    INNER JOIN renovacion r ON p.Cod_renovacion = r.Cod_renovacion
                    INNER JOIN poliza po ON r.cod_poliza = po.cod_poliza
                    INNER JOIN compania co ON po.Cod_compania = co.Cod_compania
                    INNER JOIN asegurado a ON po.CI_asegurado = a.CI
                    ORDER BY CAST(p.nro_cuota AS INTEGER) ASC;
                """
                cursor.execute(querySQL,)
                PagosBD = cursor.fetchall()
                for p in PagosBD:
                    p['dias_gracia'] = get_grace_days(p.get('Nombre_compania'))
        return PagosBD
    except Exception as e:
        print(f"Error en cobranza : {e}")
        return []
    
def pago_prima(cod):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                querySQL = "SELECT r.Frecuencia, r.Prima , r.comision, r.Cod_poliza,p.ramo,c.Nombre as compania FROM renovacion r join poliza p on r.cod_poliza=p.cod_poliza join compania c on p.Cod_compania=c.Cod_compania where r.Cod_renovacion = %s limit 1 "
                cursor.execute(querySQL, (cod,))

                PagosBD = cursor.fetchone()
                if PagosBD['ramo']=="Persona":
                    querySQL = "SELECT r.Frecuencia, r.Prima , r.comision, r.Cod_poliza,p.ramo,c.Nombre as compania,per.Producto,per.Subramo FROM renovacion r join poliza p on r.cod_poliza=p.cod_poliza join compania c on p.Cod_compania=c.Cod_compania join persona per on p.Cod_poliza=per.Cod_poliza where r.Cod_renovacion = %s limit 1 "
                    cursor.execute(querySQL, (cod,))
                    PagosBD = cursor.fetchone()
                    
        return PagosBD
    except Exception as e:
        print(f"Error en pago_prima : {e}")
        return []

    

    
def cod_poliza_F(cod):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                querySQL = "SELECT p.cod_poliza, r.comision,a.Ejecutivo from renovacion r join poliza p on p.Cod_poliza = r.cod_poliza join asegurado a on p.CI_asegurado=a.CI where Cod_renovacion=%s limit 1"
                cursor.execute(querySQL, (cod,))
                cod_poliza = cursor.fetchone()
        return cod_poliza
    except Exception as e:
        print(f"Error en cod_poliza_F : {e}")
        return []
    
def cod_poliza_P(cod):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                querySQL = "SELECT p.Cod_pago , e.cod_poliza FROM pago p INNER JOIN renovacion e ON p.Cod_renovacion = e.Cod_renovacion WHERE p.Cod_pago=%s"
                cursor.execute(querySQL, (cod,))
                cod_poliza = cursor.fetchall()
        return cod_poliza
    except Exception as e:
        print(f"Error en cod_poliza_F : {e}")
        return []
    
def lista_contrato(cod,a):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                if a==0:
                    querySQL = "SELECT r.Cod_renovacion, r.cod_poliza, r.Frecuencia,r.Fecha_vencimiento, c.Nombre as compania from renovacion r inner join poliza p on p.cod_poliza = r.Cod_poliza left join compania c on p.Cod_compania = c.Cod_compania where r.cod_poliza=%s order by r.Fecha_contrato desc, r.Cod_renovacion desc"
                    cursor.execute(querySQL, (cod,))
                    PagosBD = cursor.fetchall()
                else: 
                    querySQL = "SELECT r.Cod_renovacion, r.cod_poliza, r.Frecuencia,r.Fecha_vencimiento, c.Nombre as compania from renovacion r inner join poliza p on p.cod_poliza = r.Cod_poliza left join compania c on p.Cod_compania = c.Cod_compania where r.Cod_renovacion=%s order by r.Fecha_contrato desc, r.Cod_renovacion desc"
                    cursor.execute(querySQL, (cod,))
                    PagosBD = cursor.fetchone()

        return PagosBD
    except Exception as e:
        print(f"Error en lista_Contrato : {e}")
        return []


# Eliminar uEmpleado
def eliminarEmpleado(id_empleado, foto_empleado):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                querySQL = "DELETE FROM tbl_empleados WHERE id_empleado=%s"
                cursor.execute(querySQL, (id_empleado,))
                conexion_MySQLdb.commit()
                resultado_eliminar = cursor.rowcount

                if resultado_eliminar:
                    # Eliminadon foto_empleado desde el directorio
                    basepath = path.dirname(__file__)
                    url_File = path.join(
                        basepath, '../static/fotos_empleados', foto_empleado)

                    if path.exists(url_File):
                        remove(url_File)  # Borrar foto desde la carpeta

        return resultado_eliminar
    except Exception as e:
        print(f"Error en eliminarEmpleado : {e}")
        return []


# Eliminar usuario
def eliminarUsuario(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                cursor.execute("SELECT email_user FROM users WHERE id = %s", (id,))
                user_data = cursor.fetchone()

                if user_data:
                    # DB-level role management removed for PostgreSQL compatibility
                    # Only delete from users table
                    querySQL = "DELETE FROM users WHERE id=%s"
                    cursor.execute(querySQL, (id,))

                    conexion_MySQLdb.commit()
                    resultado_eliminar = cursor.rowcount
                    return resultado_eliminar
                else:
                    print(f"No se encontró el usuario con id {id} para eliminar.")
                    return 0

    except Exception as e:
        print(f"Error en eliminarUsuario : {e}")
        return []
    
def eliminarCompania(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                querySQL = "DELETE FROM compania WHERE Cod_compania=%s"
                cursor.execute(querySQL, (id,))
                conexion_MySQLdb.commit()
                resultado_eliminar = cursor.rowcount

        return resultado_eliminar
    except Exception as e:
        print(f"Error en eliminarCompania : {e}")
        return []
    
def eliminarAsegurado(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                querySQL = "DELETE FROM asegurado WHERE CI=%s"
                cursor.execute(querySQL, (id,))
                conexion_MySQLdb.commit()
                resultado_eliminar = cursor.rowcount

        return resultado_eliminar
    except Exception as e:
        print(f"Error en eliminarAsegurado : {e}")
        return []
    
def editarAsegurado(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                querySQL = "DELETE FROM asegurado WHERE CI=%s"
                cursor.execute(querySQL, (id,))
                conexion_MySQLdb.commit()
                resultado_editar = cursor.rowcount

        return resultado_editar
    except Exception as e:
        print(f"Error en eliminarAsegurado : {e}")
        return []
    
def eliminarEjecutivo(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                querySQL = "DELETE FROM ejecutivo WHERE cod_ejecutivo=%s"
                cursor.execute(querySQL, (id,))
                conexion_MySQLdb.commit()
                resultado_eliminar = cursor.rowcount

        return resultado_eliminar
    except Exception as e:
        print(f"Error en eliminarEjecutivo : {e}")
        return []
    

def eliminarPago(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                querySQL = "DELETE FROM pago WHERE Cod_pago=%s"
                cursor.execute(querySQL, (id,))
                conexion_MySQLdb.commit()
                resultado_eliminar = cursor.rowcount

        return resultado_eliminar
    except Exception as e:
        print(f"Error en eliminarPagos : {e}")
        return []
    
# Eliminar poliza
def revertirPago(id_pago):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                querySQL = """
                    UPDATE pago 
                    SET estado = 'EN PROCESO', 
                        Metodo_pago = NULL, 
                        tasa = NULL, 
                        fecha_pagada = NULL, 
                        recibo = NULL 
                    WHERE Cod_pago = %s
                """
                cursor.execute(querySQL, (id_pago,))
                conexion_MySQLdb.commit()
                resultado = cursor.rowcount
        return resultado
    except Exception as e:
        print(f"Error en revertirPago : {e}")
        return 0

def eliminarPoliza(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                cursor2 = conexion_MySQLdb.cursor()
                querySQL = "DELETE FROM poliza WHERE cod_poliza=%s"
                sql = "SELECT Ramo FROM poliza WHERE cod_poliza=%s"
                cursor2.execute(sql, (id,))
                result=cursor2.fetchone()
                Ramo = result['Ramo']
                if Ramo == "Persona":
                    sql = "DELETE FROM Persona WHERE cod_poliza=%s"
                elif Ramo == "Auto":
                    sql = "DELETE FROM Auto WHERE cod_poliza=%s"
                elif Ramo == "Patrimonial":
                    sql = "DELETE FROM Patrimonio WHERE cod_poliza=%s"
                elif Ramo == "Fianza":
                    sql = "DELETE FROM Fianza WHERE cod_poliza=%s"
                else:
                    sql = "DELETE FROM Viaje WHERE cod_poliza=%s"

                cursor2.execute(sql,(id,))
                cursor.execute(querySQL, (id,))
                conexion_MySQLdb.commit()
                resultado_eliminar = cursor.rowcount

        return resultado_eliminar
    except Exception as e:
        print(f"Error en eliminarPoliza : {e}")
        return []
    

def eliminarBeneficiario(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                print("holi")
                querySQL = "DELETE FROM beneficiario WHERE cod_poliza=%s"
                cursor.execute(querySQL, (id,))
                conexion_MySQLdb.commit()
                resultado_eliminar = cursor.rowcount

        return resultado_eliminar
    except Exception as e:
        print(f"Error en eliminarBeneficiario : {e}")
        return []

def sql_lista_company():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                querySQL = (f"""
                    SELECT
                         e.Cod_compania,
                         e.Nombre,
                         e.rif
                    FROM compania AS e
                    """)
                cursor.execute(querySQL,)
                companybd = cursor.fetchall()
                
        return companybd
    except Exception as e:
        print(
            f"Error en la función sql_lista_company: {e}")
        return None
    
def sql_lista_companyU(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                querySQL = ("""
                    SELECT
                         e.Cod_compania,
                         e.Nombre,
                         e.rif
                    FROM compania AS e
                    where Cod_compania=%s
                    """)
                cursor.execute(querySQL,(id,))
                companybd = cursor.fetchone()
                
        return companybd
    except Exception as e:
        print(
            f"Error en la función sql_lista_companyU: {e}")
        return None
    
def sql_lista_asegurado(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                querySQL = ("""
                    SELECT 
                        a.CI,
                        a.Nombre,
                        a.Nombre2, 
                        a.Apellido,
                        a.Apellido2,
                        a.Correo,
                        a.Telefono,
                        a.Fecha_nacimiento,
                        a.profesion,
                        a.localidad,
                        a.canal,
                        a.Tipo_CI,
                        a.Ejecutivo
                    FROM asegurado AS a
                    where a.CI=%s
                    """) 
                cursor.execute(querySQL,(id,))
                companybd = cursor.fetchone()
                
        return companybd
    except Exception as e:
        print(
            f"Error en la función sql_lista_company: {e}")
        return None
   
def sql_lista_ejecutivo():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                querySQL = (f"""
                    SELECT 
                         e.Nombre,
                         e.Apellido,
                         e.cod_ejecutivo
                    FROM ejecutivo e
                    """)
                cursor.execute(querySQL,)
                companybd = cursor.fetchall()
                
        return companybd if companybd is not None else []
    except Exception as e:
        print(f"Error en la función sql_lista_ejecutivo: {e}")
        return []
    
def sql_lista_ejecutivoU(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                querySQL = ("""
                    SELECT 
                         *
                    FROM ejecutivo e
                    where e.cod_ejecutivo=%s
                    """)
                cursor.execute(querySQL,(id,))
                companybd = cursor.fetchone()
                
        return companybd
    except Exception as e:
        print(
            f"Error en la función sql_lista_ejecutivou: {e}")
        return None
    
def lista_ejecutivosBD():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                querySQL = (f"""
                    SELECT 
                         *
                    FROM ejecutivo 
                    """)
                cursor.execute(querySQL,)
                companybd = cursor.fetchall()
                
        return companybd
    except Exception as e:
        print(
            f"Error en la función lista_ejecutivosBD: {e}")
        return None
    
def sql_lista_comisiones():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                querySQL = (f"""
                     SELECT
                        p.Ramo,
                        p.cod_poliza,
                        pag.Cod_pago,
                        e.Nombre,
                        e.cod_ejecutivo,
                        e.Apellido,
                        e.Tipo,
                        pag.monto,
                        a.Nombre AS N_asegu,
                        a.Apellido AS A_Asegu,
                        a.CI as ci_asegurado,
                        cr.comision,
                        cr.Prima,
                        (pag.monto * cr.comision / 100) AS comision_monto,
                        pag.fecha,
                        pag.moneda,
                        pag.tasa,
                        coe.comision_bono as bono,
                        c.Nombre as compania,
                        COALESCE(au.Subramo, per.Subramo) as subramo,
                        COALESCE(au.Producto, per.Producto) as producto
                    FROM
                        pago pag
                    JOIN
                        renovacion cr ON pag.Cod_renovacion = cr.cod_renovacion
                    JOIN
                        poliza p ON cr.cod_poliza = p.cod_poliza
                    JOIN
                        asegurado a ON p.CI_asegurado = a.CI
                    JOIN
                        ejecutivo e ON a.Ejecutivo = e.cod_ejecutivo
                    LEFT JOIN 
                        compania c ON p.Cod_compania = c.Cod_compania
                    LEFT JOIN 
                        auto au ON p.cod_poliza = au.Cod_poliza
                    LEFT JOIN 
                        Persona per ON p.cod_poliza = per.Cod_poliza
                    LEFT JOIN 
                        comisiones_ejecutivos coe ON e.cod_ejecutivo = coe.cod_ejecutivo AND p.Ramo = 'Auto'
                    WHERE
                        pag.estado = 'PAGADO'
                    """)
                cursor.execute(querySQL,)
                comisionbd = cursor.fetchall()
                
        # Calcular comisión de ejecutivo
        rules = sql_get_all_comisiones_config()
        if rules:
            for com in comisionbd:
                com['comision_ejecutivo'] = 0
                com['porcentaje_ejecutivo'] = 0
                for rule in rules:
                    # Comparación robusta
                    match_comp = str(rule.get('compania','')).strip().upper() == str(com.get('compania','')).strip().upper()
                    match_ramo = str(rule.get('ramo','')).strip().upper() == str(com.get('Ramo','')).strip().upper()
                    
                    rule_sub = str(rule.get('subramo') or '').strip().upper()
                    com_sub = str(com.get('subramo') or '').strip().upper()
                    match_sub = not rule_sub or rule_sub == com_sub
                    
                    match_prod = str(rule.get('producto','')).strip().upper() == str(com.get('producto','')).strip().upper()
                    match_tipo = str(rule.get('tipo_ejecutivo','')).strip().upper() == str(com.get('Tipo','')).strip().upper()

                    if match_comp and match_ramo and match_sub and match_prod and match_tipo:
                        porcentajes_str = rule.get('porcentajes')
                        if porcentajes_str:
                            try:
                                parts = [p.strip() for p in str(porcentajes_str).split(',')]
                                if parts:
                                    pct = float(parts[0])
                                    com['porcentaje_ejecutivo'] = pct
                                    monto_fraccion = float(com.get('monto') or 0)
                                    com['comision_ejecutivo'] = (monto_fraccion * pct)
                            except Exception as e:
                                print(f"Error parseando porcentajes: {e}")
                        break
        
        return comisionbd
    except Exception as e:
        import traceback
        print(f"Error en la función sql_lista_comisiones: {e}")
        traceback.print_exc()
        return None





def obtener_pagos_filtrados(pago, comision, mes=None, asegurado_id=None, ano=None, ejecutivo_id=None, fecha_inicio=None, fecha_fin=None, estado_filtro=None):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as mycursor:
                sub_query_proximo = """
                    NOT EXISTS (
                        SELECT 1 FROM pago p2 
                        WHERE p2.Cod_renovacion = p.Cod_renovacion 
                        AND p2.estado NOT IN ('PAGADO', 'ANULADO')
                        AND (
                            p2.nro_cuota < p.nro_cuota 
                            OR (p2.nro_cuota IS NULL AND (p2.fecha < p.fecha OR (p2.fecha = p.fecha AND p2.cod_pago < p.cod_pago)))
                            OR (p2.nro_cuota = p.nro_cuota AND p2.cod_pago < p.cod_pago)
                        )
                    )
                """

                base_query = """
                    SELECT 
                        p.Cod_renovacion, p.cod_pago, p.moneda, p.fecha, r.Cod_poliza, r.frecuencia, 
                        r.prima, p.monto, p.estado, po.CI_asegurado, p.nro_cuota, p.recibo,
                        p.Metodo_pago, p.fecha_pagada, p.tasa,
                        a.Nombre as Nombre_asegurado, a.Apellido as Apellido_asegurado, a.CI,
                        co.Nombre as Nombre_compania
                    FROM pago p
                    INNER JOIN renovacion r ON p.Cod_renovacion = r.Cod_renovacion
                    INNER JOIN poliza po ON r.cod_poliza = po.cod_poliza
                    INNER JOIN asegurado a ON po.CI_asegurado = a.CI
                    INNER JOIN compania co ON po.Cod_compania = co.Cod_compania
                """
                
                conditions = []
                params = []

                if mes:
                    conditions.append("EXTRACT(MONTH FROM p.fecha) = %s")
                    params.append(mes)
                
                if ano:
                    conditions.append("EXTRACT(YEAR FROM p.fecha) = %s")
                    params.append(ano)

                if asegurado_id:
                    conditions.append("po.CI_asegurado = %s")
                    params.append(asegurado_id)
                
                if ejecutivo_id:
                    conditions.append("a.Ejecutivo = %s")
                    params.append(ejecutivo_id)

                if fecha_inicio:
                    conditions.append("p.fecha >= %s")
                    params.append(fecha_inicio)
                
                if fecha_fin:
                    conditions.append("p.fecha <= %s")
                    params.append(fecha_fin)

                if estado_filtro == 'PAGADO':
                    conditions.append("p.estado = 'PAGADO'")
                elif estado_filtro == 'DISPONIBLE':
                    conditions.append(f"p.estado NOT IN ('PAGADO', 'ANULADO') AND {sub_query_proximo}")
                elif estado_filtro == 'PENDIENTE':
                    conditions.append(f"p.estado NOT IN ('PAGADO', 'ANULADO') AND NOT ({sub_query_proximo})")

                if conditions:
                    base_query += " WHERE " + " AND ".join(conditions)
                
                base_query += " ORDER BY CAST(p.nro_cuota AS INTEGER) ASC"
                
                mycursor.execute(base_query, tuple(params))
                resultado_busqueda = mycursor.fetchall()
                return resultado_busqueda

    except Exception as e:
        print(f"Ocurrió un error en obtener pagos filtrados: {e}")
        return []

def obtener_pagos_datatable(start, length, mes=None, ano=None, asegurado_id=None, estado_filtro=None):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as mycursor:
                # Asegurar columnas nuevas en comision antes de referenciarlas
                for col_name, col_type in [("monto_pago", "FLOAT"), ("bono", "FLOAT")]:
                    try:
                        mycursor.execute(f"ALTER TABLE comision ADD COLUMN {col_name} {col_type} DEFAULT NULL")
                    except:
                        pass

                # Subconsulta para determinar si un pago es el siguiente "EN PROCESO" (Disponible para cobrar)
                sub_query_proximo = """
                    NOT EXISTS (
                        SELECT 1 FROM pago p2 
                        WHERE p2.Cod_renovacion = p.Cod_renovacion 
                        AND p2.estado NOT IN ('PAGADO', 'ANULADO')
                        AND (
                            p2.nro_cuota < p.nro_cuota 
                            OR (p2.nro_cuota IS NULL AND (p2.fecha < p.fecha OR (p2.fecha = p.fecha AND p2.cod_pago < p.cod_pago)))
                            OR (p2.nro_cuota = p.nro_cuota AND p2.cod_pago < p.cod_pago)
                        )
                    )
                """

                base_joins = """
                    FROM pago p
                    INNER JOIN renovacion r ON p.Cod_renovacion = r.Cod_renovacion
                    INNER JOIN poliza po ON r.cod_poliza = po.cod_poliza
                    INNER JOIN asegurado a ON po.CI_asegurado = a.CI
                    INNER JOIN compania co ON po.Cod_compania = co.Cod_compania
                """
                
                conditions = []
                params = []

                if mes:
                    conditions.append("EXTRACT(MONTH FROM p.fecha) = %s")
                    params.append(mes)
                
                if ano:
                    conditions.append("EXTRACT(YEAR FROM p.fecha) = %s")
                    params.append(ano)

                if asegurado_id:
                    conditions.append("po.CI_asegurado = %s")
                    params.append(asegurado_id)
                
                if estado_filtro == 'PAGADO':
                    conditions.append("p.estado = 'PAGADO'")
                elif estado_filtro == 'DISPONIBLE':
                    conditions.append(f"p.estado NOT IN ('PAGADO', 'ANULADO') AND {sub_query_proximo}")
                elif estado_filtro == 'PENDIENTE':
                    conditions.append(f"p.estado NOT IN ('PAGADO', 'ANULADO') AND NOT ({sub_query_proximo})")
                elif estado_filtro == 'GRACIA':
                    # Pagos EN PROCESO cuya fecha ya venció pero aún están dentro del periodo de gracia
                    # El número de días de gracia depende de la compañía (replicamos la lógica de get_grace_days en SQL)
                    grace_case = """
                        CASE
                            WHEN UPPER(co.Nombre) LIKE '%CARACAS%' OR UPPER(co.Nombre) LIKE '%INTERNACIONAL%' OR UPPER(co.Nombre) LIKE '%OCEANICA%' THEN 10
                            WHEN UPPER(co.Nombre) LIKE '%ESTAR%' OR UPPER(co.Nombre) LIKE '%MERCANTIL%' OR UPPER(co.Nombre) LIKE '%LOYAL%'
                              OR UPPER(co.Nombre) LIKE '%VUMI%' OR UPPER(co.Nombre) LIKE '%REAL%' OR UPPER(co.Nombre) LIKE '%REGIONAL%'
                              OR UPPER(co.Nombre) LIKE '%VENEZUELA%' OR UPPER(co.Nombre) LIKE '%EVER%' THEN 30
                            ELSE 15
                        END
                    """
                    conditions.append(
                        f"p.estado NOT IN ('PAGADO', 'ANULADO')"
                        f" AND p.fecha < CURRENT_DATE"
                        f" AND DATE_ADD(p.fecha, INTERVAL ({grace_case}) DAY) >= CURRENT_DATE"
                    )

                where_clause = ""
                if conditions:
                    where_clause = " WHERE " + " AND ".join(conditions)

                # Total records (without filters, but keeping company joins if needed)
                mycursor.execute("SELECT COUNT(p.cod_pago) as total " + base_joins)
                recordsTotal = mycursor.fetchone()['total']

                # Filtered records count
                if conditions:
                    mycursor.execute("SELECT COUNT(p.cod_pago) as total " + base_joins + where_clause, tuple(params))
                    recordsFiltered = mycursor.fetchone()['total']
                else:
                    recordsFiltered = recordsTotal

                # Get data with limits, including 'es_proximo' to help frontend button disabling
                data_query = f"""
                    SELECT 
                        co.Nombre as Nombre_compania, po.cod_poliza, p.Cod_renovacion, p.cod_pago, p.moneda, p.fecha, 
                        r.Cod_poliza as r_Cod_poliza, r.frecuencia, r.prima, p.monto, p.estado, po.CI_asegurado, p.nro_cuota, p.recibo,
                        p.Metodo_pago, p.fecha_pagada, p.tasa,
                        a.Nombre as Nombre_asegurado, a.Apellido as Apellido_asegurado, a.CI,
                        ({sub_query_proximo}) as es_proximo,
                        (SELECT GROUP_CONCAT(CONCAT(COALESCE(c.descripcion, 'COMISION'), ' (', COALESCE(c.monto_d, 0), '$) - ', COALESCE(c.Estado, 'COBRADA')) SEPARATOR '||')
                         FROM comision c WHERE c.Cod_pago = p.cod_pago) as comisiones_info
                """ + base_joins + where_clause + """
                    ORDER BY p.fecha DESC
                    LIMIT %s OFFSET %s
                """
                
                params_limit = params.copy()
                params_limit.append(int(length))
                params_limit.append(int(start))

                mycursor.execute(data_query, tuple(params_limit))
                data = mycursor.fetchall()
                
                for p in data:
                    p['dias_gracia'] = get_grace_days(p.get('Nombre_compania'))
                
                return {
                    "recordsTotal": recordsTotal,
                    "recordsFiltered": recordsFiltered,
                    "data": data
                }

    except Exception as e:
        print(f"Ocurrió un error en obtener_pagos_datatable: {e}")
        return {"recordsTotal": 0, "recordsFiltered": 0, "data": []}

def obtener_polizas_datatable(start, length, tipo_filtro_fecha=None, fecha=None, anio=None, fecha_inicio=None, fecha_fin=None, estados=None, compania_id=None, ejecutivo_id=None, search_value=None):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as mycursor:
                # Subconsulta para Tiene_pago
                tiene_pago_sql = "(SELECT COUNT(*) FROM pago pay2 WHERE pay2.Cod_renovacion = r.Cod_renovacion) > 0"
                
                # Logic for estado_calculado
                estado_sql = f"""
                    CASE
                        WHEN r.estado = 'anulada' THEN 'Anulada'
                        WHEN r.estado = 'traspasada' THEN 'Traspasada'
                        WHEN r.Fecha_vencimiento <= CURRENT_DATE AND p.Ramo IN ('Persona', 'PERSONA', 'PERSONAS', 'Auto', 'AUTO', 'Patrimonial', 'PATRIMONIAL', 'Fianza', 'FIANZA', 'FIANZAS', 'Viaje', 'VIAJE', 'VIAJES') 
                             AND p.Ramo IN ('Fianza', 'FIANZA', 'FIANZAS', 'Viaje', 'VIAJE', 'VIAJES') THEN 'Finalizada'
                        WHEN r.Fecha_vencimiento > CURRENT_DATE AND {tiene_pago_sql} THEN 'Vigente'
                        WHEN r.Fecha_vencimiento > CURRENT_DATE AND NOT ({tiene_pago_sql}) THEN 'Pendiente'
                        ELSE 'Vencida'
                    END
                """

                base_from = """
                    FROM poliza p
                    INNER JOIN (
                        SELECT r1.*
                        FROM renovacion r1
                        INNER JOIN (
                            SELECT cod_poliza, MAX(Cod_renovacion) as max_id
                            FROM renovacion
                            GROUP BY cod_poliza
                        ) r2 ON r1.cod_poliza = r2.cod_poliza AND r1.Cod_renovacion = r2.max_id
                    ) r ON p.cod_poliza = r.cod_poliza
                    JOIN asegurado a ON p.CI_asegurado = a.CI
                    LEFT JOIN ejecutivo e ON a.Ejecutivo = e.cod_ejecutivo
                    LEFT JOIN compania c ON p.Cod_compania = c.Cod_compania
                """

                # Joins adicionales solo para la data (no para el conteo a menos que se filtre por producto)
                data_joins = """
                    LEFT JOIN Persona pe ON p.cod_poliza = pe.Cod_poliza AND p.Ramo IN ('Persona', 'PERSONA', 'PERSONAS')
                    LEFT JOIN Auto au ON p.cod_poliza = au.Cod_poliza AND p.Ramo IN ('Auto', 'AUTO')
                    LEFT JOIN Patrimonio pa ON p.cod_poliza = pa.Cod_poliza AND p.Ramo IN ('Patrimonial', 'PATRIMONIAL')
                    LEFT JOIN Fianza f ON p.cod_poliza = f.Cod_poliza AND p.Ramo IN ('Fianza', 'FIANZA', 'FIANZAS')
                    LEFT JOIN Viaje v ON p.cod_poliza = v.Cod_poliza AND p.Ramo IN ('Viaje', 'VIAJE', 'VIAJES')
                """

                conditions = []
                params = []

                # Búsqueda global (DataTables search bar)
                if search_value:
                    search_condition = """
                        (p.cod_poliza LIKE %s OR 
                         a.Nombre LIKE %s OR 
                         a.Apellido LIKE %s OR 
                         CAST(p.CI_asegurado AS CHAR) LIKE %s OR 
                         c.nombre LIKE %s OR 
                         p.Ramo LIKE %s)
                    """
                    conditions.append(search_condition)
                    val = f"%{search_value}%"
                    params.extend([val, val, val, val, val, val])

                # Filtros de fecha
                if tipo_filtro_fecha == 'mes' and fecha:
                    try:
                        start_date = f"{fecha}-01"
                        conditions.append("r.Fecha_contrato BETWEEN %s AND LAST_DAY(%s)")
                        params.extend([start_date, start_date])
                    except: pass
                elif tipo_filtro_fecha == 'anio' and anio:
                    conditions.append("r.Fecha_contrato BETWEEN %s AND %s")
                    params.extend([f"{anio}-01-01", f"{anio}-12-31"])
                elif tipo_filtro_fecha == 'rango' and fecha_inicio and fecha_fin:
                    conditions.append("r.Fecha_contrato BETWEEN %s AND %s")
                    params.extend([fecha_inicio, fecha_fin])

                if compania_id and compania_id != "":
                    conditions.append("p.Cod_compania = %s")
                    params.append(compania_id)
                
                if ejecutivo_id and ejecutivo_id != "":
                    conditions.append("a.Ejecutivo = %s")
                    params.append(ejecutivo_id)

                if estados and isinstance(estados, list) and len(estados) > 0 and '' not in estados:
                    placeholders = ', '.join(['%s'] * len(estados))
                    conditions.append(f"({estado_sql}) IN ({placeholders})")
                    params.extend(estados)

                where_clause = " WHERE 1=1 "
                if conditions:
                    where_clause += " AND " + " AND ".join(conditions)

                # Total records (siempre rápido)
                mycursor.execute("SELECT COUNT(*) as total FROM poliza")
                recordsTotal = mycursor.fetchone()['total']

                # Filtered records count (solo si hay filtros, si no es igual al total)
                if not conditions:
                    recordsFiltered = recordsTotal
                else:
                    count_query = "SELECT COUNT(*) as total " + base_from + where_clause
                    mycursor.execute(count_query, tuple(params))
                    recordsFiltered = mycursor.fetchone()['total']

                # Get data (solo las columnas necesarias para la tabla)
                data_query = f"""
                    SELECT
                        p.cod_poliza, p.CI_asegurado, a.Tipo_CI, a.Nombre, a.Apellido,
                        DATE_FORMAT(r.Fecha_contrato, '%Y-%m-%d') AS fecha,
                        p.Ramo, r.riesgo, r.Fecha_vencimiento, r.estado, r.Cod_renovacion,
                        c.nombre as compania_nombre,
                        e.Nombre as ejecutivo_nombre, e.Apellido as ejecutivo_apellido,
                        CASE
                            WHEN p.Ramo IN ('Persona', 'PERSONA', 'PERSONAS') THEN pe.Producto
                            WHEN p.Ramo IN ('Auto', 'AUTO') THEN au.Producto
                            WHEN p.Ramo IN ('Patrimonial', 'PATRIMONIAL') THEN pa.Producto
                            WHEN p.Ramo IN ('Fianza', 'FIANZA', 'FIANZAS') THEN f.Producto
                            WHEN p.Ramo IN ('Viaje', 'VIAJE', 'VIAJES') THEN v.Producto
                            ELSE ''
                        END as producto_nombre,
                        IF({tiene_pago_sql}, 'Sí', 'No') AS Tiene_pago,
                        ({estado_sql}) as estado_calculado
                    {base_from}
                    {data_joins}
                    {where_clause}
                    ORDER BY r.Fecha_contrato DESC
                    LIMIT %s OFFSET %s
                """
                
                params_limit = params + [int(length), int(start)]
                mycursor.execute(data_query, tuple(params_limit))
                data = mycursor.fetchall()

                return {
                    "recordsTotal": recordsTotal,
                    "recordsFiltered": recordsFiltered,
                    "data": data
                }
    except Exception as e:
        print(f"Error en obtener_polizas_datatable: {e}")
        return {"recordsTotal": 0, "recordsFiltered": 0, "data": []}

def obtener_comisiones_datatable(start, length, minDateStr=None, maxDateStr=None, ejecutivo_id=None):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as mycursor:
                base_joins = """
                    FROM pago pag
                    JOIN renovacion cr ON pag.Cod_renovacion = cr.cod_renovacion
                    JOIN poliza p ON cr.cod_poliza = p.cod_poliza
                    JOIN asegurado a ON p.CI_asegurado = a.CI
                    JOIN ejecutivo e ON a.Ejecutivo = e.cod_ejecutivo
                    LEFT JOIN compania c ON p.Cod_compania = c.Cod_compania
                    LEFT JOIN auto au ON p.cod_poliza = au.Cod_poliza
                    LEFT JOIN Persona per ON p.cod_poliza = per.Cod_poliza
                    LEFT JOIN comisiones_ejecutivos coe ON e.cod_ejecutivo = coe.cod_ejecutivo AND p.Ramo = 'Auto'
                """
                
                conditions = ["pag.estado = 'PAGADO'"]
                params = []

                if minDateStr and maxDateStr:
                    conditions.append("pag.fecha >= %s AND pag.fecha <= %s")
                    params.extend([minDateStr, maxDateStr])
                
                if ejecutivo_id:
                    conditions.append("e.cod_ejecutivo = %s")
                    params.append(ejecutivo_id)
                
                where_clause = " WHERE " + " AND ".join(conditions)

                # Total records (only 'PAGADO' filter)
                mycursor.execute("SELECT COUNT(pag.cod_pago) as total " + base_joins + " WHERE pag.estado = 'PAGADO'")
                recordsTotal = mycursor.fetchone()['total']

                # Filtered records count
                mycursor.execute("SELECT COUNT(pag.cod_pago) as total " + base_joins + where_clause, tuple(params))
                recordsFiltered = mycursor.fetchone()['total']

                # Get data with limits
                data_query = """
                     SELECT
                        p.Ramo,
                        p.cod_poliza,
                        pag.Cod_pago,
                        e.Nombre,
                        e.cod_ejecutivo,
                        e.Apellido,
                        e.Tipo,
                        pag.monto,
                        a.Nombre AS N_asegu,
                        a.Apellido AS A_Asegu,
                        a.CI as ci_asegurado,
                        cr.comision,
                        cr.Prima,
                        (pag.monto * cr.comision / 100) AS comision_monto,
                        pag.fecha,
                        pag.moneda,
                        pag.tasa,
                        coe.comision_bono as bono,
                        c.Nombre as compania,
                        COALESCE(au.Subramo, per.Subramo) as subramo,
                        COALESCE(au.Producto, per.Producto) as producto
                """ + base_joins + where_clause + """
                    ORDER BY pag.fecha DESC
                    LIMIT %s OFFSET %s
                """
                
                params_limit = params.copy()
                params_limit.append(int(length))
                params_limit.append(int(start))

                mycursor.execute(data_query, tuple(params_limit))
                data = mycursor.fetchall()
                
        # Calcular comisión de ejecutivo con prioridades (Específica > Tipo > Fallback Agente)
        rules = sql_get_all_comisiones_config()
        if rules:
            for com in data:
                com['comision_ejecutivo'] = 0
                com['porcentaje_ejecutivo'] = 0
                com['es_fallback'] = False
                com['has_rule'] = False
                
                exec_type = str(com.get('Tipo') or '').strip().upper()
                exec_cod = com.get('cod_ejecutivo')
                
                # Lista de tipos a buscar (incluyendo fallback a AGENTE para personalizados)
                search_types = [exec_type]
                if exec_type == 'PERSONALIZADO': search_types.append('AGENTE')

                found_rule = None
                for t_search in search_types:
                    # Buscamos en las reglas cargadas
                    for rule in rules:
                        r_cod = rule.get('cod_ejecutivo')
                        r_type = str(rule.get('tipo_ejecutivo','')).strip().upper()
                        
                        # Coincidencia de producto (Compañía, Ramo, Subramo, Producto)
                        r_comp = str(rule.get('compania','') or '').strip().upper()
                        c_comp = str(com.get('compania','') or '').strip().upper()
                        
                        r_ramo = str(rule.get('ramo','') or '').strip().upper()
                        c_ramo = str(com.get('Ramo','') or '').strip().upper()
                        
                        r_sub = str(rule.get('subramo') or '').strip().upper()
                        c_sub = str(com.get('subramo') or '').strip().upper()
                        
                        r_prod = str(rule.get('producto','') or '').strip().upper()
                        c_prod = str(com.get('producto','') or '').strip().upper()

                        # Si la regla tiene un campo vacío, se considera que aplica a todo (comodín)
                        # a menos que sea Compañía o Ramo que deben coincidir.
                        if r_comp == c_comp and r_ramo == c_ramo:
                            match_sub = not r_sub or r_sub == c_sub
                            match_prod = not r_prod or r_prod == c_prod
                            
                            if match_sub and match_prod:
                                # Prioridad A: Código de ejecutivo exacto (si la regla tiene ID)
                                if r_cod and exec_cod and str(r_cod) == str(exec_cod):
                                    found_rule = rule
                                    com['has_rule'] = True
                                    break
                                
                                # Prioridad B: Match por Tipo (o fallback Agente) si no hay ID en la regla
                                if (not r_cod or str(r_cod) == '0' or r_cod == '') and r_type == t_search:
                                    found_rule = rule
                                    com['has_rule'] = True
                                    if t_search == 'AGENTE' and exec_type == 'PERSONALIZADO':
                                        com['es_fallback'] = True
                                    break
                        
                    if found_rule: break
                
                if found_rule:
                    pct_str = found_rule.get('porcentajes')
                    if pct_str:
                        try:
                            parts = [p.strip() for p in str(pct_str).split(',')]
                            pct = float(parts[0])
                            com['porcentaje_ejecutivo'] = pct * 100 # Mostrar como entero (8.5%)
                            com['comision_ejecutivo'] = float(com.get('monto') or 0) * pct
                        except Exception as e:
                            print(f"Error parseando porcentajes: {e}")

        return {
            "recordsTotal": recordsTotal,
            "recordsFiltered": recordsFiltered,
            "data": data
        }

    except Exception as e:
        print(f"Ocurrió un error en obtener_comisiones_datatable: {e}")
        return {"recordsTotal": 0, "recordsFiltered": 0, "data": []}

def sql_lista_asegurados_para_filtro():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                querySQL = "SELECT CI, Nombre, Apellido FROM asegurado ORDER BY Nombre, Apellido"
                cursor.execute(querySQL)
                asegurados = cursor.fetchall()
        return asegurados
    except Exception as e:
        print(f"Error en la función sql_lista_asegurados_para_filtro: {e}")
        return None


def obtener_proyeccion_cobranza(year):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                querySQL = """
                    SELECT EXTRACT(MONTH FROM fecha) as mes, SUM(monto) as total
                    FROM pago
                    WHERE EXTRACT(YEAR FROM fecha) = %s
                    GROUP BY EXTRACT(MONTH FROM fecha)
                    ORDER BY EXTRACT(MONTH FROM fecha)
                """
                cursor.execute(querySQL, (year,))
                proyeccion = cursor.fetchall()

                # Crear un diccionario con los totales por mes
                proyeccion_dict = {item['mes']: item['total'] for item in proyeccion}

                # Crear la lista final con todos los meses
                meses = [
                    "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
                    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
                ]
                resultado = []
                for i, nombre_mes in enumerate(meses):
                    mes_num = i + 1
                    total = proyeccion_dict.get(mes_num, 0)
                    resultado.append({'mes': nombre_mes, 'total': total})

                return resultado
    except Exception as e:
        print(f"Error en la función obtener_proyeccion_cobranza: {e}")
        return []

    
def NotaCartaAval(titulo, observaciones, id,cod,tipo,fecha):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                if tipo == 1:
                    querySQL = "insert into nota_cartaAval (Cod_CartaAval,Observaciones,titulo,fecha) values (%s,%s,%s,%s)"
                elif tipo == 2:
                    querySQL = "insert into nota_Reembolso (Cod_Reembolso,Observaciones,titulo,fecha) values (%s,%s,%s,%s)"
                else:
                    querySQL = "insert into nota_Auto (Cod_Auto,Observaciones,titulo,fecha) values (%s,%s,%s,%s)"
                print("probando")
                cursor.execute(querySQL, (cod,observaciones,titulo,fecha))
                print("probando 2")
                nuevo_id = cursor.lastrowid
                conexion_MySQLdb.commit()
                resultado = nuevo_id

        return resultado
    except Exception as e:
        print(f"Error en NotaCartaAval : {e}")
        return []
    
def validarDataRegisterLogin1(name_surname, email_user, pass_user):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                querySQL = "SELECT * FROM users WHERE email_user = %s"
                cursor.execute(querySQL, (email_user,))
                userBD = cursor.fetchone()  # Obtener la primera fila de resultados

                if userBD is not None:
                    flash('el registro no fue procesado ya existe la cuenta', 'error')
                    return False
                elif not re.match(r'[^@]+@[^@]+\.[^@]+', email_user):
                    flash('el Correo es invalido', 'error')
                    return False
                elif not name_surname or not email_user or not pass_user:
                    flash('por favor llene los campos del formulario.', 'error')
                    return False
                else:
                    # La cuenta no existe y los datos del formulario son válidos, puedo realizar el Insert
                    return True
    except Exception as e:
        print(f"Error en validarDataRegisterLogin : {e}")
        return []





def recibeInsertRegisterUser1(name_surname, email_user, pass_user, permisos):
    """
    Registers a new user in the users table.
    DB-level role management removed for PostgreSQL compatibility.
    User data is managed via the users table only.
    """

    print("alo")
    respuestaValidar = validarDataRegisterLogin1(name_surname, email_user, pass_user)

    if not respuestaValidar:
        return {"success": False, "message": "Datos de registro inválidos."}

    nueva_password = generate_password_hash(pass_user, method='scrypt')
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as mycursor:
                # 1. Verificar si el usuario ya existe en la tabla 'users'
                print("prueba1")
                sql_check_app_user = "SELECT COUNT(*) AS count FROM users WHERE email_user = %s"
                mycursor.execute(sql_check_app_user, (email_user,))
                user_exists_in_app = mycursor.fetchone()['count']

                if user_exists_in_app > 0:
                    return {"success": False, "message": "Ya existe un usuario registrado con este correo electrónico."}

                # DB-level role management removed for PostgreSQL compatibility
                # User data is managed via the users table only

                # 2. Insertar el usuario en la tabla 'users' de la aplicación
                print("prueba")
                sql_insert_user = "INSERT INTO users(name_surname, email_user, pass_user, permisos) VALUES (%s, %s, %s, %s)"
                valores_insert_user = (name_surname, email_user, nueva_password, permisos)
                mycursor.execute(sql_insert_user, valores_insert_user)

                conexion_MySQLdb.commit()
                return {"success": True, "message": "Usuario registrado exitosamente."}

    except Exception as e:
        print(f"Error al crear usuario o asignar permisos: {e}")
        # En caso de error, el commit no se ejecuta, por lo que no hay cambios permanentes.
        return {"success": False, "message": f"Error interno al registrar el usuario: {e}"}



def separar_nombre_completo(nombre_completo):
    """
    Separa un nombre completo en primer apellido, segundo apellido, primer nombre y segundo nombre.
    Asume el formato: Primer Apellido, Segundo Apellido, Primer Nombre, Segundo Nombre.
    Maneja hasta 4 partes. Si hay menos, las partes restantes se devuelven como None.
    """
    partes = str(nombre_completo).strip().split()
    
    primer_apellido = None
    segundo_apellido = None
    primer_nombre = None
    segundo_nombre = None

    num_partes = len(partes)

    if num_partes >= 1:
        primer_apellido = partes[0]
    if num_partes >= 2:
        segundo_apellido = partes[1]
    if num_partes >= 3:
        primer_nombre = partes[2]
    if num_partes >= 4:
        segundo_nombre = partes[3]
    
    # Manejo de casos específicos para menos de 4 partes para una mayor robustez
    if num_partes == 1:
        # Solo Primer Apellido
        pass # Ya asignado
    elif num_partes == 2:
        # Primer Apellido, Segundo Apellido
        pass # Ya asignados
    elif num_partes == 3:
        # Primer Apellido, Segundo Apellido, Primer Nombre
        pass # Ya asignados
    # Si hay más de 4 partes, las primeras 4 se asignan según el orden.
    # El resto de las partes se ignoran por esta función simple.

    return primer_nombre, segundo_nombre, primer_apellido, segundo_apellido




def cargar_riesgo(cod_renovacion):  
    try:
        print(f"--> [DEBUG] Iniciando carga_riesgo_logica para póliza: {cod_renovacion}")

        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:

                sql = """
                    SELECT
                        p.Ramo,
                        p.cod_poliza,
                        r.Prima,
                        pag.Metodo_pago,
                        a.profesion,
                        a.localidad,
                        a.canal,
                        a.Tipo_CI
                    FROM
                        renovacion r
                    JOIN
                        poliza p ON r.Cod_poliza = p.cod_poliza
                    JOIN
                        asegurado a ON a.CI = p.CI_asegurado
                    LEFT JOIN
                        pago pag ON r.Cod_renovacion = pag.Cod_renovacion
                    WHERE
                        r.cod_renovacion = %s
                    """
                print(cod_renovacion)
                cursor.execute(sql, (cod_renovacion,))
                resultado = cursor.fetchone()
                cursor.fetchall()

                riesgo = 0.0 # Inicializar riesgo como float

                if resultado:
                    ramo = resultado.get('Ramo')
                    prima = resultado.get('Prima')
                    metodo = resultado.get('Metodo_pago')
                    profesion = resultado.get('profesion')
                    localidad = resultado.get('localidad')
                    canal = resultado.get('canal')
                    tipo_ci = resultado.get('Tipo_CI')

                    # Verificar si algún dato necesario es nulo
                    datos_necesarios = {
                        'Ramo': ramo, 'Prima': prima, 'Metodo_pago': metodo,
                        'profesion': profesion, 'localidad': localidad,
                        'canal': canal, 'Tipo_CI': tipo_ci
                    }
                    
                    nulos_encontrados = [k for k, v in datos_necesarios.items() if v is None or str(v).strip() == '']

                    if nulos_encontrados:
                        print(f"¡Atención! Datos nulos o vacíos para {cod_renovacion}: {', '.join(nulos_encontrados)}")
                        return False, f'No se pudo calcular el riesgo: los siguientes campos son nulos: {", ".join(nulos_encontrados)}.', 400
                    else:
                        print("Todos los datos necesarios para esta renovación están completos. ¡Procediendo con la acción!")

                        try:
                            primaF = float(prima)
                        except ValueError:
                            print(f"Error: El valor de prima '{prima}' no se pudo convertir a float. Asegúrate de que los datos en la base de datos son numéricos.")
                            return False, 'El valor de la prima no es un número válido.', 400

                        if tipo_ci == "V":
                            riesgo+=0.1
                        elif tipo_ci == "J":
                            riesgo+=0.2
                        else:
                            riesgo+=0.3

                        if profesion=="BR":
                            riesgo+=0.25
                        elif profesion== "MR":
                            riesgo+=0.5
                        else:
                            riesgo+=0.75

                        if primaF < 5000:
                            riesgo+=0.15
                        elif primaF < 12000:
                            riesgo+=0.30
                        else:
                            riesgo+=0.45

                        if canal=="Cara a cara":
                            riesgo+=0.1
                        elif canal=="Internet" or canal=="Correos con dominios privados":
                            riesgo+=0.2
                        else:
                            riesgo+=0.3

                        if metodo=="transferencia" or metodo=="pago movil":
                            riesgo+=0.1
                        elif metodo=="Efectivo":
                            riesgo+=0.2
                        else:
                            riesgo+=0.3

                        estados_alto_riesgo = [
                            "Amazonas", "Apure", "Barinas", "Bolívar", "Delta Amacuro",
                            "Falcón", "Mérida", "Táchira", "Zulia"
                        ]
                        if localidad=="distrito capital" or localidad=="Carabobo" or localidad=="aragua" or localidad=="la guaira" or localidad=="miranda":
                            riesgo+=0.1
                        elif localidad in estados_alto_riesgo:
                            riesgo+=0.3
                        else:
                            riesgo+=0.2

                        if ramo=="patrimonio":
                            riesgo+=0.4
                        elif ramo=="fianza":
                            riesgo+=0.6
                        else:
                            riesgo+=0.2
                        
                        print(riesgo)
                        riesgo=round(riesgo,2)
                        print(riesgo)
                        print(cod_renovacion)

                        # Asegúrate de que el riesgo final esté dentro de un rango razonable si es necesario
                        # Por ejemplo, si el riesgo no debe exceder 3.00:
                        # riesgo = min(riesgo, 3.00)
                        with conexion_MySQLdb.cursor() as cursor_update:
                            querySQL = ("""
                                UPDATE renovacion
                                SET
                                riesgo=%s
                                WHERE Cod_renovacion=%s
                                """)
                            
                            cursor_update.execute(querySQL,(riesgo, cod_renovacion)) # Usar cod_renovacion aquí
                            print("hey")
                            conexion_MySQLdb.commit()
                            print("hey")
                            print(f"--> [DEBUG] Riesgo actualizado para {cod_renovacion}: {riesgo}")
                            return True, riesgo,200 # Retorna éxito y el valor del riesgo

                else:
                    print(f"No se encontró ninguna renovación con el código {cod_renovacion}.")
                    return False, 'No se encontró la póliza para recalcular el riesgo.', 404

    except Exception as e:
        print(f"Error en la función carga_riesgo_logica: {e}")
        return False, f'Error interno del servidor: {str(e)}', 500
    



def generar_pdf_comisiones(comisiones, rangoFechas=None, ejecutivo=None, nota=None, for_preview=False):
    """
    Genera un informe de comisiones en formato PDF.

    Args:
        comisiones (list): Lista de diccionarios con los datos de las comisiones.
        rangoFechas (str, opcional): Rango de fechas filtrado.
        ejecutivo (str, opcional): Ejecutivo filtrado.
        nota (str, opcional): Una nota para agregar al final del reporte.
        for_preview (bool): Si es True, no establece la cabecera de descarga.

    Returns:
        BytesIO: Un buffer de bytes que contiene el contenido del PDF.
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
    flowables = []
    styles = getSampleStyleSheet()
    style_title = styles['h1']
    style_h2 = styles['h2']
    style_body_center = styles['BodyText']
    style_body_center.alignment = 1 # Center alignment
    style_body_left = styles['BodyText']
    style_body_left.alignment = 0 # Left alignment

    # --- 1. Añadir el logo ---
    try:
        logo_path = os.path.join(os.path.dirname(__file__), '..', 'static', 'assets', 'img', 'cabal logo cambio.png')
        if os.path.exists(logo_path):
            logo = Image(logo_path, width=1.5*inch, height=0.75*inch)
            logo.hAlign = 'LEFT'
            flowables.append(logo)
            flowables.append(Spacer(1, 12))
        else:
            print(f"Advertencia: No se encontró el logo en la ruta: {logo_path}")
    except Exception as e:
        print(f"Error al cargar el logo: {e}")

    # --- 2. Título principal ---
    titulo_principal = "Reporte de Comisiones"
    if rangoFechas:
        subtitulo_principal = f"Período: {rangoFechas}"
    else:
        subtitulo_principal = f"Fecha de Reporte: {datetime.date.today().strftime('%d/%m/%Y')}"

    flowables.append(Paragraph(titulo_principal, style_title))
    flowables.append(Paragraph(subtitulo_principal, style_body_center))
    flowables.append(Spacer(1, 12))

    # --- 3. Agrupar comisiones por ejecutivo ---
    comisiones_por_ejecutivo = {}
    for comision in comisiones:
        nombre_ejecutivo = f"{comision.get('Nombre', '')} {comision.get('Apellido', '')}".strip()
        if nombre_ejecutivo not in comisiones_por_ejecutivo:
            comisiones_por_ejecutivo[nombre_ejecutivo] = []
        comisiones_por_ejecutivo[nombre_ejecutivo].append(comision)

    # --- 4. Iterar y generar bloques por ejecutivo ---
    for nombre_ejecutivo, lista_comisiones in comisiones_por_ejecutivo.items():
        # Subtítulo para el ejecutivo
        flowables.append(Paragraph(f"Ejecutivo: {nombre_ejecutivo}", style_h2))
        flowables.append(Spacer(1, 12))

        # --- Preparar datos para la tabla del ejecutivo ---
        data_tabla = [
            ["Asegurado", "Póliza", "Ramo", "Prima ($)", "Fracción", "% Comisión", "Fecha", "Monto Comisión"]
        ]

        totalD = 0.0
        totalb = 0.0
        totalBo = 0.0
        totalBobs = 0.0

        for comision in lista_comisiones:
            try:
                asegurado = f"{comision.get('N_asegu', '')} {comision.get('A_Asegu', '')}".strip()
                cod_poliza = str(comision.get('cod_poliza', 'N/A'))
                ramo = str(comision.get('Ramo', 'N/A'))
                prima = float(comision.get('Prima', 0.0))
                monto_pago_f = float(comision.get('monto', 0.0))
                comision_porcentaje_f = float(comision.get('comision', 0.0))
                fecha_obj = comision.get('fecha')
                if isinstance(fecha_obj, str):
                    try:
                        # Formato: 'Fri, 09 May 2025 00:00:00 GMT'
                        # Nos quedamos con la parte de la fecha y hora, ignorando la zona horaria
                        fecha_sin_tz = " ".join(fecha_obj.split()[:5])
                        fecha_dt = datetime.datetime.strptime(fecha_sin_tz, '%a, %d %b %Y %H:%M:%S')
                        fecha_str = fecha_dt.strftime('%d/%m/%y')
                    except (ValueError, IndexError):
                        try:
                            # Intenta parsear la fecha en el formato esperado 'YYYY-MM-DD'
                            fecha_dt = datetime.datetime.strptime(fecha_obj, '%Y-%m-%d').date()
                            fecha_str = fecha_dt.strftime('%d/%m/%y')
                        except ValueError:
                            try:
                                # Intenta con otro formato si el primero falla
                                fecha_dt = datetime.datetime.strptime(fecha_obj, '%d-%m-%Y').date()
                                fecha_str = fecha_dt.strftime('%d/%m/%y')
                            except ValueError:
                                # Si todos los intentos de parseo fallan, usa el string original
                                fecha_str = fecha_obj
                elif isinstance(fecha_obj, (datetime.date, datetime.datetime)):
                    fecha_str = fecha_obj.strftime('%d/%m/%y')
                else:
                    fecha_str = 'N/A'
                
                # Usar comision_monto precalculada o recalcular si no existe
                comision_calculada = comision.get('comision_monto')
                if comision_calculada is None:
                    comision_calculada = (monto_pago_f * comision_porcentaje_f / 100)
                else:
                    comision_calculada = float(comision_calculada)

                bonoF = float(comision.get('comision_bono', 0.0) or 0.0)
            except (ValueError, TypeError):
                continue

            bono_comision = bonoF * monto_pago_f

            fraccion_str = f"{monto_pago_f:,.2f} $"
            comision_calc_str = f"{comision_calculada:,.2f} $"

            data_tabla.append([
                asegurado,
                cod_poliza,
                ramo,
                f"{prima:,.2f}",
                fraccion_str,
                f"{comision_porcentaje_f}%",
                fecha_str,
                comision_calc_str
            ])

            totalD += comision_calculada
            totalBo += bono_comision

        # --- Crear y Estilizar la Tabla ---
        tabla = Table(data_tabla, colWidths=[1.8*inch, 1*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.8*inch, 1.2*inch])
        estilo_tabla = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#800080")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ])
        tabla.setStyle(estilo_tabla)
        flowables.append(tabla)
        flowables.append(Spacer(1, 12))

        # --- Tabla de Totales y Nota ---
        data_totales = [
            ['Concepto', 'Monto'],
            ['Total Comisiones ($)', f"{totalD:,.2f} $"],
        ]
        if totalBo > 0:
            data_totales.append(['Total Bonos ($)', f"{totalBo:,.2f} $"])

        tabla_totales = Table(data_totales, colWidths=[120, 100])
        estilo_totales = TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ])
        tabla_totales.setStyle(estilo_totales)

        # --- Nueva Lógica para la Nota ---
        if nota and nota.strip():
            # 1. Realiza la sustitución de la cadena
            nota_formateada = nota.replace('\n', '<br/>')

            # 2. Usa la variable limpia en la f-string
            nota_parrafo = Paragraph(f"<b>Nota:</b><br/>{nota_formateada}", style_body_left)
            # Crear una tabla contenedora de 2 columnas para alinear totales y nota
            tabla_contenedor_final = Table(
                [[nota_parrafo, tabla_totales]], 
                colWidths=[4.5*inch, 2.5*inch],
                style=TableStyle([('VALIGN', (0,0), (-1,-1), 'TOP')]) # Alinear al top
            )
            flowables.append(tabla_contenedor_final)
        else:
            # Si no hay nota, solo añadir la tabla de totales alineada a la derecha
            wrapper_table = Table([[tabla_totales]], hAlign='RIGHT')
            flowables.append(wrapper_table)

        flowables.append(Spacer(1, 36)) # Espacio mayor entre bloques de ejecutivos

    doc.build(flowables)
    buffer.seek(0)
    return buffer


def generar_recibo_pdf_ejecutivo(comisiones, rangoFechas=None, ejecutivo_nombre=None, nota=None):
    """
    Genera un recibo de comisiones en formato PDF estilo 'Recibos de Comisiones',
    similar al formato de referencia con encabezado de empresa, datos del ejecutivo,
    tabla de detalles por póliza y totales al final.

    Args:
        comisiones (list): Lista de diccionarios con los datos de las comisiones.
        rangoFechas (str, opcional): Rango de fechas filtrado.
        ejecutivo_nombre (str, opcional): Nombre del ejecutivo.
        nota (str, opcional): Nota adicional.

    Returns:
        BytesIO: Buffer con el contenido del PDF.
    """
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, HRFlowable
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from io import BytesIO
    import datetime, os

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=letter,
        rightMargin=40, leftMargin=40,
        topMargin=40, bottomMargin=30
    )
    flowables = []
    styles = getSampleStyleSheet()

    # --- Estilos personalizados ---
    style_empresa = ParagraphStyle('empresa', fontSize=13, fontName='Helvetica-Bold', textColor=colors.HexColor('#1a1a2e'), spaceAfter=2, leading=16)
    style_titulo = ParagraphStyle('titulo', fontSize=11, fontName='Helvetica-Bold', textColor=colors.HexColor('#4a0080'), alignment=TA_CENTER, spaceAfter=4, leading=14)
    style_subtitulo = ParagraphStyle('subtitulo', fontSize=8.5, fontName='Helvetica', textColor=colors.HexColor('#555555'), alignment=TA_CENTER, spaceAfter=2)
    style_info = ParagraphStyle('info', fontSize=8, fontName='Helvetica', textColor=colors.HexColor('#333333'), leading=11)
    style_info_bold = ParagraphStyle('infobold', fontSize=8, fontName='Helvetica-Bold', textColor=colors.HexColor('#333333'), leading=11)
    style_ejecutivo = ParagraphStyle('ejecutivo', fontSize=9.5, fontName='Helvetica-Bold', textColor=colors.HexColor('#4a0080'), spaceBefore=6, spaceAfter=4)
    style_total = ParagraphStyle('total', fontSize=9, fontName='Helvetica-Bold', textColor=colors.HexColor('#1a1a2e'), alignment=TA_RIGHT)
    style_nota = ParagraphStyle('nota', fontSize=8, fontName='Helvetica-Oblique', textColor=colors.HexColor('#555555'), spaceBefore=4)
    style_small = ParagraphStyle('small', fontSize=7.5, fontName='Helvetica', textColor=colors.HexColor('#555555'))

    hoy = datetime.date.today().strftime('%d/%m/%Y')
    hora = datetime.datetime.now().strftime('%I:%M:%S %p')

    # --- Encabezado: Logo + Info empresa ---
    logo_path = os.path.join(os.path.dirname(__file__), '..', 'static', 'assets', 'img', 'cabal logo cambio.png')
    logo_cell = ''
    try:
        if os.path.exists(logo_path):
            logo_cell = Image(logo_path, width=1.3*inch, height=0.65*inch)
            logo_cell.hAlign = 'LEFT'
    except Exception:
        logo_cell = Paragraph('CABAL', style_empresa)

    empresa_info = [
        Paragraph('Correduría de Seguros', style_empresa),
        Paragraph('RECIBOS DE COMISIONES', style_titulo),
    ]
    if rangoFechas:
        empresa_info.append(Paragraph(f'Período: {rangoFechas}', style_subtitulo))
    else:
        empresa_info.append(Paragraph(f'Fecha de Reporte: {hoy}', style_subtitulo))

    fecha_info = [
        Paragraph(f'{hoy}', style_info_bold),
        Paragraph(f'{hora}', style_info),
        Paragraph('Page 1 of 1', style_small),
    ]

    header_table = Table(
        [[logo_cell, empresa_info, fecha_info]],
        colWidths=[1.5*inch, 4.5*inch, 1.5*inch]
    )
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ALIGN', (2, 0), (2, 0), 'RIGHT'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    flowables.append(header_table)
    flowables.append(HRFlowable(width='100%', thickness=1.5, color=colors.HexColor('#4a0080'), spaceAfter=6))

    # --- Si hay filtro de ejecutivo, mostrar info ---
    if ejecutivo_nombre:
        flowables.append(Paragraph(f'Ejecutivo: <b>{ejecutivo_nombre}</b>', style_info))
        flowables.append(Spacer(1, 4))

    # Agrupar por ejecutivo (si se muestran varios)
    comisiones_por_ejecutivo = {}
    for c in comisiones:
        nombre = f"{c.get('Nombre', '')} {c.get('Apellido', '')}".strip()
        if nombre not in comisiones_por_ejecutivo:
            comisiones_por_ejecutivo[nombre] = []
        comisiones_por_ejecutivo[nombre].append(c)

    # Colores de cabecera de tabla
    COLOR_HEADER_BG = colors.HexColor('#5c0080')
    COLOR_HEADER_FG = colors.white
    COLOR_ROW_ALT   = colors.HexColor('#f3eaff')
    COLOR_ROW_NORM  = colors.white
    COLOR_SUBTOTAL  = colors.HexColor('#e8d5ff')
    COLOR_TOTAL_BG  = colors.HexColor('#d0a8ff')

    # Anchos de columna de la tabla principal (sum = 7.5 inch aprox)
    COL_WIDTHS = [1.5*inch, 0.8*inch, 0.75*inch, 0.7*inch, 0.75*inch, 0.7*inch, 0.7*inch, 0.7*inch, 0.65*inch]

    grand_total_comision = 0.0
    grand_total_ejecutivo = 0.0
    grand_n_docs = 0

    for nombre_ejecutivo, lista in comisiones_por_ejecutivo.items():
        # Bloque de encabezado del ejecutivo
        ejec_header = Table(
            [[Paragraph(f'  Ejecutivo:  {nombre_ejecutivo}', style_ejecutivo)]],
            colWidths=[7.5*inch]
        )
        ejec_header.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#ede0ff')),
            ('LINEBELOW', (0, 0), (-1, -1), 1.0, colors.HexColor('#4a0080')),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ]))
        flowables.append(ejec_header)
        flowables.append(Spacer(1, 3))

        # Cabecera de la tabla de datos
        cabecera = [
            Paragraph('Asegurado', style_info_bold),
            Paragraph('Póliza', style_info_bold),
            Paragraph('Compañía', style_info_bold),
            Paragraph('Ramo', style_info_bold),
            Paragraph('Prima ($)', style_info_bold),
            Paragraph('Fracción', style_info_bold),
            Paragraph('% Com.', style_info_bold),
            Paragraph('Com. ($)', style_info_bold),
            Paragraph('Fecha', style_info_bold),
        ]

        data_tabla = [cabecera]
        subtotal_comision = 0.0
        subtotal_ejecutivo_com = 0.0
        n_docs = 0

        for idx, c in enumerate(lista):
            try:
                asegurado = f"{c.get('N_asegu', '')} {c.get('A_Asegu', '')}".strip() or 'N/A'
                cod_poliza = str(c.get('cod_poliza', 'N/A'))
                compania = str(c.get('compania', 'N/A') or 'N/A')
                ramo = str(c.get('Ramo', 'N/A') or 'N/A')
                prima = float(c.get('Prima', 0) or 0)
                fraccion = float(c.get('monto', 0) or 0)
                pct_comision = float(c.get('comision', 0) or 0)

                # Monto comisión de la póliza
                comision_monto = c.get('comision_monto')
                if comision_monto is None:
                    comision_monto = fraccion * pct_comision / 100
                else:
                    comision_monto = float(comision_monto)

                # Comisión del ejecutivo
                com_ejec = float(c.get('comision_ejecutivo', 0) or 0)

                # Fecha
                fecha_obj = c.get('fecha')
                if isinstance(fecha_obj, str):
                    try:
                        fecha_sin_tz = " ".join(fecha_obj.split()[:5])
                        fecha_dt = datetime.datetime.strptime(fecha_sin_tz, '%a, %d %b %Y %H:%M:%S')
                        fecha_str = fecha_dt.strftime('%d/%m/%Y')
                    except (ValueError, IndexError):
                        try:
                            fecha_str = datetime.datetime.strptime(fecha_obj, '%Y-%m-%d').strftime('%d/%m/%Y')
                        except ValueError:
                            fecha_str = fecha_obj
                elif isinstance(fecha_obj, (datetime.date, datetime.datetime)):
                    fecha_str = fecha_obj.strftime('%d/%m/%Y')
                else:
                    fecha_str = 'N/A'

                subtotal_comision += comision_monto
                subtotal_ejecutivo_com += com_ejec
                n_docs += 1

                bg = COLOR_ROW_ALT if idx % 2 == 0 else COLOR_ROW_NORM
                fila = [
                    Paragraph(asegurado, style_info),
                    Paragraph(cod_poliza, style_info),
                    Paragraph(compania, style_info),
                    Paragraph(ramo, style_info),
                    Paragraph(f'{prima:,.2f}', style_info),
                    Paragraph(f'{fraccion:,.2f}', style_info),
                    Paragraph(f'{pct_comision:.1f}%', style_info),
                    Paragraph(f'{comision_monto:,.2f}', style_info),
                    Paragraph(fecha_str, style_info),
                ]
                data_tabla.append(fila)

            except (ValueError, TypeError):
                continue

        # Fila de subtotal por ejecutivo
        subtotal_fila = [
            Paragraph(f'<b>{n_docs} Documento(s)</b>', style_info_bold),
            Paragraph('Total del Ejecutivo:', style_info_bold),
            '', '', '',
            Paragraph(f'<b>{subtotal_comision:,.2f} $</b>', style_info_bold),
            '',
            Paragraph(f'<b>{subtotal_ejecutivo_com:,.2f} $</b>', style_info_bold),
            '',
        ]
        data_tabla.append(subtotal_fila)

        tabla = Table(data_tabla, colWidths=COL_WIDTHS, repeatRows=1)
        n_rows = len(data_tabla)

        style_table = TableStyle([
            # Cabecera
            ('BACKGROUND', (0, 0), (-1, 0), COLOR_HEADER_BG),
            ('TEXTCOLOR', (0, 0), (-1, 0), COLOR_HEADER_FG),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 7.5),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('TOPPADDING', (0, 0), (-1, 0), 5),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 5),
            # Datos
            ('FONTSIZE', (0, 1), (-1, -1), 7.5),
            ('TOPPADDING', (0, 1), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 3),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            # Grilla
            ('GRID', (0, 0), (-1, n_rows - 2), 0.5, colors.HexColor('#cccccc')),
            ('LINEABOVE', (0, n_rows - 1), (-1, n_rows - 1), 1.0, colors.HexColor('#4a0080')),
            # Fila de subtotal
            ('BACKGROUND', (0, n_rows - 1), (-1, n_rows - 1), COLOR_SUBTOTAL),
            ('FONTNAME', (0, n_rows - 1), (-1, n_rows - 1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, n_rows - 1), (-1, n_rows - 1), 7.5),
        ])

        # Colores alternados de filas
        for i in range(1, n_rows - 1):
            bg = COLOR_ROW_ALT if i % 2 == 1 else COLOR_ROW_NORM
            style_table.add('BACKGROUND', (0, i), (-1, i), bg)

        tabla.setStyle(style_table)
        flowables.append(tabla)
        flowables.append(Spacer(1, 8))

        grand_total_comision += subtotal_comision
        grand_total_ejecutivo += subtotal_ejecutivo_com
        grand_n_docs += n_docs

    # --- Totales Generales ---
    flowables.append(HRFlowable(width='100%', thickness=1.5, color=colors.HexColor('#4a0080'), spaceBefore=4, spaceAfter=4))

    total_general_data = [
        [
            Paragraph(f'<b>{grand_n_docs} Documentos en Total</b>', style_info_bold),
            Paragraph('Total General:', style_info_bold),
            '', '', '',
            Paragraph(f'<b>{grand_total_comision:,.2f} $</b>', style_info_bold),
            '',
            Paragraph(f'<b>{grand_total_ejecutivo:,.2f} $</b>', style_info_bold),
            '',
        ]
    ]
    total_table = Table(total_general_data, colWidths=COL_WIDTHS)
    total_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), COLOR_TOTAL_BG),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LINEBELOW', (0, 0), (-1, -1), 1.0, colors.HexColor('#4a0080')),
    ]))
    flowables.append(total_table)

    # --- Leyenda de columnas ---
    flowables.append(Spacer(1, 8))
    leyenda_data = [
        [
            Paragraph('<b>Columnas:</b> Asegurado | Póliza | Compañía | Ramo | Prima Anual | Fracción Cobrada | % Comisión Póliza | Comisión Póliza ($) | Fecha', style_small),
        ]
    ]
    leyenda_table = Table(leyenda_data, colWidths=[7.5*inch])
    leyenda_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f8f0ff')),
        ('LINEABOVE', (0, 0), (-1, 0), 0.5, colors.HexColor('#cccccc')),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    flowables.append(leyenda_table)

    # --- Nota adicional ---
    if nota and nota.strip():
        flowables.append(Spacer(1, 10))
        nota_formateada = nota.replace('\n', '<br/>')
        flowables.append(Paragraph(f'<b>Nota:</b><br/>{nota_formateada}', style_nota))

    doc.build(flowables)
    buffer.seek(0)
    return buffer


def obtener_filtros_dashboard():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                # Obtener Compañías
                cursor.execute("SELECT Cod_compania, Nombre FROM compania ORDER BY Nombre")
                companias = cursor.fetchall()

                # Obtener Ramos
                cursor.execute("SELECT DISTINCT Ramo FROM poliza WHERE Ramo IS NOT NULL ORDER BY Ramo")
                ramos = cursor.fetchall()

                # Obtener Ejecutivos
                cursor.execute("SELECT cod_ejecutivo, CONCAT(Nombre, ' ', Apellido) as Nombre_Completo FROM ejecutivo ORDER BY Nombre")
                ejecutivos = cursor.fetchall()

                # Obtener Años (de los pagos)
                cursor.execute("SELECT DISTINCT EXTRACT(YEAR FROM fecha) as ano FROM pago WHERE fecha IS NOT NULL ORDER BY ano DESC")
                anos = cursor.fetchall()

                return {
                    'companias': companias,
                    'ramos': ramos,
                    'ejecutivos': ejecutivos,
                    'anos': anos
                }
    except Exception as e:
        print(f"Error en obtener_filtros_dashboard: {e}")
        return {}


def obtener_datos_dashboard(filtros):
    """
    Agrega datos para el dashboard basándose en filtros.
    Filtros posibles: ano, mes, compania_id, ramo, ejecutivo_id.

    NUEVA LÓGICA (2026):
    - Pólizas Pagadas = pólizas (nuevas + renovadas) cuya cuota 1 esté PAGADA
      dentro del rango seleccionado, por fecha_pagada de la cuota 1.
    - Facturación = SUM(renovacion.Prima) de esas pólizas pagadas.
    - Ejecutivos: mostrar Nombre+Apellido con crecimiento YoY.
    - Tipo de Venta: intentar mostrar nombre de departamento si existe tabla.
    """
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:

                # ── Construcción dinámica de filtros ────────────────────────────
                ano_actual_f = int(filtros.get('ano') or datetime.datetime.now().year)
                ano_anterior = ano_actual_f - 1

                # Filtros cuota 1 pagada (base de métricas principales)
                cuota1_clauses = [
                    "p.nro_cuota = 1",
                    "UPPER(p.estado) = 'PAGADO'",
                    "EXTRACT(YEAR FROM p.fecha_pagada) = %s"
                ]
                cuota1_params = [ano_actual_f]

                if filtros.get('mes'):
                    cuota1_clauses.append("EXTRACT(MONTH FROM p.fecha_pagada) = %s")
                    cuota1_params.append(int(filtros['mes']))
                if filtros.get('compania_id'):
                    cuota1_clauses.append("po.Cod_compania = %s")
                    cuota1_params.append(filtros['compania_id'])
                if filtros.get('ramo'):
                    cuota1_clauses.append("po.Ramo = %s")
                    cuota1_params.append(filtros['ramo'])
                if filtros.get('ejecutivo_id'):
                    cuota1_clauses.append("e.cod_ejecutivo = %s")
                    cuota1_params.append(filtros['ejecutivo_id'])

                cuota1_where = "WHERE " + " AND ".join(cuota1_clauses)

                # Filtros año anterior
                cuota1_clauses_ant = [
                    "p.nro_cuota = 1",
                    "UPPER(p.estado) = 'PAGADO'",
                    "EXTRACT(YEAR FROM p.fecha_pagada) = %s"
                ]
                cuota1_params_ant = [ano_anterior]
                if filtros.get('mes'):
                    cuota1_clauses_ant.append("EXTRACT(MONTH FROM p.fecha_pagada) = %s")
                    cuota1_params_ant.append(int(filtros['mes']))
                if filtros.get('compania_id'):
                    cuota1_clauses_ant.append("po.Cod_compania = %s")
                    cuota1_params_ant.append(filtros['compania_id'])
                if filtros.get('ramo'):
                    cuota1_clauses_ant.append("po.Ramo = %s")
                    cuota1_params_ant.append(filtros['ramo'])
                if filtros.get('ejecutivo_id'):
                    cuota1_clauses_ant.append("e.cod_ejecutivo = %s")
                    cuota1_params_ant.append(filtros['ejecutivo_id'])
                cuota1_where_ant = "WHERE " + " AND ".join(cuota1_clauses_ant)

                # Base JOIN
                base_join = """
                    FROM pago p
                    INNER JOIN renovacion r  ON p.Cod_renovacion = r.Cod_renovacion
                    INNER JOIN poliza po     ON r.cod_poliza = po.cod_poliza
                    INNER JOIN asegurado a   ON po.CI_asegurado = a.CI
                    LEFT  JOIN ejecutivo e   ON a.Ejecutivo = e.cod_ejecutivo
                """

                # ── 1. KPIs: Facturación (Prima renovacion) y Pólizas Pagadas ──
                kpi_query = f"""
                    SELECT
                        SUM(r.Prima) AS facturacion_total,
                        COUNT(DISTINCT po.cod_poliza) AS polizas_pagadas
                    {base_join}
                    {cuota1_where}
                """
                cursor.execute(kpi_query, tuple(cuota1_params))
                kpis_row = cursor.fetchone()

                facturacion = float(kpis_row['facturacion_total'] or 0)
                polizas_count = int(kpis_row['polizas_pagadas'] or 0)
                ticket_promedio = round(facturacion / polizas_count, 2) if polizas_count > 0 else 0.0

                # ── 2. Distribución por Ramo ──────────────────────────────────
                ramo_query = f"""
                    SELECT po.Ramo, SUM(r.Prima) AS total
                    {base_join}
                    {cuota1_where}
                    GROUP BY po.Ramo
                    ORDER BY total DESC
                """
                cursor.execute(ramo_query, tuple(cuota1_params))
                ventas_ramo = cursor.fetchall()

                # ── 3. Distribución por Tipo de Venta (nombre departamento) ──
                try:
                    tipo_venta_query = f"""
                        SELECT
                            COALESCE(dept.nombre, CAST(po.Tipo_venta AS CHAR)) AS nombre_tipo,
                            SUM(r.Prima) AS total
                        {base_join}
                        LEFT JOIN departamento dept ON po.Tipo_venta = dept.id
                        {cuota1_where}
                        GROUP BY po.Tipo_venta, dept.nombre
                        ORDER BY total DESC
                    """
                    cursor.execute(tipo_venta_query, tuple(cuota1_params))
                    ventas_tipo = cursor.fetchall()
                except Exception:
                    tipo_venta_query_fb = f"""
                        SELECT
                            CAST(po.Tipo_venta AS CHAR) AS nombre_tipo,
                            SUM(r.Prima) AS total
                        {base_join}
                        {cuota1_where}
                        GROUP BY po.Tipo_venta
                        ORDER BY total DESC
                    """
                    cursor.execute(tipo_venta_query_fb, tuple(cuota1_params))
                    ventas_tipo = cursor.fetchall()

                # ── 4. Tendencia Mensual (fecha_pagada cuota 1) ───────────────
                tendencia_query = f"""
                    SELECT EXTRACT(MONTH FROM p.fecha_pagada) AS mes, SUM(r.Prima) AS total
                    {base_join}
                    {cuota1_where}
                    GROUP BY EXTRACT(MONTH FROM p.fecha_pagada)
                    ORDER BY EXTRACT(MONTH FROM p.fecha_pagada)
                """
                cursor.execute(tendencia_query, tuple(cuota1_params))
                tendencia_mensual = cursor.fetchall()

                tendencia_query_ant = f"""
                    SELECT EXTRACT(MONTH FROM p.fecha_pagada) AS mes, SUM(r.Prima) AS total
                    {base_join}
                    {cuota1_where_ant}
                    GROUP BY EXTRACT(MONTH FROM p.fecha_pagada)
                    ORDER BY EXTRACT(MONTH FROM p.fecha_pagada)
                """
                cursor.execute(tendencia_query_ant, tuple(cuota1_params_ant))
                tendencia_mensual_ant = cursor.fetchall()

                # ── 5. Top Ejecutivos con YoY ─────────────────────────────────
                ejecutivos_query = f"""
                    SELECT
                        e.cod_ejecutivo,
                        CONCAT(COALESCE(e.Nombre,''), ' ', COALESCE(e.Apellido,'')) AS nombre_ejecutivo,
                        SUM(r.Prima) AS total_actual,
                        COUNT(DISTINCT po.cod_poliza) AS polizas_count
                    {base_join}
                    {cuota1_where}
                    GROUP BY e.cod_ejecutivo, e.Nombre, e.Apellido
                    ORDER BY total_actual DESC
                    LIMIT 10
                """
                cursor.execute(ejecutivos_query, tuple(cuota1_params))
                top_ejecutivos_raw = cursor.fetchall()

                # Totales del año anterior para TODOS los ejecutivos en una sola query
                # (antes: una query por cada uno de los top-10 dentro del loop)
                ant_ejecutivos_query = f"""
                    SELECT e.cod_ejecutivo, SUM(r.Prima) AS total_ant
                    {base_join}
                    {cuota1_where_ant}
                    GROUP BY e.cod_ejecutivo
                """
                cursor.execute(ant_ejecutivos_query, tuple(cuota1_params_ant))
                total_ant_by_ejecutivo = {
                    r['cod_ejecutivo']: float(r['total_ant'] or 0) for r in cursor.fetchall()
                }

                top_ejecutivos = []
                for ejec in top_ejecutivos_raw:
                    cod = ejec['cod_ejecutivo']
                    total_actual = float(ejec['total_actual'] or 0)
                    total_ant_ejec = total_ant_by_ejecutivo.get(cod, 0.0)

                    crecimiento_prima = round(total_actual - total_ant_ejec, 2)
                    yoy_pct = round(
                        ((total_actual - total_ant_ejec) / total_ant_ejec * 100)
                        if total_ant_ejec > 0 else 0.0, 2
                    )
                    top_ejecutivos.append({
                        'Ejecutivo':       (ejec['nombre_ejecutivo'] or 'Sin nombre').strip(),
                        'total':           round(total_actual, 2),
                        'polizas':         int(ejec['polizas_count'] or 0),
                        'total_anterior':  round(total_ant_ejec, 2),
                        'crecimiento_prima': crecimiento_prima,
                        'yoy_pct':         yoy_pct
                    })

                # ── 6. Top Aseguradoras ───────────────────────────────────────
                aseguradoras_query = f"""
                    SELECT c.Nombre AS compania, SUM(r.Prima) AS total
                    {base_join}
                    INNER JOIN compania c ON po.Cod_compania = c.Cod_compania
                    {cuota1_where}
                    GROUP BY c.Nombre
                    ORDER BY total DESC
                    LIMIT 10
                """
                cursor.execute(aseguradoras_query, tuple(cuota1_params))
                top_aseguradoras = cursor.fetchall()

                # ── 7. Treemap ────────────────────────────────────────────────
                treemap_query = f"""
                    SELECT c.Nombre AS compania, po.Ramo, SUM(r.Prima) AS total
                    {base_join}
                    INNER JOIN compania c ON po.Cod_compania = c.Cod_compania
                    {cuota1_where}
                    GROUP BY c.Nombre, po.Ramo
                    ORDER BY total DESC
                """
                cursor.execute(treemap_query, tuple(cuota1_params))
                treemap_raw = cursor.fetchall()
                treemap_data = [
                    {'x': f"{row['compania']} - {row['Ramo']}", 'y': round(float(row['total'] or 0), 2)}
                    for row in treemap_raw
                ]

                # ── 8. Crecimiento global YoY ─────────────────────────────────
                growth_query = f"""
                    SELECT SUM(r.Prima) AS total
                    {base_join}
                    {cuota1_where_ant}
                """
                cursor.execute(growth_query, tuple(cuota1_params_ant))
                total_ant_row = cursor.fetchone()
                total_ant = float(total_ant_row['total'] or 0) if total_ant_row else 0.0
                crecimiento_porcentaje = round(
                    ((facturacion - total_ant) / total_ant * 100) if total_ant > 0 else 0.0, 2
                )

                # ── 9. Eficiencia de Cobro ────────────────────────────────────
                try:
                    efic_clauses = ["EXTRACT(YEAR FROM p.fecha) = %s"]
                    efic_params  = [ano_actual_f]
                    if filtros.get('mes'):
                        efic_clauses.append("EXTRACT(MONTH FROM p.fecha) = %s")
                        efic_params.append(int(filtros['mes']))
                    efic_where = "WHERE " + " AND ".join(efic_clauses)

                    cobradas_month_clause = "EXTRACT(YEAR FROM p.fecha_pagada) = %s"
                    cobradas_params = [ano_actual_f]
                    if filtros.get('mes'):
                        cobradas_month_clause += " AND EXTRACT(MONTH FROM p.fecha_pagada) = %s"
                        cobradas_params.append(int(filtros['mes']))

                    efficiency_query = f"""
                        SELECT
                            COUNT(*) AS total_emitidas,
                            SUM(CASE WHEN UPPER(p.estado) = 'PAGADO'
                                         AND {cobradas_month_clause}
                                THEN 1 ELSE 0 END) AS total_cobradas
                        FROM pago p
                        INNER JOIN renovacion r ON p.Cod_renovacion = r.Cod_renovacion
                        INNER JOIN poliza po    ON r.cod_poliza = po.cod_poliza
                        {efic_where}
                    """
                    cursor.execute(efficiency_query, tuple(efic_params + cobradas_params))
                    eff_data = cursor.fetchone()
                    total_emitidas = int(eff_data['total_emitidas'] or 1)
                    total_cobradas_int = int(eff_data['total_cobradas'] or 0)
                    eficiencia = round((total_cobradas_int / total_emitidas) * 100, 2)
                except Exception:
                    eficiencia = 0.0

                # ── 10. Recaudación Externa ───────────────────────────────────
                if filtros.get('mes'):
                    try:
                        recaud_ext_query = """
                            SELECT COUNT(*) AS ext_count, SUM(p.monto) AS ext_monto
                            FROM pago p
                            WHERE UPPER(p.estado) = 'PAGADO'
                              AND EXTRACT(YEAR FROM p.fecha_pagada)  = %s
                              AND EXTRACT(MONTH FROM p.fecha_pagada) = %s
                              AND (EXTRACT(MONTH FROM p.fecha) != %s OR EXTRACT(YEAR FROM p.fecha) != %s)
                        """
                        cursor.execute(recaud_ext_query, (
                            ano_actual_f, int(filtros['mes']),
                            int(filtros['mes']), ano_actual_f
                        ))
                        ext_row = cursor.fetchone()
                        recaudacion_externa = {
                            'cantidad': int(ext_row['ext_count'] or 0),
                            'monto':    round(float(ext_row['ext_monto'] or 0), 2)
                        }
                    except Exception:
                        recaudacion_externa = {'cantidad': 0, 'monto': 0.0}
                else:
                    recaudacion_externa = {'cantidad': 0, 'monto': 0.0}

                # ── 11. Riesgo de Pólizas ─────────────────────────────────────
                risk_query = f"""
                    SELECT
                        CASE
                            WHEN r.riesgo BETWEEN 1.0  AND 1.6  THEN 'Bajo'
                            WHEN r.riesgo BETWEEN 1.61 AND 2.2  THEN 'Medio'
                            WHEN r.riesgo BETWEEN 2.21 AND 3.0  THEN 'Alto'
                            ELSE 'No Definido'
                        END AS nivel,
                        COUNT(DISTINCT r.Cod_renovacion) AS count_pol
                    {base_join}
                    {cuota1_where}
                    GROUP BY nivel
                """
                cursor.execute(risk_query, tuple(cuota1_params))
                riesgo_dist_raw = cursor.fetchall()
                riesgo_dist = [{'nivel': r['nivel'], 'count': int(r['count_pol'] or 0)} for r in riesgo_dist_raw]

                # ── 12. Siniestralidad ────────────────────────────────────────
                try:
                    claims_query = """
                        SELECT 'Carta Aval' AS tipo, COUNT(DISTINCT ca.Cod_CartaAval) AS total
                        FROM Carta_aval ca
                        WHERE EXTRACT(YEAR FROM ca.Fecha_noti) = %s
                        UNION ALL
                        SELECT 'Reembolso' AS tipo, COUNT(DISTINCT re.cod_reembolso) AS total
                        FROM Reembolso re
                        WHERE EXTRACT(YEAR FROM re.Fecha_ocurrencia) = %s
                        UNION ALL
                        SELECT 'Siniestro Auto' AS tipo, COUNT(DISTINCT sa.Cod_siniestroA) AS total
                        FROM AutomovilSiniestro sa
                        WHERE EXTRACT(YEAR FROM sa.Fecha_ocurrencia) = %s
                    """
                    cursor.execute(claims_query, (ano_actual_f, ano_actual_f, ano_actual_f))
                    siniestros_tipo = cursor.fetchall()
                except Exception:
                    siniestros_tipo = []

                # ── Serialización ─────────────────────────────────────────────
                def sanitize(val):
                    if val is None: return 0
                    if isinstance(val, (list, tuple)):
                        return [sanitize(i) for i in val]
                    if isinstance(val, dict):
                        return {k: sanitize(v) for k, v in val.items()}
                    from decimal import Decimal
                    if isinstance(val, Decimal):
                        return round(float(val), 2)
                    if isinstance(val, float):
                        return round(val, 2)
                    return val

                data_return = {
                    'kpis': {
                        'facturacion':              round(facturacion, 2),
                        'polizas_pagadas':          polizas_count,
                        'ticket_promedio':          round(ticket_promedio, 2),
                        'eficiencia':               round(eficiencia, 2),
                        'crecimiento':              round(crecimiento_porcentaje, 2),
                        'retencion':                0.0,
                        'recaudacion_externa':      recaudacion_externa
                    },
                    'ventas_ramo':                  sanitize(ventas_ramo),
                    'ventas_tipo':                  sanitize(ventas_tipo),
                    'tendencia_mensual':            sanitize(tendencia_mensual),
                    'tendencia_mensual_anterior':   sanitize(tendencia_mensual_ant),
                    'top_ejecutivos':               sanitize(top_ejecutivos),
                    'top_aseguradoras':             sanitize(top_aseguradoras),
                    'treemap':                      treemap_data,
                    'riesgo_dist':                  sanitize(riesgo_dist),
                    'siniestros_tipo':              sanitize(siniestros_tipo)
                }

                return data_return

    except Exception as e:
        import traceback
        print(f"Error en obtener_datos_dashboard: {e}")
        traceback.print_exc()
        return {}


# Función para generar el Excel
def generar_excel_comisiones(comisiones, rangoFechas=None, ejecutivo=None):
    """
    Genera un informe de comisiones en formato Excel.

    Args:
        comisiones (list): Lista de diccionarios con los datos de las comisiones.

    Returns:
        BytesIO: Un buffer de bytes que contiene el contenido del Excel.
    """
    wb = Workbook() 
    ws = wb.active 
    ws.title = "Reporte de Comisiones" 

    # Estilos comunes 
    bold_font = Font(bold=True) 
    centered_alignment = Alignment(horizontal='center', vertical='center') 
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')) 
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") # Amarillo 
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid") # Rojo 

    # --- Obtener el nombre del ejecutivo ---

    ejecutivo_nombre = None 
    if ejecutivo and ejecutivo.strip():
        ejecutivo_nombre = ejecutivo.strip() 

    ws.merge_cells('A1:AC1')  
    if ejecutivo_nombre:
        titulo_excel = f"Reporte de Comisiones de {ejecutivo_nombre}"
    else:
        titulo_excel = "Reporte de Comisiones"
    if rangoFechas:
        titulo_excel += f" | Fechas: {rangoFechas}"
    ws['A1'] = titulo_excel
    ws['A1'].font = Font(bold=True, size=16) 
    ws['A1'].alignment = centered_alignment 

    # Fila 5: Nombre del ejecutivo
    ws['B5'] = ejecutivo_nombre   
    ws['B5'].font = bold_font 

    # Fila 7: Encabezados de la tabla
    excel_headers = [ 
        "POLIZA", "Nro. RECIBO/FACTURA", "TOMADOR", "ASEGURADO", "FECHA EMISION", "FECHA DE COBRO", 
        "SEGUROS", "RAMO", "PRODUCTO", "PLAN", "TIPO DE POLIZA", "", "TIPO DE POLIZA", 
        "FRECUENCIA", "NRO CUOTAS", "$/Bs", "PRIMA ANUAL", "PRIMA COBRADA $", "PRIMA COBRADA BS", 
        "", "", "% Comisión Ref", "Comision Ref $", "Comision Ref Bs" 
    ] 
    
    # Escribir los encabezados en la fila 7 
    for col_idx, header_text in enumerate(excel_headers): 
        cell = ws.cell(row=7, column=col_idx + 2) # Empieza en columna B (2)   
        cell.value = header_text 
        cell.font = bold_font 
        cell.alignment = centered_alignment 
        cell.border = thin_border 

    # Ajustar anchos de columna para los encabezados importantes 
    column_widths = { 
        'B': 10, 'C': 20, 'D': 30, 'E': 30, 'F': 15, 'G': 15, 'H': 15, 'I': 15, 
        'J': 15, 'K': 10, 'L': 15, 'M': 15, 'N': 15, 'O': 15, 'P': 10, 'Q': 15, 
        'R': 18, 'S': 18, 'V': 15, 'W': 18, 'X': 18, 
    } 
    for col_letter, width in column_widths.items(): 
        ws.column_dimensions[col_letter].width = width 

    current_excel_row = 8 # Empezar a escribir datos desde la fila 8 

    totalD = 0 
    totalb = 0 
    totalBobs = 0 
    totalBo = 0 

    for comision in comisiones: 
        # Extraer datos con manejo de valores por defecto 
        asegurado_nombre = str(comision.get('N_asegu', '')) 
        asegurado_apellido = str(comision.get('A_Asegu', '')) 
        asegurado_completo = f"{asegurado_nombre} {asegurado_apellido}".strip() 

        ramo = str(comision.get('Ramo', 'N/A')) 
        cod_poliza = str(comision.get('cod_poliza', 'N/A')) 
        fecha_emision = str(comision.get('fecha_emision', 'N/A'))  
        fecha_cobro = str(comision.get('fecha', 'N/A')) 
        seguros = str(comision.get('Seguros', 'N/A'))  
        producto = str(comision.get('Producto', 'N/A')) 
        plan = str(comision.get('Plan', 'N/A')) 
        tipo_poliza = str(comision.get('Tipo_Poliza', 'N/A')) 
        frecuencia = str(comision.get('Frecuencia', 'N/A')) 
        nro_cuotas = str(comision.get('Nro_Cuotas', 'N/A')) 
        
        moneda = str(comision.get('moneda', '$'))  
        
        prima_anual = comision.get('Prima', 0.0)  
        monto_pago = comision.get('monto', 0.0)  
        comision_porcentaje = comision.get('comision', 0.0)  

        # Cálculos de comisión 
        try: 
            monto_pago_float = float(monto_pago) 
            comision_porcentaje_float = float(comision_porcentaje) / 100 # Convertir a decimal 
            comision_calculada = round(monto_pago_float * comision_porcentaje_float, 2) 
        except (ValueError, TypeError): 
            comision_calculada = 0.0 

        comision_bono_val = comision.get('comision_bono') 
        bonoF = 0.0
        if comision_bono_val is not None:
            try:
                bonoF = float(comision_bono_val)
            except ValueError:
                pass
        
        bono_comision = 0.0
        try:
            bono_comision = bonoF * monto_pago_float
        except ValueError:
            pass

        # Llenar la fila de datos en Excel 
        ws.cell(row=current_excel_row, column=2, value=cod_poliza).border = thin_border # POLIZA 
        ws.cell(row=current_excel_row, column=3, value=cod_poliza).border = thin_border # Nro. RECIBO/FACTURA 
        ws.cell(row=current_excel_row, column=4, value="").border = thin_border # TOMADOR 
        ws.cell(row=current_excel_row, column=5, value=asegurado_completo).border = thin_border # ASEGURADO  
        ws.cell(row=current_excel_row, column=6, value=fecha_emision).border = thin_border # FECHA EMISION 
        ws.cell(row=current_excel_row, column=7, value=fecha_cobro).border = thin_border # FECHA DE COBRO 
        ws.cell(row=current_excel_row, column=8, value=seguros).border = thin_border # SEGUROS 
        ws.cell(row=current_excel_row, column=9, value=ramo).border = thin_border # RAMO 
        ws.cell(row=current_excel_row, column=10, value=producto).border = thin_border # PRODUCTO 
        ws.cell(row=current_excel_row, column=11, value=plan).border = thin_border # PLAN 
        ws.cell(row=current_excel_row, column=12, value=tipo_poliza).border = thin_border # TIPO DE POLIZA 
        ws.cell(row=current_excel_row, column=13, value="").border = thin_border # COLUMNA L (vacía) 
        ws.cell(row=current_excel_row, column=14, value=frecuencia).border = thin_border # FRECUENCIA 
        ws.cell(row=current_excel_row, column=15, value=nro_cuotas).border = thin_border # NRO CUOTAS 
        ws.cell(row=current_excel_row, column=16, value=moneda).border = thin_border # $/Bs 

        # Prima Anual y Prima Cobrada (depende de la moneda) 
        if moneda == "$": 
            ws.cell(row=current_excel_row, column=17, value=prima_anual).border = thin_border # PRIMA ANUAL 
            ws.cell(row=current_excel_row, column=18, value=monto_pago).border = thin_border # PRIMA COBRADA $ 
            ws.cell(row=current_excel_row, column=19, value=0.0).border = thin_border # PRIMA COBRADA BS 
            ws.cell(row=current_excel_row, column=18).number_format = '"$"#,##0.00' 
            ws.cell(row=current_excel_row, column=17).number_format = '"$"#,##0.00' 
            ws.cell(row=current_excel_row, column=19).number_format = '"Bs"#,##0.00' 
        else: # Asumo que es "Bs" 
            ws.cell(row=current_excel_row, column=17, value=0.0).border = thin_border # PRIMA ANUAL 
            ws.cell(row=current_excel_row, column=18, value=0.0).border = thin_border # PRIMA COBRADA $ 
            ws.cell(row=current_excel_row, column=19, value=monto_pago).border = thin_border # PRIMA COBRADA BS 
            ws.cell(row=current_excel_row, column=18).number_format = '"$"#,##0.00' 
            ws.cell(row=current_excel_row, column=17).number_format = '"$"#,##0.00' 
            ws.cell(row=current_excel_row, column=19).number_format = '"Bs"#,##0.00' 

        # Comisiones y bonos (a partir de la columna V) 
        ws.cell(row=current_excel_row, column=22, value=comision_porcentaje).border = thin_border # % Comisión Ref 
        ws.cell(row=current_excel_row, column=22).number_format = '0.00%' # Formato de porcentaje 

        if moneda == "$": 
            ws.cell(row=current_excel_row, column=23, value=comision_calculada).border = thin_border # Comision Ref $ 
            ws.cell(row=current_excel_row, column=24, value=0.0).border = thin_border # Comision Ref Bs 
            ws.cell(row=current_excel_row, column=23).number_format = '"$"#,##0.00' 
            ws.cell(row=current_excel_row, column=24).number_format = '"Bs"#,##0.00' 
        else: # Asumo que es "Bs" 
            ws.cell(row=current_excel_row, column=23, value=0.0).border = thin_border # Comision Ref $ 
            ws.cell(row=current_excel_row, column=24, value=comision_calculada).border = thin_border # Comision Ref Bs 
            ws.cell(row=current_excel_row, column=23).number_format = '"$"#,##0.00' 
            ws.cell(row=current_excel_row, column=24).number_format = '"Bs"#,##0.00' 

        # Lógica de cálculo de totales para Excel (duplicada para consistencia)
        if moneda != "BS": 
            totalD += comision_calculada
            totalBo += bono_comision
        else: 
            totalb += comision_calculada
            totalBobs += bono_comision

        current_excel_row += 1 

    # Totales y Sumas (basado en la estructura del CSV) 
    # Fila "Suma total"  
    ws.merge_cells(start_row=current_excel_row, start_column=2, end_row=current_excel_row, end_column=3) # Bx:Cx 
    ws.cell(row=current_excel_row, column=2, value="Suma total").font = bold_font  
    # Primas Cobradas Totales 
    ws.cell(row=current_excel_row, column=18, value=totalD).font = bold_font   
    ws.cell(row=current_excel_row, column=18).number_format = '"$"#,##0.00'   
    ws.cell(row=current_excel_row, column=19, value=totalb).font = bold_font   
    ws.cell(row=current_excel_row, column=19).number_format = '"Bs"#,##0.00'   
    ws.cell(row=current_excel_row, column=20, value="Bs").font = bold_font # Columna T del CSV  

    current_excel_row += 1 

    # Fila de "Comisión $" y "Comisión Bs"  
    ws.merge_cells(start_row=current_excel_row + 1, start_column=21, end_row=current_excel_row + 1, end_column=22) # U:V del CSV  
    ws.cell(row=current_excel_row + 1, column=21, value="COMISIÓN $").font = bold_font   
    ws.cell(row=current_excel_row + 1, column=23, value="COMISIÓN BS").font = bold_font   

    ws.cell(row=current_excel_row + 2, column=21, value=round(totalD, 2)).font = bold_font   
    ws.cell(row=current_excel_row + 2, column=21).number_format = '"$"#,##0.00'   
    ws.cell(row=current_excel_row + 2, column=23, value=round(totalb, 2)).font = bold_font  
    ws.cell(row=current_excel_row + 2, column=23).number_format = '"Bs"#,##0.00'   
    ws.cell(row=current_excel_row + 2, column=21).fill = yellow_fill # Fondo amarillo  
    ws.cell(row=current_excel_row + 2, column=23).fill = yellow_fill # Fondo amarillo  

    current_excel_row += 3 # Mover a la fila después de las comisiones principales 

    # Parte de la "PORCION IPP 4,5%" y Bonos
    ws.merge_cells(start_row=current_excel_row, start_column=16, end_row=current_excel_row, end_column=18) # P:R del CSV  
    ws.cell(row=current_excel_row, column=16, value="PORCION IPP 4,5%").font = bold_font   
    ws.cell(row=current_excel_row, column=16).fill = yellow_fill   # Fondo amarillo 

    ws.cell(row=current_excel_row + 1, column=21, value=round(totalBo, 2)).font = bold_font   
    ws.cell(row=current_excel_row + 1, column=21).number_format = '"$"#,##0.00'   
    ws.cell(row=current_excel_row + 1, column=21).fill = red_fill   # Fondo rojo 
    
    ws.cell(row=current_excel_row + 1, column=23, value=round(totalBobs, 2)).font = bold_font   
    ws.cell(row=current_excel_row + 1, column=23).number_format = '"Bs"#,##0.00'   
    ws.cell(row=current_excel_row + 1, column=23).fill = red_fill   # Fondo rojo 

    # Guardar el libro de trabajo en un buffer de bytes 
    buffer_excel = BytesIO() 
    wb.save(buffer_excel) 
    buffer_excel.seek(0)
    return buffer_excel

def sql_lista_riesgos():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                querySQL = ("""
                    SELECT
                        r.Cod_renovacion,
                        a.CI,
                        a.Nombre,
                        a.Apellido,
                        r.riesgo,
                        p.Ramo,
                        r.Prima,
                        pag.Metodo_pago,
                        a.profesion,
                        a.localidad,
                        a.canal,
                        a.Tipo_CI
                    FROM
                        renovacion r
                    JOIN
                        poliza p ON r.Cod_poliza = p.cod_poliza
                    JOIN
                        asegurado a ON p.CI_asegurado = a.CI
                    LEFT JOIN
                        pago pag ON r.Cod_renovacion = pag.Cod_renovacion
                    ORDER BY
                        r.Cod_renovacion DESC;
                    """)
                cursor.execute(querySQL,)
                riesgosBD = cursor.fetchall()
        return riesgosBD
    except Exception as e:
        print(
            f"Errro en la función sql_lista_riesgos: {e}")
        return None


def anular_poliza_db(cod_poliza):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                querySQL = """
                    UPDATE renovacion
                    SET estado = 'anulada'
                    WHERE Cod_renovacion = (
                        SELECT Cod_renovacion
                        FROM (
                            SELECT Cod_renovacion
                            FROM renovacion
                            WHERE cod_poliza = %s
                            ORDER BY Cod_renovacion DESC
                            LIMIT 1
                        ) AS subquery
                    )
                """
                cursor.execute(querySQL, (cod_poliza,))
                conexion_MySQLdb.commit()
                return cursor.rowcount
    except Exception as e:
        print(f"Error en anula_poliza_db: {e}")
        return None

def traspasar_poliza_db(cod_poliza):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                querySQL = """
                    UPDATE renovacion
                    SET estado = 'traspasada'
                    WHERE Cod_renovacion = (
                        SELECT Cod_renovacion
                        FROM (
                            SELECT Cod_renovacion
                            FROM renovacion
                            WHERE cod_poliza = %s
                            ORDER BY Cod_renovacion DESC
                            LIMIT 1
                        ) AS subquery
                    )
                """
                cursor.execute(querySQL, (cod_poliza,))
                conexion_MySQLdb.commit()
                return cursor.rowcount
    except Exception as e:
        print(f"Error en transpasa_poliza_db: {e}")
        return None

def obtener_pagos_para_comisiones():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                querySQL = """
                    SELECT 
                        DATE_FORMAT(p.fecha_pagada, '%Y-%m-%d') AS "FECHA COBRO RECIBO",
                        r.cod_poliza AS "NRO. POLIZA",
                        p.cod_pago AS "NRO. RECIBO",
                        a.Ejecutivo as cod_ejecutivo,
                        CONCAT(a.Nombre, ' ', a.Apellido) AS "NOMBRE CLIENTE",
                        r.Prima AS "MONTO PRIMA",
                        p.monto AS "MONTO USD",
                        r.comision,
                        p.procesado,
                        p.tasa AS "TASA EGRESO",
                        'PAGO DE COMISION' AS "DESCRIPCIÓN MOVIMIENTO", -- Placeholder, puedes ajustar esto con un JOIN a una tabla de tipos de movimiento si existe
                        p.monto AS "MONTO PAGADO",
                        p.estado
                    FROM pago p
                    JOIN renovacion r ON p.Cod_renovacion = r.Cod_renovacion
                    JOIN poliza po ON r.cod_poliza = po.cod_poliza
                    JOIN asegurado a ON po.CI_asegurado = a.CI
                    WHERE p.estado = 'PAGADO'
                    ORDER BY p.fecha_pagada DESC;
                """
                cursor.execute(querySQL)
                pagos = cursor.fetchall()
                return pagos
    except Exception as e:
        print(f"Error en la función obtener_pagos_para_comisiones: {e}")
        return []

def obtener_comisiones_pagadas():
    """
    Obtiene una lista de comisiones asociadas a pagos que ya han sido marcados como 'PAGADO'.
    """
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                querySQL = """
                    SELECT 
                        DATE_FORMAT(p.fecha_pagada, '%Y-%m-%d') AS "FECHA COBRO RECIBO",
                        r.cod_poliza AS "NRO. POLIZA",
                        r.Cod_renovacion AS "COD. RENOVACION",
                        p.cod_pago AS "NRO. RECIBO",
                        e.cod_ejecutivo AS "COD. EJECUTIVO",
                        CONCAT(a.Nombre, ' ', a.Apellido) AS "NOMBRE CLIENTE",
                        CONCAT(e.Nombre, ' ', e.Apellido) AS "NOMBRE EJECUTIVO",
                        r.Prima AS "MONTO PRIMA",
                        p.monto AS "MONTO USD",
                        p.procesado,
                        p.tasa AS "TASA PAGO",
                        r.comision AS "COMISION %",
                        c.bono AS "BONO COMISION",
                        c.tasa AS "TASA COMISION",
                        c.monto_bs AS "MONTO BS COMISION",
                        c.monto_d AS "MONTO D",
                        c.moneda AS "MONEDA COMISION",
                        p.monto AS "MONTO PAGADO"
                    FROM pago p
                    JOIN renovacion r ON p.Cod_renovacion = r.Cod_renovacion
                    JOIN poliza po ON r.cod_poliza = po.cod_poliza
                    JOIN asegurado a ON po.CI_asegurado = a.CI
                    JOIN ejecutivo e ON a.Ejecutivo = e.cod_ejecutivo
                    LEFT JOIN comision c ON p.cod_pago = c.Cod_pago
                    WHERE p.estado = 'PAGADO'
                    ORDER BY p.fecha_pagada DESC;
                """
                cursor.execute(querySQL)
                comisiones = cursor.fetchall()
                return comisiones
    except Exception as e:
        print(f"Error en la función obtener_comisiones_pagadas: {e}")
        return []

def asignar_comisiones_faltantes():
    """
    Busca todas las renovaciones sin comisión asignada e intenta calcular y guardarla.
    """
    success_count = 0
    failure_count = 0
    failed_renovations = [] # Lista para almacenar los cod_renovacion que fallaron
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                # Obtener todas las renovaciones donde la comisión es NULL o 0
                sql_get_missing = "SELECT Cod_renovacion FROM renovacion WHERE comision IS NULL OR comision = 0"
                cursor.execute(sql_get_missing)
                renovaciones_faltantes = cursor.fetchall()

                if not renovaciones_faltantes:
                    return {'success': True, 'message': 'No hay comisiones pendientes por asignar.', 'assigned': 0, 'failed': 0, 'failed_list': []}

                # Iterar sobre cada renovación y llamar a la lógica de asignación individual
                for renovacion in renovaciones_faltantes:
                    cod_renovacion = renovacion['Cod_renovacion']
                    print(f"DEBUG: Procesando Cod_renovacion: {cod_renovacion}")
                    try:
                        resultado_asignacion = asignar_comision(cod_renovacion)
                        if resultado_asignacion is not None:
                            success_count += 1
                            print(f"DEBUG: Comisión asignada con éxito para {cod_renovacion}.")
                        else:
                            failure_count += 1
                            failed_renovations.append(cod_renovacion)
                            print(f"DEBUG: Falló la asignación de comisión para {cod_renovacion} (asignar_comision retornó None).")
                    except Exception as e:
                        failure_count += 1
                        failed_renovations.append(cod_renovacion)
                        print(f"ERROR: Excepción al llamar asignar_comision para {cod_renovacion}: {str(e)}")

        message = f'Proceso de asignación masiva completado. Asignadas: {success_count}, Fallaron: {failure_count}.'
        if failure_count > 0:
            message += f' Renovaciones fallidas: {", ".join(map(str, failed_renovations))}'
        
        return {'success': True, 'message': message, 'assigned': success_count, 'failed': failure_count, 'failed_list': failed_renovations}
    except Exception as e:
        print(f"ERROR: Se produjo un error en la función asignar_comisiones_faltantes: {str(e)}")
        return {'success': False, 'message': f"Error interno del servidor: {str(e)}", 'assigned': 0, 'failed': len(renovaciones_faltantes), 'failed_list': [r['Cod_renovacion'] for r in renovaciones_faltantes]}
        return {'success': True, 'message': 'Proceso de asignación masiva completado.', 'assigned': success_count, 'failed': failure_count}
    except Exception as e:
        print(f"Error en la función asignar_comisiones_faltantes: {e}")
        return {'success': False, 'message': str(e), 'assigned': 0, 'failed': 0}

def limpiar_y_convertir_a_float(valor_str):
    """
    Función auxiliar para limpiar y convertir una cadena de texto a float.
    Maneja formatos como '1.234,56' y '138.13'.
    """
    if not valor_str:
        return None
    try:
        s = str(valor_str).strip()
        if ',' in s:
            s = s.replace('.', '').replace(',', '.')
        return float(s)
    except (ValueError, TypeError):
        return None

def crear_bloque_comision(data):
    """
    Crea un nuevo registro en bloque_pago_comision y devuelve el ID.
    Data debe contener: numero_egreso, referencia_bancaria, fecha_movimiento, monto_total, codigo_banco
    """
    try:
        egreso = data.get('numero_egreso')
        ref = data.get('referencia_bancaria')
        fecha = data.get('fecha_movimiento')
        monto_total = data.get('monto_total', 0)
        compania = data.get('compania')
        codigo_banco = data.get('codigo_banco')
        
        if not all([egreso, ref, fecha]):
            return {'success': False, 'message': 'Faltan datos para crear el bloque.'}

        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                # Asegurar columna codigo_banco
                try:
                    cursor.execute("ALTER TABLE bloque_pago_comision ADD COLUMN codigo_banco VARCHAR(50) DEFAULT NULL")
                except:
                    pass

                sql = """
                    INSERT INTO bloque_pago_comision
                    (numero_egreso, referencia_bancaria, fecha_movimiento, monto_total, compania, codigo_banco, fecha_creacion)
                    VALUES (%s, %s, %s, %s, %s, %s, NOW())
                """
                cursor.execute(sql, (egreso, ref, fecha, monto_total, compania, codigo_banco))
                conexion_MySQLdb.commit()
                return {'success': True, 'id_bloque': cursor.lastrowid}
    except Exception as e:
        print(f"Error creando bloque comision: {e}")
        return {'success': False, 'message': str(e)}

def obtener_bloques_comision():
    """
    Obtiene todos los bloques de pago de comisión registrados.
    """
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                # Obtenemos los bloques y calculamos un total aproximado sumando las comisiones asociadas
                sql = """
                    SELECT 
                        b.id_bloque,
                        b.numero_egreso,
                        b.referencia_bancaria,
                        b.compania,
                        DATE_FORMAT(b.fecha_movimiento, '%Y-%m-%d') as fecha_movimiento,
                        DATE_FORMAT(b.fecha_creacion, '%Y-%m-%d %H:%i') as fecha_creacion,
                        (SELECT COUNT(*) FROM comision c WHERE c.id_bloque = b.id_bloque) as cantidad_pagos,
                        (SELECT COALESCE(SUM(c.monto_bs), 0) + COALESCE(SUM(c.bono), 0) FROM comision c WHERE c.id_bloque = b.id_bloque AND c.moneda IN ('Bs', 'BOLIVARES')) as total_bs,
                        (SELECT COALESCE(SUM(c.monto_d), 0) + COALESCE(SUM(c.bono), 0) FROM comision c WHERE c.id_bloque = b.id_bloque AND c.moneda IN ('$', 'DOLARES')) as total_usd,
                        (SELECT moneda FROM comision c WHERE c.id_bloque = b.id_bloque LIMIT 1) as moneda_bloque
                    FROM bloque_pago_comision b
                    ORDER BY b.fecha_creacion DESC
                """
                cursor.execute(sql)
                bloques = cursor.fetchall()
                return bloques
    except Exception as e:
        print(f"Error en obtener_bloques_comision: {e}")
        return []



def insertar_poliza_pendiente(data):
    try:
        # Helper to get value from multiple possible keys (case-insensitive and handling legacy encodings)
        def get_val(item, keys, default=None):
            for k in keys:
                if k in item:
                    return item[k]
            return default

        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                # Normalizing keys early to check existence
                nro_poliza = get_val(data, ['NRO. POLIZA', 'nro_poliza', 'poliza', 'Póliza'])
                
                # VALIDACIÓN: Si ya existe esta póliza pendiente de corrección, no la duplicamos
                if nro_poliza:
                    cursor.execute("SELECT id FROM polizas_pendientes WHERE nro_poliza = %s AND estado = 'PENDIENTE'", (nro_poliza,))
                    if cursor.fetchone():
                        return True

                sql = """
                    INSERT INTO polizas_pendientes 
                    (nro_poliza, nro_recibo, nombre_cliente, monto_prima, monto_comision, 
                    porcentaje_comision, fecha_cobro, compania, descripcion, moneda, 
                    tasa, monto_pagado, id_bloque)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """
                
                nro_recibo = get_val(data, ['NRO. RECIBO', 'nro_recibo', 'recibo', 'Recibo'])
                nombre_cliente = get_val(data, ['NOMBRE CLIENTE', 'nombre_cliente', 'asegurado', 'Cliente'])
                monto_prima = limpiar_y_convertir_a_float(get_val(data, ['MONTO PRIMA', 'monto_prima', 'prima']))
                
                # Comision USD (monto_usd_excel is the standard name in session)
                monto_comision = limpiar_y_convertir_a_float(get_val(data, ['monto_usd_excel', 'MONTO COMISIÃ“N USD', 'monto_comision_usd']))
                if data.get('moneda') == 'BOLIVARES':
                    monto_comision = 0

                porcentaje = limpiar_y_convertir_a_float(get_val(data, ['predefined_comision_porcentaje', 'porcentaje_comision', 'comision']))
                fecha = get_val(data, ['FECHA COBRO RECIBO', 'fecha_cobro', 'fecha'])
                
                # Handle potentially malformed dates from string (DD/MM/YYYY or DD-MM-YYYY)
                if isinstance(fecha, str):
                    try:
                        fecha_str = fecha.replace('-', '/')
                        if '/' in fecha_str:
                            parts = fecha_str.split('/')
                            if len(parts) == 3:
                                # Si el primer bloque es día (DD)
                                if len(parts[0]) <= 2 and len(parts[2]) >= 2:
                                    anho = parts[2] if len(parts[2]) == 4 else f"20{parts[2]}"
                                    # Convertimos a formato MySQL YYYY-MM-DD
                                    fecha = f"{anho}-{parts[1].zfill(2)}-{parts[0].zfill(2)}"
                    except:
                        pass

                compania = get_val(data, ['compania', 'COMPANIA'])
                descripcion = get_val(data, ['DESCRIPCION MOVIMIENTO', 'DESCRIPCIÃ“N MOVIMIENTO', 'DESCRIPCIÓN MOVIMIENTO', 'descripcion'])
                moneda = get_val(data, ['moneda', 'MONEDA'])
                tasa = limpiar_y_convertir_a_float(get_val(data, ['tasa_excel', 'TASA CAMBIO', 'tasa']))
                monto_pagado = limpiar_y_convertir_a_float(get_val(data, ['monto_pagado_excel', 'monto_pagado']))
                id_bloque = get_val(data, ['id_bloque'])

                cursor.execute(sql, (
                    nro_poliza,
                    nro_recibo,
                    nombre_cliente,
                    monto_prima,
                    monto_comision,
                    porcentaje,
                    fecha,
                    compania,
                    descripcion,
                    moneda,
                    tasa,
                    monto_pagado,
                    id_bloque
                ))
                conexion_MySQLdb.commit()
                return True
    except Exception as e:
        print(f"Error en insertar_poliza_pendiente: {e}")
        return False

def obtener_polizas_pendientes():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                cursor.execute("SELECT * FROM polizas_pendientes WHERE estado = 'PENDIENTE' ORDER BY fecha_registro DESC")
                return cursor.fetchall()
    except Exception as e:
        print(f"Error en obtener_polizas_pendientes: {e}")
        return []

def obtener_conteo_polizas_pendientes():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                cursor.execute("SELECT COUNT(*) as total FROM polizas_pendientes WHERE estado = 'PENDIENTE'")
                resultado = cursor.fetchone()
                return resultado['total'] if resultado else 0
    except Exception as e:
        print(f"Error en obtener_conteo_polizas_pendientes: {e}")
        return 0

def eliminar_poliza_pendiente(id_pendiente):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                cursor.execute("DELETE FROM polizas_pendientes WHERE id = %s", (id_pendiente,))
                conexion_MySQLdb.commit()
                return True
    except Exception as e:
        print(f"Error en eliminar_poliza_pendiente: {e}")
        return False

def obtener_detalle_poliza_pendiente(id_pendiente):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                cursor.execute("SELECT * FROM polizas_pendientes WHERE id = %s", (id_pendiente,))
                return cursor.fetchone()
    except Exception as e:
        print(f"Error en obtener_detalle_poliza_pendiente: {e}")
        return None

def sql_reporte_sudaseg(mes, ano):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                querySQL = """
                    SELECT 
                        'COMISIÓN' AS descripcion,
                        '' AS identificador_anexo,
                        c.Nombre AS nombre_aseguradora,
                        a.CI AS rif_asegurado,
                        p.Fecha_emision AS fecha_emision_recibo,
                        pag.fecha AS fecha_cobro_recibo,
                        pag.cod_pago AS numero_recibo,
                        p.cod_poliza AS numero_poliza,
                        ren.Prima AS monto_prima,
                        COALESCE(com.monto_bs, '') AS monto_comision,
                        COALESCE(com.comision_porcentaje, '') AS porcentaje_commission,
                        p.Ramo AS ramos,
                        bpc.fecha_movimiento AS fecha_deposito_transferencia,
                        bpc.referencia_bancaria AS numero_cheque_transferencia,
                        COALESCE(bpc.codigo_banco, '') AS codigo_entidad_bancaria,
                        pag.fecha_pagada AS fecha_pago
                    FROM pago pag
                    JOIN renovacion ren ON pag.Cod_renovacion = ren.Cod_renovacion
                    JOIN poliza p ON ren.cod_poliza = p.cod_poliza
                    JOIN asegurado a ON p.CI_asegurado = a.CI
                    LEFT JOIN compania c ON p.Cod_compania = c.Cod_compania
                    LEFT JOIN comision com ON pag.Cod_pago = com.Cod_pago
                    LEFT JOIN bloque_pago_comision bpc ON com.id_bloque = bpc.id_bloque
                    WHERE EXTRACT(MONTH FROM pag.fecha) = %s AND EXTRACT(YEAR FROM pag.fecha) = %s
                    AND pag.estado = 'PAGADO'
                """
                cursor.execute(querySQL, (mes, ano))
                reporte = cursor.fetchall()
        return reporte
    except Exception as e:
        print(f"Error en sql_reporte_sudaseg: {e}")
        return []

def procesar_comision_cobrada(data):
    """
    Procesa una comisión cobrada, la inserta en la base de datos y actualiza el estado del pago.
    """
    recibo_id = data.get('reciboId')
    cod_ejecutivo = data.get('codEjecutivo')
    bono = data.get('bono')
    tasa = data.get('tasa')
    comision_bs = data.get('comisionBs')
    comision_porcentaje = data.get('comisionPorcentaje')
    monto_dolar = data.get('montoDolar')
    moneda = data.get('moneda')
    nro_recibo_ext = data.get('nro_recibo_externo')
    src_descripcion = data.get('descripcion')
    cod_poliza = data.get('codPoliza')

    # Si es dólares, la tasa puede ser None
    if moneda in ('$', 'DOLARES') and (not tasa or float(tasa) == 0):
        tasa = None

    # Relax validation for special Retention/Adjustment/Pending cases
    if (not recibo_id or (not cod_ejecutivo and str(cod_ejecutivo) != '0')) and recibo_id not in ('RETENCION', 'POLIZA_OK', 'PENDIENTE', 'AJUSTE_GLOBAL'):
        return {'success': False, 'message': 'Faltan datos requeridos (reciboId o codEjecutivo).'}

    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                # Asegurar columnas (Auto-healing avanzado)
                columnas = [
                    ("moneda", "VARCHAR(10)"),
                    ("monto_d", "FLOAT"),
                    ("monto_pago", "FLOAT"), # Nueva columna para almacenar el monto del pago original
                    ("id_bloque", "INT"),
                    ("nro_recibo_externo", "VARCHAR(100)"),
                    ("descripcion", "TEXT"),
                    ("comision_porcentaje", "FLOAT"),
                    ("cod_poliza", "VARCHAR(100)")
                ]
                for col_name, col_type in columnas:
                    try:
                        cursor.execute(f"ALTER TABLE comision ADD COLUMN {col_name} {col_type} DEFAULT NULL")
                    except:
                        pass
                
                # Determine value for Cod_pago (handle non-numeric string for INT column)
                try:
                    db_cod_pago = int(recibo_id)
                except (ValueError, TypeError):
                    db_cod_pago = 0 # Use 0 for retentions or unlinked records if no numeric ID exists
                
                # Obtener info del pago para calcular comisión del ejecutivo y guardar monto original
                monto_pago_val = 0
                if db_cod_pago > 0:
                    sql_info = """
                        SELECT p.monto, r.comision, r.Cod_renovacion, e.cod_ejecutivo as ejecutivo_poliza
                        FROM pago p
                        JOIN renovacion r ON p.Cod_renovacion = r.Cod_renovacion
                        JOIN poliza po ON r.cod_poliza = po.cod_poliza
                        JOIN asegurado a ON po.CI_asegurado = a.CI
                        JOIN ejecutivo e ON a.Ejecutivo = e.cod_ejecutivo
                        WHERE p.cod_pago = %s
                    """
                    cursor.execute(sql_info, (db_cod_pago,))
                    pago_data = cursor.fetchone()
                    if pago_data:
                        monto_pago_val = pago_data['monto']
                        perc_ejecutivo = pago_data['comision']
                        
                        # Si no hay porcentaje asignado, intentamos asignarlo ahora
                        if perc_ejecutivo is None:
                            perc_ejecutivo = asignar_comision(pago_data['Cod_renovacion'])
                        
                        # Si el bono viene en 0, lo calculamos
                        if (not bono or float(bono) == 0) and perc_ejecutivo:
                            bono = float(monto_pago_val) * float(perc_ejecutivo)
                        
                        # Si no viene ejecutivo, usamos el de la póliza
                        if not cod_ejecutivo or str(cod_ejecutivo) == '0':
                            cod_ejecutivo = pago_data['ejecutivo_poliza']

                # Insertar la nueva comisión
                sql = """
                    INSERT INTO comision (
                        Cod_pago, cod_ejecutivo, bono, tasa, 
                        monto_bs, monto_d, monto_pago, moneda, Estado, id_bloque,
                        nro_recibo_externo, descripcion, comision_porcentaje, cod_poliza
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, "COBRADA", %s, %s, %s, %s, %s)
                """
                valores = (
                    db_cod_pago, cod_ejecutivo, bono, tasa, 
                    comision_bs, monto_dolar, monto_pago_val, moneda, data.get('id_bloque'),
                    nro_recibo_ext, src_descripcion, comision_porcentaje, cod_poliza
                )
                cursor.execute(sql, valores)
                conexion_MySQLdb.commit()
                
                return {
                    'success': True, 
                    'processed': False,
                    'message': f'Registro de comisión procesado exitosamente.',
                    'reciboId': recibo_id
                }
    except Exception as e:
        # Idealmente, aquí deberías registrar el error en un sistema de logging
        # import logging
        # logging.error(f"Error en procesar_comision_cobrada: {e}")
        return {'success': False, 'message': f'Error de base de datos: {str(e)}'}

def actualizar_comision_cobrada(data):
    """
    Actualiza una comisión procesada en la tabla 'comision'.
    """
    try:
        cod_pago = data.get('reciboId')
        bono_str = data.get('bono')
        tasa_str = data.get('tasa')
        monto_bs_str = data.get('comisionBs')
        comision_porcentaje_str = data.get('comisionPorcentaje')
        monto_dolar_str = data.get('montoDolar')
        moneda = data.get('moneda')

        bono = limpiar_y_convertir_a_float(bono_str)
        tasa = limpiar_y_convertir_a_float(tasa_str)
        monto_bs = limpiar_y_convertir_a_float(monto_bs_str)
        monto_d = limpiar_y_convertir_a_float(monto_dolar_str)

        # Si es dólares, la tasa puede ser None
        if moneda == '$' and (not tasa or tasa == 0):
            tasa = None

        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                # Obtener Cod_renovacion desde Cod_pago
                sql_get_renovacion = "SELECT Cod_renovacion FROM pago WHERE Cod_pago = %s"
                cursor.execute(sql_get_renovacion, (cod_pago,))
                renovacion_info = cursor.fetchone()

                # Actualiza la tabla 'renovacion' con el nuevo porcentaje de comisión
                if renovacion_info and comision_porcentaje_str:
                    cod_renovacion = renovacion_info['Cod_renovacion']
                    comision_porcentaje_float = limpiar_y_convertir_a_float(comision_porcentaje_str)
                    if comision_porcentaje_float is not None:
                        comision_decimal = comision_porcentaje_float / 100
                        sql_update_renovacion = "UPDATE renovacion SET comision = %s WHERE Cod_renovacion = %s"
                        cursor.execute(sql_update_renovacion, (comision_decimal, cod_renovacion))

                # Actualiza la tabla comision
                # Asegurar columnas (Auto-healing)
                try:
                    cursor.execute("ALTER TABLE comision ADD COLUMN moneda VARCHAR(10) DEFAULT NULL")
                except:
                    pass
                try:
                    cursor.execute("ALTER TABLE comision ADD COLUMN monto_d FLOAT DEFAULT NULL")
                except:
                    pass

                sql = "UPDATE comision SET bono = %s, tasa = %s, monto_bs = %s, monto_d = %s, moneda = %s WHERE Cod_pago = %s"
                valores = (bono, tasa, monto_bs, monto_d, moneda, cod_pago)
                cursor.execute(sql, valores)
                conexion_MySQLdb.commit()
                return {'success': True, 'message': 'Comisión actualizada correctamente.'}
    except Exception as e:
        print(f"Error en la función actualizar_comision_cobrada: {e}")
        return {'success': False, 'message': str(e)}
    
def sql_get_all_comisiones_config():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                querySQL = """
                    SELECT 
                        c.*, 
                        e.Nombre AS NombreEjecutivo, 
                        e.Apellido AS ApellidoEjecutivo
                    FROM comisiones_config c
                    LEFT JOIN ejecutivo e ON c.cod_ejecutivo = e.cod_ejecutivo
                """
                cursor.execute(querySQL)
                comisiones = cursor.fetchall()
        return comisiones
    except Exception as e:
        print(f"Error en la función sql_get_all_comisiones_config: {e}")
        return None

def _validate_porcentajes_format(porcentajes_str):
    if not porcentajes_str:
        return False
    parts = porcentajes_str.split(',')
    for part in parts:
        try:
            float(part.strip())
        except ValueError:
            return False
    return True

def sql_add_comision_config(data):
    try:
        porcentajes_str = data['porcentajes']
        if not _validate_porcentajes_format(porcentajes_str):
            print(f"Error: Formato de porcentajes inválido: {porcentajes_str}")
            return None # Or raise a more specific error

        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                sql = """
                    INSERT INTO comisiones_config
                    (compania, ramo, subramo, producto, tipo_ejecutivo, porcentajes, cod_ejecutivo)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                """

                # Handle empty cod_ejecutivo
                cod_ejecutivo = data.get('cod_ejecutivo')
                if cod_ejecutivo == '':
                    cod_ejecutivo = None

                valores = (data['compania'], data['ramo'], data['subramo'], data['producto'], data['tipo_ejecutivo'], porcentajes_str, cod_ejecutivo)
                cursor.execute(sql, valores)
                conexion_MySQLdb.commit()
                return cursor.lastrowid
    except Exception as e:
        print(f"Error en la función sql_add_comision_config: {e}")
        return None

def sql_update_comision_config(id, data):
    try:
        porcentajes_str = data['porcentajes']
        if not _validate_porcentajes_format(porcentajes_str):
            print(f"Error: Formato de porcentajes inválido: {porcentajes_str}")
            return None # Or raise a more specific error

        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                sql = """
                    UPDATE comisiones_config 
                    SET compania=%s, ramo=%s, subramo=%s, producto=%s, tipo_ejecutivo=%s, porcentajes=%s, cod_ejecutivo=%s 
                    WHERE id=%s
                """

                # Handle empty cod_ejecutivo
                cod_ejecutivo = data.get('cod_ejecutivo')
                if cod_ejecutivo == '':
                    cod_ejecutivo = None

                valores = (data['compania'], data['ramo'], data['subramo'], data['producto'], data['tipo_ejecutivo'], porcentajes_str, cod_ejecutivo, id)
                cursor.execute(sql, valores)
                conexion_MySQLdb.commit()
                return cursor.rowcount
    except Exception as e:
        print(f"Error en la función sql_update_comision_config: {e}")
        return None

def sql_delete_comision_config(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                sql = "DELETE FROM comisiones_config WHERE id=%s"
                cursor.execute(sql, (id,))
                conexion_MySQLdb.commit()
                return cursor.rowcount
    except Exception as e:
        print(f"Error en la función sql_delete_comision_config: {e}")
        return None

def generar_plantilla_excel(tipo):

    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Plantilla"
    
    if tipo == 'generico':
        headers = ['CEDULA DEL ASEGURADO', 'ASEGURADO', 'FECHA EMISION', 'POLIZA', 'RAMO', 'PRIMA TOTAL (SIN IGTF)']
    elif tipo == 'mercantil':
        headers = [
            'POLIZA', 'SEGUROS', 'RAMO', 'PRODUCTO', 'SUB RAMO', 'TOMADOR', 
            'CORREDOR / REFERIDOR', 'ID', 'NACIONALIDAD', 'PRIMER NOMBRE', 
            'SEGUNDO NOMBRE', 'PRIMER APELLIDO', 'SEGUNDO APELLIDO', 'SUMA ASEGURADA', 
            'FECHA EMISION', 'PRIMA TOTAL (SIN IGTF)', 'INICIAL / FRACCION (SIN IGTF)', 
            'FRECUENCIA', '$/BS', 'METODO DE PAGO', 'TASA DE CAMBIO', 'ESTATUS', 
            'NRO CUOTAS', 'PLACA', 'MARCA', 'MODELO', 'AÑO', 'CORREO ASEGURADO', 
            'NRO CUOTAS', 'PLACA', 'MARCA', 'MODELO', 'AÑO', 'CORREO ASEGURADO', 
            'TELEFONO ASEGURADO', 'FECHA DE COBRO'
        ]
    elif tipo == 'siniestros':
        headers = [
            'POLIZA', 'CEDULA_ASEGURADO', 'FECHA_SINIESTRO', 'MONTO_SINIESTRO', 
            'ESTATUS_SINIESTRO', 'DESCRIPCION', 'RAMO'
        ]
    else:
        return None
        
    ws.append(headers)
    
    # Ajustar ancho de columnas básico
    for i, _ in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = 20
        
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def obtener_detalles_bloque(id_bloque):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                # Recuperar detalles de la tabla comision + joins para data extra
                # NOTA: Usamos COALESCE para intentar sacar la póliza y asegurado tanto 
                # por el vínculo directo (cod_poliza de la tabla comision) como por el vínculo histórico (vía Cod_pago)
                sql = """
                    SELECT 
                        c.nro_recibo_externo as recibo,
                        c.descripcion,
                        COALESCE(p_direct.cod_poliza, p_from_pay.cod_poliza, 'MANUAL') as cod_poliza,
                        COALESCE(
                            CONCAT_WS(' ', a_direct.Nombre, a_direct.Apellido),
                            CONCAT_WS(' ', a_from_pay.Nombre, a_from_pay.Apellido),
                            '-'
                        ) as asegurado,
                        COALESCE(CONCAT_WS(' ', e.Nombre, e.Apellido), '-') as ejecutivo,
                        c.monto_d, 
                        c.monto_bs,
                        c.tasa,
                        c.moneda,
                        c.Cod_pago
                    FROM comision c
                    LEFT JOIN poliza p_direct ON c.cod_poliza = p_direct.cod_poliza
                    LEFT JOIN asegurado a_direct ON p_direct.CI_asegurado = a_direct.CI
                    LEFT JOIN pago pay ON c.Cod_pago = pay.cod_pago
                    LEFT JOIN renovacion ren ON pay.Cod_renovacion = ren.Cod_renovacion
                    LEFT JOIN poliza p_from_pay ON ren.cod_poliza = p_from_pay.cod_poliza
                    LEFT JOIN asegurado a_from_pay ON p_from_pay.CI_asegurado = a_from_pay.CI
                    LEFT JOIN ejecutivo e ON c.cod_ejecutivo = e.cod_ejecutivo
                    WHERE c.id_bloque = %s
                """
                cursor.execute(sql, (id_bloque,))
                detalles = cursor.fetchall()
                
                # Normalizar datos para retenciones o faltantes
                formatted_detalles = []
                for d in detalles:
                    recibo_val = str(d['recibo'])
                    if not d['recibo'] or recibo_val == 'MANUAL' or d['Cod_pago'] == 0:
                        d['recibo'] = 'MANUAL / RETENCIÓN' if recibo_val == 'MANUAL' else (d['recibo'] or 'RETENCION')
                        
                    formatted_detalles.append(d)
                    
                return formatted_detalles
    except Exception as e:
        print(f"Error en obtener_detalles_bloque: {e}")
        return []

def procesar_registro_desde_pendiente(dataForm):
    try:
        id_pendiente = dataForm.get('id_pendiente')
        with connectionBD() as conn:
            with conn.cursor() as cursor:
                # 0. Recuperar detalle pendiente
                cursor.execute("SELECT * FROM polizas_pendientes WHERE id = %s", (id_pendiente,))
                detalle = cursor.fetchone()
                if not detalle:
                    return {'success': False, 'message': 'El registro pendiente ya no existe.'}

                # 1. Asegurado
                cedula_str = dataForm.get('Cedula_asegurado', '').strip()
                if not cedula_str:
                     return {'success': False, 'message': 'La cédula es obligatoria.'}
                try:
                    cedula = int(cedula_str)
                except:
                    return {'success': False, 'message': 'Cédula inválida.'}

                cursor.execute("SELECT CI FROM asegurado WHERE CI = %s", (cedula,))
                if not cursor.fetchone():
                    sql_asig = """
                        INSERT INTO asegurado 
                        (CI, Nombre, Nombre2, Apellido, Apellido2, Tipo_CI, Correo, Fecha_nacimiento, Telefono, Ejecutivo, profesion, localidad, canal) 
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """
                    telefono = f"{dataForm.get('prefijo_telefono', '')}{dataForm.get('telefono_asegurado_numero', '')}"
                    cursor.execute(sql_asig, (
                        cedula, dataForm.get('nombre_asegurado'), dataForm.get('nombre_asegurado2'),
                        dataForm.get('apellido_asegurado'), dataForm.get('apellido_asegurado2'),
                        dataForm.get('tipo'), dataForm.get('email_asegurado'), dataForm.get('fecha_nacimiento') or None,
                        telefono, dataForm.get('Ejecutivo'), dataForm.get('Profesion'),
                        dataForm.get('localidad'), dataForm.get('Canal')
                    ))

                # 2. Póliza
                cod_poliza = dataForm.get('numero_poliza')
                if not cod_poliza:
                     return {'success': False, 'message': 'Número de póliza obligatorio.'}

                cursor.execute("SELECT cod_poliza FROM poliza WHERE cod_poliza = %s", (cod_poliza,))
                if cursor.fetchone():
                    return {'success': False, 'message': f'La póliza {cod_poliza} ya existe.'}

                cursor.execute("SELECT Cod_compania FROM compania WHERE Nombre = %s LIMIT 1", (dataForm.get('company'),))
                res_cia = cursor.fetchone()
                cod_cia = res_cia['Cod_compania'] if res_cia else None

                sql_pol = "INSERT INTO poliza (cod_poliza, CI_asegurado, Fecha_emision, Cod_compania, Tomador, Ramo, Tipo_venta) VALUES (%s, %s, %s, %s, %s, %s, %s)"
                cursor.execute(sql_pol, (cod_poliza, cedula, dataForm.get('fecha_emision'), cod_cia, dataForm.get('Tomador'), dataForm.get('Ramo'), dataForm.get('Tipo_venta')))

                # 3. Renovación
                prima_str = str(dataForm.get('prima', '0')).replace('.', '').replace(',', '.')
                try:
                    prima_float = float(prima_str)
                except:
                    prima_float = 0.0

                fecha_emision_raw = dataForm.get('fecha_emision')
                try:
                    fecha_emision_dt = datetime.datetime.strptime(fecha_emision_raw, '%Y-%m-%d')
                except:
                    fecha_emision_dt = datetime.datetime.now()
                
                fecha_vencimiento = fecha_emision_dt + timedelta(days=365)
                
                sql_renov = "INSERT INTO renovacion (Cod_poliza, Prima, Frecuencia, Fecha_contrato, cobertura, Fecha_vencimiento) VALUES (%s, %s, %s, %s, %s, %s)"
                cursor.execute(sql_renov, (cod_poliza, prima_float, dataForm.get('Frecuencia') or 1, fecha_emision_raw, dataForm.get('Cobertura'), fecha_vencimiento))
                cod_renovacion = cursor.lastrowid

                # 4. Ramo Específico
                ramo = dataForm.get('Ramo')
                if ramo == 'Persona':
                    cursor.execute("INSERT INTO Persona (Cod_poliza, Producto, Subramo) VALUES (%s, %s, %s) ON CONFLICT DO NOTHING", (cod_poliza, dataForm.get('producto'), dataForm.get('SubRamo')))
                elif ramo == 'Auto':
                    cursor.execute("INSERT INTO Auto (Cod_poliza, modelo, Producto, placa, año, marca, Subramo) VALUES (%s, %s, %s, %s, %s, %s, %s)", 
                                   (cod_poliza, dataForm.get('Modelo'), dataForm.get('producto'), dataForm.get('placa'), dataForm.get('Ano') or 0, dataForm.get('Marca'), dataForm.get('SubRamo')))
                elif ramo == 'Patrimonial':
                    cursor.execute("INSERT INTO Patrimonio (Cod_poliza, direccion, Producto, Subramo) VALUES (%s, %s, %s, %s)", (cod_poliza, dataForm.get('direccion'), dataForm.get('producto'), dataForm.get('SubRamo')))
                elif ramo == 'Fianza':
                    cursor.execute("INSERT INTO Fianza (Cod_poliza, Producto, Subramo) VALUES (%s, %s, %s)", (cod_poliza, dataForm.get('producto'), dataForm.get('SubRamo')))
                elif ramo == 'Viaje':
                    cursor.execute("INSERT INTO Viaje (Cod_poliza, cod_pasaporte, Producto, Subramo) VALUES (%s, %s, %s, %s)", (cod_poliza, dataForm.get('pasaporte'), dataForm.get('producto'), dataForm.get('SubRamo')))

                # 5. Registrar Pagos Pendientes basados en Frecuencia
                try:
                    frecuencia = int(dataForm.get('Frecuencia', 1))
                except:
                    frecuencia = 1
                
                frecuencia_map = {
                    1: {'pagos': 1, 'intervalo': 12}, 
                    2: {'pagos': 12, 'intervalo': 1}, 
                    3: {'pagos': 4, 'intervalo': 3},  
                    4: {'pagos': 2, 'intervalo': 6},  
                    5: {'pagos': 3, 'intervalo': 4},  
                    6: {'pagos': 6, 'intervalo': 2}   
                }
                config_pago = frecuencia_map.get(frecuencia, {'pagos': 1, 'intervalo': 12})
                num_pagos_total = config_pago['pagos']
                intervalo_meses = config_pago['intervalo']
                
                compania_up = str(dataForm.get('company') or '').upper()
                subramo_up = str(dataForm.get('SubRamo') or '').upper()
                if compania_up == "MERCANTIL PANAMÁ" and subramo_up == "SALUD" and frecuencia == 2:
                    num_pagos_total = 10
                
                monto_cuota = prima_float / num_pagos_total if num_pagos_total > 0 else prima_float
                
                mon_excel = str(detalle.get('moneda', '')).upper()
                moneda_pago = '$' if 'DOLAR' in mon_excel or '$' in mon_excel else 'Bs'
                
                from dateutil.relativedelta import relativedelta
                sql_ins_pago = "INSERT INTO pago (Cod_renovacion, Moneda, fecha, monto, estado, nro_cuota) VALUES (%s, %s, %s, %s, %s, %s)"
                new_cod_pago = None

                for i in range(1, num_pagos_total + 1):
                    offset_meses = (i - 1) * intervalo_meses
                    fecha_siguiente_pago = fecha_emision_dt + relativedelta(months=offset_meses)

                    cursor.execute(sql_ins_pago, (
                        cod_renovacion,
                        moneda_pago,
                        fecha_siguiente_pago.strftime('%Y-%m-%d'),
                        monto_cuota,
                        'EN PROCESO',
                        i
                    ))
                    if i == 1:
                        new_cod_pago = cursor.lastrowid
                
                # 6. Vincular Comisión y Borrar Pendiente
                sql_upd_com = "UPDATE comision SET Cod_pago = %s, cod_ejecutivo = %s WHERE id_bloque = %s AND nro_recibo_externo = %s AND (Cod_pago = 0 OR Cod_pago IS NULL)"
                if new_cod_pago:
                    cursor.execute(sql_upd_com, (new_cod_pago, dataForm.get('Ejecutivo') or 0, detalle['id_bloque'], detalle['nro_recibo']))
                
                cursor.execute("DELETE FROM polizas_pendientes WHERE id = %s", (id_pendiente,))
                
                conn.commit()
                return {'success': True, 'message': 'Póliza registrada y comisión vinculada correctamente.', 'data': cod_poliza}

    except Exception as e:
        print(f"Error en procesar_registro_desde_pendiente: {e}")
        return {'success': False, 'message': str(e)}

def procesar_siniestros_excel(file):
    import pandas as pd
    from io import BytesIO
    import math
    import datetime
    import unicodedata
    
    def normalize_str(s):
        if not isinstance(s, str):
            return str(s)
        s = ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
        return s.strip().upper()

    try:
        xls = pd.ExcelFile(file)
        sheets = xls.sheet_names
        target_sheets = ["Reembolso", "carta aval", "automovil", "patrimonial", "notificaciones de viajes"]
        
        data = {}
        
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                # Precargar todas las pólizas una sola vez (antes: hasta 3 queries por fila del Excel)
                cursor.execute("SELECT Cod_poliza, Cod_compania FROM poliza")
                polizas_by_exact = {}
                polizas_by_clean = {}
                polizas_by_clean_nozero = {}
                for p in cursor.fetchall():
                    cod = p['Cod_poliza']
                    polizas_by_exact[cod] = p
                    clean = cod.replace("-", "").replace(" ", "")
                    polizas_by_clean.setdefault(clean, p)
                    nozero = clean[1:] if clean.startswith("0") else clean
                    polizas_by_clean_nozero.setdefault(nozero, p)

                for sheet_name in sheets:
                    # check case-insensitive match using "in" to allow variations like "Reembolsos"
                    is_match = any(target.lower() in sheet_name.lower() for target in target_sheets)

                    # Si la hoja coincide o si es la única hoja en el archivo (asumimos Reembolso por defecto)
                    if is_match or len(sheets) == 1:
                        try:
                            # Read without assuming row 0 is header
                            df = pd.read_excel(xls, sheet_name=sheet_name, header=None)
                            
                            # Find header row
                            header_idx = 0
                            for i, row in df.head(15).iterrows():
                                row_str = ' '.join(str(x).upper() for x in row.values if pd.notna(x))
                                if 'POLIZA' in row_str or 'PÓLIZA' in row_str or 'CEDULA' in row_str or 'CÉDULA' in row_str or 'SINIESTRO' in row_str:
                                    header_idx = i
                                    break
                            
                            # Assign headers and drop previous rows
                            df.columns = df.iloc[header_idx]
                            df = df.iloc[header_idx+1:].reset_index(drop=True)
                            
                            # Drop columns that are completely NaN or have no name
                            df = df.loc[:, df.columns.notna()]
                            df.columns = [str(c).strip() if pd.notna(c) else "" for c in df.columns]
                            df.dropna(how='all', inplace=True)
                            
                            # Convert dates to string representation
                            for col in df.select_dtypes(include=['datetime64', 'datetimetz']).columns:
                                df[col] = df[col].dt.strftime('%Y-%m-%d')
                            
                            df = df.replace({pd.NA: None, float('nan'): None})
                            sheet_data = df.to_dict('records')
                            
                            seen_in_file = set() # (company_id, siniestro_code)

                            # Determinar la tabla una sola vez por hoja (antes: recalculado en cada fila)
                            tbl = "reembolso"
                            cn = sheet_name.lower()
                            if "carta aval" in cn: tbl = "carta_aval"
                            elif "automovil" in cn: tbl = "automovilsiniestro"
                            if not is_match and len(sheets) == 1: tbl = "reembolso"

                            # Precargar siniestros ya existentes para esta tabla (antes: 1 query por fila)
                            cursor.execute(f"""
                                SELECT t.codigo_siniestro AS codigo, p.Cod_compania AS compania
                                FROM {tbl} t JOIN poliza p ON t.Cod_poliza = p.Cod_poliza
                            """)
                            siniestros_existentes = set(
                                (r['compania'], str(r['codigo']).strip()) for r in cursor.fetchall()
                            )

                            for row in sheet_data:
                                row_search = {}
                                for k, v in row.items():
                                    if isinstance(v, float) and (math.isnan(v) or pd.isna(v)): row[k] = None
                                    elif isinstance(v, float) and v.is_integer(): row[k] = str(int(v))
                                    elif isinstance(v, (datetime.datetime, datetime.date)): row[k] = str(v)
                                    if k: row_search[normalize_str(k)] = row[k]

                                poliza = row_search.get('POLIZA') or row_search.get('NUMERO DE POLIZA') or next((v for k,v in row_search.items() if 'POLIZA' in k), None)
                                siniestro_cod = row_search.get('SINIESTRO') or row_search.get('CODIGO SINIESTRO') or row_search.get('CODIGO_SINIESTRO') or next((v for k,v in row_search.items() if 'SINIESTRO' in k), None)

                                row['existe_poliza'] = False
                                row['ya_existe_siniestro'] = False
                                row['duplicate_in_file'] = False
                                company_id = None

                                if poliza is not None:
                                    poliza_str = str(poliza).strip()
                                    if poliza_str:
                                        p_data = polizas_by_exact.get(poliza_str)

                                        # Si no hace match directo, intentamos limpiando guiones y espacios
                                        if not p_data:
                                            poliza_clean = poliza_str.replace("-", "").replace(" ", "")
                                            p_data = polizas_by_clean.get(poliza_clean)

                                            # Si sigue sin hacer match, intentamos quitando un '0' inicial si existe
                                            if not p_data:
                                                poliza_sin_cero = poliza_clean[1:] if poliza_clean.startswith("0") else poliza_clean
                                                p_data = polizas_by_clean_nozero.get(poliza_sin_cero)

                                        if p_data:
                                            row['existe_poliza'] = True
                                            row['Poliza_Match_DB'] = p_data['Cod_poliza']
                                            company_id = p_data['Cod_compania']

                                # Check for duplicate claim code in same company
                                if company_id and siniestro_cod:
                                    sc_str = str(siniestro_cod).strip()
                                    if sc_str:
                                        # Check if already in the DB for THIS company
                                        if (company_id, sc_str) in siniestros_existentes:
                                            row['ya_existe_siniestro'] = True

                                        # Check if duplicate within the same Excel file
                                        file_key = (company_id, sc_str)
                                        if file_key in seen_in_file:
                                            row['duplicate_in_file'] = True
                                        seen_in_file.add(file_key)

                            final_sheet_name = sheet_name if is_match else "Reembolso"
                            data[final_sheet_name] = sheet_data
                        except Exception as e:
                            print(f"Error procesando hoja {sheet_name}: {str(e)}")
                            raise e
                            
        return {'success': True, 'data': data}
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return {'success': False, 'error': f'Error al procesar el archivo Excel: {str(e)}'}

def normalize_key_siniestro(s):
    import unicodedata
    if not isinstance(s, str): return str(s)
    s = ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
    return s.strip().upper()

def guardar_siniestro_individual_db(sheet_name, row):
    import math
    try:
        target_table = None
        clean_name = sheet_name.lower().strip()
        
        if "reembolso" in clean_name: target_table = "reembolso"
        elif "carta aval" in clean_name: target_table = "carta_aval"
        elif "automovil" in clean_name: target_table = "automovilsiniestro"
        elif "patrimonial" in clean_name: target_table = "reembolso"
        elif "notificaciones de viajes" in clean_name: target_table = "reembolso"
        
        if not target_table:
            return {'success': False, 'message': f'Hoja "{sheet_name}" no reconocida.'}

        poliza = "Desconocida"
        row_map = {normalize_key_siniestro(k): v for k, v in row.items()}
        
        poliza = row.get('Poliza_Match_DB') or row.get('Poliza') or row_map.get('POLIZA') or row_map.get('NUMERO DE POLIZA')
        if not poliza:
            for k_norm, val in row_map.items():
                if 'POLIZA' in k_norm:
                    poliza = val
                    break

        if not poliza:
            return {'success': False, 'message': 'Póliza no detectada en la fila.'}
        
        # Se usan las funciones globales normalize_moneda y clean_monto

        if target_table == "reembolso":
            sql = """INSERT INTO reembolso 
                    (Cod_poliza, Diagnostico, Estado, Fecha_ocurrencia, Fecha_noti, Fecha_max, 
                        Moneda, Monto_solicitado, Monto_pagado, Fecha_pago, Correo, codigo_siniestro, monto_dolares, Tipo_Atencion)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
                    
            params = (
                str(poliza),
                row_map.get('ESPECIALIDAD') or row_map.get('DIAGNOSTICO'),
                row_map.get('ESTATUS') or row_map.get('ESTADO'),
                row_map.get('FECHA DE OCURRENCIA') or row_map.get('FECHA_OCURRENCIA'),
                row_map.get('FECHA DE NOTIFICACION') or row_map.get('FECHA_NOTI'),
                row_map.get('FECHA LIMITE (MAX 30 DIAS)') or row_map.get('FECHA_MAX'),
                normalize_moneda(row_map.get('MONEDA')),
                clean_monto(row_map.get('MONTO SOLICITADO') or row_map.get('MONTO_SOLICITADO')),
                clean_monto(row_map.get('MONTO PAGADO') or row_map.get('MONTO_PAGADO')),
                row_map.get('FECHA DE PAGO') or row_map.get('FECHA_PAGO'),
                row_map.get('CORREO'),
                str(row_map.get('SINIESTRO') or row_map.get('CODIGO_SINIESTRO') or ''),
                clean_monto(row_map.get('MONTO EN DOLARES') or row_map.get('MONTO_DOLARES')),
                row_map.get('TIPO DE ATENCION') or row_map.get('TIPO_ATENCION')
            )
            
        elif target_table == "carta_aval":
            sql = """INSERT INTO carta_aval 
                    (Cod_poliza, Diagnostico, Procedimiento, Estado, Moneda, Monto_solicitado, Monto_aprobado, 
                        Fecha_noti, Fecha_apro, Correo, codigo_siniestro, Tipo_Atencion,
                        MES, NEGOCIO, TITULAR, RECLAMANTE, REFERIDOR, ANALISTA, OBSERVACION)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
            
            diag_proc = row_map.get('DIAGNOSTICO / PROCEDIMIENTO') or row_map.get('DIAGNOSTICO') or row_map.get('ESPECIALIDAD')
            
            params = (
                str(poliza),
                diag_proc,
                row_map.get('PROCEDIMIENTO') or diag_proc,
                row_map.get('ESTATUS') or row_map.get('ESTADO'),
                normalize_moneda(row_map.get('MONEDA')),
                clean_monto(row_map.get('MONTO SOLICITADO') or row_map.get('MONTO_SOLICITADO')),
                clean_monto(row_map.get('MONTO APROBADO') or row_map.get('MONTO_PAGADO')),
                row_map.get('FECHA DE NOTIFICACION') or row_map.get('FECHA_NOTI'),
                row_map.get('FECHA DE APROBACION') or row_map.get('FECHA_APRO') or row_map.get('FECHA_PAGO'),
                row_map.get('CORREO'),
                str(row_map.get('SINIESTRO') or row_map.get('CODIGO_SINIESTRO') or ''),
                row_map.get('TIPO DE ATENCION') or row_map.get('TIPO_ATENCION'),
                row_map.get('MES'),
                row_map.get('NEGOCIO'),
                row_map.get('TITULAR'),
                row_map.get('RECLAMANTE'),
                row_map.get('REFERIDOR'),
                row_map.get('ANALISTA'),
                row_map.get('OBSERVACION')
            )
            
        elif target_table == "automovilsiniestro":
            sql = """INSERT INTO automovilsiniestro 
                    (Cod_poliza, Fecha_ocurrencia, Fecha_noti, Estado, Monto_orden, Correo, Descripcion, codigo_siniestro)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)"""
            
            params = (
                str(poliza),
                row_map.get('FECHA DE OCURRENCIA') or row_map.get('FECHA_OCURRENCIA'),
                row_map.get('FECHA DE NOTIFICACION') or row_map.get('FECHA_NOTI'),
                row_map.get('ESTATUS') or row_map.get('ESTADO'),
                clean_monto(row_map.get('MONTO PAGADO') or row_map.get('MONTO_PAGADO') or row_map.get('MONTO_ORDEN')),
                row_map.get('CORREO'),
                row_map.get('OBSERVACIONES') or row_map.get('DESCRIPCION'),
                str(row_map.get('SINIESTRO') or row_map.get('CODIGO_SINIESTRO') or '')
            )

        # Cleanup params
        final_params = []
        for p in params:
            if isinstance(p, float) and math.isnan(p): final_params.append(None)
            elif p == "" or p == "nan": final_params.append(None)
            else: final_params.append(p)

        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                cursor.execute(sql, tuple(final_params))
                conexion_MySQLdb.commit()
                return {'success': True, 'message': 'Siniestro registrado.'}

    except Exception as e:
        return {'success': False, 'message': str(e)}

def guardar_siniestros_batch(siniestros_data):
    import math
    import datetime
    from flask import session
    
    def normalize_key(s):
        import unicodedata
        if not isinstance(s, str): return str(s)
        s = ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
        return s.strip().upper()

    try:
        inserted_count = 0
        errors = []
        
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                print(f"DEBUG: siniestros_data keys: {list(siniestros_data.keys())}")
                for sheet_name, rows in siniestros_data.items():
                    target_table = None
                    clean_name = sheet_name.lower().strip()
                    
                    if "reembolso" in clean_name: target_table = "reembolso"
                    elif "carta aval" in clean_name: target_table = "carta_aval"
                    elif "automovil" in clean_name: target_table = "automovilsiniestro"
                    elif "patrimonial" in clean_name: target_table = "reembolso" # Fallback if no specific patrimonial table
                    elif "notificaciones de viajes" in clean_name: target_table = "reembolso" # Fallback
                    
                    print(f"DEBUG: Hoja: '{sheet_name}', Target Table: '{target_table}', Filas: {len(rows)}")
                    if not target_table: continue
                    
                    for row in rows:
                        if not row: continue
                        try:
                            # 1. Map row to DB fields
                            poliza = "Desconocida"
                            row_map = {normalize_key_siniestro(k): v for k, v in row.items()}
                            print(f"DEBUG: Procesando fila. Map keys: {list(row_map.keys())}")
                            
                            # Common fields
                            poliza = row.get('Poliza_Match_DB') or row.get('Poliza') or row_map.get('POLIZA') or row_map.get('NUMERO DE POLIZA')
                            if not poliza:
                                for k_norm, val in row_map.items():
                                    if 'POLIZA' in k_norm:
                                        poliza = val
                                        break

                            print(f"DEBUG: Poliza detectada: '{poliza}'")
                            if not poliza:
                                print(f"DEBUG: Fila saltada por falta de POLIZA.")
                                continue
                            
                            # Use core function instead of duplicating logic? 
                            # Let's keep it mostly as is for stability or refactor later.
                            # For now, it's safer to keep batch separate.
                            
                            # Data preparation based on table
                            # Se usan las funciones globales normalize_moneda y clean_monto

                            if target_table == "reembolso":
                                sql = """INSERT INTO reembolso 
                                        (Cod_poliza, Diagnostico, Estado, Fecha_ocurrencia, Fecha_noti, Fecha_max, 
                                         Moneda, Monto_solicitado, Monto_pagado, Fecha_pago, Correo, codigo_siniestro, monto_dolares)
                                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
                                        
                                params = (
                                    str(poliza),
                                    row_map.get('ESPECIALIDAD') or row_map.get('DIAGNOSTICO') or row_map.get('DIAGNOSTICO / PROCEDIMIENTO'),
                                    row_map.get('ESTATUS') or row_map.get('ESTADO'),
                                    row_map.get('FECHA DE OCURRENCIA') or row_map.get('FECHA_OCURRENCIA'),
                                    row_map.get('FECHA DE NOTIFICACION') or row_map.get('FECHA_NOTI'),
                                    row_map.get('FECHA LIMITE (MAX 30 DIAS)') or row_map.get('FECHA_MAX'),
                                    normalize_moneda(row_map.get('MONEDA')),
                                    clean_monto(row_map.get('MONTO SOLICITADO') or row_map.get('MONTO_SOLICITADO')),
                                    clean_monto(row_map.get('MONTO PAGADO') or row_map.get('MONTO_PAGADO')),
                                    row_map.get('FECHA DE PAGO') or row_map.get('FECHA_PAGO'),
                                    row_map.get('CORREO'),
                                    str(row_map.get('SINIESTRO') or row_map.get('CODIGO_SINIESTRO') or ''),
                                    clean_monto(row_map.get('MONTO EN DOLARES') or row_map.get('MONTO_DOLARES'))
                                )
                                
                            elif target_table == "carta_aval":
                                sql = """INSERT INTO carta_aval 
                                        (Cod_poliza, Diagnostico, Procedimiento, Estado, Moneda, Monto_solicitado, Monto_aprobado, 
                                         Fecha_noti, Fecha_apro, Correo, codigo_siniestro, Tipo_Atencion,
                                         MES, NEGOCIO, TITULAR, RECLAMANTE, REFERIDOR, ANALISTA, OBSERVACION)
                                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
                                
                                diag_proc = row_map.get('DIAGNOSTICO / PROCEDIMIENTO') or row_map.get('DIAGNOSTICO') or row_map.get('ESPECIALIDAD')
                                
                                params = (
                                    str(poliza),
                                    diag_proc,
                                    row_map.get('PROCEDIMIENTO') or diag_proc,
                                    row_map.get('ESTATUS') or row_map.get('ESTADO'),
                                    normalize_moneda(row_map.get('MONEDA')),
                                    clean_monto(row_map.get('MONTO SOLICITADO') or row_map.get('MONTO_SOLICITADO')),
                                    clean_monto(row_map.get('MONTO APROBADO') or row_map.get('MONTO_PAGADO')),
                                    row_map.get('FECHA DE NOTIFICACION') or row_map.get('FECHA_NOTI'),
                                    row_map.get('FECHA DE APROBACION') or row_map.get('FECHA_APRO') or row_map.get('FECHA_PAGO'),
                                    row_map.get('CORREO'),
                                    str(row_map.get('SINIESTRO') or row_map.get('CODIGO_SINIESTRO') or ''),
                                    row_map.get('TIPO DE ATENCION') or row_map.get('TIPO_ATENCION'),
                                    row_map.get('MES'),
                                    row_map.get('NEGOCIO'),
                                    row_map.get('TITULAR'),
                                    row_map.get('RECLAMANTE'),
                                    row_map.get('REFERIDOR'),
                                    row_map.get('ANALISTA'),
                                    row_map.get('OBSERVACION')
                                )
                                
                            elif target_table == "automovilsiniestro":
                                sql = """INSERT INTO automovilsiniestro 
                                        (Cod_poliza, Fecha_ocurrencia, Fecha_noti, Estado, Monto_orden, Correo, Descripcion, codigo_siniestro)
                                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)"""
                                
                                params = (
                                    str(poliza),
                                    row_map.get('FECHA DE OCURRENCIA') or row_map.get('FECHA_OCURRENCIA'),
                                    row_map.get('FECHA DE NOTIFICACION') or row_map.get('FECHA_NOTI'),
                                    row_map.get('ESTATUS') or row_map.get('ESTADO'),
                                    row_map.get('MONTO PAGADO') or row_map.get('MONTO_PAGADO') or row_map.get('MONTO_ORDEN'),
                                    row_map.get('CORREO'),
                                    row_map.get('OBSERVACIONES') or row_map.get('DESCRIPCION'),
                                    str(row_map.get('SINIESTRO') or row_map.get('CODIGO_SINIESTRO') or '')
                                )
                            
                            # Cleanup params: replace NaN/empty with None
                            final_params = []
                            for p in params:
                                if isinstance(p, float) and math.isnan(p): final_params.append(None)
                                elif p == "" or p == "nan": final_params.append(None)
                                else: final_params.append(p)

                            cursor.execute(sql, tuple(final_params))
                            inserted_count += 1
                            
                        except Exception as row_e:
                            print(f"Error en fila de {sheet_name}: {row_e}")
                            errors.append(f"Hoja {sheet_name}, Póliza {poliza}: {str(row_e)}")
                
                conexion_MySQLdb.commit()
                
        return {
            'success': True, 
            'count': inserted_count, 
            'errors': errors
        }
    except Exception as e:
        print(f"Error en guardar_siniestros_batch: {e}")
        return {'success': False, 'error': str(e)}

# ---- Catálogo de Productos ----

def sql_lista_catalogo_producto():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                querySQL = ("""
                    SELECT
                         cp.id,
                         cp.Cod_compania,
                         c.Nombre,
                         cp.base,
                         cp.Ramo,
                         cp.Subramo,
                         cp.Producto,
                         cp.activo
                    FROM catalogo_producto AS cp
                    INNER JOIN compania AS c ON c.Cod_compania = cp.Cod_compania
                    ORDER BY c.Nombre, cp.Ramo, cp.Subramo, cp.Producto
                    """)
                cursor.execute(querySQL,)
                catalogobd = cursor.fetchall()

        return catalogobd
    except Exception as e:
        print(
            f"Error en la función sql_lista_catalogo_producto: {e}")
        return None

def sql_catalogo_producto_detalle(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                querySQL = ("""
                    SELECT
                         cp.id,
                         cp.Cod_compania,
                         c.Nombre,
                         cp.base,
                         cp.Ramo,
                         cp.Subramo,
                         cp.Producto,
                         cp.activo
                    FROM catalogo_producto AS cp
                    INNER JOIN compania AS c ON c.Cod_compania = cp.Cod_compania
                    WHERE cp.id = %s
                    """)
                cursor.execute(querySQL, (id,))
                catalogobd = cursor.fetchone()

        return catalogobd
    except Exception as e:
        print(
            f"Error en la función sql_catalogo_producto_detalle: {e}")
        return None

def sql_lista_company_para_catalogo():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                querySQL = ("""
                    SELECT
                         Cod_compania,
                         Nombre
                    FROM compania
                    ORDER BY Nombre
                    """)
                cursor.execute(querySQL,)
                companybd = cursor.fetchall()

        return companybd
    except Exception as e:
        print(
            f"Error en la función sql_lista_company_para_catalogo: {e}")
        return None

def sql_insertar_catalogo_producto(data):
    try:
        if not data.get('Cod_compania') or not data.get('base') or not data.get('Ramo') or not data.get('Subramo') or not data.get('Producto'):
            return {'success': False, 'message': 'Todos los campos son obligatorios.'}

        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                sql = "INSERT INTO catalogo_producto (Cod_compania, base, Ramo, Subramo, Producto) VALUES (%s, %s, %s, %s, %s)"
                valores = (
                    data.get('Cod_compania'),
                    data.get('base'),
                    data.get('Ramo'),
                    data.get('Subramo'),
                    data.get('Producto'),
                )
                cursor.execute(sql, valores)
                conexion_MySQLdb.commit()
                return {'success': True, 'message': 'Producto registrado exitosamente.'}

    except Exception as e:
        return {'success': False, 'message': f'Error inesperado en sql_insertar_catalogo_producto: {str(e)}'}

def sql_actualizar_catalogo_producto(id, data):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                querySQL = """
                        UPDATE catalogo_producto
                        SET
                            Cod_compania = %s,
                            base = %s,
                            Ramo = %s,
                            Subramo = %s,
                            Producto = %s
                        WHERE id = %s
                    """
                values = (
                    data.get('Cod_compania'),
                    data.get('base'),
                    data.get('Ramo'),
                    data.get('Subramo'),
                    data.get('Producto'),
                    id,
                )
                cursor.execute(querySQL, values)
                conexion_MySQLdb.commit()
                return {'success': True, 'message': 'Producto actualizado exitosamente.'}
    except Exception as e:
        print(f"Ocurrió un error en sql_actualizar_catalogo_producto: {e}")
        return {'success': False, 'message': f'Error inesperado en sql_actualizar_catalogo_producto: {str(e)}'}

def sql_toggle_activo_catalogo_producto(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                querySQL = "UPDATE catalogo_producto SET activo = NOT activo WHERE id = %s"
                cursor.execute(querySQL, (id,))
                conexion_MySQLdb.commit()
                if cursor.rowcount:
                    return {'success': True, 'message': 'Estado del producto actualizado correctamente.'}
                else:
                    return {'success': False, 'message': 'No se encontró el producto indicado.'}
    except Exception as e:
        print(f"Error en sql_toggle_activo_catalogo_producto: {e}")
        return {'success': False, 'message': f'Error inesperado en sql_toggle_activo_catalogo_producto: {str(e)}'}

def sql_catalogo_producto_activo_por_base(base):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor() as cursor:
                querySQL = ("""
                    SELECT
                         c.Nombre,
                         cp.Ramo,
                         cp.Subramo,
                         cp.Producto
                    FROM catalogo_producto AS cp
                    INNER JOIN compania AS c ON c.Cod_compania = cp.Cod_compania
                    WHERE cp.activo = 1 AND cp.base = %s
                    ORDER BY c.Nombre, cp.Ramo, cp.Subramo, cp.Producto
                    """)
                cursor.execute(querySQL, (base,))
                catalogobd = cursor.fetchall()

        return catalogobd
    except Exception as e:
        print(
            f"Error en la función sql_catalogo_producto_activo_por_base: {e}")
        return None